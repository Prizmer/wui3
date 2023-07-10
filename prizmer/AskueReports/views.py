# coding -*- coding: utf-8 -*-

#from django.shortcuts import render

from django.shortcuts import render, HttpResponse
from django.db import connection
import io
from openpyxl import Workbook
#from openpyxl.compat import range
import datetime
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
import common_sql
import re
#from excel_response import ExcelResponse
from django.shortcuts import redirect
#from datetime import datetime, date, time
import decimal
# для работы с xml
from lxml import etree
from django.conf import settings
#from io import StringIO
import time as pt
import zipfile
from datetime import date, datetime, timedelta
import random 

separator = getattr(settings, 'SEPARATOR', ',') #'.' #separator = '.' or ','


def get_val_by_round(val, ROUND_SIZE, separator):
    new_val = f"{val:.{ROUND_SIZE}f}" #<--- для питона старше 3.6
    return new_val

def zagotovka(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    ws['B5'] = 'Заготовка'
    ws['B5'].style = "ali_grey"
    #response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'zagotovka' 
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

# Стили
ali_grey   = NamedStyle(name = "ali_grey", fill=PatternFill(fill_type='solid', start_color='DCDCDC'), border=Border(left=Side(border_style='thin',color='FF000000'), bottom=Side(border_style='thin',color='FF000000'), right=Side(border_style='thin',color='FF000000'), top=Side(border_style='thin',color='FF000000')), alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True))
ali_white  = NamedStyle(name = "ali_white", border=Border(left=Side(border_style='thin',color='FF000000'), bottom=Side(border_style='thin',color='FF000000'), right=Side(border_style='thin',color='FF000000'), top=Side(border_style='thin',color='FF000000')), alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True))
ali_blue   = NamedStyle(name = "ali_blue", fill=PatternFill(fill_type='solid', start_color='E6E6FA'), border=Border(left=Side(border_style='thin',color='FF000000'), bottom=Side(border_style='thin',color='FF000000'), right=Side(border_style='thin',color='FF000000'), top=Side(border_style='thin',color='FF000000')), alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True))
ali_pink   = NamedStyle(name = "ali_pink", fill=PatternFill(fill_type='solid', start_color='FFF0F5'), border=Border(left=Side(border_style='thin',color='FF000000'), bottom=Side(border_style='thin',color='FF000000'), right=Side(border_style='thin',color='FF000000'), top=Side(border_style='thin',color='FF000000')), alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True))

ali_yellow = NamedStyle(name = "ali_yellow", fill=PatternFill(fill_type='solid', start_color='EEEE00'), border=Border(left=Side(border_style='thin',color='FF000000'), bottom=Side(border_style='thin',color='FF000000'), right=Side(border_style='thin',color='FF000000'), top=Side(border_style='thin',color='FF000000')), alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True))
ali_white_size_18  = NamedStyle(name = "ali_white_size_18", font=Font(size=18))
# Конец описания стилей

def translate(name):
 
    #Заменяем пробелы и преобразуем строку к нижнему регистру
    name = name.replace(' ','-').lower()
 
    #
    transtable = (
        ## Большие буквы
        ("Щ", "Sch"),
        ("Щ", "SCH"),
        # two-symbol
        ("Ё", "Yo"),
        ("Ё", "YO"),
        ("Ж", "Zh"),
        ("Ж", "ZH"),
        ("Ц", "Ts"),
        ("Ц", "TS"),
        ("Ч", "Ch"),
        ("Ч", "CH"),
        ("Ш", "Sh"),
        ("Ш", "SH"),
        ("Ы", "Yi"),
        ("Ы", "YI"),
        ("Ю", "Yu"),
        ("Ю", "YU"),
        ("Я", "Ya"),
        ("Я", "YA"),
        # one-symbol
        ("А", "A"),
        ("Б", "B"),
        ("В", "V"),
        ("Г", "G"),
        ("Д", "D"),
        ("Е", "E"),
        ("З", "Z"),
        ("И", "I"),
        ("Й", "J"),
        ("К", "K"),
        ("Л", "L"),
        ("М", "M"),
        ("Н", "N"),
        ("О", "O"),
        ("П", "P"),
        ("Р", "R"),
        ("С", "S"),
        ("Т", "T"),
        ("У", "U"),
        ("Ф", "F"),
        ("Х", "H"),
        ("Э", "E"),
        ("Ъ", "`"),
        ("Ь", "'"),
        ## Маленькие буквы
        # three-symbols
        ("щ", "sch"),
        # two-symbols
        ("ё", "yo"),
        ("ж", "zh"),
        ("ц", "ts"),
        ("ч", "ch"),
        ("ш", "sh"),
        ("ы", "yi"),
        ("ю", "yu"),
        ("я", "ya"),
        # one-symbol
        ("а", "a"),
        ("б", "b"),
        ("в", "v"),
        ("г", "g"),
        ("д", "d"),
        ("е", "e"),
        ("з", "z"),
        ("и", "i"),
        ("й", "j"),
        ("к", "k"),
        ("л", "l"),
        ("м", "m"),
        ("н", "n"),
        ("о", "o"),
        ("п", "p"),
        ("р", "r"),
        ("с", "s"),
        ("т", "t"),
        ("у", "u"),
        ("ф", "f"),
        ("х", "h"),
        ("ъ", "`"),
        ("ь", "'"),
        ("э", "e"),
        ("№", "#"),
    )
    #перебираем символы в таблице и заменяем
    for symb_in, symb_out in transtable:
        name = name.replace(symb_in, symb_out)
    #возвращаем переменную
    return name

def get_k_t_n_by_serial_number(serial_number):
    """Получаем Ктн по серийному номеру счтчика"""
    simpleq = connection.cursor()
    simpleq.execute("""SELECT 
                          link_abonents_taken_params.coefficient_2
                        FROM 
                          public.meters, 
                          public.link_abonents_taken_params, 
                          public.taken_params
                        WHERE 
                          link_abonents_taken_params.guid_taken_params = taken_params.guid AND
                          taken_params.guid_meters = meters.guid AND
                          meters.factory_number_manual = %s
                        ORDER BY
                          link_abonents_taken_params.coefficient_2 DESC
                        LIMIT 1;""", [serial_number])
    simpleq = simpleq.fetchall()
    return simpleq[0][0]
    

    
def get_k_t_t_by_serial_number(serial_number):
    """Получаем Ктт по серийному номеру счтчика"""
    simpleq = connection.cursor()
    simpleq.execute("""SELECT 
                          link_abonents_taken_params.coefficient
                        FROM 
                          public.meters, 
                          public.link_abonents_taken_params, 
                          public.taken_params
                        WHERE 
                          link_abonents_taken_params.guid_taken_params = taken_params.guid AND
                          taken_params.guid_meters = meters.guid AND
                          meters.factory_number_manual = %s
                        ORDER BY
                          link_abonents_taken_params.coefficient DESC
                        LIMIT 1;""", [serial_number])
    simpleq = simpleq.fetchall()
    return simpleq[0][0]
    
def get_k_a_by_serial_number(serial_number):
    """Получаем Коэффициент А по серийному номеру счтчика"""
    simpleq = connection.cursor()
    simpleq.execute("""SELECT 
                          link_abonents_taken_params.coefficient_3
                        FROM 
                          public.meters, 
                          public.link_abonents_taken_params, 
                          public.taken_params
                        WHERE 
                          link_abonents_taken_params.guid_taken_params = taken_params.guid AND
                          taken_params.guid_meters = meters.guid AND
                          meters.factory_number_manual = %s
                        ORDER BY
                          link_abonents_taken_params.coefficient DESC
                        LIMIT 1;""", [serial_number])
    simpleq = simpleq.fetchall()
    return simpleq[0][0]


def report_3_tarifa_k(request): # Отчет по А+ и R+ с коэффициентами

    response = io.StringIO()    
    wb = Workbook()

    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)    
    ws = wb.active                   

# Шапка отчета   
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Срез показаний счетчика с коэффициентами за период' + ' ' + str(request.session["electric_data_start"]) + " - " + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = "ali_grey"
    ws['D4'].style = "ali_grey"
    ws['E4'].style = "ali_grey"
    
    ws['C5'] = 'Ктт'
    ws['C5'].style = "ali_grey"
    ws['D5'] = 'Ктн'
    ws['D5'].style = "ali_grey"
    ws['E5'] = 'А'
    ws['E5'].style = "ali_grey" 
    
    # Сумма A+
    ws.merge_cells('F4:I4')
    ws['F4'] = 'Суммарные показания/энергия A+'
    ws['F4'].style = "ali_grey"
    ws['G4'].style = "ali_grey"
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['F5'] = 'Показания на ' + str(request.session["electric_data_start"])
    ws['F5'].style = "ali_grey"
    ws['G5'] = 'Показания на ' + str(request.session["electric_data_end"])
    ws['G5'].style = "ali_grey"
    ws['H5'] = 'Разность показаний'
    ws['H5'].style = "ali_grey"
    ws['I5'] = 'Энергия кВт*ч'
    ws['I5'].style = "ali_grey"
    
    # Сумма R+
    ws.merge_cells('J4:M4')
    ws['J4'] = 'Суммарные показания/энергия R+'
    ws['J4'].style = "ali_grey"
    ws['K4'].style = "ali_grey"
    ws['L4'].style = "ali_grey"
    ws['M4'].style = "ali_grey"
    ws['J5'] = 'Показания на ' + str(request.session["electric_data_start"])
    ws['J5'].style = "ali_grey"
    ws['K5'] = 'Показания на ' + str(request.session["electric_data_end"])
    ws['K5'].style = "ali_grey"
    ws['L5'] = 'Разность показаний'
    ws['L5'].style = "ali_grey"
    ws['M5'] = 'Энергия кВт*ч'
    ws['M5'].style = "ali_grey"
    
   
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17
    
# Заполняем таблицу данными ----------------------------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------------------------------------------------------
    is_abonent_level    = re.compile(r'abonent')
    is_object_level     = re.compile(r'level')    
    obj_title           = request.session["obj_title"]
    electric_data_start = request.session["electric_data_start"]
    electric_data_end   = request.session["electric_data_end"]
    obj_key             = request.session["obj_key"]
    data_table = []
    
#--------------------------------------------------------------------------------------------------------------------------------------------------------------
    if bool(is_object_level.search(obj_key)): # Если это объект, то формируем список абонентов
        cursor_abonents_list = connection.cursor()
        cursor_abonents_list.execute("""
                                  SELECT 
                                   abonents.name
                                  FROM 
                                   public.objects, 
                                   public.abonents
                                  WHERE 
                                   objects.guid = abonents.guid_objects AND
                                   objects.name = %s
                                  ORDER BY
                                   abonents.name ASC;""",[obj_title])
        abonents_list = cursor_abonents_list.fetchall()

    elif bool(is_abonent_level.search(obj_key)):   # Если это отдельный абонент, то делаем выборку для одного абонента, а за имя родительсого объекта берем завод.
        abonents_list = [(obj_title,)]
        obj_title = "Завод"
     
    for x in range(len(abonents_list)):
        # delta for groups abonents 'start date' A+
        cursor_t0_aplus_delta_start_temp = connection.cursor()
        cursor_t0_aplus_delta_start_temp.execute("""
                    SELECT 
                      daily_values.date, 
                      daily_values.value, 
                      abonents.name, 
                      daily_values.id_taken_params, 
                      objects.name, 
                      names_params.name, 
                      meters.factory_number_manual, 
                      resources.name,
                      link_abonents_taken_params.coefficient,
                      link_abonents_taken_params.coefficient_2,
                      link_abonents_taken_params.coefficient_3
                    FROM 
                      public.daily_values, 
                      public.link_abonents_taken_params, 
                      public.taken_params, 
                      public.abonents, 
                      public.objects, 
                      public.names_params, 
                      public.params, 
                      public.meters, 
                      public.resources
                    WHERE 
                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                      taken_params.id = daily_values.id_taken_params AND
                      taken_params.guid_params = params.guid AND
                      taken_params.guid_meters = meters.guid AND
                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                      objects.guid = abonents.guid_objects AND
                      names_params.guid = params.guid_names_params AND
                      resources.guid = names_params.guid_resources AND
                      abonents.name = %s AND 
                      objects.name = %s AND 
                      names_params.name = 'T0 A+' AND 
                      daily_values.date = %s AND 
                      resources.name = 'Электричество'
                      ORDER BY
                      objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
        data_table_t0_aplus_delta_start_temp = cursor_t0_aplus_delta_start_temp.fetchall()
    
    # delta for groups abonents 'end date' A+
        cursor_t0_aplus_delta_end_temp = connection.cursor()
        cursor_t0_aplus_delta_end_temp.execute("""
                    SELECT 
                      daily_values.date, 
                      daily_values.value, 
                      abonents.name, 
                      daily_values.id_taken_params, 
                      objects.name, 
                      names_params.name, 
                      meters.factory_number_manual, 
                      resources.name,
                      link_abonents_taken_params.coefficient,
                      link_abonents_taken_params.coefficient_2,
                      link_abonents_taken_params.coefficient_3
                    FROM 
                      public.daily_values, 
                      public.link_abonents_taken_params, 
                      public.taken_params, 
                      public.abonents, 
                      public.objects, 
                      public.names_params, 
                      public.params, 
                      public.meters, 
                      public.resources
                    WHERE 
                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                      taken_params.id = daily_values.id_taken_params AND
                      taken_params.guid_params = params.guid AND
                      taken_params.guid_meters = meters.guid AND
                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                      objects.guid = abonents.guid_objects AND
                      names_params.guid = params.guid_names_params AND
                      resources.guid = names_params.guid_resources AND
                      abonents.name = %s AND 
                      objects.name = %s AND 
                      names_params.name = 'T0 A+' AND 
                      daily_values.date = %s AND 
                      resources.name = 'Электричество'
                      ORDER BY
                      objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
        data_table_t0_aplus_delta_end_temp = cursor_t0_aplus_delta_end_temp.fetchall()
        
    # delta for groups abonents 'start date' R+
        cursor_t0_rplus_delta_start_temp = connection.cursor()
        cursor_t0_rplus_delta_start_temp.execute("""
                    SELECT 
                      daily_values.date, 
                      daily_values.value, 
                      abonents.name, 
                      daily_values.id_taken_params, 
                      objects.name, 
                      names_params.name, 
                      meters.factory_number_manual, 
                      resources.name,
                      link_abonents_taken_params.coefficient,
                      link_abonents_taken_params.coefficient_2,
                      link_abonents_taken_params.coefficient_3
                    FROM 
                      public.daily_values, 
                      public.link_abonents_taken_params, 
                      public.taken_params, 
                      public.abonents, 
                      public.objects, 
                      public.names_params, 
                      public.params, 
                      public.meters, 
                      public.resources
                    WHERE 
                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                      taken_params.id = daily_values.id_taken_params AND
                      taken_params.guid_params = params.guid AND
                      taken_params.guid_meters = meters.guid AND
                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                      objects.guid = abonents.guid_objects AND
                      names_params.guid = params.guid_names_params AND
                      resources.guid = names_params.guid_resources AND
                      abonents.name = %s AND 
                      objects.name = %s AND 
                      names_params.name = 'T0 R+' AND 
                      daily_values.date = %s AND 
                      resources.name = 'Электричество'
                      ORDER BY
                      objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
        data_table_t0_rplus_delta_start_temp = cursor_t0_rplus_delta_start_temp.fetchall()
    
    # delta for groups abonents 'end date' R+
        cursor_t0_rplus_delta_end_temp = connection.cursor()
        cursor_t0_rplus_delta_end_temp.execute("""
                    SELECT 
                      daily_values.date, 
                      daily_values.value, 
                      abonents.name, 
                      daily_values.id_taken_params, 
                      objects.name, 
                      names_params.name, 
                      meters.factory_number_manual, 
                      resources.name,
                      link_abonents_taken_params.coefficient,
                      link_abonents_taken_params.coefficient_2,
                      link_abonents_taken_params.coefficient_3
                    FROM 
                      public.daily_values, 
                      public.link_abonents_taken_params, 
                      public.taken_params, 
                      public.abonents, 
                      public.objects, 
                      public.names_params, 
                      public.params, 
                      public.meters, 
                      public.resources
                    WHERE 
                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                      taken_params.id = daily_values.id_taken_params AND
                      taken_params.guid_params = params.guid AND
                      taken_params.guid_meters = meters.guid AND
                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                      objects.guid = abonents.guid_objects AND
                      names_params.guid = params.guid_names_params AND
                      resources.guid = names_params.guid_resources AND
                      abonents.name = %s AND 
                      objects.name = %s AND 
                      names_params.name = 'T0 R+' AND 
                      daily_values.date = %s AND 
                      resources.name = 'Электричество'
                      ORDER BY
                      objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
        data_table_t0_rplus_delta_end_temp = cursor_t0_rplus_delta_end_temp.fetchall()
    
       
        data_table_temp = []
        data_table_temp.append(abonents_list[x][0]) # наименование канала
        try:# заводской номер
            data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][6])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:# T0 A+ нач
            data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")        
        try:# T0 A+ кон
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:# расход T0 A+
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1] - data_table_t0_aplus_delta_start_temp[0][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try: #k_1
            data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][8])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try: #k_2
            data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][9])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:#k_3
            data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][10])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")

        try:# T0 R+ нач
            data_table_temp.append(data_table_t0_rplus_delta_start_temp[0][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")        
        try:# T0 R+ кон
            data_table_temp.append(data_table_t0_rplus_delta_end_temp[0][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:# расход T0 R+
            data_table_temp.append(data_table_t0_rplus_delta_end_temp[0][1] - data_table_t0_rplus_delta_start_temp[0][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
                      
        data_table.append(data_table_temp)
        
    for row in range(6, len(abonents_list)+6):
        ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0]) # наименование канала
        ws.cell('A%s'%(row)).style = "ali_grey"
        
        ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
        ws.cell('B%s'%(row)).style = "ali_white"
        
        ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][2])  # T0 A+ нач
        ws.cell('F%s'%(row)).style = "ali_white"
               
        ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][3])  # T0 A+ кон
        ws.cell('G%s'%(row)).style = "ali_white"
        
        ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][4]) # расход T0 A+
        ws.cell('H%s'%(row)).style = "ali_white"
        
        ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][5]) # коэффициент 1  Ктт
        ws.cell('C%s'%(row)).style = "ali_white"
        
        ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][6]) # коэффициент 2  Ктн
        ws.cell('D%s'%(row)).style = "ali_white"
                
        ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][7]) # коэффициент 3  Постоянная счётчика
        ws.cell('E%s'%(row)).style = "ali_white"
        
        
        ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][8])  # T0 R+ нач
        ws.cell('J%s'%(row)).style = "ali_white"
               
        ws.cell('K%s'%(row)).value = '%s' % (data_table[row-6][9])  # T0 R+ кон
        ws.cell('K%s'%(row)).style = "ali_white"
        
        ws.cell('L%s'%(row)).value = '%s' % (data_table[row-6][10]) # расход T0 R+
        ws.cell('L%s'%(row)).style = "ali_white"
        
        try:
            ws.cell('I%s'%(row)).value = '%s' % (float(data_table[row-6][4])*float(data_table[row-6][5])*float(data_table[row-6][6])) # T0 R+ энергия с учёток коэффициентов
            ws.cell('I%s'%(row)).style = "ali_yellow"
        except UnicodeEncodeError:
            ws.cell('I%s'%(row)).value = '%s' % '-'
            ws.cell('I%s'%(row)).style = "ali_yellow"
        except TypeError:
            ws.cell('I%s'%(row)).value = '%s' % '-'
            ws.cell('I%s'%(row)).style = "ali_yellow"
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % (float(data_table[row-6][10])*float(data_table[row-6][5])*float(data_table[row-6][6])) # T0 R+ энергия с учёток коэффициентов
            ws.cell('M%s'%(row)).style = "ali_yellow"
        except UnicodeEncodeError:
            ws.cell('M%s'%(row)).value = '%s' % '-'
            ws.cell('M%s'%(row)).style = "ali_yellow"
        except TypeError:
            ws.cell('M%s'%(row)).value = '%s' % '-'
            ws.cell('M%s'%(row)).style = "ali_yellow"
                    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
#    response['Content-Disposition'] = "attachment; filename=electric.xlsx"
    output_name = 'otchet za period ' + electric_data_start + '-' + electric_data_end+ translate(obj_title)
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   

    return response
    
    
def pokazania(request): # Показания по общему тарифу по А+ и R+
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

# Шапка отчета   
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Срез показаний с коэффициентами на дату' + ' ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = "ali_grey"
    ws['D4'].style = "ali_grey"
    ws['E4'].style = "ali_grey"
    
    ws['C5'] = 'Ктт'
    ws['C5'].style = "ali_grey"
    ws['D5'] = 'Ктн'
    ws['D5'].style = "ali_grey"
    ws['E5'] = 'А'
    ws['E5'].style = "ali_grey" 
    
    # Сумма
    ws.merge_cells('F4:I4')
    ws['F4'] = 'Суммарные показания/энергия'
    ws['F4'].style = "ali_grey"
    ws['G4'].style = "ali_grey"
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['F5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['G5'].style = "ali_yellow"
    
    ws['H5'] = 'Показания R+ на ' + str(request.session["electric_data_end"])
    ws['H5'].style = "ali_grey"
    
    ws['I5'] = 'Энергия R+ на ' + str(request.session["electric_data_end"])
    ws['I5'].style = "ali_yellow"
    
   
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
# Заполняем таблицу данными ----------------------------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------------------------------------------------------
    is_abonent_level    = re.compile(r'abonent')
    is_object_level     = re.compile(r'level')    
    obj_title           = request.session["obj_title"]
    electric_data_end   = request.session["electric_data_end"]
    obj_key             = request.session["obj_key"]
    data_table = []
    
#--------------------------------------------------------------------------------------------------------------------------------------------------------------
    if bool(is_object_level.search(obj_key)): # Если это объект, то формируем список абонентов
        cursor_abonents_list = connection.cursor()
        cursor_abonents_list.execute("""
                                  SELECT 
                                   abonents.name
                                  FROM 
                                   public.objects, 
                                   public.abonents
                                  WHERE 
                                   objects.guid = abonents.guid_objects AND
                                   objects.name = %s
                                  ORDER BY
                                   abonents.name ASC;""",[obj_title])
        abonents_list = cursor_abonents_list.fetchall()

    elif bool(is_abonent_level.search(obj_key)):   # Если это отдельный абонент, то делаем выборку для одного абонента, а за имя родительсого объекта берем завод.
        abonents_list = [(obj_title,)]
        obj_title = "Завод"

    for x in range(len(abonents_list)):    
    # delta A+ for groups abonents 'end date'
        cursor_t0_aplus_delta_end_temp = connection.cursor()
        cursor_t0_aplus_delta_end_temp.execute("""
                    SELECT 
                      daily_values.date, 
                      daily_values.value, 
                      abonents.name, 
                      daily_values.id_taken_params, 
                      objects.name, 
                      names_params.name, 
                      meters.factory_number_manual, 
                      resources.name,
                      link_abonents_taken_params.coefficient,
                      link_abonents_taken_params.coefficient_2,
                      link_abonents_taken_params.coefficient_3
                    FROM 
                      public.daily_values, 
                      public.link_abonents_taken_params, 
                      public.taken_params, 
                      public.abonents, 
                      public.objects, 
                      public.names_params, 
                      public.params, 
                      public.meters, 
                      public.resources
                    WHERE 
                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                      taken_params.id = daily_values.id_taken_params AND
                      taken_params.guid_params = params.guid AND
                      taken_params.guid_meters = meters.guid AND
                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                      objects.guid = abonents.guid_objects AND
                      names_params.guid = params.guid_names_params AND
                      resources.guid = names_params.guid_resources AND
                      abonents.name = %s AND 
                      objects.name = %s AND 
                      names_params.name = 'T0 A+' AND 
                      daily_values.date = %s AND 
                      resources.name = 'Электричество'
                      ORDER BY
                      objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
        data_table_t0_aplus_delta_end_temp = cursor_t0_aplus_delta_end_temp.fetchall()
        
            # delta R+ for groups abonents 'end date'
        cursor_t0_rplus_delta_end_temp = connection.cursor()
        cursor_t0_rplus_delta_end_temp.execute("""
                    SELECT 
                      daily_values.date, 
                      daily_values.value, 
                      abonents.name, 
                      daily_values.id_taken_params, 
                      objects.name, 
                      names_params.name, 
                      meters.factory_number_manual, 
                      resources.name,
                      link_abonents_taken_params.coefficient,
                      link_abonents_taken_params.coefficient_2,
                      link_abonents_taken_params.coefficient_3
                    FROM 
                      public.daily_values, 
                      public.link_abonents_taken_params, 
                      public.taken_params, 
                      public.abonents, 
                      public.objects, 
                      public.names_params, 
                      public.params, 
                      public.meters, 
                      public.resources
                    WHERE 
                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                      taken_params.id = daily_values.id_taken_params AND
                      taken_params.guid_params = params.guid AND
                      taken_params.guid_meters = meters.guid AND
                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                      objects.guid = abonents.guid_objects AND
                      names_params.guid = params.guid_names_params AND
                      resources.guid = names_params.guid_resources AND
                      abonents.name = %s AND 
                      objects.name = %s AND 
                      names_params.name = 'T0 R+' AND 
                      daily_values.date = %s AND 
                      resources.name = 'Электричество'
                      ORDER BY
                      objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
        data_table_t0_rplus_delta_end_temp = cursor_t0_rplus_delta_end_temp.fetchall()
           
        data_table_temp = []
        data_table_temp.append(abonents_list[x][0]) # наименование канала
        try:# заводской номер
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][6])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:# T0 A+ кон
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")        
        try: #k_1
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][8])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try: #k_2
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][9])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:#k_3
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][10])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:# T0 R+ кон
            data_table_temp.append(data_table_t0_rplus_delta_end_temp[0][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
                       
        data_table.append(data_table_temp)
        
    for row in range(6, len(abonents_list)+6):
        ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0]) # наименование канала
        ws.cell('A%s'%(row)).style = "ali_grey"
        
        ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
        ws.cell('B%s'%(row)).style = "ali_white"
        
        ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][2])  # T0 A+ кон
        ws.cell('F%s'%(row)).style = "ali_white"
        
        ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3]) # коэффициент 1  Ктт
        ws.cell('C%s'%(row)).style = "ali_white"
        
        ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4]) # коэффициент 2  Ктн
        ws.cell('D%s'%(row)).style = "ali_white"
                
        ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5]) # коэффициент 3  Постоянная счётчика
        ws.cell('E%s'%(row)).style = "ali_white"

        ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][6]) # T0 R+ кон
        ws.cell('H%s'%(row)).style = "ali_white"
                
        try:
            ws.cell('G%s'%(row)).value = '%s' % (float(data_table[row-6][2])*float(data_table[row-6][3])*float(data_table[row-6][4])) # T0 A+ энергия с учёток коэффициентов
            ws.cell('G%s'%(row)).style = "ali_yellow"
        except UnicodeEncodeError:
            ws.cell('G%s'%(row)).value = '%s' % '-'
            ws.cell('G%s'%(row)).style = "ali_yellow"
        except TypeError:
            ws.cell('G%s'%(row)).value = '%s' % '-'
            ws.cell('G%s'%(row)).style = "ali_yellow"
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (float(data_table[row-6][6])*float(data_table[row-6][3])*float(data_table[row-6][4])) # T0 R+ энергия с учёток коэффициентов
            ws.cell('I%s'%(row)).style = "ali_yellow"
        except UnicodeEncodeError:
            ws.cell('I%s'%(row)).value = '%s' % '-'
            ws.cell('I%s'%(row)).style = "ali_yellow"
        except TypeError:
            ws.cell('I%s'%(row)).value = '%s' % '-'
            ws.cell('I%s'%(row)).style = "ali_yellow"

    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel") 
#    response['Content-Disposition'] = 'attachment; filename=eclectric.xlsx'
    output_name = 'otchet po pokazaniyam za ' + electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    
    return response


def profil_30_min(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    meters_name         = request.session["obj_title"]
    electric_data_end   = request.session["electric_data_end"]
    
    ws.merge_cells('A2:F2')
    ws['A2'] = 'Профиль мощности по абоненту ' + str(request.session['obj_title']) + ' за ' + str(request.session["electric_data_end"])
    ws['A3'] = 'Ктн = ' + str(common_sql.get_k_t_n(meters_name))
    ws['B3'] = 'Ктт = ' + str(common_sql.get_k_t_t(meters_name))
    ws['B5'] = 'Дата'
    ws['B5'].style = "ali_grey"
    ws['C5'] = 'Время'
    ws['C5'].style = "ali_grey"
    ws['D5'] = 'Наименование'
    ws['D5'].style = "ali_grey"
    ws['E5'] = 'A+ кВт'
    ws['E5'].style = "ali_grey"
    ws['F5'] = 'R+ кВАр'
    ws['F5'].style = "ali_grey"
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['D'].width = 30

    
    a_plus = connection.cursor()
    a_plus.execute("""SELECT 
                          various_values.date, 
                          various_values."time", 
                          various_values.value, 
                          meters.name, 
                          meters.address, 
                          names_params.name
                        FROM 
                          public.various_values, 
                          public.meters, 
                          public.params, 
                          public.taken_params, 
                          public.names_params
                        WHERE 
                          params.guid_names_params = names_params.guid AND
                          taken_params.guid_params = params.guid AND
                          taken_params.guid_meters = meters.guid AND
                          taken_params.id = various_values.id_taken_params AND
                          various_values.date = %s AND 
                          meters.name = %s AND 
                          names_params.name = 'A+ Профиль';""",[electric_data_end, meters_name])
    a_plus = a_plus.fetchall()
   
        
    r_plus = connection.cursor()
    r_plus.execute("""SELECT 
                          various_values.date, 
                          various_values."time", 
                          various_values.value, 
                          meters.name, 
                          meters.address, 
                          names_params.name
                        FROM 
                          public.various_values, 
                          public.meters, 
                          public.params, 
                          public.taken_params, 
                          public.names_params
                        WHERE 
                          params.guid_names_params = names_params.guid AND
                          taken_params.guid_params = params.guid AND
                          taken_params.guid_meters = meters.guid AND
                          taken_params.id = various_values.id_taken_params AND
                          various_values.date = %s AND 
                          meters.name = %s AND 
                          names_params.name = 'R+ Профиль';""",[electric_data_end, meters_name])
    r_plus = r_plus.fetchall()
          
    data_table = []
    for x in range(len(a_plus)):
        data_table_temp = []
        try:
            data_table_temp.append(a_plus[x][0])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(a_plus[x][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(a_plus[x][3])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(a_plus[x][2]*2*common_sql.get_k_t_n(meters_name)*common_sql.get_k_t_t(meters_name))
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(r_plus[x][2]*2*common_sql.get_k_t_n(meters_name)*common_sql.get_k_t_t(meters_name))
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        data_table.append(data_table_temp)
            
    for row in range(6, len(data_table)+6):
        ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][0]) # дата
        ws.cell('B%s'%(row)).style = "ali_grey"
        
        ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][1])  # время
        ws.cell('C%s'%(row)).style = "ali_white"
        
        ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][2])  # канал
        ws.cell('D%s'%(row)).style = "ali_white"
        
        ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][3]) # значение A+
        ws.cell('E%s'%(row)).style = "ali_white"
        
        ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][4]) # значение R+
        ws.cell('F%s'%(row)).style = "ali_white"
           
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    
    output_name = 'profil 30 min za ' + electric_data_end # формируем имя для excel отчета 
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_hour_increment(request): # Выгрузка excel по часовым приращениям
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    meters_name         = request.session["obj_title"]
    electric_data_end   = request.session["electric_data_end"]    

# Шапка очета      
    ws.merge_cells('B2:F2')
    ws['B2'] = 'Почасовой учет электроэнергии за ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('B3:F3')
    ws['B3'] = 'Абонент: ' + str(request.session['obj_title'])
    
    ws['B5'] = 'Дата'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Время'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Абонент'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Серийный номер'
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 'A+ кВт*ч'
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = '+ А+ кВт*ч'
    ws['G5'].style = "ali_grey"
    
    ws['H5'] = 'R+ кВар*ч'
    ws['H5'].style = "ali_grey"
    
    ws['I5'] = '+ R+ кВар*ч'
    ws['I5'].style = "ali_grey"
    
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20  
#-------------
    
#Запрашиваем данные для отчета
    time_list = ['00:00', '00:30','01:00', '01:30', '02:00', '02:30', '03:00', '03:30', '04:00', '04:30', '05:00', '05:30', '06:00', '06:30', '07:00', '07:30', '08:00', '08:30', '09:00', '09:30', '10:00', '10:30', '11:00', '11:30', '12:00', '12:30', '13:00', '13:30', '14:00', '14:30', '15:00', '15:30', '16:00', '16:30', '17:00', '17:30', '18:00', '18:30', '19:00', '19:30', '20:00', '20:30', '21:00', '21:30', '22:00', '22:30', '23:00', '23:30']
    meters_name         = request.session["obj_title"]
    electric_data_end   = request.session["electric_data_end"]
    
    serial_number = common_sql.get_serial_number_by_meter_name(meters_name)
        
    data_table = []
    # Добавляем первую строку в таблицу данных. Делаем запрос показаний на начало суток.
    data_table.append([electric_data_end,'00:00', meters_name, serial_number, common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 A+' ),common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 R+' ), '0', '0'])
    
    if common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 A+' ) != 'Нет данных':  # Если есть показания на начало суток выполняем почасовое приращение  
        for x in range(24):
            data_table_temp = []
            data_table_temp.append(electric_data_end)
            data_table_temp.append(time_list[(2*x)])
            data_table_temp.append(meters_name)
            data_table_temp.append(serial_number)
            data_table_temp.append(data_table[len(data_table)-1][4] + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'A+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'A+ Профиль'))
            data_table_temp.append(common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 R+' ) + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'R+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'R+ Профиль'))
            data_table_temp.append(common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'A+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'A+ Профиль'))
            data_table_temp.append(common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'R+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'R+ Профиль'))    
            if x == 0: # Убираем первую строку. Так как показания на 00:00 берем отдельным запросом
                next
            else:
                data_table.append(data_table_temp)    
#-----------------------------

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][0]) # дата
        ws.cell('B%s'%(row)).style = "ali_grey"
        
        ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][1])  # время
        ws.cell('C%s'%(row)).style = "ali_white"
        
        ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][2])  # канал
        ws.cell('D%s'%(row)).style = "ali_white"
        
        ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][3]) # Заводской номер
        ws.cell('E%s'%(row)).style = "ali_white"
        
        ws.cell('F%s'%(row)).value = '%s' % (round(data_table[row-6][4],2)) # значение A+
        ws.cell('F%s'%(row)).style = "ali_white"
        
        ws.cell('G%s'%(row)).value = '%s' % (round(float(data_table[row-6][6]),2)) # значение + A+
        ws.cell('G%s'%(row)).style = "ali_white"
        
        ws.cell('H%s'%(row)).value = '%s' % (round(data_table[row-6][5],2)) # значение R+
        ws.cell('H%s'%(row)).style = "ali_white"

        ws.cell('I%s'%(row)).value = '%s' % (round(float(data_table[row-6][7]),2)) # значение + R+
        ws.cell('I%s'%(row)).style = "ali_white"
#---------------------------
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'chasovie prirasheniya'
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
    
def pokazania_period(request): # Показания по абоненту за период
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

    meters_name         = request.session["obj_title"]
    parent_name         = request.session['obj_parent_title']
    electric_data_end   = request.session["electric_data_end"]
    electric_data_start = request.session['electric_data_start']
    data_table = []
# Шапка отчета
    ws.merge_cells('B2:F2')
    ws['B2'] = 'Ежедневные показания за период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('B3:F3')
    ws['B3'] = 'Абонент: ' + str(request.session['obj_title'])
   
    ws['B5'] = 'Дата'
    ws['B5'].style = "ali_grey"
       
    ws['C5'] = 'Абонент'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Серийный номер'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'A+ кВт*ч'
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 'R+ кВар*ч'
    ws['F5'].style = "ali_grey"
       
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20          
# Конец шапки

#Запрашиваем данные для отчета
    end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
    start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
    dates = [x for x in common_sql.daterange(start_date,
                  end_date,
                  step=datetime.timedelta(days=1),
                  inclusive=True)]

    for x in range(len(dates)):
        data_table_temp = get_data_table_by_date_daily(meters_name, parent_name, datetime.datetime.strftime(dates[x], "%d.%m.%Y"))
        if data_table_temp:
            data_table.extend(data_table_temp)
        else:
            data_table.append([datetime.datetime.strftime(dates[x], "%d.%m.%Y"),meters_name,common_sql.get_serial_number_by_meter_name(meters_name), 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д'])
   
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][0]) # дата
        ws.cell('B%s'%(row)).style = "ali_grey"
        
        ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][1])  # абонент
        ws.cell('C%s'%(row)).style = "ali_white"
        
        ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
        ws.cell('D%s'%(row)).style = "ali_white"
        
        ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][3]) # значение A+
        ws.cell('E%s'%(row)).style = "ali_white"
        
        ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][8]) # значение R+
        ws.cell('F%s'%(row)).style = "ali_white"
        
#---------------------------
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'pokazania za period' 
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_rejim_day(request): # Отчет по режимному дню
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    meters_name         = request.session["obj_title"]
#    parent_name         = request.session['obj_parent_title']
    electric_data_end   = request.session["electric_data_end"]
#    electric_data_start = request.session['electric_data_start']
    data_table = []
    #print meters_name
    
    general_k = common_sql.get_k_t_t(meters_name) * common_sql.get_k_t_n(meters_name)#поменяла функцию: в запросе нужен factory namber manual, а не name в meters
# Шапка отчета
    ws['A1'] = "ЗАО 'Кировская керамика'"
    ws['H1'] = 'Шифр'
    ws['H2'] = 'Питающий центр'
    ws['H3'] = '№ фидера'
    
    ws['D5'] = 'Протокол (первичный)'
    ws['D5'].style = "ali_white_size_18"
    
    ws['E7'] = 'трансформаторного напряжения _____ вольт'
    
    ws['B6'] = 'записей показаний электросчетчиков и вольтметров, а также определения нагрузок'
    ws['B7'] = "и тангенса 'фи' за " + str(electric_data_end) + 'г'
    ws['B9'] = 'Акт. Счетчик № '# + str(common_sql.get_serial_number_by_meter_name(meters_name)) 
    ws['E9'] = 'Реакт. Счетчик № '# + str(common_sql.get_serial_number_by_meter_name(meters_name))
    ws['B10'] = 'Расч. Коэфициент' 
    ws['E10'] = 'Расч. Коэфициент'

    ws['A11'] = 'Время' 
    ws['A12'] = 'записи, часы'
    ws['A39'] = 'суточный расход' 
    ws['A40'] = 'активной и реактивной'
    ws['A41'] = 'энергии' 
    ws['A42'] = 'Контрольная сумма'
    ws['A44'] = 'Запись показаний счетчиков производили'
    ws['A44'] = 'Запись показаний счетчиков производили'
    ws['A45'] = 'фамилия ______________ подпись ______________'
    ws['A46'] = 'фамилия ______________ подпись ______________'
    ws['A47'] = 'фамилия ______________ подпись ______________'

    ws['G44'] = 'Расчеты произвел:'
    ws['G46'] = 'фамилия ______________ подпись ______________'

    ws['D10'] = ws['G10'] = general_k # Общий коэффициент
    ws['D9'] = ws['G9'] = str(common_sql.get_serial_number_by_meter_name(meters_name)) #Серийный номер прибора #поменяла функцию: в запросе нужен factory namber manual, а не name в meters
        
    ws['A13'] = '0' 
    ws['A14'] = '1' 
    ws['A15'] = '2'
    ws['A16'] = '3' 
    ws['A17'] = '4'
    ws['A18'] = '5' 
    ws['A19'] = '6'
    ws['A20'] = '7' 
    ws['A21'] = '8'
    ws['A22'] = '9' 
    ws['A23'] = '10'
    ws['A24'] = '11' 
    ws['A25'] = '12'
    ws['A26'] = '13' 
    ws['A27'] = '14'
    ws['A28'] = '15' 
    ws['A29'] = '16'
    ws['A30'] = '17' 
    ws['A31'] = '18'
    ws['A32'] = '19' 
    ws['A33'] = '20'
    ws['A34'] = '21' 
    ws['A35'] = '22'
    ws['A36'] = '23' 
    ws['A37'] = '24'

    ws['B11'] = 'Показания' 
    ws['B12'] = 'счетчика'
    
    ws['C11'] = 'Разность' 
    ws['C12'] = 'показаний'
    
    ws['D11'] = 'расход за' 
    ws['D12'] = 'час(квт)'
    
    ws['E11'] = 'Показания' 
    ws['E12'] = 'счетчика'
    
    ws['F11'] = 'Разность' 
    ws['F12'] = 'показаний'
    
    ws['G11'] = 'расход за' 
    ws['G12'] = 'час(квт)'

    ws['H10'] = 'тангенс'
    ws['H11'] = 'фи'
    
    ws['I9'] = 'Показания'
    ws['I10'] = 'вольтметров на'
    ws['I11'] = 'стороне'
    ws['I12'] = 'в/н'
    ws['J12'] = 'н/н'
    
    ws['K9'] = 'мощность '
    ws['K10'] = 'включенных'
    ws['K11'] = 'компен.устр'
    ws['K12'] = 'кВар'
    
    for col_idx in range(1, 12):
        col = get_column_letter(col_idx)
        for row in range(13, 38):
            ws.cell('%s%s'%(col, row)).style = "ali_white"

    ws.column_dimensions['A'].width = 12            
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['I'].width = 7
    ws.column_dimensions['J'].width = 7    
    ws.row_dimensions[5].height = 30
    # Конец шапки

#Запрашиваем данные для отчета
    time_list = ['00:00', '00:30','01:00', '01:30', '02:00', '02:30', '03:00', '03:30', '04:00', '04:30', '05:00', '05:30', '06:00', '06:30', '07:00', '07:30', '08:00', '08:30', '09:00', '09:30', '10:00', '10:30', '11:00', '11:30', '12:00', '12:30', '13:00', '13:30', '14:00', '14:30', '15:00', '15:30', '16:00', '16:30', '17:00', '17:30', '18:00', '18:30', '19:00', '19:30', '20:00', '20:30', '21:00', '21:30', '22:00', '22:30', '23:00', '23:30']
    
    serial_number = common_sql.get_serial_number_by_meter_name(meters_name)
    # Добавляем первую строку в таблицу данных. Делаем запрос показаний на начало суток.
    data_table.append([electric_data_end,'00:00', meters_name, serial_number, common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 A+' ),common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 R+' ), '0', '0'])
    
    if common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 A+' ) != 'Нет данных':  # Если есть показания на начало суток выполняем почасовое приращение  
        for x in range(24):
            data_table_temp = []
            data_table_temp.append(electric_data_end) # Дата
            data_table_temp.append(time_list[(2*x)])  # Отчетный час
            data_table_temp.append(meters_name)       # Имя абонента
            data_table_temp.append(serial_number)     # Серийный номер
            data_table_temp.append(data_table[len(data_table)-1][4] + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'A+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'A+ Профиль'))     # Показиние счётчика за предыдущий час + две получасовки А+           
            data_table_temp.append(data_table[len(data_table)-1][5] + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'R+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'R+ Профиль'))     # Показиние счётчика за предыдущий час + две получасовки R+

            data_table_temp.append(common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'A+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'A+ Профиль'))                                        # Сумма двух получасовок: потребленная энергия за час А+
            data_table_temp.append(common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'R+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'R+ Профиль'))                                        # Сумма двух получасовок: потребленная энергия за час R+  
            if x == 0: # Убираем первую строку. Так как показания на 00:00 берем отдельным запросом 
                next
            else:
                data_table.append(data_table_temp)
    if data_table[23][4] and data_table[23][5]:
        data_table.append([(datetime.datetime.strptime(electric_data_end, '%d.%m.%Y') + datetime.timedelta(days=1)).strftime('%d.%m.%Y'),'00:00', meters_name, serial_number, common_sql.get_daily_value_by_meter_name(meters_name, (datetime.datetime.strptime(electric_data_end, '%d.%m.%Y') + datetime.timedelta(days=1)).strftime('%d.%m.%Y'), 'T0 A+' ),common_sql.get_daily_value_by_meter_name(meters_name, (datetime.datetime.strptime(electric_data_end, '%d.%m.%Y') + datetime.timedelta(days=1)).strftime('%d.%m.%Y'), 'T0 R+' ),common_sql.get_daily_value_by_meter_name(meters_name, (datetime.datetime.strptime(electric_data_end, '%d.%m.%Y') + datetime.timedelta(days=1)).strftime('%d.%m.%Y'), 'T0 A+' ) - data_table[23][4],common_sql.get_daily_value_by_meter_name(meters_name, (datetime.datetime.strptime(electric_data_end, '%d.%m.%Y') + datetime.timedelta(days=1)).strftime('%d.%m.%Y'), 'T0 R+' ) - data_table[23][5]])

    #------------

# Заполняем отчет значениями
    for row in range(13, len(data_table)+13):
        
        ws.cell('B%s'%(row)).value = '%s' % (round(data_table[row-13][4],4)) # значение A+
        ws.cell('B%s'%(row)).style = "ali_white"
        
        ws.cell('C%s'%(row)).value = '%s' % (round(float(data_table[row-13][6]),4)) # значение + A+
        ws.cell('C%s'%(row)).style = "ali_white"
        
        ws.cell('D%s'%(row)).value = '%s' % (round(float(data_table[row-13][6]),4)*general_k) # значение + A+
        ws.cell('D%s'%(row)).style = "ali_white"
        
        ws.cell('E%s'%(row)).value = '%s' % (round(data_table[row-13][5],4)) # значение R+
        ws.cell('E%s'%(row)).style = "ali_white"

        ws.cell('F%s'%(row)).value = '%s' % (round(float(data_table[row-13][7]),4)) # значение + R+
        ws.cell('F%s'%(row)).style = "ali_white"
        
        ws.cell('G%s'%(row)).value = '%s' % (round(float(data_table[row-13][7]),4)*general_k) # значение + R+
        ws.cell('G%s'%(row)).style = "ali_white"
        
        fi=0
        try:
            fi = round((float(data_table[row-13][7])*general_k)/(float(data_table[row-13][6])*general_k),2)
        except:
            fi = ''
        
        ws.cell('H%s'%(row)).value = '%s' % fi # значение + R+/+ A+
        ws.cell('H%s'%(row)).style = "ali_white"  
    #---------------------------      
    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'rejimniy den '+ str(electric_data_end) 
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    

def report_economic_electric(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    #--------------------------------------------------------------------------------------------------------------------      
# Шапка отчета
    ws.merge_cells('B2:F2')
    ws['B2'] = 'Таблица расчета удельного коэффициента период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('B3:F3')
    ws['B3'] = 'Абонент: Литейный цех'
   
    ws['B5'] = 'Дата'
    ws['B5'].style = "ali_grey"
       
    ws['C5'] = 'Изготовленная продукция, кг'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Затраченная A+, кВт*ч'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Удельный расход A+, кВт*ч/кг'
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 'Затраченная R+, кВар*ч'
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'Удельный расход R+, кВт*ч/кг'
    ws['G5'].style = "ali_grey"

    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18    
    ws.row_dimensions[5].height = 30    
# Конец шапки
    
#Запрашиваем данные для отчета---
    data_table = []
    
    electric_data_end   = request.session["electric_data_end"]
    electric_data_start   = request.session["electric_data_start"]

    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']    

    end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
    start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
    dates = [x for x in common_sql.daterange(start_date,
                  end_date,
                  step=datetime.timedelta(days=1),
                  inclusive=True)]
                  
    for x in range(len(dates)):
        try:
            data_table_temp = []
            delta_a_plus = 1
            delta_r_plus = 1

            try:
                delta_a_plus = common_sql.delta_sum_a_plus(dates[x+1])-common_sql.delta_sum_a_plus(dates[x])
                if delta_a_plus > 0:
                    delta_a_plus = delta_a_plus
                else:
                    delta_a_plus = 'Н/Д'
                delta_r_plus = common_sql.delta_sum_r_plus(dates[x+1])-common_sql.delta_sum_r_plus(dates[x])
                if delta_r_plus > 0:
                    delta_r_plus = delta_r_plus
                else:
                    delta_r_plus = 'Н/Д'

            except:
                delta_a_plus = 'Н/Д'
                delta_r_plus = 'Н/Д'

            data_table_temp.append(dates[x])
            data_table_temp.append(common_sql.product_sum(dates[x]))
            data_table_temp.append(delta_a_plus)
            data_table_temp.append(delta_a_plus/(common_sql.product_sum(dates[x])))
            data_table_temp.append(delta_r_plus)
            data_table_temp.append(delta_r_plus/(common_sql.product_sum(dates[x])))
        except:
            next
        data_table.append(data_table_temp)

#Конец запроса данных------------
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        ws.cell('B%s'%(row)).value = '%s' % ((data_table[row-6][0]).strftime("%d-%m-%Y")) # дата
        ws.cell('B%s'%(row)).style = "ali_grey"
        
        ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][1])  # изготовленная продукция, кг
        ws.cell('C%s'%(row)).style = "ali_white"
        
        ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][2])  # затраченная А+
        ws.cell('D%s'%(row)).style = "ali_white"
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][3])  # удельных расход А+/кг
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][4])  # затраченная R+
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][5])  # удельный расход R+/кг
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
        
        
        
#---------------------------    
    #--------------------------------------------------------------------------------------------------------------------    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    
    output_name = 'udelniy_coefficient_liteiniy_ceh' 
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response


def report_pokazaniya_water_identificators(request): # Показания по воде за дату с идентификаторами
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

    meters_name         = request.session["obj_title"]
    parent_name         = request.session['obj_parent_title']
    electric_data_end   = request.session["electric_data_end"]
#    electric_data_start = request.session['electric_data_start']
    data_table = []

# Шапка отчета
    ws.merge_cells('B2:F2')
    ws['B2'] = 'Показания по воде с идентификаторами за ' + str(request.session["electric_data_end"])
       
    ws['B5'] = 'Абонент'
    ws['B5'].style = "ali_grey"
       
    ws['C5'] = 'Идентификатор'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания'
    ws['D5'].style = "ali_grey"
    
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20        
# Конец шапки

#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_channel(meters_name, electric_data_end)
    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []        
        for x in range(len(list_of_abonents_2)):
            data_table_temp = common_sql.get_daily_water_channel(list_of_abonents_2[x], electric_data_end)
            data_table.extend(data_table_temp)
    elif(bool(is_object_level_1.search(obj_key))):
        
        list_of_objects_2 = common_sql.list_of_objects(common_sql.return_parent_guid_by_abonent_name(meters_name)) #Список квартир для объекта с пульсарами
        data_table = []
        for x in range(len(list_of_objects_2)):
            data_table_temp = [(list_of_objects_2[x][0],)]
            list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(meters_name), list_of_objects_2[x][0])
            for y in range(len(list_of_abonents_2)):
                data_table_temp2 = common_sql.get_daily_water_channel(list_of_abonents_2[y], electric_data_end)

                data_table_temp.extend(data_table_temp2)                                
                      
            data_table.extend(data_table_temp)
              
    else:
        data_table = [1,1,1,1]

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][0]) # абонент
        ws.cell('B%s'%(row)).style = "ali_grey"
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][4])  # идентификатор
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][2])  # показания м3
            ws.cell('D%s'%(row)).style = "ali_white"

        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
        
#---------------------------

    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'pokazania po vode' 
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    

def report_electric_simple_2_zones_v2(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Срез показаний с коэффициентами на дату' + ' ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = "ali_grey"
    ws['D4'].style = "ali_grey"
    ws['E4'].style = "ali_grey"
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = "ali_grey"
    ws['D5'] = 'Ктт'
    ws['D5'].style = "ali_grey"
    ws['E5'] = 'А'
    ws['E5'].style = "ali_grey" 
    
    # Сумма
    ws.merge_cells('F4:G4')
    ws['F4'] = 'Сумма'
    ws['F4'].style = "ali_grey"
    ws['G4'].style = "ali_grey"
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['F5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['G5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('H4:I4')
    ws['H4'] = 'Тариф 1'
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['H5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['H5'].style = "ali_grey"
    
    ws['I5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['I5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('J4:K4')
    ws['J4'] = 'Тариф 2'
    ws['J4'].style = "ali_grey"
    ws['K4'].style = "ali_grey"
    ws['J4'].style = "ali_grey"
    ws['K4'].style = "ali_grey"
    ws['J5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['J5'].style = "ali_grey"
    
    ws['K5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['K5'].style = "ali_yellow"
    
       
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    
    obj_parent_title    = request.session['obj_parent_title']
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    is_electric_monthly = request.session['is_electric_monthly']
    is_electric_daily   = request.session['is_electric_daily']
    is_electric_period  = request.session['is_electric_period']
    data_table = []
    if True:
        if True:            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'monthly')
                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'daily')

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # показания на начало месяца для объекта
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д']]        

#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for abonents group
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д']]
            elif (is_electric_daily == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д']]
              
            elif (is_electric_monthly == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д']]
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][9])  # Ктн
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][8])  # Ктт
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][10])  # Ка
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
                   
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][3])  # Сумма А+
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % round((data_table[row-6][3]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Сумма А+
            ws.cell('G%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('G%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][4])  # Тариф 1 А+
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % round((data_table[row-6][4]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Тариф 1 А+
            ws.cell('I%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('I%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][5])  # Тариф 2 А+
            ws.cell('J%s'%(row)).style = "ali_white"
        except:
            ws.cell('J%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % round((data_table[row-6][5]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Тариф 2 А+
            ws.cell('K%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('K%s'%(row)).style = "ali_yellow"
            next           
# Сохраняем в ecxel    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = '2_tariffa' 
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response
    

def report_electric_simple_3_zones_v2(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    wb.guess_types = True
    
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']    
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+' .Срез показаний с коэффициентами на дату' + ' ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование абонента'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = "ali_grey"
    ws['D4'].style = "ali_grey"
    ws['E4'].style = "ali_grey"
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = "ali_grey"
    ws['D5'] = 'Ктт'
    ws['D5'].style = "ali_grey"
    ws['E5'] = 'А'
    ws['E5'].style = "ali_grey" 
    
    # Сумма
    ws.merge_cells('F4:G4')
    ws['F4'] = 'Сумма'
    ws['F4'].style = "ali_grey"
    ws['G4'].style = "ali_grey"
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['F5'] = 'Показания A+ на ' + electric_data_end
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'Энергия A+ на ' + electric_data_end
    ws['G5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('H4:I4')
    ws['H4'] = 'Тариф 1'
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['H5'] = 'Показания A+ на ' +  electric_data_end
    ws['H5'].style = "ali_grey"
    
    ws['I5'] = 'Энергия A+ на ' +  electric_data_end
    ws['I5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('J4:K4')
    ws['J4'] = 'Тариф 2'
    ws['J4'].style = "ali_grey"
    ws['K4'].style = "ali_grey"
    ws['J4'].style = "ali_grey"
    ws['K4'].style = "ali_grey"
    ws['J5'] = 'Показания A+ на ' +  electric_data_end
    ws['J5'].style = "ali_grey"
    
    ws['K5'] = 'Энергия A+ на ' +  electric_data_end
    ws['K5'].style = "ali_yellow"
    
    # Тариф 3
    ws.merge_cells('L4:M4')
    ws['L4'] = 'Тариф 3'
    ws['L4'].style = "ali_grey"
    ws['M4'].style = "ali_grey"
    ws['L4'].style = "ali_grey"
    ws['M4'].style = "ali_grey"
    ws['L5'] = 'Показания A+ на ' +  electric_data_end
    ws['L5'].style = "ali_grey"
    
    ws['M5'] = 'Энергия A+ на ' +  electric_data_end
    ws['M5'].style = "ali_yellow"
         
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
    #выборка данных из БД
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    
    obj_parent_title    = request.session['obj_parent_title']
            
    obj_key             = request.session['obj_key']
    is_electric_monthly = request.session['is_electric_monthly']
    is_electric_daily   = request.session['is_electric_daily']
    is_electric_period  = request.session['is_electric_period']
    data_table = []
    if True:
        if True:            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'monthly')
                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'daily')

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # показания на начало месяца для объекта
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д']]        

#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for abonents group
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д']]
            elif (is_electric_daily == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д']]
              
            elif (is_electric_monthly == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д']]

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:            
            ws.cell('C%s'%(row)).value = '%s' % str(data_table[row-6][9]).replace('.', separator)# Ктн
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % str(data_table[row-6][8]).replace('.', separator)  # Ктт
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % str(data_table[row-6][10]).replace('.', separator)  # Ка
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
                   
        try:
            #ws.cell('F%s'%(row)).number_format = 'Comma'
            #ws.cell('F%s'%(row)).value = '%s' % str(data_table[row-6][3]).replace('.',',')  # Сумма А+
            ws.cell('F%s'%(row)).value = '%s' % str(round((data_table[row-6][3]), 3)).replace('.', separator)  # Сумма А+
            ws.cell('F%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % str(round((data_table[row-6][3]*data_table[row-6][8]*data_table[row-6][9]),3)).replace('.', separator)  # "Энергия Сумма А+
            ws.cell('G%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('G%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % str(round((data_table[row-6][4]),3)).replace('.', separator)  # Тариф 1 А+
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            val = round((data_table[row-6][4]*data_table[row-6][8]*data_table[row-6][9]),3)            
            ws.cell('I%s'%(row)).value = '%s' % str(val).replace('.', separator)  # "Энергия Тариф 1 А+
            ws.cell('I%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('I%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % str(round((data_table[row-6][5]),3)).replace('.', separator)  # Тариф 2 А+
            ws.cell('J%s'%(row)).style = "ali_white"
        except:
            ws.cell('J%s'%(row)).style = "ali_white"
            next
            
        try:
            val = round((data_table[row-6][5]*data_table[row-6][8]*data_table[row-6][9]),3)
            ws.cell('K%s'%(row)).value = '%s' % str(val).replace('.', separator) # "Энергия Тариф 2 А+
            ws.cell('K%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('K%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % str(round((data_table[row-6][6]),3)).replace('.', separator)  # Тариф 3 А+
            ws.cell('L%s'%(row)).style = "ali_white"

        except:
            ws.cell('L%s'%(row)).style = "ali_white"
            next
            
        try:
            val = round((data_table[row-6][6]*data_table[row-6][8]*data_table[row-6][9]),3)
            ws.cell('M%s'%(row)).value = '%s' % str(val).replace('.', separator)  # "Энергия Тариф 3 А+
            ws.cell('M%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('M%s'%(row)).style = "ali_yellow"
            next

# Сохраняем в ecxel  
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = '3_tariffa_'+translate(obj_title)+'_'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def report_current_3_zones_v2(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']    
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+' .Срез показаний с коэффициентами на дату' + ' ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = "ali_grey"
    ws['D4'].style = "ali_grey"
    ws['E4'].style = "ali_grey"
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = "ali_grey"
    ws['D5'] = 'Ктт'
    ws['D5'].style = "ali_grey"
    ws['E5'] = 'А'
    ws['E5'].style = "ali_grey" 
    
    # Сумма
    ws.merge_cells('F4:G4')
    ws['F4'] = 'Сумма'
    ws['F4'].style = "ali_grey"
    ws['G4'].style = "ali_grey"
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['F5'] = 'Показания A+ на ' + electric_data_end
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'Энергия A+ на ' + electric_data_end
    ws['G5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('H4:I4')
    ws['H4'] = 'Тариф 1'
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['H5'] = 'Показания A+ на ' +  electric_data_end
    ws['H5'].style = "ali_grey"
    
    ws['I5'] = 'Энергия A+ на ' +  electric_data_end
    ws['I5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('J4:K4')
    ws['J4'] = 'Тариф 2'
    ws['J4'].style = "ali_grey"
    ws['K4'].style = "ali_grey"
    ws['J4'].style = "ali_grey"
    ws['K4'].style = "ali_grey"
    ws['J5'] = 'Показания A+ на ' +  electric_data_end
    ws['J5'].style = "ali_grey"
    
    ws['K5'] = 'Энергия A+ на ' +  electric_data_end
    ws['K5'].style = "ali_yellow"
    
    # Тариф 3
    ws.merge_cells('L4:M4')
    ws['L4'] = 'Тариф 3'
    ws['L4'].style = "ali_grey"
    ws['M4'].style = "ali_grey"
    ws['L4'].style = "ali_grey"
    ws['M4'].style = "ali_grey"
    ws['L5'] = 'Показания A+ на ' +  electric_data_end
    ws['L5'].style = "ali_grey"
    
    ws['M5'] = 'Энергия A+ на ' +  electric_data_end
    ws['M5'].style = "ali_yellow"
         
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
    #выборка данных из БД
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    
    obj_parent_title    = request.session['obj_parent_title']
            
    obj_key             = request.session['obj_key']
    is_electric_monthly = request.session['is_electric_monthly']
    is_electric_daily   = request.session['is_electric_daily']
    is_electric_period  = request.session['is_electric_period']
    data_table = []
    if True:
        if True:            
            if (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'current')         
                            
            elif (bool(is_object_level.search(obj_key))): # показания на начало месяца для объекта
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'current')

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][9])  # Ктн
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][8])  # Ктт
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][10])  # Ка
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
                   
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][3])  # Сумма А+
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % round((data_table[row-6][3]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Сумма А+
            ws.cell('G%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('G%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][4])  # Тариф 1 А+
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % round((data_table[row-6][4]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Тариф 1 А+
            ws.cell('I%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('I%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][5])  # Тариф 2 А+
            ws.cell('J%s'%(row)).style = "ali_white"
        except:
            ws.cell('J%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % round((data_table[row-6][5]*data_table[row-6][8]*data_table[row-6][9]),3) # "Энергия Тариф 2 А+
            ws.cell('K%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('K%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table[row-6][6])  # Тариф 3 А+
            ws.cell('L%s'%(row)).style = "ali_white"

        except:
            ws.cell('L%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % round((data_table[row-6][6]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Тариф 3 А+
            ws.cell('M%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('M%s'%(row)).style = "ali_yellow"
            next

# Сохраняем в ecxel  
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = '3_tariffa_'+translate(obj_title)+'_'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def electric_between_report(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Значения профиля показаний за период с' + ' '+str(request.session["electric_data_start"]) +' по '+ str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:C5')
    ws['C4'] = 'Дата'
    ws['C4'].style = "ali_grey"
    ws['C5'].style = "ali_grey"
    
    # Сумма
    ws.merge_cells('D4:D5')
    ws['D4'] = 'Сумма - Показания A+ '
    ws['D4'].style = "ali_grey"
    ws['D5'].style = "ali_grey"
 
    # Дельта
    ws.merge_cells('E4:E5')
    ws['E4'] = 'Сумма - Расход за прошедшие сутки'
    ws['E4'].style = "ali_grey"
    ws['E5'].style = "ali_grey"
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_electric_daily    = request.session['is_electric_daily']
    obj_parent_title    = request.session['obj_parent_title']
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
    obj_key             = request.session['obj_key']

    data_table = []
    if True:
        if True:            
            if (is_electric_daily == '1') & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                params=['T0 A+','T1 A+','T2 A+','T3 A+', 'Электричество']
                data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
            else:
                pass
            
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][3])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][4])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # дата
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][5])  # сумма-показания
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][12])  # Расход за прошедшие сутки
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next

# Сохраняем в ecxel    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'potreblenie_electric' + str(electric_data_start) + ' - ' + str(electric_data_end)
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response
    
def electric_between_2_zones_report(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Значения профиля показаний за период с' + ' '+str(request.session["electric_data_start"]) +' по '+ str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:C5')
    ws['C4'] = 'Дата'
    ws['C4'].style = "ali_grey"
    ws['C5'].style = "ali_grey"
    
    # Сумма
    ws.merge_cells('D4:D5')
    ws['D4'] = 'Сумма - Показания T0 A+ '
    ws['D4'].style = "ali_grey"
    ws['D5'].style = "ali_grey"
 
    # Дельта
    ws.merge_cells('E4:E5')
    ws['E4'] = 'Сумма - Расход за прошедшие сутки T0'
    ws['E4'].style = "ali_grey"
    ws['E5'].style = "ali_grey"
    
        # Сумма
    ws.merge_cells('F4:F5')
    ws['F4'] = 'Сумма - Показания T1 A+ '
    ws['F4'].style = "ali_grey"
    ws['F5'].style = "ali_grey"
 
    # Дельта
    ws.merge_cells('G4:G5')
    ws['G4'] = 'Сумма - Расход за прошедшие сутки T1'
    ws['G4'].style = "ali_grey"
    ws['G5'].style = "ali_grey"
    
        # Сумма
    ws.merge_cells('H4:H5')
    ws['H4'] = 'Сумма - Показания T2 A+ '
    ws['H4'].style = "ali_grey"
    ws['H5'].style = "ali_grey"
 
    # Дельта
    ws.merge_cells('I4:I5')
    ws['I4'] = 'Сумма - Расход за прошедшие сутки T2'
    ws['I4'].style = "ali_grey"
    ws['I5'].style = "ali_grey"
    
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_electric_daily    = request.session['is_electric_daily']
    obj_parent_title    = request.session['obj_parent_title']
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
    obj_key             = request.session['obj_key']

    data_table = []
    if True:
        if True:            
            if (is_electric_daily == '1') & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                params=['T0 A+','T1 A+','T2 A+','T3 A+', 'Электричество']
                data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
                #data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end)
            else:
                pass
            
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][3])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][4])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # дата
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][5])  # сумма-показания t0
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][12])  # Расход за прошедшие сутки t0
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # сумма-показанияt1
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][13])  # Расход за прошедшие суткиt1
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # сумма-показанияt2
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][14])  # Расход за прошедшие суткиt2
            ws.cell('I%s'%(row)).style = "ali_white"
        except:
            ws.cell('I%s'%(row)).style = "ali_white"
            next

# Сохраняем в ecxel    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'potreblenie_electric_2_zones_' + str(electric_data_start) + ' - ' + str(electric_data_end)
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response
    
def electric_between_3_zones_report(request):
    
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+'. Значения профиля показаний за период с' + ' '+electric_data_start +' по '+ electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:C5')
    ws['C4'] = 'Дата'
    ws['C4'].style = "ali_grey"
    ws['C5'].style = "ali_grey"
    
    # Сумма
    ws.merge_cells('D4:D5')
    ws['D4'] = 'Показания T0 A+ '
    ws['D4'].style = "ali_grey"
    ws['D5'].style = "ali_grey"
 
    # Дельта
    ws.merge_cells('E4:E5')
    ws['E4'] = 'Расход за прошедшие сутки T0'
    ws['E4'].style = "ali_grey"
    ws['E5'].style = "ali_grey"
    
        # Сумма
    ws.merge_cells('F4:F5')
    ws['F4'] = 'Показания T1 A+ '
    ws['F4'].style = "ali_grey"
    ws['F5'].style = "ali_grey"
 
    # Дельта
    ws.merge_cells('G4:G5')
    ws['G4'] = 'Расход за прошедшие сутки T1'
    ws['G4'].style = "ali_grey"
    ws['G5'].style = "ali_grey"
    
        # Сумма
    ws.merge_cells('H4:H5')
    ws['H4'] = 'Показания T2 A+ '
    ws['H4'].style = "ali_grey"
    ws['H5'].style = "ali_grey"
 
    # Дельта
    ws.merge_cells('I4:I5')
    ws['I4'] = 'Расход за прошедшие сутки T2'
    ws['I4'].style = "ali_grey"
    ws['I5'].style = "ali_grey"
    
        # Сумма
    ws.merge_cells('J4:J5')
    ws['J4'] = 'Показания T3 A+ '
    ws['J4'].style = "ali_grey"
    ws['J5'].style = "ali_grey"
 
    # Дельта
    ws.merge_cells('K4:K5')
    ws['K4'] = 'Расход за прошедшие сутки T3'
    ws['K4'].style = "ali_grey"
    ws['K5'].style = "ali_grey"
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_electric_daily    = request.session['is_electric_daily']
    obj_parent_title    = request.session['obj_parent_title']
    
    obj_key             = request.session['obj_key']

    data_table = []
    if True:
        if True:            
            if (is_electric_daily == '1') & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                params=['T0 A+','T1 A+','T2 A+','T3 A+', 'Электричество']
                data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
            else:
                pass
            
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][3])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][4])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # дата
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % get_val(data_table[row-6][5])  # сумма-показания t0
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            # print 'val ', data_table[row-6][12]
            # print type(data_table[row-6][12])
            # print 'just str ' ,unicode(data_table[row-6][12])
            # print '%s' % unicode(data_table[row-6][12]).replace('.', separator)

            ws.cell('E%s'%(row)).value = '%s' % get_val(data_table[row-6][12]) #str(data_table[row-6][12]).replace('.', separator)   # Расход за прошедшие сутки t0
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % get_val(data_table[row-6][6])   # сумма-показанияt1
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % get_val(data_table[row-6][13])   # Расход за прошедшие суткиt1
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('H%s'%(row)).value = '%s' % get_val(data_table[row-6][7])  # сумма-показанияt2
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % get_val(data_table[row-6][14])  # Расход за прошедшие суткиt2
            ws.cell('I%s'%(row)).style = "ali_white"
        except:
            ws.cell('I%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('J%s'%(row)).value = '%s' % get_val(data_table[row-6][8])   # сумма-показанияt3
            ws.cell('J%s'%(row)).style = "ali_white"
        except:
            ws.cell('J%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % get_val(data_table[row-6][15])   # Расход за прошедшие суткиt3
            ws.cell('K%s'%(row)).style = "ali_white"
        except:
            ws.cell('K%s'%(row)).style = "ali_white"
            next

# Сохраняем в ecxel    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'potreblenie_electric_3_zones_'+translate(obj_title)+'_' + electric_data_start + ' - ' +electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def get_val(val):
    if val == 'Н/Д': 
        return val
    else: 
        return str(val).replace('.', separator)
        


def report_electric_potreblenie_2_zones(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление электроэнергии в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = "ali_grey"
    ws['D4'].style = "ali_grey"
    ws['E4'].style = "ali_grey"
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = "ali_grey"
    ws['D5'] = 'Ктт'
    ws['D5'].style = "ali_grey"
    ws['E5'] = 'А'
    ws['E5'].style = "ali_grey" 
    
    # Сумма
    ws.merge_cells('F3:I3')
    ws['F3'] = 'Сумма A+, кВт*ч'
    ws['F3'].style = "ali_grey"
    ws['G3'].style = "ali_grey"
    ws['H3'].style = "ali_grey"
    ws['I3'].style = "ali_grey"
    
    ws.merge_cells('F4:G4')
    ws['F4'] = 'На ' + str(request.session["electric_data_start"])
    ws['F4'].style = "ali_grey"
    
    ws.merge_cells('H4:I4')
    ws['H4'] = 'На ' + str(request.session["electric_data_end"])
    ws['H4'].style = "ali_grey"

    ws['F5'] = 'Показания'
    ws['F5'].style = "ali_grey"     
    ws['G5'] = 'Энергия'
    ws['G5'].style = "ali_yellow"
    
    ws['H5'] = 'Показания'
    ws['H5'].style = "ali_grey"     
    ws['I5'] = 'Энергия'
    ws['I5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('J3:M3')
    ws['J3'] = 'Тариф 1 A+, кВт*ч'
    ws['J3'].style = "ali_grey"
    ws['K3'].style = "ali_grey"
    ws['L3'].style = "ali_grey"
    ws['M3'].style = "ali_grey"
    
    ws.merge_cells('J4:K4')
    ws['J4'] = 'На ' + str(request.session["electric_data_start"])
    ws['J4'].style = "ali_grey"
    
    ws.merge_cells('L4:M4')
    ws['L4'] = 'На ' + str(request.session["electric_data_end"])
    ws['L4'].style = "ali_grey"

    ws['J5'] = 'Показания'
    ws['J5'].style = "ali_grey"     
    ws['K5'] = 'Энергия'
    ws['K5'].style = "ali_yellow"
    
    ws['L5'] = 'Показания'
    ws['L5'].style = "ali_grey"     
    ws['M5'] = 'Энергия'
    ws['M5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('N3:Q3')
    ws['N3'] = 'Тариф 2 A+, кВт*ч'
    ws['N3'].style = "ali_grey"
    ws['O3'].style = "ali_grey"
    ws['O3'].style = "ali_grey"
    ws['Q3'].style = "ali_grey"
    
    ws.merge_cells('N4:O4')
    ws['N4'] = 'На ' + str(request.session["electric_data_start"])
    ws['N4'].style = "ali_grey"
    
    ws.merge_cells('P4:Q4')
    ws['P4'] = 'На ' + str(request.session["electric_data_end"])
    ws['P4'].style = "ali_grey"

    ws['N5'] = 'Показания'
    ws['N5'].style = "ali_grey"     
    ws['O5'] = 'Энергия'
    ws['O5'].style = "ali_yellow"
    
    ws['P5'] = 'Показания'
    ws['P5'].style = "ali_grey"     
    ws['Q5'] = 'Энергия'
    ws['Q5'].style = "ali_yellow"
         
    # Расход
    ws.merge_cells('R3:W3')
    ws['R3'] = 'Расход А+, кВт*ч'
    ws['R3'].style = "ali_grey"
    ws['W3'].style = "ali_grey"
        # Расход Т0
    ws.merge_cells('R4:S4')
    ws['R4'] = 'Сумма'
    ws['R4'].style = "ali_grey"
    ws['R5'] = 'Показания'
    ws['R5'].style = "ali_grey"
    ws['S5'] = 'Энергия'
    ws['S5'].style = "ali_yellow"
        # Расход Т1
    ws.merge_cells('T4:U4')
    ws['T4'] = 'Tариф 1'
    ws['T4'].style = "ali_grey"
    ws['T5'] = 'Показания'
    ws['T5'].style = "ali_grey"
    ws['U5'] = 'Энергия'
    ws['U5'].style = "ali_yellow"
        # Расход Т2
    ws.merge_cells('V4:W4')
    ws['V4'] = 'Tариф 2'
    ws['V4'].style = "ali_grey"
    ws['V5'] = 'Показания'
    ws['V5'].style = "ali_grey"
    ws['W5'] = 'Энергия'
    ws['W5'].style = "ali_yellow"
  
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']                        
    obj_key             = request.session['obj_key']
    is_electric_monthly = request.session['is_electric_monthly']
    is_electric_daily   = request.session['is_electric_daily']
    data_table_end   = []
    data_table_start = []
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
                     
                 
    if (bool(is_abonent_level.search(obj_key))):
        if (is_electric_monthly == "1"):
            data_table_end   = common_sql.get_data_table_by_date_monthly_2_zones(meters_name, parent_name, electric_data_end)
            data_table_start = common_sql.get_data_table_by_date_monthly_2_zones(meters_name, parent_name, electric_data_start)
        elif (is_electric_daily == "1"):
            data_table_end   = common_sql.get_data_table_by_date_daily_2_zones(meters_name, parent_name, electric_data_end)
            data_table_start = common_sql.get_data_table_by_date_daily_2_zones(meters_name, parent_name, electric_data_start)


    elif (bool(is_object_level_2.search(obj_key))):
            list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
            data_table_end   = []
            data_table_start = []
            for x in range(len(list_of_abonents_2)):
                if (is_electric_monthly == "1"):                
                    data_table_temp_end = common_sql.get_data_table_by_date_monthly_2_zones(list_of_abonents_2[x], meters_name, electric_data_end)
                    data_table_temp_start = common_sql.get_data_table_by_date_monthly_2_zones(list_of_abonents_2[x], meters_name, electric_data_start)

                elif (is_electric_daily == "1"):                
                    data_table_temp_end = common_sql.get_data_table_by_date_daily_2_zones(list_of_abonents_2[x], meters_name, electric_data_end)
                    data_table_temp_start = common_sql.get_data_table_by_date_daily_2_zones(list_of_abonents_2[x], meters_name, electric_data_start)

                if data_table_temp_end and data_table_start:
                    data_table_end.extend(data_table_temp_end)
                    data_table_start.extend(data_table_temp_start)
                    
                else:
                    data_table_start.extend([['', list_of_abonents_2[x],'Н/Д','Н/Д','Н/Д','Н/Д','Н/Д','Н/Д','Н/Д']])
                    data_table_end.extend([['', list_of_abonents_2[x],'Н/Д','Н/Д','Н/Д','Н/Д','Н/Д','Н/Д','Н/Д']])
#Запрашиваем данные для отчета конец
                    
# Заполняем отчет значениями
    for row in range(6, len(data_table_end)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table_end[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table_end[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (get_k_t_n_by_serial_number(data_table_end[row-6][2]))  # Ктн
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Ктт
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (get_k_a_by_serial_number(data_table_end[row-6][2]))  # Ка
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
                   
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table_end[row-6][3])  # Сумма А+ на конец интервала
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table_end[row-6][3]*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Энергия Сумма А+ на конец интервала
            ws.cell('I%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('I%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table_start[row-6][3])  # Сумма А+ на начало интервала
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table_start[row-6][3]*get_k_t_n_by_serial_number(data_table_start[row-6][2])*get_k_t_t_by_serial_number(data_table_start[row-6][2]))  # Энергия Сумма А+ на начало интервала
            ws.cell('G%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('G%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table_end[row-6][4])  # Тариф 1 А+ на конец интервала
            ws.cell('L%s'%(row)).style = "ali_white"
        except:
            ws.cell('L%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % (data_table_end[row-6][4]*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Энергия Тариф 1 А+ на конец интервала
            ws.cell('M%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('M%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table_start[row-6][4])  # Тариф 1 А+ на начало интервала
            ws.cell('J%s'%(row)).style = "ali_white"
        except:
            ws.cell('J%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table_start[row-6][4]*get_k_t_n_by_serial_number(data_table_start[row-6][2])*get_k_t_t_by_serial_number(data_table_start[row-6][2]))  # Энергия Тариф 1 А+ на начало интервала
            ws.cell('K%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('K%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table_end[row-6][5])  # Тариф 2 А+ на конец интервала
            ws.cell('P%s'%(row)).style = "ali_white"
        except:
            ws.cell('P%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('Q%s'%(row)).value = '%s' % (data_table_end[row-6][5]*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Энергия Тариф 2 А+ на конец интервала
            ws.cell('Q%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('Q%s'%(row)).style = "ali_yellow"
            next

        try:
            ws.cell('N%s'%(row)).value = '%s' % (data_table_start[row-6][5])  # Тариф 2 А+ на начало интервала
            ws.cell('N%s'%(row)).style = "ali_white"
        except:
            ws.cell('N%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table_start[row-6][5]*get_k_t_n_by_serial_number(data_table_start[row-6][2])*get_k_t_t_by_serial_number(data_table_start[row-6][2]))  # Энергия Тариф 2 А+ на начало интервала
            ws.cell('O%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('O%s'%(row)).style = "ali_yellow"
            next
            
        # Расход
        try:
            ws.cell('R%s'%(row)).value = '%s' % (data_table_end[row-6][3] - data_table_start[row-6][3] )  # Расход Сумма А+
            ws.cell('R%s'%(row)).style = "ali_white"
        except:
            ws.cell('R%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('S%s'%(row)).value = '%s' % ((data_table_end[row-6][3] - data_table_start[row-6][3])*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Расход Сумма Энергия А+
            ws.cell('S%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('S%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('T%s'%(row)).value = '%s' % (data_table_end[row-6][4] - data_table_start[row-6][4] )  # Расход Тариф 1 А+
            ws.cell('T%s'%(row)).style = "ali_white"
        except:
            ws.cell('T%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('U%s'%(row)).value = '%s' % ((data_table_end[row-6][4] - data_table_start[row-6][4])*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Расход Тариф 1 Энергия А+
            ws.cell('U%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('U%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('V%s'%(row)).value = '%s' % (data_table_end[row-6][5] - data_table_start[row-6][5] )  # Расход Тариф 2 А+
            ws.cell('V%s'%(row)).style = "ali_white"
        except:
            ws.cell('V%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('W%s'%(row)).value = '%s' % ((data_table_end[row-6][5] - data_table_start[row-6][5])*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Расход Тариф 2 Энергия А+
            ws.cell('W%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('W%s'%(row)).style = "ali_yellow"
            next            
# Конец наполнения отчёта
            
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    
    output_name = 'rashod_2_zones ' + str(electric_data_start) + ' - ' + str(electric_data_end)
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response



def report_electric_potreblenie_3_zones_v2(request):
    SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')
    ROUND_SIZE = getattr(settings, 'ROUND_SIZE', 3)
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+'. Потребление электроэнергии в период с ' + electric_data_start + ' по ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование абонента'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = "ali_grey"
    ws['D4'].style = "ali_grey"
    ws['E4'].style = "ali_grey"
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = "ali_grey"
    ws['D5'] = 'Ктт'
    ws['D5'].style = "ali_grey"
    ws['E5'] = 'А'
    ws['E5'].style = "ali_grey" 
    
    # Сумма
    ws.merge_cells('F3:I3')
    ws['F3'] = 'Сумма A+, кВт*ч'
    ws['F3'].style = "ali_grey"
    ws['G3'].style = "ali_grey"
    ws['H3'].style = "ali_grey"
    ws['I3'].style = "ali_grey"
    
    ws.merge_cells('F4:G4')
    ws['F4'] = 'На ' + str(request.session["electric_data_start"])
    ws['F4'].style = "ali_grey"
    
    ws.merge_cells('H4:I4')
    ws['H4'] = 'На ' + str(request.session["electric_data_end"])
    ws['H4'].style = "ali_grey"

    ws['F5'] = 'Показания'
    ws['F5'].style = "ali_grey"     
    ws['G5'] = 'Энергия'
    ws['G5'].style = "ali_yellow"
    
    ws['H5'] = 'Показания'
    ws['H5'].style = "ali_grey"     
    ws['I5'] = 'Энергия'
    ws['I5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('J3:M3')
    ws['J3'] = 'Тариф 1 A+, кВт*ч'
    ws['J3'].style = "ali_grey"
    ws['K3'].style = "ali_grey"
    ws['L3'].style = "ali_grey"
    ws['M3'].style = "ali_grey"
    
    ws.merge_cells('J4:K4')
    ws['J4'] = 'На ' + str(request.session["electric_data_start"])
    ws['J4'].style = "ali_grey"
    
    ws.merge_cells('L4:M4')
    ws['L4'] = 'На ' + str(request.session["electric_data_end"])
    ws['L4'].style = "ali_grey"

    ws['J5'] = 'Показания'
    ws['J5'].style = "ali_grey"     
    ws['K5'] = 'Энергия'
    ws['K5'].style = "ali_yellow"
    
    ws['L5'] = 'Показания'
    ws['L5'].style = "ali_grey"     
    ws['M5'] = 'Энергия'
    ws['M5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('N3:Q3')
    ws['N3'] = 'Тариф 2 A+, кВт*ч'
    ws['N3'].style = "ali_grey"
    ws['O3'].style = "ali_grey"
    ws['O3'].style = "ali_grey"
    ws['Q3'].style = "ali_grey"
    
    ws.merge_cells('N4:O4')
    ws['N4'] = 'На ' + str(request.session["electric_data_start"])
    ws['N4'].style = "ali_grey"
    
    ws.merge_cells('P4:Q4')
    ws['P4'] = 'На ' + str(request.session["electric_data_end"])
    ws['P4'].style = "ali_grey"

    ws['N5'] = 'Показания'
    ws['N5'].style = "ali_grey"     
    ws['O5'] = 'Энергия'
    ws['O5'].style = "ali_yellow"
    
    ws['P5'] = 'Показания'
    ws['P5'].style = "ali_grey"     
    ws['Q5'] = 'Энергия'
    ws['Q5'].style = "ali_yellow"
    
    # Тариф 3
    ws.merge_cells('R3:U3')
    ws['R3'] = 'Тариф 3 A+, кВт*ч'
    ws['R3'].style = "ali_grey"
    ws['S3'].style = "ali_grey"
    ws['T3'].style = "ali_grey"
    ws['U3'].style = "ali_grey"
    
    ws.merge_cells('R4:S4')
    ws['R4'] = 'На ' + str(request.session["electric_data_start"])
    ws['R4'].style = "ali_grey"
    
    ws.merge_cells('T4:U4')
    ws['T4'] = 'На ' + str(request.session["electric_data_end"])
    ws['T4'].style = "ali_grey"

    ws['R5'] = 'Показания'
    ws['R5'].style = "ali_grey"     
    ws['S5'] = 'Энергия'
    ws['S5'].style = "ali_yellow"
    
    ws['T5'] = 'Показания'
    ws['T5'].style = "ali_grey"     
    ws['U5'] = 'Энергия'
    ws['U5'].style = "ali_yellow"
         
    # Расход
    ws.merge_cells('V3:AC3')
    ws['V3'] = 'Расход А+, кВт*ч'
    ws['V3'].style = "ali_grey"
    ws['AC3'].style = "ali_grey"
        # Расход Т0
    ws.merge_cells('V4:W4')
    ws['V4'] = 'Сумма'
    ws['V4'].style = "ali_grey"
    ws['V5'] = 'Показания'
    ws['V5'].style = "ali_grey"
    ws['W5'] = 'Энергия'
    ws['W5'].style = "ali_yellow"
        # Расход Т1
    ws.merge_cells('X4:Y4')
    ws['X4'] = 'Tариф 1'
    ws['X4'].style = "ali_grey"
    ws['X5'] = 'Показания'
    ws['X5'].style = "ali_grey"
    ws['Y5'] = 'Энергия'
    ws['Y5'].style = "ali_yellow"
        # Расход Т2
    ws.merge_cells('Z4:AA4')
    ws['Z4'] = 'Tариф 2'
    ws['Z4'].style = "ali_grey"
    ws['Z5'] = 'Показания'
    ws['Z5'].style = "ali_grey"
    ws['AA5'] = 'Энергия'
    ws['AA5'].style = "ali_yellow"
        # Расход Т3
    ws.merge_cells('AB4:AC4')
    ws['AB4'] = 'Tариф 3'
    ws['AB4'].style = "ali_grey"
    ws['AC4'].style = "ali_grey"
    ws['AB5'] = 'Показания'
    ws['AB5'].style = "ali_grey"
    ws['AC5'] = 'Энергия'
    ws['AC5'].style = "ali_yellow"

    if SHOW_LIC_NUM:
            ws.merge_cells('AD4:AD5')
            ws['AD4'] = 'Лицевой номер абонента'
            ws['AD4'].style = "ali_grey"
            ws['AD5'].style = "ali_grey"
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    
    obj_parent_title    = request.session['obj_parent_title']
    
    obj_key             = request.session['obj_key']
    is_electric_delta  = request.session['is_electric_delta']
    is_electric_monthly=request.session['is_electric_monthly']
    data_table = []
    if True:
        if True:                        
            res='Электричество'
            
            if (is_electric_monthly=="1"):
                dm='monthly'
            else:
                dm='daily'
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                    isAbon=True                    
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
                
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                    isAbon=False
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
            #*********************************************************************************************************************************************************************
            elif (is_electric_delta == '1') &(bool(is_group_level.search(obj_key))):
                    data_table=common_sql.get_data_table_electric_period_for_group(obj_title,obj_parent_title, electric_data_start, electric_data_end, res)
                    request.session["data_table_export"] = data_table
#Запрашиваем данные для отчета конец
                    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' %  str(round(float(data_table[row-6][23]), 1)).replace('.', separator)  # Ктн
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % str(round(float(data_table[row-6][20]), 1)).replace('.', separator)  # Ктт
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % str(round(float(data_table[row-6][24]), 1)).replace('.', separator)  # Ка
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
                   
        try:            
            ws.cell('H%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][7]), ROUND_SIZE, separator)  # Сумма А+ на конец интервала
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][7]), ROUND_SIZE, separator)  # Энергия Сумма А+ на конец интервала
            ws.cell('I%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('I%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s ' % get_val_by_round(float(data_table[row-6][2]), ROUND_SIZE, separator)  # '%s' % (data_table[row-6][2])  # Сумма А+ на начало интервала
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][2]), ROUND_SIZE, separator)  # Энергия Сумма А+ на начало интервала
            ws.cell('G%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('G%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][8]), ROUND_SIZE, separator)   # Тариф 1 А+ на конец интервала
            ws.cell('L%s'%(row)).style = "ali_white"
        except:
            ws.cell('L%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][8]), ROUND_SIZE, separator)  # Энергия Тариф 1 А+ на конец интервала
            ws.cell('M%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('M%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][3]), ROUND_SIZE, separator)  # Тариф 1 А+ на начало интервала
            ws.cell('J%s'%(row)).style = "ali_white"
        except:
            ws.cell('J%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][3]), ROUND_SIZE, separator)  # Энергия Тариф 1 А+ на начало интервала
            ws.cell('K%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('K%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][9]), ROUND_SIZE, separator)   # Тариф 2 А+ на конец интервала
            ws.cell('P%s'%(row)).style = "ali_white"
        except:
            ws.cell('P%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('Q%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][9]), ROUND_SIZE, separator)  # Энергия Тариф 2 А+ на конец интервала
            ws.cell('Q%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('Q%s'%(row)).style = "ali_yellow"
            next

        try:
            ws.cell('N%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][4]), ROUND_SIZE, separator)  # Тариф 2 А+ на начало интервала
            ws.cell('N%s'%(row)).style = "ali_white"
        except:
            ws.cell('N%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][4]), ROUND_SIZE, separator)  # Энергия Тариф 2 А+ на начало интервала
            ws.cell('O%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('O%s'%(row)).style = "ali_yellow"
            next            
            
        try:
            ws.cell('T%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][10]), ROUND_SIZE, separator)  # Тариф 3 А+ на конец интервала
            ws.cell('T%s'%(row)).style = "ali_white"
        except:
            ws.cell('T%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('U%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][10]), ROUND_SIZE, separator)  # "Энергия Тариф 3 А+ на конец интервала
            ws.cell('U%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('U%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('R%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][5]), ROUND_SIZE, separator)  # Тариф 3 А+ на начало интервала
            ws.cell('R%s'%(row)).style = "ali_white"
        except:
            ws.cell('R%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('S%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][5]), ROUND_SIZE, separator)  # "Энергия Тариф 3 А+ на начало интервала
            ws.cell('S%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('S%s'%(row)).style = "ali_yellow"
            next
        # Расход
        try:
            ws.cell('V%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][12]), ROUND_SIZE, separator)  # Расход Сумма А+
            ws.cell('V%s'%(row)).style = "ali_white"
        except:
            ws.cell('V%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('W%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][12]), ROUND_SIZE, separator)  # Расход Сумма Энергия А+
            ws.cell('W%s'%(row)).style = "ali_yellow"            
        except:
            ws.cell('W%s'%(row)).style = "ali_yellow"
            next
            
        try:
            #val = format(data_table[row-6][13],'.3f')
            ws.cell('X%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][13]), ROUND_SIZE, separator)    # Расход Тариф 1 А+
            ws.cell('X%s'%(row)).style = "ali_white"
        except:
            ws.cell('X%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('Y%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][13]), ROUND_SIZE, separator)    # Расход Тариф 1 Энергия А+
            ws.cell('Y%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('Y%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('Z%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][14]), ROUND_SIZE, separator)  # Расход Тариф 2 А+
            ws.cell('Z%s'%(row)).style = "ali_white"
        except:
            ws.cell('Z%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('AA%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][14]), ROUND_SIZE, separator)  # Расход Тариф 2 Энергия А+
            ws.cell('AA%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('AA%s'%(row)).style = "ali_yellow"
            next
            
        try:
            #val = format(data_table[row-6][15],'.3f')
            ws.cell('AB%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][15]), ROUND_SIZE, separator)  # Расход Тариф 3 А+
            ws.cell('AB%s'%(row)).style = "ali_white"
        except:
            ws.cell('AB%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('AC%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][15]), ROUND_SIZE, separator)  # Расход Тариф 3 Энергия А+
            ws.cell('AC%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('AC%s'%(row)).style = "ali_yellow"
            next

        if SHOW_LIC_NUM:
            try:                
                ws.cell('AD%s'%(row)).value = '%s' % (data_table[row-6][25])   # Лицевой нмоер абонента
                ws.cell('AD%s'%(row)).style = "ali_white"
            except:
                ws.cell('AD%s'%(row)).style = "ali_white"
                next
# Конец наполнения отчёта
            
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    
    output_name = 'rashod_3_zones_'+translate(obj_title)+'_' + str(electric_data_start) + '-' + str(electric_data_end)
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_electric_potreblenie_2_zones_v2(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление электроэнергии в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = "ali_grey"
    ws['D4'].style = "ali_grey"
    ws['E4'].style = "ali_grey"
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = "ali_grey"
    ws['D5'] = 'Ктт'
    ws['D5'].style = "ali_grey"
    ws['E5'] = 'А'
    ws['E5'].style = "ali_grey" 
    
    # Сумма
    ws.merge_cells('F3:I3')
    ws['F3'] = 'Сумма A+, кВт*ч'
    ws['F3'].style = "ali_grey"
    ws['G3'].style = "ali_grey"
    ws['H3'].style = "ali_grey"
    ws['I3'].style = "ali_grey"
    
    ws.merge_cells('F4:G4')
    ws['F4'] = 'На ' + str(request.session["electric_data_start"])
    ws['F4'].style = "ali_grey"
    
    ws.merge_cells('H4:I4')
    ws['H4'] = 'На ' + str(request.session["electric_data_end"])
    ws['H4'].style = "ali_grey"

    ws['F5'] = 'Показания'
    ws['F5'].style = "ali_grey"     
    ws['G5'] = 'Энергия'
    ws['G5'].style = "ali_yellow"
    
    ws['H5'] = 'Показания'
    ws['H5'].style = "ali_grey"     
    ws['I5'] = 'Энергия'
    ws['I5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('J3:M3')
    ws['J3'] = 'Тариф 1 A+, кВт*ч'
    ws['J3'].style = "ali_grey"
    ws['K3'].style = "ali_grey"
    ws['L3'].style = "ali_grey"
    ws['M3'].style = "ali_grey"
    
    ws.merge_cells('J4:K4')
    ws['J4'] = 'На ' + str(request.session["electric_data_start"])
    ws['J4'].style = "ali_grey"
    
    ws.merge_cells('L4:M4')
    ws['L4'] = 'На ' + str(request.session["electric_data_end"])
    ws['L4'].style = "ali_grey"

    ws['J5'] = 'Показания'
    ws['J5'].style = "ali_grey"     
    ws['K5'] = 'Энергия'
    ws['K5'].style = "ali_yellow"
    
    ws['L5'] = 'Показания'
    ws['L5'].style = "ali_grey"     
    ws['M5'] = 'Энергия'
    ws['M5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('N3:Q3')
    ws['N3'] = 'Тариф 2 A+, кВт*ч'
    ws['N3'].style = "ali_grey"
    ws['O3'].style = "ali_grey"
    ws['O3'].style = "ali_grey"
    ws['Q3'].style = "ali_grey"
    
    ws.merge_cells('N4:O4')
    ws['N4'] = 'На ' + str(request.session["electric_data_start"])
    ws['N4'].style = "ali_grey"
    
    ws.merge_cells('P4:Q4')
    ws['P4'] = 'На ' + str(request.session["electric_data_end"])
    ws['P4'].style = "ali_grey"

    ws['N5'] = 'Показания'
    ws['N5'].style = "ali_grey"     
    ws['O5'] = 'Энергия'
    ws['O5'].style = "ali_yellow"
    
    ws['P5'] = 'Показания'
    ws['P5'].style = "ali_grey"     
    ws['Q5'] = 'Энергия'
    ws['Q5'].style = "ali_yellow"
    

         
    # Расход
    ws.merge_cells('R3:W3')
    ws['R3'] = 'Расход А+, кВт*ч'
    ws['R3'].style = "ali_grey"
    ws['W3'].style = "ali_grey"
        # Расход Т0
    ws.merge_cells('R4:S4')
    ws['R4'] = 'Сумма'
    ws['R4'].style = "ali_grey"
    ws['R5'] = 'Показания'
    ws['R5'].style = "ali_grey"
    ws['S5'] = 'Энергия'
    ws['S5'].style = "ali_yellow"
        # Расход Т1
    ws.merge_cells('T4:U4')
    ws['T4'] = 'Tариф 1'
    ws['T4'].style = "ali_grey"
    ws['T5'] = 'Показания'
    ws['T5'].style = "ali_grey"
    ws['U5'] = 'Энергия'
    ws['U5'].style = "ali_yellow"
        # Расход Т2
    ws.merge_cells('V4:W4')
    ws['V4'] = 'Tариф 2'
    ws['V4'].style = "ali_grey"
    ws['V5'] = 'Показания'
    ws['V5'].style = "ali_grey"
    ws['W5'] = 'Энергия'
    ws['W5'].style = "ali_yellow"
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    
    obj_parent_title    = request.session['obj_parent_title']
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
    obj_key             = request.session['obj_key']
    is_electric_delta  = request.session['is_electric_delta']
    is_electric_monthly=request.session['is_electric_monthly']
    data_table = []
    if True:
        if True:                        
            res='Электричество'
            
            if (is_electric_monthly=="1"):
                dm='monthly'
            else:
                dm='daily'
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                    isAbon=True                    
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
                
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                    isAbon=False
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
            #*********************************************************************************************************************************************************************
            elif (is_electric_delta == '1') &(bool(is_group_level.search(obj_key))):
                    data_table=common_sql.get_data_table_electric_period_for_group(obj_title,obj_parent_title, electric_data_start, electric_data_end, res)
                    request.session["data_table_export"] = data_table
#Запрашиваем данные для отчета конец
                    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][23])  # Ктн
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][20])  # Ктт
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][24])  # Ка
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
                   
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # Сумма А+ на конец интервала
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][7]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Сумма А+ на конец интервала
            ws.cell('I%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('I%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][2])  # Сумма А+ на начало интервала
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][2]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Сумма А+ на начало интервала
            ws.cell('G%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('G%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table[row-6][8])  # Тариф 1 А+ на конец интервала
            ws.cell('L%s'%(row)).style = "ali_white"
        except:
            ws.cell('L%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % (data_table[row-6][8]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Тариф 1 А+ на конец интервала
            ws.cell('M%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('M%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][3])  # Тариф 1 А+ на начало интервала
            ws.cell('J%s'%(row)).style = "ali_white"
        except:
            ws.cell('J%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table[row-6][3]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Тариф 1 А+ на начало интервала
            ws.cell('K%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('K%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table[row-6][9])  # Тариф 2 А+ на конец интервала
            ws.cell('P%s'%(row)).style = "ali_white"
        except:
            ws.cell('P%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('Q%s'%(row)).value = '%s' % (data_table[row-6][9]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Тариф 2 А+ на конец интервала
            ws.cell('Q%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('Q%s'%(row)).style = "ali_yellow"
            next

        try:
            ws.cell('N%s'%(row)).value = '%s' % (data_table[row-6][4])  # Тариф 2 А+ на начало интервала
            ws.cell('N%s'%(row)).style = "ali_white"
        except:
            ws.cell('N%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table[row-6][4]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Тариф 2 А+ на начало интервала
            ws.cell('O%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('O%s'%(row)).style = "ali_yellow"
            next
            

        # Расход
        try:
            ws.cell('R%s'%(row)).value = '%s' % (data_table[row-6][12])  # Расход Сумма А+
            ws.cell('R%s'%(row)).style = "ali_white"
        except:
            ws.cell('R%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('S%s'%(row)).value = '%s' % (data_table[row-6][12]*data_table[row-6][20]*data_table[row-6][23])  # Расход Сумма Энергия А+
            ws.cell('S%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('S%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('T%s'%(row)).value = '%s' % (data_table[row-6][13])  # Расход Тариф 1 А+
            ws.cell('T%s'%(row)).style = "ali_white"
        except:
            ws.cell('T%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('U%s'%(row)).value = '%s' % (data_table[row-6][13]*data_table[row-6][20]*data_table[row-6][23])  # Расход Тариф 1 Энергия А+
            ws.cell('U%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('U%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('V%s'%(row)).value = '%s' % (data_table[row-6][14])  # Расход Тариф 2 А+
            ws.cell('V%s'%(row)).style = "ali_white"
        except:
            ws.cell('V%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('W%s'%(row)).value = '%s' % (data_table[row-6][14]*data_table[row-6][20]*data_table[row-6][23])  # Расход Тариф 2 Энергия А+
            ws.cell('W%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('W%s'%(row)).style = "ali_yellow"
            next

# Конец наполнения отчёта
            
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    
    output_name = 'rashod_2_zones ' + str(electric_data_start) + ' - ' + str(electric_data_end)
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response




def pokazaniya_heat_report_v2(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Показания теплосчётчиков на ' + str(request.session["electric_data_end"])
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Заводской номер'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Показания, Гкал'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания, м3'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Время работы, ч'
    ws['E5'].style = "ali_grey"

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']

    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            #request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            #request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']  

    list_except = []
                     
    data_table = []
    list_except = []
    if (bool(is_abonent_level.search(obj_key))):     
        data_table = common_sql.get_data_table_by_date_heat_v2(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_heat_v2(meters_name, parent_name, electric_data_end, False)
        for row in data_table:
            for x in list_except:
                if x==row[2]:
                    data_table.remove(x)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по расходу воды
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5])  # время работы
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17 
                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'heat_report_'+translate(parent_name)+'_'+translate(meters_name)+'-'+electric_data_end

    #output_name = u'heat_report' 
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pokazaniya_sayany(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    meters_name           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = meters_name+'. Показания теплосчётчиков Саяны на ' + electric_data_end 
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Номер счётчика'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Дата'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания Q1, Гкал'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Показания M1, т'
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 't1, C˚'
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 't2, C˚'
    ws['G5'].style = "ali_grey"

# ниже не переделывала
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')

    
    parent_name    = request.session['obj_parent_title']
    
    obj_key             = request.session['obj_key']

    
    #print parent_name,meters_name,electric_data_end, obj_key
    
    data_table = []
#    if request.is_ajax():
#        #if request.method == 'GET':
#            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
#            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
#            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
#            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # Дата
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу Q1
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу Q2
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания по теплу t1
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания по теплу t2
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
            
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17 
                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'heat_sayany_report_'+translate(meters_name)+'_'+electric_data_end 
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pokazaniya_sayany_archive(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Показания теплосчётчиков Саяны на ' + str(request.session["electric_data_end"])
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Номер счётчика'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Дата'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания Q1'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Показания Q2'
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 't1'
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 't2'
    ws['G5'].style = "ali_grey"

# ниже не переделывала
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')

    
    parent_name    = request.session['obj_parent_title']
    meters_name           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']

    
    data_table = []

    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # Дата
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу Q1
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу Q2
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания по теплу t1
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания по теплу t2
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
            
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17 

    import zipfile

    o=io.StringIO()
    
    zf = zipfile.ZipFile(response, mode='w', compression=zipfile.ZIP_DEFLATED)
    zf.writestr('README.txt', 'test msg')
    #wb.save(o)
    zf.writestr('test.xlsx',o.getvalue())
    zf.close()
    response=HttpResponse(response.getvalue())
    response['Content-Type'] = 'application/x-zip-compressed'
    response['Content-Disposition'] = "attachment; filename=\"sayani_test.zip\""
    
    return response
    
def report_sayany_last(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

    meters_name           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = meters_name+'. Показания теплосчётчиков Саяны на ' + electric_data_end
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Номер счётчика'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Дата'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания Q1, Гкал'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Показания М1, т'
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 't1, C˚'
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 't2, C˚'
    ws['G5'].style = "ali_grey"

# ниже не переделывала
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')

    
    parent_name    = request.session['obj_parent_title']
    meters_name           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']

    
    data_table = []
#    if request.is_ajax():
#        #if request.method == 'GET':
#            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
#            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
#            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
#            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, False)
    
    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        if (data_table[i][3] is None):
            data_table[i][0]=electric_data_end
            dt=common_sql.get_data_table_by_date_heat_sayany_v2(data_table[i][1], meters_name, None, True)
            if (len(dt)>0):
                data_table[i]=dt[0]
        data_table[i]=tuple(data_table[i])
    
    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # Дата
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу Q1
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу Q2
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания по теплу t1
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания по теплу t2
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
            
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17 
                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'heat_sayany_report_'+translate(meters_name)+'_' +electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_heat_potreblenie_sayany(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
#Шапка
    ws.merge_cells('A2:G2')
    ws['A2'] = meters_name+'. Потребление по теплосчётчикам Sayany в период с ' + electric_data_start + ' по ' +electric_data_end
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Показания Q1 на '  + electric_data_start+', Гкал'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания Q1 на '  + electric_data_end+', Гкал'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Потребление Q1, Гкал'
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 'Показания M1 на '  + electric_data_start+', т'
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'Показания M1 на '  + electric_data_end+', т'
    ws['G5'].style = "ali_grey"
    
    ws['H5'] = 'Потребление M1, т'
    ws['H5'].style = "ali_grey"

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
                        
    obj_key             = request.session['obj_key']

    data_table = []
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_period_heat_sayany(meters_name, parent_name,electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_period_heat_sayany(meters_name, parent_name,electric_data_start, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания Q1 на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания Q1 на конец
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление Q1
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания M1 на начало
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания M1 на конец
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # Потребление M1
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
#    ws.column_dimensions['C'].width = 35
#    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['H'].width = 18
#    ws.column_dimensions['F'].width = 18
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    
    output_name = 'potreblenie_heat_report_'+translate(meters_name)+'_'+electric_data_start+'-'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_water_potreblenie_pulsar(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    meters_name         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')
    electric_data_start   = request.GET.get('electric_data_start')
    
#Шапка
    ws.merge_cells('A2:G2')
    ws['A2'] = meters_name+'. Потребление по импульсным водосчётчикам в период с ' + electric_data_start + ' по ' +electric_data_end
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Тип ресурса '
    ws['C5'].style = "ali_grey"
    
    ws['d5'] = 'Показания на '  + electric_data_start+', м3'
    ws['d5'].style = "ali_grey"
    
    ws['e5'] = 'Показания на '  + electric_data_end+', м3'
    ws['e5'].style = "ali_grey"
    
    ws['f5'] = 'Потребление, м3'
    ws['f5'].style = "ali_grey"
    
    ws['g5'] = 'Лицевой номер '
    ws['g5'].style = "ali_grey"    
  
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'level2')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level1')

    parent_name         = request.GET.get('obj_parent_title')
    obj_key             = request.GET.get('obj_key')
    data_table=[]
    
    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_water_period_pulsar(meters_name, parent_name,electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_water_period_pulsar(meters_name, parent_name,electric_data_start, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
    # for i in range(len(data_table)):
    #     data_table[i]=list(data_table[i])
    #     num=data_table[i][3]
    #     if ('ХВС, №' in num) or ('ГВС, №' in num):
    #         num=num.replace(u'ХВС, №', ' ')
    #         num=num.replace(u'ГВС, №', ' ')
    #         data_table[i][3]=num
    #     data_table[i]=tuple(data_table[i])

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][3])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][4])  # Тип ресурса
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % get_val(data_table[row-6][5])  # Показания на начало
            ws.cell('d%s'%(row)).style = "ali_white"
        except:
            ws.cell('d%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % get_val(data_table[row-6][6])  # Показания  на конец
            ws.cell('e%s'%(row)).style = "ali_white"
        except:
            ws.cell('e%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('f%s'%(row)).value = '%s' % get_val(data_table[row-6][7])  # Потребление
            ws.cell('f%s'%(row)).style = "ali_white"
        except:
            ws.cell('f%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('g%s'%(row)).value = '%s' % (data_table[row-6][1])  # Лицевой номер
            ws.cell('g%s'%(row)).style = "ali_white"
        except:
            ws.cell('g%s'%(row)).style = "ali_white"
            next


    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    #ws.column_dimensions['H'].width = 25
    ws.column_dimensions['F'].width = 18
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    
    output_name = 'potreblenie_water_report_'+translate(meters_name)+'_'+electric_data_start+'-'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def pokazaniya_water_current_report(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Последние считанные показания по ХВС и ГВС на ' + str(datetime.now())
    
    ws['A5'] = 'Дата'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Время'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Абонент'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Заводской номер счётчика'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Показания ХВС, м3'
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 'Показания ГВС, м3'
    ws['F5'].style = "ali_grey"

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']

    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    
    data_table=[]
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_current_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end,  True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table_temp=common_sql.get_current_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end,  False)
        for row in data_table_temp:
            if row[4]=='Н/Д' and row[5]=='Н/Д':
                row2=common_sql.get_current_water_gvs_hvs(str(row[2]), str(row[6]) , electric_data_end, True)
                if len(row2)==0:
                    r=[str(electric_data_end), 'Н/Д', str(row[2]),str(row[3]), 'Н/Д', 'Н/Д']
                    data_table.append(r)
                else:
                    data_table.append(row2[0])
            else:
                data_table.append(row)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Дата
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # время
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # абонент
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # заводской номер
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # хвс
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # гвс
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 17 
                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'water_report_current' 
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def pokazaniya_water_daily_report(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Показания импульсные по ХВС и ГВС за ' + str(request.session["electric_data_end"])
    
    ws['A5'] = 'Дата'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Абонент'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Заводской номер счётчика'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания ХВС, м3'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Показания ГВС, м3'
    ws['E5'].style = "ali_grey"

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end, 'daily', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table=common_sql.get_daily_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end, 'daily', False)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Дата
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  #  абонент
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  #хвс
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # гвс
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next

    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25 
                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'water_report' 
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

#_________________________________________________________________________
def report_all_res_by_date(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]

#Шапка
    ws['A1'] = 'Лицевой счет'
    ws['A1'].style = "ali_grey"
    
    ws['B1'] = 'Дата начала работы'
    ws['B1'].style = "ali_grey"
    
    ws['C1'] = 'Номер прибора'
    ws['C1'].style = "ali_grey"
    
    ws['D1'] = 'тип прибора'
    ws['D1'].style = "ali_grey"
    
    ws['E1'] = 'Наименование ПУ'
    ws['E1'].style = "ali_grey"
    
    ws['F1'] = 'Показания'
    ws['F1'].style = "ali_grey"
    
    ws['G1'] = 'Дата Окончания' 
    ws['G1'].style = "ali_grey"
    
    ws['H1'] = 'Дата'
    ws['H1'].style = "ali_grey"

    ws['I1'] = 'Квартирный' 
    ws['I1'].style = "ali_grey"
    
    ws['J1'] = 'Адрес'
    ws['J1'].style = "ali_grey"
    
    ws['K1'] = 'Операция'
    ws['K1'].style = "ali_grey"
    
    ws['L1'] = 'Заменяемый счетчик'
    ws['L1'].style = "ali_grey"
    
    ws['M1'] = 'Показания заменяемого счетчика'
    ws['M1'].style = "ali_grey"

    ws['N1'] = 'Улица'
    ws['N1'].style = "ali_grey"

    ws['O1'] = 'Дом'
    ws['O1'].style = "ali_grey"

    ws['P1'] = 'Квартира'
    ws['P1'].style = "ali_grey"
#Запрашиваем данные для отчета

           
    
    data_table = []
    data_table = common_sql.get_data_table_report_all_res_by_date(electric_data_end)

    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        for j in range(1,len(data_table[i])):
            if (data_table[i][j] == None) or (data_table[i][j] is None):
                data_table[i][j]=''


# Заполняем отчет значениями
    for row in range(2, len(data_table)+2):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-2][0])  # лицевой счёт
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            a=str(data_table[row-2][1])
            b=datetime.datetime.strptime(a,'%Y-%m-%d')
            d=datetime.datetime.strftime(b,"%d.%m.%Y")
            ws.cell('B%s'%(row)).value = '%s' % d  # начальная дата
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-2][2])  # номер счётчика
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-2][3])  # тип энергоресурса
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-2][4])  # наименование ПУ
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-2][5])  # показания на начальную дату
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % electric_data_end  # дата снятия показаний
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
                        
        try:
            ws.cell('K%s'%(row)).value = 'ВводПоказаний'
            ws.cell('K%s'%(row)).style = "ali_white"
        except:
            ws.cell('K%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table[row-2][8])  # Дом-корпус
            ws.cell('O%s'%(row)).style = "ali_white"
        except:
            ws.cell('O%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table[row-2][7])  # № квартиры
            ws.cell('P%s'%(row)).style = "ali_white"
        except:
            ws.cell('P%s'%(row)).style = "ali_white"
            next
            

#    ws.row_dimensions[5].height = 41
#    ws.column_dimensions['A'].width = 10 
#    ws.column_dimensions['B'].width = 10 
#    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['j'].width = 15
    ws.column_dimensions['k'].width = 15
                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_all_resources_'+str(electric_data_end)
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_electric_all_by_date(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]

#Шапка
    
    ws['A1'] = 'Лицевой счет'
    ws['A1'].style = "ali_grey"
    
    ws['B1'] = 'Дата начала работы'
    ws['B1'].style = "ali_grey"
    
    ws['C1'] = 'Номер прибора'
    ws['C1'].style = "ali_grey"
    
    ws['D1'] = 'тип прибора'
    ws['D1'].style = "ali_grey"
    
    ws['E1'] = 'Наименование ПУ'
    ws['E1'].style = "ali_grey"
    
    ws['F1'] = 'Показания'
    ws['F1'].style = "ali_grey"
    
    ws['G1'] = 'Дата Окончания' 
    ws['G1'].style = "ali_grey"
    
    ws['H1'] = 'Дата'
    ws['H1'].style = "ali_grey"

    ws['I1'] = 'Квартирный' 
    ws['I1'].style = "ali_grey"
    
    ws['J1'] = 'Адрес'
    ws['J1'].style = "ali_grey"
    
    ws['K1'] = 'Операция'
    ws['K1'].style = "ali_grey"
    
    ws['L1'] = 'Заменяемый счетчик'
    ws['L1'].style = "ali_grey"
    
    ws['M1'] = 'Показания заменяемого счетчика'
    ws['M1'].style = "ali_grey"

    ws['N1'] = 'Улица'
    ws['N1'].style = "ali_grey"

    ws['O1'] = 'Дом'
    ws['O1'].style = "ali_grey"

    ws['P1'] = 'Квартира'
    ws['P1'].style = "ali_grey"
#Запрашиваем данные для отчета

           
    
    data_table = []
    data_table = common_sql.get_data_table_report_electric_res_by_date(electric_data_end)

    #zamenyem None na N/D vezde
#    if len(data_table)>0: 
#        data_table=common_sql.ChangeNull(data_table, None)
    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        for j in range(1,len(data_table[i])):
            if (data_table[i][j] == None) or (data_table[i][j] is None):
                data_table[i][j]=''
        data_table[i]=tuple(data_table[i])


# Заполняем отчет значениями
    for row in range(2, len(data_table)+2):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-2][0])  # лицевой счёт
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            a=str(data_table[row-2][1])
            b=datetime.datetime.strptime(a,'%Y-%m-%d')
            d=datetime.datetime.strftime(b,"%d.%m.%Y")

            ws.cell('B%s'%(row)).value = '%s' %d  # начальная дата
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-2][2])  # номер счётчика
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-2][3])  # тип энергоресурса
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-2][9])  # наименование ПУ
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-2][5])  # показания на начальную дату
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            #a=str(data_table[row-2][6])
            a=str(electric_data_end)
            #b=datetime.datetime.strptime(a,'%Y-%m-%d')
            #d=datetime.datetime.strftime(b,"%d.%m.%Y")
            ws.cell('H%s'%(row)).value = '%s' % a  # дата снятия показаний
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('K%s'%(row)).value = 'ВводПоказаний'
            ws.cell('K%s'%(row)).style = "ali_white"
        except:
            ws.cell('K%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table[row-2][8])  # Дом-корпус
            ws.cell('O%s'%(row)).style = "ali_white"
        except:
            ws.cell('O%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table[row-2][7])  # № квартиры
            ws.cell('P%s'%(row)).style = "ali_white"
        except:
            ws.cell('P%s'%(row)).style = "ali_white"
            next
            

#    ws.row_dimensions[5].height = 41
#    ws.column_dimensions['A'].width = 10 
#    ws.column_dimensions['B'].width = 10 
#    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['j'].width = 15
    ws.column_dimensions['k'].width = 15
                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_electric_resources_'+str(electric_data_end)
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_water_all_by_date(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]

#Шапка
    
    ws['A1'] = 'Лицевой счет'
    ws['A1'].style = "ali_grey"
    
    ws['B1'] = 'Дата начала работы'
    ws['B1'].style = "ali_grey"
    
    ws['C1'] = 'Номер прибора'
    ws['C1'].style = "ali_grey"
    
    ws['D1'] = 'тип прибора'
    ws['D1'].style = "ali_grey"
    
    ws['E1'] = 'Наименование ПУ'
    ws['E1'].style = "ali_grey"
    
    ws['F1'] = 'Показания'
    ws['F1'].style = "ali_grey"
    
    ws['G1'] = 'Дата Окончания' 
    ws['G1'].style = "ali_grey"
    
    ws['H1'] = 'Дата'
    ws['H1'].style = "ali_grey"

    ws['I1'] = 'Квартирный' 
    ws['I1'].style = "ali_grey"
    
    ws['J1'] = 'Адрес'
    ws['J1'].style = "ali_grey"
    
    ws['K1'] = 'Операция'
    ws['K1'].style = "ali_grey"
    
    ws['L1'] = 'Заменяемый счетчик'
    ws['L1'].style = "ali_grey"
    
    ws['M1'] = 'Показания заменяемого счетчика'
    ws['M1'].style = "ali_grey"

    ws['N1'] = 'Улица'
    ws['N1'].style = "ali_grey"

    ws['O1'] = 'Дом'
    ws['O1'].style = "ali_grey"

    ws['P1'] = 'Квартира'
    ws['P1'].style = "ali_grey"
#Запрашиваем данные для отчета

           
    
    data_table = []
    data_table = common_sql.get_data_table_report_water_res_by_date(electric_data_end)

    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        for j in range(1,len(data_table[i])):
            if (data_table[i][j] == None) or (data_table[i][j] is None):
                data_table[i][j]=''

# Заполняем отчет значениями
    for row in range(2, len(data_table)+2):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-2][0])  # лицевой счёт
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            a=str(data_table[row-2][1])
            b=datetime.datetime.strptime(a,'%Y-%m-%d')
            d=datetime.datetime.strftime(b,"%d.%m.%Y")
            ws.cell('B%s'%(row)).value = '%s' % d  # начальная дата
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-2][2])  # номер счётчика
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-2][3])  # тип энергоресурса
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-2][4])  # наименование ПУ
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-2][5])  # показания на начальную дату
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            d=electric_data_end
            ws.cell('H%s'%(row)).value = '%s' % d  # дата снятия показаний
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
                        
        try:
            ws.cell('K%s'%(row)).value = 'ВводПоказаний'
            ws.cell('K%s'%(row)).style = "ali_white"
        except:
            ws.cell('K%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table[row-2][8])  # Дом-корпус
            ws.cell('O%s'%(row)).style = "ali_white"
        except:
            ws.cell('O%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table[row-2][7])  # № квартиры
            ws.cell('P%s'%(row)).style = "ali_white"
        except:
            ws.cell('P%s'%(row)).style = "ali_white"
            next
            

#    ws.row_dimensions[5].height = 41
#    ws.column_dimensions['A'].width = 10 
#    ws.column_dimensions['D'].width = 30 
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['j'].width = 15
    ws.column_dimensions['k'].width = 15
                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_water_resources_'+str(electric_data_end)
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_heat_all_by_date(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]

#Шапка
    
    ws['A1'] = 'Лицевой счет'
    ws['A1'].style = "ali_grey"
    
    ws['B1'] = 'Дата начала работы'
    ws['B1'].style = "ali_grey"
    
    ws['C1'] = 'Номер прибора'
    ws['C1'].style = "ali_grey"
    
    ws['D1'] = 'тип прибора'
    ws['D1'].style = "ali_grey"
    
    ws['E1'] = 'Наименование ПУ'
    ws['E1'].style = "ali_grey"
    
    ws['F1'] = 'Показания'
    ws['F1'].style = "ali_grey"
    
    ws['G1'] = 'Дата Окончания' 
    ws['G1'].style = "ali_grey"
    
    ws['H1'] = 'Дата'
    ws['H1'].style = "ali_grey"

    ws['I1'] = 'Квартирный' 
    ws['I1'].style = "ali_grey"
    
    ws['J1'] = 'Адрес'
    ws['J1'].style = "ali_grey"
    
    ws['K1'] = 'Операция'
    ws['K1'].style = "ali_grey"
    
    ws['L1'] = 'Заменяемый счетчик'
    ws['L1'].style = "ali_grey"
    
    ws['M1'] = 'Показания заменяемого счетчика'
    ws['M1'].style = "ali_grey"

    ws['N1'] = 'Улица'
    ws['N1'].style = "ali_grey"

    ws['O1'] = 'Дом'
    ws['O1'].style = "ali_grey"

    ws['P1'] = 'Квартира'
    ws['P1'].style = "ali_grey"
#Запрашиваем данные для отчета

    
    data_table = common_sql.get_data_table_report_heat_res_by_date(electric_data_end)
         
    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        if (data_table[i][5] is None):            
            dt=common_sql.get_data_table_by_date_heat_sayany_for_buhgaltery(data_table[i][0], data_table[i][8], electric_data_end)         
            if (len(dt)>0):                
                data_table[i]=dt[0]
        data_table[i]=tuple(data_table[i])
    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

# Заполняем отчет значениями
    for row in range(2, len(data_table)+2):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-2][0])  # лицевой счёт
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            a=str(data_table[row-2][1])
            b=datetime.datetime.strptime(a,'%Y-%m-%d')
            d=datetime.datetime.strftime(b,"%d.%m.%Y")
            ws.cell('B%s'%(row)).value = '%s' % d  # начальная дата
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-2][2])  # номер счётчика
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-2][3])  # тип энергоресурса
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-2][4])  # наименование ПУ
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-2][5])  # показания на начальную дату
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-2][6]) # дата снятия показаний
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
                        
        try:
            ws.cell('K%s'%(row)).value = 'ВводПоказаний'
            ws.cell('K%s'%(row)).style = "ali_white"
        except:
            ws.cell('K%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table[row-2][8])  # Дом-корпус
            ws.cell('O%s'%(row)).style = "ali_white"
        except:
            ws.cell('O%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table[row-2][7])  # № квартиры
            ws.cell('P%s'%(row)).style = "ali_white"
        except:
            ws.cell('P%s'%(row)).style = "ali_white"
            next
            

#    ws.row_dimensions[5].height = 41
#    ws.column_dimensions['A'].width = 10 
#    ws.column_dimensions['B'].width = 10 
#    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['j'].width = 15
    ws.column_dimensions['k'].width = 15
                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_heat_resources_'+str(electric_data_end)
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_resources_all(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    electric_data_start = request.session["electric_data_start"]
    electric_data_end   = request.session["electric_data_end"]
#Шапка
    ws.merge_cells('A2:G2')
    ws['A2'] = 'Филиград: 1, 2, 3 корпуса. Показания по энергоресурсам за период c '+electric_data_start +' по '+electric_data_end
    
    ws['A5'] = 'Лицевой'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Дата начала'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Номер прибора'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Тип прибора'
    ws['D5'].style = "ali_grey"
    
    #ws['E5'] = 'Показания на '+electric_data_end
    ws['E5'] = 'текущие'
    ws['E5'].style = "ali_grey"
    
    #ws['F5'] = 'Показания на '+electric_data_start
    ws['F5'] = 'предыдущие'
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'разница' 
    ws['G5'].style = "ali_grey"
    
    ws['H5'] = 'Дата'#установки
    ws['H5'].style = "ali_grey"

    ws['I5'] = 'Дата' # съёма
    ws['I5'].style = "ali_grey"
    
    ws['J5'] = 'Квар'
    ws['J5'].style = "ali_grey"

    ws['K5'] = 'Адр'
    ws['K5'].style = "ali_grey"

    ws['L5'] = 'Операция'
    ws['L5'].style = "ali_grey"
    
    ws['M5'] = 'Квар'
    ws['M5'].style = "ali_grey"
    
#Запрашиваем данные для отчета

           
    
    data_table = []
    data_table = common_sql.get_data_table_report_all_res_period3(electric_data_start, electric_data_end)

    #zamenyem None na N/D vezde
    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        for j in range(1,len(data_table[i])):
            if (data_table[i][j] == None) or (data_table[i][j] is None):
                data_table[i][j]=''
        data_table[i]=tuple(data_table[i])


# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # лицевой счёт
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            a=str(data_table[row-6][1])
            b=datetime.datetime.strptime(a,'%Y-%m-%d')
            d=datetime.datetime.strftime(b,"%d.%m.%Y")
            ws.cell('B%s'%(row)).value = '%s' %d  # начальная дата
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # номер счётчика
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # тип энергоресурса
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5])  # показания на конченую дату
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # показания на начальную дату
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][7])  # Дельта
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
            
        try:
            a=str(data_table[row-6][8])
            b=datetime.datetime.strptime(a,'%Y-%m-%d')
            d=datetime.datetime.strftime(b,"%d.%m.%Y")
            ws.cell('H%s'%(row)).value = '%s' %d  # Дата установки приборов
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            a=str(data_table[row-6][9])
            b=datetime.datetime.strptime(a,'%Y-%m-%d')
            d=datetime.datetime.strftime(b,"%d.%m.%Y")
            ws.cell('I%s'%(row)).value = '%s' %d  # Дата конечная
            ws.cell('I%s'%(row)).style = "ali_white"
        except:
            ws.cell('I%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('L%s'%(row)).value = 'ВводПоказаний'  # Операция
            ws.cell('L%s'%(row)).style = "ali_white"
        except:
            ws.cell('L%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('M%s'%(row)).value = '%s' % (data_table[row-6][10])  # Абонент
            ws.cell('M%s'%(row)).style = "ali_white"
        except:
            ws.cell('M%s'%(row)).style = "ali_white"
            next

    ws.row_dimensions[5].height = 41
#    ws.column_dimensions['A'].width = 10 
    ws.column_dimensions['L'].width = 20 
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
#    ws.column_dimensions['j'].width = 15
                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_all_resources_'+ str(electric_data_start)+'_'+str(electric_data_end)
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def pokazaniya_heat_current_report_v2(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Показания теплосчётчиков. Последние считанные данные'
    
    ws['A5'] = 'Дата'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Время'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Абонент'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Заводской номер'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Показания, Гкал'
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 'Показания, м3'
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'Время работы, ч'
    ws['G5'].style = "ali_grey"
    
    ws['H5'] = 'Твхода, С'
    ws['H5'].style = "ali_grey"

    ws['I5'] = 'Твыхода, С'
    ws['I5'].style = "ali_grey"
    
    ws['J5'] = 'Разница Т, С'
    ws['J5'].style = "ali_grey"
    
    ws['K5'] = 'Код ошибки'
    ws['K5'].style = "ali_grey"
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    obj_key             = request.session['obj_key']

    list_except = []
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']

    data_table=[]
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_current_heat_v2(obj_title, parent_name, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_current_heat_v2(obj_title, parent_name, False)
        for row in data_table:
            for x in list_except:
                if x==row[2]:
                    data_table.remove(x)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по расходу воды
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Абонент
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # заводской номер
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания по теплу
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # Показания по расходу воды
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][8])  # заводской номер
            ws.cell('I%s'%(row)).style = "ali_white"
        except:
            ws.cell('I%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][9])  # Показания по теплу
            ws.cell('J%s'%(row)).style = "ali_white"
        except:
            ws.cell('J%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table[row-6][10])  # Показания по расходу воды
            ws.cell('K%s'%(row)).style = "ali_white"
        except:
            ws.cell('K%s'%(row)).style = "ali_white"

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35
                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'heat_current_report' 
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_potreblenie_heat_v2(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление тепловой энергии в период с ' + str(electric_data_start)  + ' по ' + str(electric_data_end)
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Показания на '  + str(electric_data_start)  + ', Гкал'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания на '  + str(electric_data_end) + ', Гкал'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Потребление, Гкал'
    ws['E5'].style = "ali_grey"
    
    ws['f5'] = 'Объём на '  + str(electric_data_start)  + ', м3'
    ws['f5'].style = "ali_grey"
    
    ws['g5'] = 'Объём на '  + str(electric_data_end) + ', м3'
    ws['g5'].style = "ali_grey"
    
    ws['h5'] = 'Потребленный объём, м3'
    ws['h5'].style = "ali_grey"
    
    ws['i5'] = 'Время работы на '  + str(electric_data_start) + ', ч'
    ws['i5'].style = "ali_grey"
    
    ws['j5'] = 'Время работы на '  + str(electric_data_end) + ', ч'
    ws['j5'].style = "ali_grey"
    
    ws['k5'] = 'Время работы с ' + str(electric_data_start) + ' по ' + str(electric_data_end) + ', ч'
    ws['k5'].style = "ali_grey"
    
#    ws['F5'] = 'Время работы с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"]) + ' ,ч'
#    ws['F5'].style = "ali_grey"
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
#    electric_data_end   = request.session['electric_data_end']
#    electric_data_start   = request.session['electric_data_start']                        
    obj_key             = request.session['obj_key']
    #is_electric_monthly = request.session['is_electric_monthly']
    #is_electric_daily   = request.session['is_electric_daily']
    #data_table_end   = []
    #data_table_start = []
    list_except = []
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_for_period_v3(meters_name, parent_name, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_for_period_v3(meters_name, parent_name, electric_data_start, electric_data_end, False)
    else:
        data_table = []

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5])  # Потребление
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # Объём на начало
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][7])  #Объём на конец
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('h%s'%(row)).value = '%s' % (data_table[row-6][8])  # Объём - дельта
            ws.cell('h%s'%(row)).style = "ali_white"
        except:
            ws.cell('h%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('i%s'%(row)).value = '%s' % (data_table[row-6][9])  # Время работы на начао
            ws.cell('i%s'%(row)).style = "ali_white"
        except:
            ws.cell('i%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('j%s'%(row)).value = '%s' % (data_table[row-6][10])  # Время работы на конец 
            ws.cell('j%s'%(row)).style = "ali_white"
        except:
            ws.cell('j%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('k%s'%(row)).value = '%s' % (data_table[row-6][11])  # Время работы - дельта
            ws.cell('k%s'%(row)).style = "ali_white"
        except:
            ws.cell('k%s'%(row)).style = "ali_white"
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
#    ws.column_dimensions['F'].width = 18
#____________
   
#------------

                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    #output_name = u'potreblenie_heat_report'
    output_name = 'potreblenie_heat_report_'+translate(parent_name)+'_'+translate(meters_name)+'_'+electric_data_start+'-'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_water_potreblenie_tekon_hvs(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Текон. Потребление воды ХВС в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_start"])
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Потребление, м3'
    ws['E5'].style = "ali_grey"

#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_start = request.session['electric_data_start']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name, electric_data_start, electric_data_end, 'Канал 1',  'Tekon_hvs', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name,electric_data_start, electric_data_end, 'Канал 1',  'Tekon_hvs', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
#        try:
#            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # Время работы
#            ws.cell('F%s'%(row)).style = "ali_white"
#        except:
#            ws.cell('F%s'%(row)).style = "ali_white"
#            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
#    ws.column_dimensions['F'].width = 18
#____________
   
#------------

                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'potreblenie_water_tekon_hvs_report'
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_water_tekon_hvs(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Текон. Потребление воды ХВС на ' +str(request.session["electric_data_end"])
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['C5'].style = "ali_grey"
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']    
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, 'Канал 1',  'Tekon_hvs', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, 'Канал 1',  'Tekon_hvs', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35

#    ws.column_dimensions['F'].width = 18
#____________
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'pokazaniya_na_datu_water_tekon_hvs_report'
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_water_potreblenie_tekon_gvs(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Текон. Потребление воды ГВС в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_start"])
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Потребление, м3'
    ws['E5'].style = "ali_grey"
    
#    ws['F5'] = 'Время работы с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"]) + ' ,ч'
#    ws['F5'].style = "ali_grey"
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_start = request.session['electric_data_start']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name, electric_data_start, electric_data_end, 'Канал 2',  'Tekon_gvs', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name,electric_data_start, electric_data_end, 'Канал 2',  'Tekon_gvs', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
#        try:
#            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # Время работы
#            ws.cell('F%s'%(row)).style = "ali_white"
#        except:
#            ws.cell('F%s'%(row)).style = "ali_white"
#            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
#    ws.column_dimensions['F'].width = 18
#____________
   
#------------

                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'potreblenie_water_tekon_gvs_report'
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_tekon_heat_potreblenie(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Текон. Потребление воды ГВС в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_start"])
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Потребление, м3'
    ws['E5'].style = "ali_grey"
    
#    ws['F5'] = 'Время работы с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"]) + ' ,ч'
#    ws['F5'].style = "ali_grey"
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_start = request.session['electric_data_start']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name, electric_data_start, electric_data_end, 'Канал 3',  'Tekon_heat', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name,electric_data_start, electric_data_end, 'Канал 3',  'Tekon_heat', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
#        try:
#            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # Время работы
#            ws.cell('F%s'%(row)).style = "ali_white"
#        except:
#            ws.cell('F%s'%(row)).style = "ali_white"
#            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
#    ws.column_dimensions['F'].width = 18
#____________
   
#------------

                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'potreblenie_heat_tekon_report'
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_water_tekon_gvs(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Текон. Потребление воды ГВС на ' +str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['C5'].style = "ali_grey"

    
#    ws['F5'] = 'Время работы с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"]) + ' ,ч'
#    ws['F5'].style = "ali_grey"
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']    
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, 'Канал 2',  'Tekon_gvs', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, 'Канал 2',  'Tekon_gvs', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35

#    ws.column_dimensions['F'].width = 18
#____________
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'pokazaniya_na_datu_water_tekon_gvs_'+translate(parent_name)+'_'+translate(meters_name)+'_'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_tekon_heat_by_date(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Текон. Потребление воды ГВС на ' +str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['C5'].style = "ali_grey"

    
#    ws['F5'] = 'Время работы с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"]) + ' ,ч'
#    ws['F5'].style = "ali_grey"
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']    
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, 'Канал 3',  'Tekon_heat', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, 'Канал 3',  'Tekon_heat', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35

#    ws.column_dimensions['F'].width = 18
#____________
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'pokazaniya_na_datu_heat_tekon_'+translate(parent_name)+'_'+translate(meters_name)+'_'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_water_by_date(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    meters_name         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')
#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = meters_name+'. Показания по воде на ' + electric_data_end
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Номер счётчика'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Пульсар'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Канал'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Показания'
    ws['E5'].style = "ali_grey"
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'level2')
    is_object_level_2 = re.compile(r'level1')
    
    parent_name         = request.GET.get('obj_parent_title')
    #meters_name         = request.session['obj_title']
    electric_data_end   = request.GET.get('electric_data_end')           
    obj_key             = request.GET.get('obj_key')
    dc='daily'
    data_table = []
    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_water_by_date(meters_name, parent_name, electric_data_end, True, dc)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_water_by_date(meters_name, parent_name, electric_data_end, False, dc)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # пульсар
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # канал
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % get_val(data_table[row-6][5])  # Показания
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
            
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'water_'+translate(parent_name)+'_'+translate(meters_name)+'_'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

    
def report_forma_80020(request):
    import zipfile
    response = io.BytesIO()
    #Запрашиваем данные для отчета
    
    group_80020_name    = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']                        
   

    # Формируем список дат на основе начальной и конечной даты полученной от web-календарей
    end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
    start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
    list_of_dates = [x for x in common_sql.daterange(start_date,
                  end_date,
                  step=datetime.timedelta(days=1),
                  inclusive=True)]
    
    info_group_80020 = common_sql.get_info_group_80020(group_80020_name)
    inn_sender_from_base      = info_group_80020[0][0]
    name_sender_from_base     = info_group_80020[0][1]
    inn_postavshik_from_base  = info_group_80020[0][2]
    name_postavshik_from_base = info_group_80020[0][3]
    dogovor_number_from_base  = info_group_80020[0][4]
    
    #Узнаем GUID счётчиков, которые входят в группу 80020
    meters_guid_list = common_sql.get_meters_guid_list_by_group_name(group_80020_name)

   
    #Создаем архив
    zf = zipfile.ZipFile(response, mode='w', compression=zipfile.ZIP_DEFLATED)
    
    for dates in range(len(list_of_dates)):
    #Формируем файл xml по форме Мосэнергосбыт 80020   
        
        # Создание корневого элемента message
        root = etree.Element('message', {'class': '80020', 'version': '2', 'number': '1' })
        
        # Добавление дочерних элементов - <datetime> <sender> <area> в <root>
        datetimeElt = etree.SubElement(root, 'datetime')
        senderElt = etree.SubElement(root, 'sender')
        areaElt = etree.SubElement(root, 'area')
        
        # Присваиваем значения в <day> <timestamp> <daylightsavingtime>
        day = etree.SubElement(datetimeElt, 'day')
        timestamp = etree.SubElement(datetimeElt, 'timestamp')
        daylightsavingtime = etree.SubElement(datetimeElt, 'daylightsavingtime')
        
        day.text = str(list_of_dates[dates].strftime('%Y%m%d'))
        timestamp.text = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
        daylightsavingtime.text = '0'
        
        # Присваиваем значения <name> <inn> для <sender>
        inn_sender = etree.SubElement(senderElt, 'inn')
        name_sender = etree.SubElement(senderElt, 'name')
        
        inn_sender.text  = inn_sender_from_base
        name_sender.text = name_sender_from_base
        
        # Присваиваем значения для <area>
        inn_area = etree.SubElement(areaElt, 'inn')
        name_abonent_area = etree.SubElement(areaElt, 'name_abonent')
        
        inn_area.text = inn_postavshik_from_base
        name_abonent_area.text = name_postavshik_from_base
        
        # Добавление дочерних элементов
        for x in range(len(meters_guid_list)):
            info_measuring_point = common_sql.get_info_measuring_point_in_group_80020(meters_guid_list[x])
            my_measure_code = str(info_measuring_point[0][0])
            my_measure_name = str(info_measuring_point[0][1])
            measurepointElt = etree.SubElement(areaElt, 'measuringpoint', code = my_measure_code , name = my_measure_name )
        
            list_of_taken_params = []
            #Получаем считываемые параметры по guid счётчика.
             #A+
            name_of_type_meters = common_sql.get_name_of_type_meter_by_guid(meters_guid_list[x])
            if name_of_type_meters[0][0] == u'Меркурий 230-УМ':
                guid_params = u'922ad57c-8f5e-4f00-a78d-e3ba89ef859f'
            elif name_of_type_meters[0][0] == u'Меркурий 230':
                guid_params = u'6af9ddce-437a-4e07-bd70-6cf9dcc10b31'
            elif name_of_type_meters[0][0] == u'СЭТ-4ТМ.03М':
                guid_params = u'4f505e17-7d71-4cf8-9880-c6ce33612e6e'
            else:
                pass
            result = common_sql.get_taken_param_by_guid_meters_and_guid_params(meters_guid_list[x], guid_params)
            result_list = []
            result_list.append(str(result[0][0]))
            result_list.append(str(result[0][1]))
            list_of_taken_params.append(result_list)
            #R+
            if name_of_type_meters[0][0] == u'Меркурий 230-УМ':
                 guid_params = u'61101fa3-a96a-4934-9482-e32036c12829'
            elif name_of_type_meters[0][0] == u'Меркурий 230':
                 guid_params = u'66e997c0-8128-40a7-ae65-7e8993fbea61'
            elif name_of_type_meters[0][0] == u'СЭТ-4ТМ.03М':
                 guid_params = u'55abd40d-fb3c-4100-88f2-46d79be7733a'
            else:
                pass
            
            result = common_sql.get_taken_param_by_guid_meters_and_guid_params(meters_guid_list[x], guid_params)
            result_list = []
            result_list.append(str(result[0][0]))
            result_list.append(str(result[0][1]))
            list_of_taken_params.append(result_list)
            
            for y in range(len(list_of_taken_params)):
                my_dict_of_profil = {0: ['0000', '0030', 0, 1],
                                     1: ['0030', '0100', 0, 1],
                                     2: ['0100', '0130', 0, 1],
                                     3: ['0130', '0200', 0, 1],
                                     4: ['0200', '0230', 0, 1],
                                     5: ['0230', '0300', 0, 1],
                                     6: ['0300', '0330', 0, 1],
                                     7: ['0330', '0400', 0, 1],
                                     8: ['0400', '0430', 0, 1],
                                     9: ['0430', '0500', 0, 1],
                                     10: ['0500', '0530', 0, 1],
                                     11: ['0530', '0600', 0, 1],
                                     12: ['0600', '0630', 0, 1],
                                     13: ['0630', '0700', 0, 1],
                                     14: ['0700', '0730', 0, 1],
                                     15: ['0730', '0800', 0, 1],
                                     16: ['0800', '0830', 0, 1],
                                     17: ['0830', '0900', 0, 1],
                                     18: ['0900', '0930', 0, 1],
                                     19: ['0930', '1000', 0, 1],
                                     20: ['1000', '1030', 0, 1],
                                     21: ['1030', '1100', 0, 1],
                                     22: ['1100', '1130', 0, 1],                                
                                     23: ['1130', '1200', 0, 1],
                                     24: ['1200', '1230', 0, 1],
                                     25: ['1230', '1300', 0, 1],
                                     26: ['1300', '1330', 0, 1],
                                     27: ['1330', '1400', 0, 1],
                                     28: ['1400', '1430', 0, 1],
                                     29: ['1430', '1500', 0, 1],
                                     30: ['1500', '1530', 0, 1],
                                     31: ['1530', '1600', 0, 1],
                                     32: ['1600', '1630', 0, 1],
                                     33: ['1630', '1700', 0, 1],
                                     34: ['1700', '1730', 0, 1],
                                     35: ['1730', '1800', 0, 1],
                                     36: ['1800', '1830', 0, 1],
                                     37: ['1830', '1900', 0, 1],
                                     38: ['1900', '1930', 0, 1],
                                     39: ['1930', '2000', 0, 1],
                                     40: ['2000', '2030', 0, 1],                                
                                     41: ['2030', '2100', 0, 1],
                                     42: ['2100', '2130', 0, 1],
                                     43: ['2130', '2200', 0, 1],
                                     44: ['2200', '2230', 0, 1],
                                     45: ['2230', '2300', 0, 1],
                                     46: ['2300', '2330', 0, 1],
                                     47: ['2330', '0000', 0, 1] }
                time_table = ['00:00:00', '00:30:00',
                              '01:00:00', '01:30:00',
                              '02:00:00', '02:30:00',
                              '03:00:00', '03:30:00',
                              '04:00:00', '04:30:00',
                              '05:00:00', '05:30:00',
                              '06:00:00', '06:30:00',
                              '07:00:00', '07:30:00',
                              '08:00:00', '08:30:00',
                              '09:00:00', '09:30:00',
                              '10:00:00', '10:30:00',
                              '11:00:00', '11:30:00',
                              '12:00:00', '12:30:00',
                              '13:00:00', '13:30:00',
                              '14:00:00', '14:30:00',
                              '15:00:00', '15:30:00',
                              '16:00:00', '16:30:00',
                              '17:00:00', '17:30:00',
                              '18:00:00', '18:30:00',
                              '19:00:00', '19:30:00',
                              '20:00:00', '20:30:00',
                              '21:00:00', '21:30:00',
                              '22:00:00', '22:30:00',
                              '23:00:00', '23:30:00']
                if list_of_taken_params[y][1] == 'A+ Профиль':
                    my_measuring_channel_code = '01'
                    for time in range(len(time_table)):
                        result_30_min = common_sql.get_30_min_value_by_meters_number_param_names_and_datetime(list_of_taken_params[y][0], list_of_taken_params[y][1], list_of_dates[dates].strftime('%Y-%m-%d'), time_table[time])
                        if result_30_min:
                            my_dict_of_profil[time]=[my_dict_of_profil[time][0], my_dict_of_profil[time][1], float(result_30_min[0][4])*common_sql.get_k_t_t_by_factory_number_manual(list_of_taken_params[y][0])*common_sql.get_k_t_n_by_factory_number_manual(list_of_taken_params[y][0]), 0] #Квт.ч
                            #print u'------ Есть значение',list_of_taken_params[y][1] , my_dict_of_profil[time][0], my_dict_of_profil[time][1]
                            
                        else:
                            #print u'Оставляем по умолчанию', list_of_taken_params[y][1], my_dict_of_profil[time][0], my_dict_of_profil[time][1]
                            pass
                    
                elif list_of_taken_params[y][1] == 'R+ Профиль':
                    my_measuring_channel_code = '03'
                    for time in range(len(time_table)):
                        result_30_min = common_sql.get_30_min_value_by_meters_number_param_names_and_datetime(list_of_taken_params[y][0], list_of_taken_params[y][1], list_of_dates[dates].strftime('%Y-%m-%d'), time_table[time])
                        if result_30_min:
                            my_dict_of_profil[time]=[my_dict_of_profil[time][0], my_dict_of_profil[time][1], float(result_30_min[0][4])*common_sql.get_k_t_t_by_factory_number_manual(list_of_taken_params[y][0])*common_sql.get_k_t_n_by_factory_number_manual(list_of_taken_params[y][0]), 0] # Квар.ч
                            #print u'------ Есть значение',list_of_taken_params[y][1], my_dict_of_profil[time][0], my_dict_of_profil[time][1]
                            
                        else:
                           # print u'Оставляем по умолчанию', list_of_taken_params[y][1], my_dict_of_profil[time][0], my_dict_of_profil[time][1]
                           pass
                else:
                    #print u'Нет совпадений параметров'
                    my_measuring_channel_code = ''
                       
                my_measuring_channel_desc = str(list_of_taken_params[y][0]) + ' ' + str(list_of_taken_params[y][1])            
                measuringchannelElt = etree.SubElement(measurepointElt, 'measuringchannel', code = my_measuring_channel_code, desc = my_measuring_channel_desc )
                
                                                  
                for z in range (len(my_dict_of_profil)):
                    periodElt = etree.SubElement(measuringchannelElt, 'period', start = str(my_dict_of_profil[z][0]), end = str(my_dict_of_profil[z][1]))
                    value  = etree.SubElement(periodElt, 'value', status = str(my_dict_of_profil[z][3]))
                    value.text = str(my_dict_of_profil[z][2])
            
        
        # Создание и сохранение документа
        doc = etree.ElementTree(root) 
        myxml_IO=io.BytesIO()   
        doc.write(myxml_IO, xml_declaration=True, encoding='UTF-8')
        # Формируем имя документа
        name_of_document = '80020'
        name_of_file_80020 =name_of_document + '_'+ inn_sender_from_base + '_' + str(list_of_dates[dates].strftime('%Y%m%d')) + '_1' + '.xml'
        zf.writestr(name_of_file_80020, myxml_IO.getvalue())
        print(name_of_file_80020)
    zf.close()
    
    response=HttpResponse(response.getvalue())
    response['Content-Type'] = 'application/x-zip-compressed'
    
    output_name = '80020_' + str(dogovor_number_from_base) + '_' + start_date.strftime('%Y%m%d') + '-'+ end_date.strftime('%Y%m%d')
    file_ext = 'zip'
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
  
    return response

def report_forma_80020_fast(request):
    """Отчёт для формирования формы 80020 для МосЭнергоСбыта"""

    #TODO: сделать логирование формирования данного отчёта
    #print("Вызвали отчёт быстрого формирования 80020")

    # Для подсчёта времени работы отчёта
    #time_start_report_log = pt.time()

    response = io.BytesIO()
    
    #Запрашиваем данные для отчета
    group_80020_name    = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']                        
   
    # Формируем список дат на основе начальной и конечной даты полученной от web-календарей
    end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
    start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
    list_of_dates = [x for x in common_sql.daterange(start_date,
                  end_date,
                  step=datetime.timedelta(days=1),
                  inclusive=True)]
    
    info_group_80020 = common_sql.get_info_group_80020(group_80020_name)
    inn_sender_from_base      = info_group_80020[0][0]
    name_sender_from_base     = info_group_80020[0][1]
    inn_postavshik_from_base  = info_group_80020[0][2]
    name_postavshik_from_base = info_group_80020[0][3]
    dogovor_number_from_base  = info_group_80020[0][4]
    
    #Узнаем GUID счётчиков, которые входят в группу 80020
    meters_guid_list = common_sql.get_meters_guid_list_by_group_name(group_80020_name)
      
    #Создаем архив
    zf = zipfile.ZipFile(response, mode='w', compression=zipfile.ZIP_DEFLATED)

    # Проходимся по датам... на каждую дату создаём свой xml файл и у паковываем его в архив
    for dates in range(len(list_of_dates)):
    #Формируем файл xml по форме Мосэнергосбыт 80020   
        
        # Создание корневого элемента message
        root = etree.Element('message', {'class': '80020', 'version': '2', 'number': '1' })
        
        # Добавление дочерних элементов - <datetime> <sender> <area> в <root>
        datetimeElt = etree.SubElement(root, 'datetime')
        senderElt = etree.SubElement(root, 'sender')
        areaElt = etree.SubElement(root, 'area')
        
        # Присваиваем значения в <day> <timestamp> <daylightsavingtime>
        day = etree.SubElement(datetimeElt, 'day')
        timestamp = etree.SubElement(datetimeElt, 'timestamp')
        daylightsavingtime = etree.SubElement(datetimeElt, 'daylightsavingtime')
        
        day.text = str(list_of_dates[dates].strftime('%Y%m%d'))
        timestamp.text = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
        daylightsavingtime.text = '0'
        
        # Присваиваем значения <name> <inn> для <sender>
        inn_sender = etree.SubElement(senderElt, 'inn')
        name_sender = etree.SubElement(senderElt, 'name')
        
        inn_sender.text  = inn_sender_from_base
        name_sender.text = name_sender_from_base
        
        # Присваиваем значения для <area>
        inn_area = etree.SubElement(areaElt, 'inn')
        name_abonent_area = etree.SubElement(areaElt, 'name_abonent')
        
        inn_area.text = inn_postavshik_from_base
        name_abonent_area.text = name_postavshik_from_base


        # Добавление дочерних элементов
        
        for meter in range(len(meters_guid_list)):
            # Запрашиваем таблицу с данными по профилям мощности за одни сутки по guid счётчика
            new_data_table_of_a_plus_r_plus = common_sql.get_30_by_meter_for_period(meters_guid_list[meter][0], list_of_dates[dates].date(), list_of_dates[dates].date())
            
            # Присваиваем значения code и name для счётчика
            my_measure_code = str(new_data_table_of_a_plus_r_plus[0][8])
            my_measure_name = str(new_data_table_of_a_plus_r_plus[0][9])
            
            # Создаём элемент в xml для счётчика
            measurepointElt = etree.SubElement(areaElt, 'measuringpoint', code = my_measure_code , name = my_measure_name )
        

            # Каналы
            # A+ профиль-------------------------------------------------------------------------------        
            my_measuring_channel_desc = str(my_measure_name + ' А+ Профиль ')            
            measuringchannelElt = etree.SubElement(measurepointElt, 'measuringchannel', code = '01', desc = my_measuring_channel_desc )
            
            # Диапазон времени с 00:00 до 23:30                                    
            for z in range(47):
                if new_data_table_of_a_plus_r_plus[z][11] is not None:
                    periodElt = etree.SubElement(measuringchannelElt, 'period', start = str(new_data_table_of_a_plus_r_plus[z][10]).replace(':',''), end = str(new_data_table_of_a_plus_r_plus[z+1][10]).replace(':',''))
                    value  = etree.SubElement(periodElt, 'value', status = '0')
                    value.text = str(new_data_table_of_a_plus_r_plus[z][11])
                else:
                    periodElt = etree.SubElement(measuringchannelElt, 'period', start = str(new_data_table_of_a_plus_r_plus[z][10]).replace(':',''), end = str(new_data_table_of_a_plus_r_plus[z+1][10]).replace(':',''))
                    value  = etree.SubElement(periodElt, 'value', status = '1')
                    value.text = "0"
                    value.attrib['status'] = '1'


            # Диапазон времени с 23:30 до 00:00  
            if new_data_table_of_a_plus_r_plus[47][11] is not None:
                periodElt = etree.SubElement(measuringchannelElt, 'period', start = str(new_data_table_of_a_plus_r_plus[47][10]).replace(':',''), end = str(new_data_table_of_a_plus_r_plus[0][10]).replace(':','') )
                value  = etree.SubElement(periodElt, 'value', status = '0')
                value.text = str(new_data_table_of_a_plus_r_plus[47][11])
            else:
                periodElt = etree.SubElement(measuringchannelElt, 'period', start = str(new_data_table_of_a_plus_r_plus[47][10]).replace(':',''), end = str(new_data_table_of_a_plus_r_plus[0][10]).replace(':','') )
                value  = etree.SubElement(periodElt, 'value', status = '1')
                value.text = "0"
                value.attrib['status'] = '1'
            
            # R+ профиль-------------------------------------------------------------------------------
            my_measuring_channel_desc = str(my_measure_name + ' R+ Профиль ')            
            measuringchannelElt = etree.SubElement(measurepointElt, 'measuringchannel', code = '03', desc = my_measuring_channel_desc )
            
            # Диапазон времени с 00:00 до 23:30                                    
            for z in range(47):
                if new_data_table_of_a_plus_r_plus[z][11] is not None:
                    periodElt = etree.SubElement(measuringchannelElt, 'period', start = str(new_data_table_of_a_plus_r_plus[z][10]).replace(':',''), end = str(new_data_table_of_a_plus_r_plus[z+1][10]).replace(':',''))
                    value  = etree.SubElement(periodElt, 'value', status = '0')
                    value.text = str(new_data_table_of_a_plus_r_plus[z][12])
                else:
                    periodElt = etree.SubElement(measuringchannelElt, 'period', start = str(new_data_table_of_a_plus_r_plus[z][10]).replace(':',''), end = str(new_data_table_of_a_plus_r_plus[z+1][10]).replace(':',''))
                    value  = etree.SubElement(periodElt, 'value', status = '1')
                    value.text = "0"
                    value.attrib['status'] = '1'


            # Диапазон времени с 23:30 до 00:00  
            if new_data_table_of_a_plus_r_plus[47][11] is not None:
                periodElt = etree.SubElement(measuringchannelElt, 'period', start = str(new_data_table_of_a_plus_r_plus[47][10]).replace(':',''), end = str(new_data_table_of_a_plus_r_plus[0][10]).replace(':','') )
                value  = etree.SubElement(periodElt, 'value', status = '0')
                value.text = str(new_data_table_of_a_plus_r_plus[47][12])
            else:
                periodElt = etree.SubElement(measuringchannelElt, 'period', start = str(new_data_table_of_a_plus_r_plus[47][10]).replace(':',''), end = str(new_data_table_of_a_plus_r_plus[0][10]).replace(':','') )
                value  = etree.SubElement(periodElt, 'value', status = '1')
                value.text = "0"
                value.attrib['status'] = '1'            
        
        # Создание и сохранение документа
        doc = etree.ElementTree(root) 
        myxml_IO=io.BytesIO()   
        doc.write(myxml_IO, xml_declaration=True, encoding='UTF-8')

        # Формируем имя документа
        name_of_document = '80020'
        name_of_file_80020 =name_of_document + '_'+ inn_sender_from_base + '_' + str(list_of_dates[dates].strftime('%Y%m%d')) + '_1' + '.xml'
        zf.writestr(name_of_file_80020, myxml_IO.getvalue())
        
        #TODO: сделать информатор для пользователя, сколько файлов уже готово...   
        #print(name_of_file_80020)
    zf.close()
    
    #Подсчёт времени работы отчёта
    #print(time_start_report_log - pt.time())

    response=HttpResponse(response.getvalue())
    response['Content-Type'] = 'application/x-zip-compressed'
    
    #Формируем имя выгружаемого архива с получасовками
    output_name = '80020_' + str(dogovor_number_from_base) + '_' + start_date.strftime('%Y%m%d') + '-'+ end_date.strftime('%Y%m%d')
    file_ext = 'zip'
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
  
    return response


def report_elf_hvs_by_date(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Эльф. Потребление воды ХВС на ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик Эльф'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Счётчик ХВС'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['D5'].style = "ali_grey"
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         =  request.session['obj_parent_title']
    obj_title         =  request.session['obj_title']
    electric_data_end   =  request.session['electric_data_end']            
    obj_key             =  request.session['obj_key']
    
    data_table = []
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_elf(obj_title, obj_parent_title , electric_data_end, '1','attr1', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_daily_water_elf(obj_title, obj_parent_title , electric_data_end, '1', 'attr1',False)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # счётчик эльф
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # счётчик хвс
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания 
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next


    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17

#------------

                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'elf_hvs_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_end
    print(output_name)
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    
    return response
    
def report_elf_hvs_potreblenie(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление воды ХВС в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик Эльф'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Счётчик ХВС'
    ws['C5'].style = "ali_grey"
        
    ws['D5'] = 'Показания на ' + str(request.session["electric_data_start"])
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Показания на '  + str(request.session["electric_data_end"]) 
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 'Потребление'
    ws['F5'].style = "ali_grey"

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    electric_data_start = request.session['electric_data_start']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_elf_period(obj_title, obj_parent_title , electric_data_end, electric_data_start,'1','attr1', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_daily_water_elf_period(obj_title, obj_parent_title , electric_data_end, electric_data_start,'1','attr1', False)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][3])  # Потребление
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Время работы
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 18
#____________
   
#------------

                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'potreblenie_elf_hvs_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_start+'_'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_elf_gvs_potreblenie(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление воды ГВС в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик Эльф'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Счётчик ГВС'
    ws['C5'].style = "ali_grey"
        
    ws['D5'] = 'Показания на '  + str(request.session["electric_data_start"])
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Показания на '  +  str(request.session["electric_data_end"])
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 'Потребление'
    ws['F5'].style = "ali_grey"

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    electric_data_start = request.session['electric_data_start']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_elf_period(obj_title, obj_parent_title , electric_data_end, electric_data_start,'2','attr2', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_daily_water_elf_period(obj_title, obj_parent_title , electric_data_end, electric_data_start,'2','attr2', False)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][3])  # Потребление
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Время работы
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 18
#____________
   
#------------

                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'potreblenie_elf_gvs_report'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_start+'_'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_elf_gvs_by_date(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Эльф. Потребление воды ГВС на ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик Эльф'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Счётчик ГВС'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['D5'].style = "ali_grey"
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         =  request.session['obj_parent_title']
    obj_title         =  request.session['obj_title']
    electric_data_end   =  request.session['electric_data_end']            
    obj_key             =  request.session['obj_key']
    
    data_table = []
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_elf(obj_title, obj_parent_title , electric_data_end, '2','attr2', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_daily_water_elf(obj_title, obj_parent_title , electric_data_end, '2', 'attr2',False)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # счётчик эльф
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # счётчик гвс
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания 
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next


    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17

#------------

                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'elf_gvs_report_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pulsar_water_period(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Пульсар. Потребление воды на ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Тип счётчика'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Стояк'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Счётчик'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Показания на '  + str(request.session["electric_data_start"])+', м3'
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 'Показания на '  + str(request.session["electric_data_end"])+', м3'
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'Разница, м3'
    ws['G5'].style = "ali_grey"

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         =  request.session['obj_parent_title']
    obj_title         =  request.session['obj_title']
    electric_data_end   =  request.session['electric_data_end']   
    electric_data_start  =  request.session['electric_data_start']            
    obj_key             =  request.session['obj_key']
    
    data_table = []
                     
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_for_period(obj_parent_title, obj_title, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_for_period(obj_parent_title, obj_title, electric_data_start,electric_data_end, False)

    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # тип
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # стояк
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # счётчик 
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # значения на начало
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # значения на конец
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # дельта
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next


    ws.row_dimensions[5].height = 63
    ws.column_dimensions['A'].width = 17 
#    ws.column_dimensions['B'].width = 17 
#    ws.column_dimensions['C'].width = 17
#    ws.column_dimensions['D'].width = 17

#------------

                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'pulsar_water_period_report_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_start+'-'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pulsar_water_daily(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Пульсар. Потребление воды на ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Тип счётчика'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Стояк'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Счётчик'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Показания на '  + str(request.session["electric_data_end"])+', м3'
    ws['E5'].style = "ali_grey"
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         =  request.session['obj_parent_title']
    obj_title         =  request.session['obj_title']
    electric_data_end   =  request.session['electric_data_end']            
    obj_key             =  request.session['obj_key']
    
    data_table = []
    
    
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_daily(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_daily(obj_parent_title, obj_title, electric_data_end, False)

    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # тип
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # стояк
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # счётчик
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5])  # показаня
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next


    ws.row_dimensions[5].height = 63
    ws.column_dimensions['A'].width = 23 
#    ws.column_dimensions['B'].width = 17 
#    ws.column_dimensions['C'].width = 17
#    ws.column_dimensions['D'].width = 17

#------------
                   
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'pulsar_water_report_'+translate(obj_title)+'_'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pulsar_water_daily_row(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]

#Шапка
    ws['A1'] = 'Абонент'
    ws['A1'].style = "ali_grey"
    ws.merge_cells('A1:A2')
    
    ws['B1'] = 'Стояк 1'
    ws['B1'].style = "ali_grey"
    ws.merge_cells('B1:E1')
    
    ws['B2'] = 'Счётчик ХВС'
    ws['B2'].style = "ali_grey"
    
    ws['C2'] = 'Значение ХВС, м3'
    ws['C2'].style = "ali_grey"   
    
    
    ws['D2'] = 'Счётчик ГВС'
    ws['D2'].style = "ali_grey"
    
    ws['E2'] = 'Значение ГВС, м3'
    ws['E2'].style = "ali_grey"
    
    
    ws['F1'] = 'Стояк 2'
    ws['F1'].style = "ali_grey"
    ws.merge_cells('F1:I1')
    
    ws['F2'] = 'Счётчик ХВС'
    ws['F2'].style = "ali_grey"
    
    ws['G2'] = 'Значение ХВС, м3'
    ws['G2'].style = "ali_grey"   
    
    
    ws['H2'] = 'Счётчик ГВС'
    ws['H2'].style = "ali_grey"
    
    ws['I2'] = 'Значение ГВС, м3'
    ws['I2'].style = "ali_grey"
    
    
    ws['J1'] = 'Стояк 3'
    ws['J1'].style = "ali_grey"
    ws.merge_cells('J1:M1')
    
    ws['J2'] = 'Счётчик ХВС'
    ws['J2'].style = "ali_grey"
    
    ws['K2'] = 'Значение ХВС, м3'
    ws['K2'].style = "ali_grey"   
    
    
    ws['L2'] = 'Счётчик ГВС'
    ws['L2'].style = "ali_grey"
    
    ws['M2'] = 'Значение ГВС, м3'
    ws['M2'].style = "ali_grey"
    
    ws['N1'] = 'Сумма ХВС, м3'
    ws['N1'].style = "ali_grey"
    ws.merge_cells('N1:N2')
    
    ws['O1'] = 'Сумма ГВС, м3'
    ws['O1'].style = "ali_grey"
    ws.merge_cells('O1:O2')
#Запрашиваем данные для отчета

    
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
             
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_daily_row(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_daily_row(obj_parent_title, obj_title, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull_for_pulsar(data_table)

    
# Заполняем отчет значениями
    for row in range(3, len(data_table)+3):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-3][1])  # абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-3][2])  # nomer hvs
            ws.cell('B%s'%(row)).style = "ali_blue"
        except:
            ws.cell('B%s'%(row)).style = "ali_blue"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-3][3])  # value hvs
            ws.cell('C%s'%(row)).style = "ali_blue"
        except:
            ws.cell('C%s'%(row)).style = "ali_blue"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-3][4])  # gvs num
            ws.cell('D%s'%(row)).style = "ali_pink"
        except:
            ws.cell('D%s'%(row)).style = "ali_pink"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-3][5])  # value gvs
            ws.cell('E%s'%(row)).style = "ali_pink"
        except:
            ws.cell('E%s'%(row)).style = "ali_pink"
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-3][6])  # num hvs
            ws.cell('F%s'%(row)).style = "ali_blue"
        except:
            ws.cell('F%s'%(row)).style = "ali_blue"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-3][7]) # val hvs
            ws.cell('G%s'%(row)).style = "ali_blue"
        except:
            ws.cell('G%s'%(row)).style = "ali_blue"
            next
                        
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-3][8])  # gvs num
            ws.cell('H%s'%(row)).style = "ali_pink"
        except:
            ws.cell('H%s'%(row)).style = "ali_pink"
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-3][9])  # gvs val
            ws.cell('I%s'%(row)).style = "ali_pink"
        except:
            ws.cell('I%s'%(row)).style = "ali_pink"
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-3][10])  # hvs num
            ws.cell('J%s'%(row)).style = "ali_blue"
        except:
            ws.cell('J%s'%(row)).style = "ali_blue"
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table[row-3][11])  # hvs val
            ws.cell('K%s'%(row)).style = "ali_blue"
        except:
            ws.cell('K%s'%(row)).style = "ali_blue"
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table[row-3][12])  # hvs num
            ws.cell('L%s'%(row)).style = "ali_pink"
        except:
            ws.cell('L%s'%(row)).style = "ali_pink"
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % (data_table[row-3][13])  # hvs val
            ws.cell('M%s'%(row)).style = "ali_pink"
        except:
            ws.cell('M%s'%(row)).style = "ali_pink"
            next
            
        try:
            ws.cell('N%s'%(row)).value = '%s' % (data_table[row-3][14])  # hvs num
            ws.cell('N%s'%(row)).style = "ali_blue"
        except:
            ws.cell('N%s'%(row)).style = "ali_blue"
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table[row-3][15])  # hvs val
            ws.cell('O%s'%(row)).style = "ali_pink"
        except:
            ws.cell('O%s'%(row)).style = "ali_pink"
            next

#    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
        
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_water_pulsar_row_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+str(electric_data_end)
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pulsar_heat_daily(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Пульсар. Показания по теплу на ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Энергия, Гкал'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Объем, м3'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Температура входа, С'
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 'Температура выхода, С'
    ws['F5'].style = "ali_grey"
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         =  request.session['obj_parent_title']
    obj_title         =  request.session['obj_title']
    electric_data_end   =  request.session['electric_data_end']            
    obj_key             =  request.session['obj_key']
    
    data_table = []
                     
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_daily_pulsar_teplo(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_daily_pulsar_teplo(obj_parent_title, obj_title, electric_data_end, False)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # тип
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % get_val(data_table[row-6][3])  # стояк
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value =  '%s' % get_val(data_table[row-6][4])  # счётчик
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % get_val(data_table[row-6][5])  # показаня
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % get_val(data_table[row-6][6])  # показаня
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next

    ws.row_dimensions[5].height = 63
    ws.column_dimensions['A'].width = 20 
    ws.column_dimensions['C'].width = 23 
    ws.column_dimensions['E'].width = 17 
    ws.column_dimensions['F'].width = 17
    ws.column_dimensions['D'].width = 17

#------------
            
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'pulsar_heat_report_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pulsar_heat_period(request):
    #SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')
    ROUND_SIZE = getattr(settings, 'ROUND_SIZE', 3)
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]
    electric_data_start   = request.session["electric_data_start"]

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Пульсар. Потребление тепла с ' +str(electric_data_start)+' по '+ str(electric_data_end)
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Показания Энергии на ' + str(electric_data_start)+', Гкал '
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания Энергии на ' + str(electric_data_end)+', Гкал '
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Потребление Энергии, Гкал'
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 'Показания Объёма на ' + str(electric_data_start)+', м3'
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'Показания Объёма на ' + str(electric_data_end)+', м3'
    ws['G5'].style = "ali_grey"
    
    ws['H5'] = 'Потребление Объёма, м3'
    ws['H5'].style = "ali_grey"
#Запрашиваем данные для отчета

    
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
          
    obj_key             = request.session['obj_key']
             
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title,electric_data_end, electric_data_start, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title, electric_data_end,electric_data_start, False)
               
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][2]),ROUND_SIZE, separator)  # '%s' % get_val(round(float(data_table[row-6][2]),7)) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][3]),ROUND_SIZE, separator)  # '%s' % get_val(round(float(data_table[row-6][3]),7)) # '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][4]),ROUND_SIZE, separator)  # '%s' % get_val(round(float(data_table[row-6][4]),7))  # '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][5]),ROUND_SIZE, separator)  #  '%s' % get_val(round(float(data_table[row-6][5]),7)) # '%s' % (data_table[row-6][5])  # Время работы
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][6]),ROUND_SIZE, separator)  # '%s' % get_val(round(float(data_table[row-6][6]),7)) # '%s' % (data_table[row-6][6])  # Время работы
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('H%s'%(row)).value =  '%s' % get_val_by_round(float(data_table[row-6][7]),ROUND_SIZE, separator)  # '%s' % get_val(round(float(data_table[row-6][7]),7))  #'%s' % (data_table[row-6][7])  # Время работы
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
     
    ws.row_dimensions[5].height = 54
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['H'].width = 15
        
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_heat_pulsar_period_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+str(electric_data_start)+'-'+str(electric_data_end)
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pulsar_heat_period_2(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]
    electric_data_start   = request.session["electric_data_start"]

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Пульсар. Потребление энергии с ' +str(electric_data_start)+' по '+ str(electric_data_end)
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Показания Энергии на ' + str(electric_data_start)+', Гкал '
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания Энергии на ' + str(electric_data_end)+', Гкал '
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Потребление Энергии, Гкал'
    ws['E5'].style = "ali_grey"
    
#Запрашиваем данные для отчета

    
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
          
    obj_key             = request.session['obj_key']
    data_table=[]      
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title,electric_data_end, electric_data_start, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title, electric_data_end,electric_data_start, False)
               
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value =  '%s' % (format(data_table[row-6][2],'.7f'))   # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value =  '%s' % (format(data_table[row-6][3],'.7f'))   # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value =  '%s' % (format(data_table[row-6][4],'.7f'))  # Потребление
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        

            
     
#    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['E'].width = 17 
        
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_heat_pulsar_energy_period_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+str(electric_data_end)+'-'+str(electric_data_start)
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_pulsar_heat_daily_2(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Пульсар. Показания по теплу на ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Энергия, Гкал'
    ws['C5'].style = "ali_grey"
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         =  request.session['obj_parent_title']
    obj_title         =  request.session['obj_title']
    electric_data_end   =  request.session['electric_data_end']            
    obj_key             =  request.session['obj_key']
    
    data_table = []
                     
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_daily_pulsar_teplo(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_daily_pulsar_teplo(obj_parent_title, obj_title, electric_data_end, False)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # тип
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value =  '%s' % (format(data_table[row-6][3],'.7f'))   # стояк
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
        

    ws.row_dimensions[5].height = 63
    ws.column_dimensions['A'].width = 20 
    ws.column_dimensions['C'].width = 23 
    ws.column_dimensions['E'].width = 17 
#    ws.column_dimensions['F'].width = 17
#    ws.column_dimensions['D'].width = 17

#------------
      
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'pulsar_heat_energy_report_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_end
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_heat_elf_period_2(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    obj_parent_title         = request.GET.get('obj_parent_title')
    obj_title         = request.GET.get('obj_title')          


    electric_data_end   = request.GET.get("electric_data_end")
    electric_data_start   = request.GET.get("electric_data_start")


#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Теплосчётчик Эльф. Потребление тепла в период c ' +str(electric_data_start)+' по '+ str(electric_data_end)
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Показания Энергии на ' + str(electric_data_start)+', Гкал '
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания Энергии на ' + str(electric_data_end)+', Гкал '
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Потребление Энергии, Гкал'
    ws['E5'].style = "ali_grey"
    
    ws['f5'] = 'Показания Объёма на ' + str(electric_data_start)+', м3'
    ws['f5'].style = "ali_grey"
    
    ws['g5'] = 'Показания Объёма на ' + str(electric_data_end)+', м3'
    ws['g5'].style = "ali_grey"       
    
    ws['h5'] = 'Расход объёма, м3'
    ws['h5'].style = "ali_grey"
    
#Запрашиваем данные для отчета

    
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')    
    obj_parent_title         = request.GET.get('obj_parent_title')
    obj_title         = request.GET.get('obj_title')          
    obj_key             = request.GET.get('obj_key')
#    print obj_parent_title
#    print obj_title
    data_table=[]      
    if (bool(is_abonent_level.search(obj_key))):
         data_table = common_sql.get_data_table_elf_period(obj_parent_title, obj_title, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
         data_table = common_sql.get_data_table_elf_period(obj_parent_title, obj_title, electric_data_start,electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)

    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (format(data_table[row-6][3],'.7f'))  # Показания по теплу на конец
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (format(data_table[row-6][2],'.7f'))  # Показания по теплу на начало
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (format(data_table[row-6][4],'.7f'))  # Потребление
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('f%s'%(row)).value = '%s' % (format(data_table[row-6][6],'.7f'))  # Показания по теплу на конец
            ws.cell('f%s'%(row)).style = "ali_white"
        except:
            ws.cell('f%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('g%s'%(row)).value = '%s' % (format(data_table[row-6][5],'.7f'))  # Показания по теплу на начало
            ws.cell('g%s'%(row)).style = "ali_white"
        except:
            ws.cell('g%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('h%s'%(row)).value = '%s' % (format(data_table[row-6][7],'.7f'))  # Потребление
            ws.cell('h%s'%(row)).style = "ali_white"
        except:
            ws.cell('h%s'%(row)).style = "ali_white"
            next

            
     
#    ws.row_dimensions[5].height = 41
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 18 
        
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    
    output_name = 'elf_heat_energy_report_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_start+'-'+electric_data_end
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response


def report_heat_elf_period(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    electric_data_start = request.session['electric_data_start'] #request.session['electric_data_start']
    electric_data_end   = request.session['electric_data_end'] 
#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Теплосчётчик Эльф. Потребление тепла в период с ' + str(electric_data_start) + ' по ' + str(electric_data_end)
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Показания Энергии на ' + str(electric_data_end)+', Гкал'
    ws['C5'].style = "ali_grey"
        
    ws['D5'] = 'Показания Энергии на ' + str(electric_data_start)+', Гкал'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Потребление Энергии, Гкал'
    ws['E5'].style = "ali_grey"
    
    ws['f5'] = 'Показания Объёма на ' + str(electric_data_end)+', м3'
    ws['f5'].style = "ali_grey"
        
    ws['g5'] = 'Показания Объёма на ' + str(electric_data_start)+', м3'
    ws['g5'].style = "ali_grey"
    
    ws['h5'] = 'Потребленённый Объём, м3'
    ws['h5'].style = "ali_grey"

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    #electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    #electric_data_start = request.session['electric_data_start']
    
#    print unicode(request.session.items())
#    print obj_parent_title
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        pass
        # какая-то фигня - в obj_parent_title, obj_title передаются одинаковые значения
        #data_table = common_sql.get_data_table_elf_period(obj_parent_title, obj_title, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_elf_period(obj_parent_title, obj_title, electric_data_start,electric_data_end, False)
    
  
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Время работы
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('g%s'%(row)).value = '%s' % (data_table[row-6][6])  # Время работы
            ws.cell('g%s'%(row)).style = "ali_white"
        except:
            ws.cell('g%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('h%s'%(row)).value = '%s' % (data_table[row-6][7])  # Время работы
            ws.cell('h%s'%(row)).style = "ali_white"
        except:
            ws.cell('h%s'%(row)).style = "ali_white"
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 18
#____________
                   
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'potreblenie_elf_report_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_start+'_'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
  

def report_heat_elf_daily(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Теплосчётчик Эльф. Показания по теплу на ' + str(request.session["electric_data_end"])
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Показания Энергии на ' + str(request.session["electric_data_end"])+', Гкал'
    ws['C5'].style = "ali_grey"
        
    ws['D5'] = 'Показания Объёма на ' + str(request.session["electric_data_end"])+', м3'
    ws['D5'].style = "ali_grey"


    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title    = request.GET.get('obj_parent_title')
    obj_title           = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')            
    obj_key             = request.GET.get('obj_key')

    
#    print unicode(request.session.items())
#    print obj_parent_title
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_elf_heat_daily(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_elf_heat_daily(obj_parent_title, obj_title, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)
  
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
       

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
#____________
                   
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'pokazaniya_teplo_elf_report_'+translate(obj_parent_title)+'_'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_heat_water_elf_daily(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    obj_parent_title    = request.GET.get('obj_parent_title')
    obj_title           = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')            
    obj_key             = request.GET.get('obj_key')
#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Теплосчётчик Эльф. Показания по теплу и воде на ' + str(electric_data_end)
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Показания Энергии на ' + str(electric_data_end)+', Гкал'
    ws['C5'].style = "ali_grey"
        
    ws['D5'] = 'Показания Объёма на ' + str(electric_data_end)+', м3'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Счётчик ХВС'
    ws['E5'].style = "ali_grey"
    
    ws['f5'] = 'Значение на '+ str(electric_data_end)
    ws['f5'].style = "ali_grey"
        
    ws['g5'] = 'Счётчик ГВС'
    ws['g5'].style = "ali_grey"
    
    ws['h5'] = 'Значение на '+ str(electric_data_end)
    ws['h5'].style = "ali_grey"

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
#    obj_parent_title         = request.session['obj_parent_title']
#    obj_title         = request.session['obj_title']
#    electric_data_end   = request.session['electric_data_end']            
#    obj_key             = request.session['obj_key']
    #electric_data_start = request.session['electric_data_start']
    
#    print unicode(request.session.items())
#    print obj_parent_title
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_elf_heat_water_daily(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_elf_heat_water_daily(obj_parent_title, obj_title, electric_data_end, False)
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)
  
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Время работы
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('g%s'%(row)).value = '%s' % (data_table[row-6][6])  # Время работы
            ws.cell('g%s'%(row)).style = "ali_white"
        except:
            ws.cell('g%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('h%s'%(row)).value = '%s' % (data_table[row-6][7])  # Время работы
            ws.cell('h%s'%(row)).style = "ali_white"
        except:
            ws.cell('h%s'%(row)).style = "ali_white"
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 18
#____________
                   
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'pokazaniya_elf_report_'+translate(obj_parent_title)+'_'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_water_pulsar_potreblenie_skladochnaya(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    obj_parent_title         = request.GET.get('obj_parent_title')
    obj_title                = request.GET.get('obj_title')
    electric_data_end        = request.GET.get("electric_data_end")
    electric_data_start      = request.GET.get("electric_data_start")


#Шапка
    
#    ws.merge_cells('A2:E2')
#    ws['A2'] = 'Теплосчётчик Пульсар. Потребление воды в период c ' +unicode(electric_data_start)+' по '+ unicode(electric_data_end)
    ws.merge_cells('B2:N2')
    ws['B2'] ='Шапка отчёта'
    ws['B2'].style = "ali_grey"
    ws['C2'].style = "ali_grey"
    ws['D2'].style = "ali_grey"
    ws['E2'].style = "ali_grey"
    ws['F2'].style = "ali_grey"
    ws['G2'].style = "ali_grey"
    ws['H2'].style = "ali_grey"
    ws['I2'].style = "ali_grey"
    ws['J2'].style = "ali_grey"
    ws['K2'].style = "ali_grey"
    ws['L2'].style = "ali_grey"
    ws['M2'].style = "ali_grey"
    ws['N2'].style = "ali_grey"
    
    ws.merge_cells('B3:N3')
    ws['B3'] ='Шапка отчёта'
    ws['B3'].style = "ali_grey"
    ws['C3'].style = "ali_grey"
    ws['D3'].style = "ali_grey"
    ws['E3'].style = "ali_grey"
    ws['F3'].style = "ali_grey"
    ws['G3'].style = "ali_grey"
    ws['H3'].style = "ali_grey"
    ws['I3'].style = "ali_grey"
    ws['J3'].style = "ali_grey"
    ws['K3'].style = "ali_grey"
    ws['L3'].style = "ali_grey"
    ws['M3'].style = "ali_grey"
    ws['N3'].style = "ali_grey"

    ws.merge_cells('B4:B5')
    ws['B4'] = 'Квартира'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'ГВС'
    ws['C4'].style = "ali_grey"
    
    ws['C5'] = 'Счётчик'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Нач, показание Т1 (м^3)'
    ws['D5'].style = "ali_grey"

    ws['E5'] = 'Дата/Время'
    ws['E5'].style = "ali_grey"
    
    ws.merge_cells('F4:H4')
    ws['F4'] = 'ГВС'
    ws['F4'].style = "ali_grey"
    
    ws['F5'] = 'Кон, показание Т1 (м^3)'
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'Дата/Время'
    ws['G5'].style = "ali_grey"

    ws['H5'] = 'Разница ГВС'
    ws['H5'].style = "ali_grey"
    
    ws.merge_cells('I4:K4')
    ws['I4'] = 'ХВС'
    ws['I4'].style = "ali_grey"
    
    ws['I5'] = 'Счётчик'
    ws['I5'].style = "ali_grey"
    
    ws['J5'] = 'Нач, показание Т1 (м^3)'
    ws['J5'].style = "ali_grey"

    ws['K5'] = 'Дата/Время'
    ws['K5'].style = "ali_grey"
    
    ws.merge_cells('L4:N4')
    ws['L4'] = 'ХВС'
    ws['L4'].style = "ali_grey"
    ws['N4'].style = "ali_grey"
    
    ws['L5'] = 'Кон, показание Т1 (м^3)'
    ws['L5'].style = "ali_grey"
    
    ws['M5'] = 'Дата/Время'
    ws['M5'].style = "ali_grey"

    ws['N5'] = 'Разница ХВС'
    ws['N5'].style = "ali_grey"
    
#Запрашиваем данные для отчета

    
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')   
    obj_parent_title         = request.GET.get('obj_parent_title')
    obj_title                = request.GET.get('obj_title')
    electric_data_end        = request.GET.get("electric_data_end")
    electric_data_start      = request.GET.get("electric_data_start")             
    obj_key                 = request.GET.get('obj_key')

    data_table=[]      
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_for_period_Skladochnaya(obj_parent_title, obj_title, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_for_period_Skladochnaya(obj_parent_title, obj_title, electric_data_start,electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)

    
 #Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % ((data_table[row-6][3]).strftime("%d-%m-%Y"))  # заводской номер
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][4])  # заводской номер
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % ((data_table[row-6][5]).strftime("%d-%m-%Y"))  # заводской номер
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][6])  # заводской номер
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][7])  # заводской номер
            ws.cell('I%s'%(row)).style = "ali_white"
        except:
            ws.cell('I%s'%(row)).style = "ali_white"
            next
            
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][8])  # заводской номер
            ws.cell('J%s'%(row)).style = "ali_white"
        except:
            ws.cell('J%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % ((data_table[row-6][9]).strftime("%d-%m-%Y"))  # заводской номер
            ws.cell('K%s'%(row)).style = "ali_white"
        except:
            ws.cell('K%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table[row-6][10])  # заводской номер
            ws.cell('L%s'%(row)).style = "ali_white"
        except:
            ws.cell('L%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % ((data_table[row-6][11]).strftime("%d-%m-%Y")) # заводской номер
            ws.cell('M%s'%(row)).style = "ali_white"
        except:
            ws.cell('M%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('N%s'%(row)).value = '%s' % (data_table[row-6][12])  # заводской номер
            ws.cell('N%s'%(row)).style = "ali_white"
        except:
            ws.cell('N%s'%(row)).style = "ali_white"
            next

     
#    ws.row_dimensions[5].height = 41
    #ws.row_dimensions[5].height = 41
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 17
    #ws.column_dimensions['F'].width = 35
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['H'].width = 18 
    
    ws.column_dimensions['J'].width = 25    
    ws.column_dimensions['K'].width = 17
    #ws.column_dimensions['L'].width = 35
    ws.column_dimensions['L'].width = 25
    ws.column_dimensions['N'].width = 18 
        
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    
    output_name = 'pulsar_water_report_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_start+'-'+electric_data_end
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
    
def report_rejim_electro(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    obj_parent_title         = request.GET.get('obj_parent_title')
    obj_title                = request.GET.get('obj_title')
    electric_data_end        = request.GET.get("electric_data_end")    


#Шапка
 
    ws['A2'] = "Адрес"
    #ws['H1'] = u'Шифр'
    ws['G2'] = 'Абонент'
    ws['G3'] = 'Питающий центр'
    #ws['G3'] = u'№ фидера'
    
    ws['D5'] = 'Протокол (первичный)'
    ws['D5'].style = "ali_white_size_18"
    
    ws['E7'] = 'трансформаторного напряжения _____ вольт'
    
    ws['B6'] = 'записей показаний электросчетчиков и вольтметров, а также определения нагрузок'
    ws['B7'] = "и тангенса 'фи' за " + str(electric_data_end) + 'г'
    ws['B9'] = 'Счётчик Акт. и Реакт. энергии № '# + str(common_sql.get_serial_number_by_meter_name(meters_name)) 
    #ws['E9'] = u'Реакт. Счетчик № '# + str(common_sql.get_serial_number_by_meter_name(meters_name))
    ws['B10'] = 'КТТ' 
    ws['B11'] = 'КТН'
    ws['B12'] = 'Расч. коэффициент трансформации'

    #ws['A11'] = u'Время' 
    ws['A17'] = 'час'
    ws['B14'] = '30-минутные срезы мощности (график мощности'
    ws['B15'] = '              в памяти счетчика)'
    ws['B16'] = ' активная, Вт'
    ws['B17'] = 'отпуск'
    ws['C17'] = 'приём'
    ws['D16'] = ' реактивная, ВАр'
    ws['D17'] = 'отпуск'
    ws['E17'] = 'приём'
    
    ws['F15'] = ' Расход энергии'
    ws['F16'] = ' активная, '
    ws['G16'] = ' реактивная, '
    ws['F17'] = '   КВтч'
    ws['G17'] = '   КВАрч'
    
    ws['H15'] = 'тангенс'
    ws['I15'] = 'косинус'
    ws['H17'] = '    фи'
    ws['I17'] = '    фи'
    
    ws['J15'] = ' Полная'
    ws['J16'] = ' мощность'  
    ws['J17'] = '   КВа'
    
    ws['K15'] = ' Показания'
    ws['K16'] = ' вольтметров '    
    ws['K17'] = '   в/н'
    ws['L17'] = '   н/н'
    
    ws['M15'] = ' Мощность'
    ws['M16'] = 'компен.устр'  
    ws['M17'] = '   КВаh'
    
    ws['A18'] = '0-1' 
    ws['A19'] = '1-2' 
    ws['A20'] = '2-3'
    ws['A21'] = '3-4' 
    ws['A22'] = '4-5'
    ws['A23'] = '5-6' 
    ws['A24'] = '6-7'
    ws['A25'] = '7-8' 
    ws['A26'] = '8-9'
    ws['A27'] = '9-10' 
    ws['A28'] = '10-11'
    ws['A29'] = '11-12' 
    ws['A30'] = '12-13'
    ws['A31'] = '13-14' 
    ws['A32'] = '14-15'
    ws['A33'] = '15-16' 
    ws['A34'] = '16-17'
    ws['A35'] = '17-18' 
    ws['A36'] = '18-19'
    ws['A37'] = '19-20' 
    ws['A38'] = '20-21'
    ws['A39'] = '21-22' 
    ws['A40'] = '22-23'
    ws['A41'] = '23-24' 
    #ws['A37'] = u'24'
    
    

    ws.column_dimensions['A'].width = 7            
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 11
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['F'].width = 11
    ws.column_dimensions['G'].width = 11
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 11  
    ws.column_dimensions['K'].width = 7 
    ws.column_dimensions['L'].width = 7 
    ws.row_dimensions[5].height = 30
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')      
    obj_key                 = request.GET.get('obj_key')
    dt_activ=[]
    dt_reactiv=[]
    if (bool(is_abonent_level.search(obj_key))):
        dt_activ=common_sql.get_data_table_rejim( obj_title, electric_data_end, 'A+ Профиль')
        dt_reactiv=common_sql.get_data_table_rejim( obj_title, electric_data_end, 'R+ Профиль')
    
    if len(dt_activ)>0: 
#Заполняем отчет значениями
        try:
            ws.cell('E9').value = '%s' % (obj_title)  # Абонент грщ
            ws.cell('E9').style = "ali_white"
        except:
            ws.cell('E9').style = "ali_white"
            next
    
        try:
            ws.cell('C10').value = '%s' % (dt_activ[0][1])  # ктт
            ws.cell('C10').style = "ali_white"
        except:
            ws.cell('C10').style = "ali_white"
            next
            
        try:
            ws.cell('C11').value = '%s' % (dt_activ[0][2])  # ктн
            ws.cell('C11').style = "ali_white"
        except:
            ws.cell('C11').style = "ali_white"
            next
            
        try:
            ws.cell('E12').value = '%s' % (dt_activ[0][2]*dt_activ[0][1])  # ктн
            ws.cell('E12').style = "ali_white"
        except:
            ws.cell('E12').style = "ali_white"
            next
          
        try:
            t_pred=dt_activ[0][0]+dt_activ[0][7]
            ws.cell('F18').value = '%s' % (t_pred) #сумма первой получасовки/2 и значения на полнось на начало месяца
            ws.cell('F18').style = "ali_white"
        except:
            ws.cell('F18').style = "ali_white"
                
        for row in range(18, 42): 
            n=row-11
            
            try:
                ws.cell('C%s'%(row)).value = '%s' % (dt_activ[0][n]*2) # сумма получасовок за час
                ws.cell('C%s'%(row)).style = "ali_white"
            except:
                ws.cell('C%s'%(row)).style = "ali_white"
                
            try:
                ws.cell('E%s'%(row)).value = '%s' % (dt_reactiv[0][n]*2) # сумма получасовок за час
                ws.cell('E%s'%(row)).style = "ali_white"
            except:
                ws.cell('E%s'%(row)).style = "ali_white"
        
        
        try:
            t_pred_act=dt_activ[0][0]+dt_activ[0][7]
            ws.cell('F18').value = '%s' % (t_pred_act) #сумма первой получасовки/2 и значения на полнось на начало месяца
            ws.cell('F18').style = "ali_white"
        except:
            ws.cell('F18').style = "ali_white"
            
        try:
            t_pred_react=dt_reactiv[0][0]+dt_reactiv[0][7]
            ws.cell('G18').value = '%s' % (t_pred_react) #сумма первой получасовки/2 и значения на полнось на начало месяца
            ws.cell('G18').style = "ali_white"
        except:
            ws.cell('G18').style = "ali_white"
            
        for row in range(19, 42): 
            n=row-11
            try:
                t_pred_act+=dt_activ[0][n]
                ws.cell('F%s'%(row)).value = '%s' % (t_pred_act) # 
                ws.cell('F%s'%(row)).style = "ali_white"
            except:
                ws.cell('F%s'%(row)).style = "ali_white"
                
            try:
                t_pred_react+=dt_reactiv[0][n]                
                ws.cell('G%s'%(row)).value = '%s' % (t_pred_react) # 
                ws.cell('G%s'%(row)).style = "ali_white"
            except:
                ws.cell('G%s'%(row)).style = "ali_white"
            
        
    
    
    
    
    
    for col_idx in range(1, 14):
        col = get_column_letter(col_idx)
        for row in range(18, 42):
            ws.cell('%s%s'%(col, row)).style = "ali_white"
            
            
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    
    output_name = 'protokol_rejim_den_'+translate(obj_parent_title)+'_'+'_'+electric_data_end
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_heat_karat_potreblenie(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    obj_parent_title         = request.GET.get('obj_parent_title')
    obj_title                = request.GET.get('obj_title')
    electric_data_end        = request.GET.get("electric_data_end")
    electric_data_start      = request.GET.get("electric_data_start")
    obj_key                  = request.GET.get('obj_key')
#Шапка    
#Шапка
    ws.merge_cells('A2:E2')
    ws['A1'] = str(obj_parent_title) + " " + str(obj_title)
    ws['A2'] = 'Карат 307. Потребление тепла в период с ' + str(electric_data_start) + ' по ' + str(electric_data_end)
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Тепловая Энергия на '  + str(electric_data_start)
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Тепловая Энергия на '  + str(electric_data_end)
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Потребление Энергии, Гкал'
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 'Масса на '  + str(electric_data_start)
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'Масса на '  + str(electric_data_end)
    ws['G5'].style = "ali_grey"
    
    ws['H5'] = 'Потребление Массы, кг'
    ws['H5'].style = "ali_grey"
    
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_karat_potreblenie(obj_parent_title, obj_title, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_karat_potreblenie(obj_parent_title, obj_title,electric_data_start, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания по теплу на начало
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания по теплу на конец
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # Потребление
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
    ws.row_dimensions[5].height = 41
    
    ws.column_dimensions['A'].width = 20 
    ws.column_dimensions['B'].width = 17 
#    ws.column_dimensions['C'].width = 20
#    ws.column_dimensions['D'].width = 35
#    ws.column_dimensions['E'].width = 18
#    ws.column_dimensions['F'].width = 18
             
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'potreblenie_heat_karat_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+ str(electric_data_start) + "-" + str(electric_data_end)
    file_ext = 'xlsx'
 
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_heat_karat_daily(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    obj_parent_title         = request.GET.get('obj_parent_title')
    obj_title                = request.GET.get('obj_title')
    electric_data_end        = request.GET.get("electric_data_end")
    #electric_data_start      = request.GET.get("electric_data_start")
    obj_key                  = request.GET.get('obj_key')
#Шапка    
#Шапка
    ws.merge_cells('A2:E2')
    ws['A1'] = str(obj_parent_title) + " " + str(obj_title)
    ws['A2'] = 'Карат 307. Показания на ' + str(electric_data_end)
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Тепловая Энергия, Гкал'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Масса, кг '
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Температура вход'
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 'Температура выход'
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'Время наработки, мин.'
    ws['G5'].style = "ali_grey"
    
    ws['H5'] = 'Код ошибки (0-ошибок нет)'
    ws['H5'].style = "ali_grey"
    
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table=[]
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_karat_heat_water_daily(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_karat_heat_water_daily(obj_parent_title, obj_title, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][2])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][3])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу 
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания по массе
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][6])  # tin
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][7])  # tout
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][8])  # twork
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][9])  # ошибки
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
    ws.row_dimensions[5].height = 41
    
    ws.column_dimensions['A'].width = 20 
    ws.column_dimensions['B'].width = 17 
#    ws.column_dimensions['C'].width = 20
#    ws.column_dimensions['D'].width = 35
#    ws.column_dimensions['E'].width = 18
#    ws.column_dimensions['F'].width = 18
             
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'heat_karat_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+ str(electric_data_end)
    file_ext = 'xlsx'
 
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_empty_alert(request):
   
    #response = io.StringIO()
    response = HttpResponse('Отчёт не предусмотрен', content_type="text/plain")
    output_name = 'empty'
    file_ext = 'txt'
 
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)  
    return response


def get_month_russian_name_by_num(m):
    if m == 1:
         M='январь'
    elif m==2:
         M='февраль'
    elif m==3:
         M='март'
    elif m==4:
         M='апрель'
    elif m==5:
         M='май'
    elif m==6:
         M='июнь'
    elif m==7:
         M='июль'
    elif m==8:
         M='август'
    elif m==9:
         M='сентябрь'
    elif m==10:
         M='октябрь'
    elif m==11:
         M='ноябрь'
    elif m==12:
         M='декабрь'
    else:
         M='ошибка'
    return M
    
def add_3columns_to_dt(data_table,data_range,n1,n2,n3):

    for i in range(0,len(data_table)):
        data_table[i]=list(data_table[i])
        data_table[i].append(data_range[i][n1])
        data_table[i].append(data_range[i][n2])
        data_table[i].append(data_range[i][n3])
        data_table[i]=tuple(data_table[i])
    return data_table  
    
def report_water_elf_potreblenie_monthly_with_delta(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    # запрашиваем данные
    data_table = []
    electric_data_end        = request.GET.get("electric_data_end")
    electric_data_start      = request.GET.get("electric_data_start")
    
    dt_date = []
    dt_range = []
    dt_date = common_sql.generate_monthly_range(electric_data_start,electric_data_end)
    double_dates = []
    dt_row4 = []
    dt_row5 = []
    rash = 'РАСХОД'
    znach = 'Значения'
    pred = 'предыдущие'
    tek = 'текущие'
   
    for row in range(0,len(dt_date)):
       data_start = dt_date[row][0].strftime("%d.%m.%Y")
       if (row+1)<len(dt_date):
           data_end = dt_date[row+1][0].strftime("%d.%m.%Y")
       else:
           data_end = dt_date[row][0].strftime("%d.%m.%Y")
       
       double_dates.append(data_start[3:])
       double_dates.append(data_end[3:])
       double_dates.append(get_month_russian_name_by_num(datetime.datetime.strptime(data_end, '%d.%m.%Y').month))
       
       dt_row4.append(znach)
       dt_row4.append(znach)
       dt_row4.append(rash)
       
       dt_row5.append(pred)
       dt_row5.append(tek)
       dt_row5.append('___')
       
       dt_range = common_sql.get_data_table_elf_period_monthly(data_start, data_end)       
       if row == 0:         
           data_table=dt_range           
       else:                          
           data_table=add_3columns_to_dt(data_table,dt_range,4,5,6)
                         
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)    
        #val_num= len(data_table[0]) - 6   
        
    #count_month=range(1,len(dt_date)-1)
    if len(data_table)>0o3:
        double_dates.pop()
        double_dates.pop()
        double_dates.pop()


#Шапка
    ws.merge_cells('A2:J2')
    ws['A2'] = 'Потребление по водосчётчикам Эльф по месяцам с ' + dt_date[0][0].strftime("%d.%m.%Y") + ' по ' + dt_date[-1][0].strftime("%d.%m.%Y")
    
    ws.merge_cells('A4:A6')
    ws['A4'] = '№ помещения'
    ws['A4'].style = "ali_grey"
    
    ws.merge_cells('B4:B6')
    ws['B4'] = 'Тип прибора'
    ws['B4'].style = "ali_grey"
    
    ws.merge_cells('C4:C6')
    ws['C4'] = 'Номер ПУ'
    ws['C4'].style = "ali_grey"    

    
    chrNum=68
    nextLitera=False
    if len(double_dates) < 49: #выход за пределы дат - и выполнятся долго будет и нужна спец.обработка букв в эксель, это будет около 1,5 года        
        for i in range(0,len(double_dates)): 
            if chrNum == 91:            
                nextLitera=True
                chrNum=65
            if not nextLitera:            
                bukva = chr(chrNum)
                if chrNum != 90:
                    bukva_next= chr(chrNum+1)
                else:
                     bukva_next= 'AA'
            else:            
                bukva = 'A'+chr(chrNum)
                bukva_next= 'A'+chr(chrNum+1)
            if not bool(i % 3): 
                ws.merge_cells(bukva+'4:'+bukva_next+'4')
            print(bukva, chrNum)
            ws[bukva+'4'] = dt_row4[i]               
            ws[bukva+'4'].style = "ali_grey"
            
            ws[bukva+'5'] = dt_row5[i]               
            ws[bukva+'5'].style = "ali_grey"        
            
            ws[bukva+'6'] =  double_dates[i]  #получаем букву экселевского столбца и перебираем их черех аццкие коды
            ws[bukva+'6'].style = "ali_grey"
            
            ws.column_dimensions[bukva].width = 15
            chrNum+=1
            
            
    # Заполняем отчет значениями
        for row in range(7, len(data_table)+7):
            try:
                ws.cell('A%s'%(row)).value = '%s' % (data_table[row-7][0][9:])  # номре помещения
                ws.cell('A%s'%(row)).style = "ali_white"
            except:
                ws.cell('A%s'%(row)).style = "ali_white"
                next
            
            try:
                ws.cell('B%s'%(row)).value = '%s' % (data_table[row-7][3])  # тип прибора
                ws.cell('B%s'%(row)).style = "ali_white"
            except:
                ws.cell('B%s'%(row)).style = "ali_white"
                next
                
            try:
                ws.cell('C%s'%(row)).value = '%s' % (data_table[row-7][2])  # номер счётчика
                ws.cell('C%s'%(row)).style = "ali_white"
            except:
                ws.cell('C%s'%(row)).style = "ali_white"
                next
               
               
            chrNum=68
            nextLitera = False
            for dt_col in range(4,len(data_table[row-7])-3):
                 if chrNum == 91:
                    nextLitera=True
                    chrNum=65
                 if not nextLitera:
                    bukva = chr(chrNum)                
                 else:               
                    bukva = 'A'+chr(chrNum)                
                 try:
                    ws.cell(bukva+'%s'%(row)).value = '%s' % (data_table[row-7][dt_col])  #дельта
                    ws.cell(bukva+'%s'%(row)).style = "ali_white"
                 except:
                    ws.cell(bukva+'%s'%(row)).style = "ali_white"
                    next
                 chrNum+=1

#    ws.column_dimensions['D'].width = 17
#    ws.column_dimensions['E'].width = 17
#    ws.column_dimensions['F'].width = 18
#____________

    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'potreblenie_elf_'+electric_data_start+'_'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_water_elf_daily(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Показания по водосчётчикам Эльф на' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик Эльф'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Счётчик воды'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Тип счётчика'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['E5'].style = "ali_grey"

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         =  request.session['obj_parent_title']
    obj_title         =  request.session['obj_title']
    electric_data_end   =  request.session['electric_data_end']            
    obj_key             =  request.session['obj_key']
    
    data_table = []
                     
    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_water_daily_elf(obj_title, obj_parent_title,  electric_data_end, True)
      
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_water_daily_elf(obj_title, obj_parent_title,  electric_data_end, False)
       

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # счётчик эльф
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # счётчик хвс
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания 
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания 
            ws.cell('e%s'%(row)).style = "ali_white"
        except:
            ws.cell('e%s'%(row)).style = "ali_white"
            next


    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17

#------------

                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'elf_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_end
    print(output_name)
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    
    return response
    
def report_water_elf_potreblenie(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление по водосчётчикам Эльф за период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик Эльф'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Счётчик  воды'
    ws['C5'].style = "ali_grey"
    
    ws['d5'] = 'Тип счётчика'
    ws['d5'].style = "ali_grey"
        
    ws['e5'] = 'Показания на '  + str(request.session["electric_data_start"])
    ws['e5'].style = "ali_grey"
    
    ws['f5'] = 'Показания на '  +  str(request.session["electric_data_end"])
    ws['f5'].style = "ali_grey"
    
    ws['g5'] = 'Потребление'
    ws['g5'].style = "ali_grey"

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    electric_data_start = request.session['electric_data_start']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_water_period_elf(obj_title, obj_parent_title, electric_data_start, electric_data_end, True)
      
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_water_period_elf(obj_title, obj_parent_title, electric_data_start, electric_data_end, False)
       

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
                  
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Время работы
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('g%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания по теплу на начало
            ws.cell('g%s'%(row)).style = "ali_white"
        except:
            ws.cell('g%s'%(row)).style = "ali_white"
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 17
    ws.column_dimensions['g'].width = 18
#____________
   

    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'potreblenie_elf_report'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_start+'_'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_balance_period_electric(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    
    ws = wb.active
    ws.title = "percent_noBalance"
    
    obj_parent_title         =  request.session['obj_parent_title']
    obj_title         =         request.session['obj_title']
    electric_data_start   =     request.session['electric_data_start'] 
    electric_data_end   =       request.session['electric_data_end'] 
    is_abonent_level =          re.compile(r'abonent')
    obj_key             =       request.session['obj_key']
#Шапка
    ws.merge_cells('A1:E1')
    ws['A1'] = 'Балансная группа: ' + obj_title
    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Баланс по электричеству c ' + str(electric_data_start) + ' по ' + str(electric_data_end)
    
    ws['A5'] = 'Дата'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Небаланс, кВт*ч'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = '% '
    ws['C5'].style = "ali_grey"
    
       
#Запрашиваем данные для отчета
          
    dtAll=[]
    dt_type_abon=common_sql.GetSimpleTable('types_abonents',"","")
    
    for i in range(0,len(dt_type_abon)):
         guid_type_abon=dt_type_abon[i][0]         
         if not(bool(is_abonent_level.search(obj_key))):
             data_table = common_sql.get_data_table_balance_electric_perid(obj_parent_title, obj_title,electric_data_start, electric_data_end,guid_type_abon)
             #type_abon=translate(dt_type_abon[i][1])             
             if len(data_table)>0: data_table=common_sql.ChangeNull(data_table, None)          
             dtAll.append(data_table)
             
    dt_delta=[]   
   
    if len(dtAll)>0:
        for j in range(1,len(dtAll[0])):
            sumD=0
            vv=0
            for i in range(0,len(dtAll)):                
                if (dtAll[i][j][6] == 'Н/Д' or dtAll[i][j][6] == None  or dtAll[i][j][6] == 'None'): 
                    if (j+1)<len(dtAll[0]): j+=1                             
                if dtAll[i][j][1] == True:                   
                    sumD+=decimal.Decimal(dtAll[i][j][6])
                    vv=decimal.Decimal(dtAll[i][j][6])                    
                else:
                    sumD-=decimal.Decimal(dtAll[i][j][6])
            #считаем проценты
            percent=0           
            if (vv > decimal.Decimal(0)):
                percent=sumD*100/vv
            #print data_table[j][5]
            dt_delta.append([data_table[j][5],sumD, decimal.Decimal(percent)])

        
# Заполняем отчет значениями
    for row in range(6, len(dt_delta)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (dt_delta[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % str(dt_delta[row-6][1]).replace('.',separator)  # счётчик эльф
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (round(dt_delta[row-6][2],2))  # счётчик хвс
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
       
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
   
    #___________________________________________________
   
    
    
    for val in dtAll: 
        #print val[1], val[2], val[3]
        ws2 = wb.create_sheet(title="balance_detail_"+translate(val[0][2]))
        ws2.merge_cells('A1:E1')
        ws2['A1'] = 'Балансная группа: ' + obj_title
        
        ws2.merge_cells('A2:E2')
        ws2['A2'] = 'Баланс по электричеству c' + str(electric_data_start) + ' по ' + str(electric_data_end)
        
        ws2['A5'] = 'Тип'
        ws2['A5'].style = "ali_grey"
        
        ws2['B5'] = 'Сумма T0, кВт*ч'
        ws2['B5'].style = "ali_grey"
        
        ws2['C5'] = 'Дата '
        ws2['C5'].style = "ali_grey"
        
        ws2['d5'] = 'Потребление T0, кВт*ч'
        ws2['d5'].style = "ali_grey"
        
        ws2['e5'] = 'Опрошено счётчиков'
        ws2['e5'].style = "ali_grey"
        #print val
        for row in range(6, len(val)+6):
            try:
                ws2.cell('A%s'%(row)).value = '%s' % (val[row-6][2])  # Абонент
                ws2.cell('A%s'%(row)).style = "ali_white"
            except:
                ws2.cell('A%s'%(row)).style = "ali_white"
                next
            
            try:
                ws2.cell('B%s'%(row)).value = '%s' % (val[row-6][3])  # счётчик эльф
                ws2.cell('B%s'%(row)).style = "ali_white"
            except:
                ws2.cell('B%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('C%s'%(row)).value = '%s' % (val[row-6][5])  # счётчик хвс
                ws2.cell('C%s'%(row)).style = "ali_white"
            except:
                ws2.cell('C%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('d%s'%(row)).value = '%s' % (val[row-6][6])  # счётчик хвс
                ws2.cell('d%s'%(row)).style = "ali_white"
            except:
                ws2.cell('d%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('e%s'%(row)).value = '%s' % (val[row-6][7])  # счётчик хвс
                ws2.cell('e%s'%(row)).style = "ali_white"
            except:
                ws2.cell('e%s'%(row)).style = "ali_white"
                next
       
        ws2.row_dimensions[5].height = 41
        ws2.column_dimensions['A'].width = 17 
        ws2.column_dimensions['B'].width = 17 
        ws2.column_dimensions['C'].width = 17    
    
#------------

                    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'balance_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_start+'-'+electric_data_end
    print(output_name)
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    
    return response
    
def report_all_res_by_date_v2(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    electric_data_end = ''
    obj_title           = request.session['obj_title']
    obj_key             = request.session['obj_key']
    obj_parent_title    = request.session['obj_parent_title']       
    electric_data_end   = request.session['electric_data_end']  
    

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Отчёт по всем ресурсам на ' + str(electric_data_end) + '. ' + str(obj_parent_title) + ' ' + str(obj_title)
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Тип показаний'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Счётчик'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания Энергии на ' + str(electric_data_end)
    ws['D5'].style = "ali_grey"
    
  
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')   
    
    decimal.getcontext().prec = 3   
               
    data_table=[]
    obj_parent_title_water=""
    obj_parent_title_electric=""
    if (bool(is_abonent_level.search(obj_key))):      
            obj_parent_title_water=str(obj_parent_title)+" Вода"
            obj_parent_title_electric=obj_parent_title
            #print  obj_parent_title_water, obj_parent_title_electric, len( obj_parent_title_electric)
            data_table = common_sql.get_data_table_all_res_for_abon(obj_parent_title_water, obj_parent_title_electric, obj_title, electric_data_end)
    if (bool(is_object_level.search(obj_key))):
        #здесь условно на уровне абонента
        n=str(obj_parent_title).find('Вода')
        #print obj_parent_title, n
        if (n>0):
            obj_parent_title_water=obj_parent_title
            obj_parent_title_electric=obj_parent_title[0:n-1]
            #print  obj_parent_title_water, obj_parent_title_electric, len( obj_parent_title_electric)
        data_table = common_sql.get_data_table_all_res_for_abon(obj_parent_title_water, obj_parent_title_electric, obj_title, electric_data_end)
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None) 

    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (round(float(data_table[row-6][3]),3)) # '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
    ws.row_dimensions[5].height = 51
    ws.column_dimensions['A'].width = 30 
    ws.column_dimensions['b'].width = 20 
    ws.column_dimensions['c'].width = 20 
   
    #ws.column_dimensions['H'].width = 15
    
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_all_res_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+str(electric_data_end)
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_electric_res_status(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]  

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Статистика опроса по электричеству на ' + str(electric_data_end)
    

    ws['A5'] = 'Объект'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Опрошено счётчиков'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Всего счётчиков'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Процент опроса'
    ws['D5'].style = "ali_grey"
    
    ws['e5'] = 'Не опрошено счётчиков'
    ws['e5'].style = "ali_grey"
    
  
#Запрашиваем данные для отчета
    
    obj_title           = request.session['obj_title']
    obj_parent_title    = request.session['obj_parent_title']       
    electric_data_end   = request.session['electric_data_end']  
               
    dt_objects = common_sql.get_res_objects('electric')

    dtAll_statistic=[]
    dtAll_no_data_meters=[]
    for obj in dt_objects:
        dt_statistic= common_sql.get_electric_count(obj[0],  electric_data_end)
        dt_no_data_meters=common_sql.get_electric_no_data(obj[0],  electric_data_end)
        dtAll_statistic.append(dt_statistic)
        if len(dt_no_data_meters)>0: 
            dt_no_data_meters=common_sql.ChangeNull(dt_no_data_meters, None)
            dtAll_no_data_meters.append(dt_no_data_meters)

    
# Заполняем отчет значениями
    for row in range(6, len(dtAll_statistic)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][2]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][3]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('d%s'%(row)).style = "ali_white"
        except:
            ws.cell('d%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][4]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('e%s'%(row)).style = "ali_white"
        except:
            ws.cell('e%s'%(row)).style = "ali_white"
            next
            
    ws.row_dimensions[5].height = 51
    ws.column_dimensions['A'].width = 30 
    ws.column_dimensions['b'].width = 20 
    ws.column_dimensions['c'].width = 20 
   
    #ws.column_dimensions['H'].width = 
    #___________________________________________________
    
    #print val[1], val[2], val[3]
    ws2 = wb.create_sheet(title="electric_detail")
    ws2.merge_cells('A1:E1')
           
    ws2.merge_cells('A2:E2')
    ws2['A2'] = 'Не ответившие счётчики на ' + str(electric_data_end)
    
    ws2['A5'] = 'Объект'
    ws2['A5'].style = "ali_grey"
    
    ws2['B5'] = 'Абонент'
    ws2['B5'].style = "ali_grey"
    
    ws2['C5'] = 'Счётчик '
    ws2['C5'].style = "ali_grey"
    
    ws2['d5'] = 'T0, кВт*ч'
    ws2['d5'].style = "ali_grey"
    
    ws2['e5'] = 'T1, кВт*ч'
    ws2['e5'].style = "ali_grey"
    
    ws2['f5'] = 'T2, кВт*ч'
    ws2['f5'].style = "ali_grey"
    
    ws2['g5'] = 'T3, кВт*ч'
    ws2['g5'].style = "ali_grey"

    ws2['h5'] = 'Сетевой адрес'
    ws2['h5'].style = "ali_grey"

    ws2['i5'] = 'IP адрес'
    ws2['i5'].style = "ali_grey"

    ws2['j5'] = 'IP порт'
    ws2['j5'].style = "ali_grey"

    ws2['k5'] = 'Тип прибора'
    ws2['k5'].style = "ali_grey"
    
        #print val
    row = 5
    for value in dtAll_no_data_meters:                 
        for val in value: 
            row+=1
            try:
                ws2.cell('A%s'%(row)).value = '%s' % (val[0])  # Абонент
                ws2.cell('A%s'%(row)).style = "ali_white"
            except:
                ws2.cell('A%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('b%s'%(row)).value = '%s' % (val[1])  # Абонент
                ws2.cell('b%s'%(row)).style = "ali_white"
            except:
                ws2.cell('b%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('c%s'%(row)).value = '%s' % (val[2])  # Абонент
                ws2.cell('c%s'%(row)).style = "ali_white"
            except:
                ws2.cell('c%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('d%s'%(row)).value = '%s' % (val[3])  # Абонент
                ws2.cell('d%s'%(row)).style = "ali_white"
            except:
                ws2.cell('d%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('e%s'%(row)).value = '%s' % (val[4])  # Абонент
                ws2.cell('e%s'%(row)).style = "ali_white"
            except:
                ws2.cell('e%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('f%s'%(row)).value = '%s' % (val[5])  # Абонент
                ws2.cell('f%s'%(row)).style = "ali_white"
            except:
                ws2.cell('f%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('g%s'%(row)).value = '%s' % (val[6])  # Абонент
                ws2.cell('g%s'%(row)).style = "ali_white"
            except:
                ws2.cell('g%s'%(row)).style = "ali_white"
                next
            
            try:
                ws2.cell('h%s'%(row)).value = '%s' % (val[7])  # Абонент
                ws2.cell('h%s'%(row)).style = "ali_white"
            except:
                ws2.cell('h%s'%(row)).style = "ali_white"
                next

            try:
                ws2.cell('i%s'%(row)).value = '%s' % (val[8])  # Абонент
                ws2.cell('i%s'%(row)).style = "ali_white"
            except:
                ws2.cell('i%s'%(row)).style = "ali_white"
                next

            try:
                ws2.cell('j%s'%(row)).value = '%s' % (val[9])  # Абонент
                ws2.cell('j%s'%(row)).style = "ali_white"
            except:
                ws2.cell('j%s'%(row)).style = "ali_white"
                next

            try:
                ws2.cell('k%s'%(row)).value = '%s' % (val[10])  # Абонент
                ws2.cell('k%s'%(row)).style = "ali_white"
            except:
                ws2.cell('k%s'%(row)).style = "ali_white"
                next
   
    ws2.row_dimensions[5].height = 51
    ws2.column_dimensions['A'].width = 30 
    ws2.column_dimensions['b'].width = 20 
    ws2.column_dimensions['c'].width = 20 
    ws2.column_dimensions['i'].width = 20 
    ws2.column_dimensions['k'].width = 20 
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_electric_statistic_'+str(electric_data_end)
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_heat_res_status(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]  

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Статистика опроса по теплу на ' + str(electric_data_end)
    

    ws['A5'] = 'Объект'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Опрошено счётчиков'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Всего счётчиков'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Процент опроса'
    ws['D5'].style = "ali_grey"
    
    ws['e5'] = 'Не опрошено счётчиков'
    ws['e5'].style = "ali_grey"
    
  
#Запрашиваем данные для отчета
    
    obj_title           = request.session['obj_title']
    obj_parent_title    = request.session['obj_parent_title']       
    electric_data_end   = request.session['electric_data_end']  
               
    dt_objects = common_sql.get_res_objects('heat')
    #print 'print len(dt_objects) ', len(dt_objects)
    dtAll_statistic=[]
    dtAll_no_data_meters=[]
    for obj in dt_objects:
        #print obj[0]
        dt_statistic= common_sql.get_heat_count(obj[0],  electric_data_end)
        dt_no_data_meters=common_sql.get_heat_no_data(obj[0],  electric_data_end)
        dtAll_statistic.append(dt_statistic)
        if len(dt_no_data_meters)>0: 
            dt_no_data_meters=common_sql.ChangeNull(dt_no_data_meters, None)
            dtAll_no_data_meters.append(dt_no_data_meters)

    
# Заполняем отчет значениями
    for row in range(6, len(dtAll_statistic)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][2]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][3]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('d%s'%(row)).style = "ali_white"
        except:
            ws.cell('d%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][4]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('e%s'%(row)).style = "ali_white"
        except:
            ws.cell('e%s'%(row)).style = "ali_white"
            next
            
    ws.row_dimensions[5].height = 51
    ws.column_dimensions['A'].width = 30 
    ws.column_dimensions['b'].width = 20 
    ws.column_dimensions['c'].width = 20 
   
    #ws.column_dimensions['H'].width = 
    #___________________________________________________
    
    #print val[1], val[2], val[3]
    ws2 = wb.create_sheet(title="heat_detail")
    ws2.merge_cells('A1:E1')
           
    ws2.merge_cells('A2:E2')
    ws2['A2'] = 'Не ответившие счётчики на ' + str(electric_data_end)
    
    ws2['A5'] = 'Объект'
    ws2['A5'].style = "ali_grey"
    
    ws2['B5'] = 'Абонент'
    ws2['B5'].style = "ali_grey"
    
    ws2['C5'] = 'Счётчик '
    ws2['C5'].style = "ali_grey"
    
    # ws2['d5'] = 'Энергия, Гкал'
    # ws2['d5'].style = "ali_grey"
    
    # # ws2['e5'] = 'Объем, м3'
    # # ws2['e5'].style = "ali_grey"
    
    # ws2['f5'] = 'Температура входа, С'
    # ws2['f5'].style = "ali_grey"
    
    # ws2['g5'] = 'Температура выхода, С'
    # ws2['g5'].style = "ali_grey"
    
        #print val
    row = 5
    for value in dtAll_no_data_meters:                 
        for val in value: 
            row+=1
            try:
                ws2.cell('A%s'%(row)).value = '%s' % (val[0])  # Абонент
                ws2.cell('A%s'%(row)).style = "ali_white"
            except:
                ws2.cell('A%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('b%s'%(row)).value = '%s' % (val[1])  # Абонент
                ws2.cell('b%s'%(row)).style = "ali_white"
            except:
                ws2.cell('b%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('c%s'%(row)).value = '%s' % (val[2])  # Абонент
                ws2.cell('c%s'%(row)).style = "ali_white"
            except:
                ws2.cell('c%s'%(row)).style = "ali_white"
                next
                
            # try:
            #     ws2.cell('d%s'%(row)).value = '%s' % (val[3])  # Абонент
            #     ws2.cell('d%s'%(row)).style = "ali_white"
            # except:
            #     ws2.cell('d%s'%(row)).style = "ali_white"
            #     next
                
            # try:
            #     ws2.cell('e%s'%(row)).value = '%s' % (val[5])  # Абонент
            #     ws2.cell('e%s'%(row)).style = "ali_white"
            # except:
            #     ws2.cell('e%s'%(row)).style = "ali_white"
            #     next
                
            # try:
            #     ws2.cell('f%s'%(row)).value = '%s' % (val[6])  # Абонент
            #     ws2.cell('f%s'%(row)).style = "ali_white"
            # except:
            #     ws2.cell('f%s'%(row)).style = "ali_white"
            #     next
                
            # try:
            #     ws2.cell('g%s'%(row)).value = '%s' % (val[6])  # Абонент
            #     ws2.cell('g%s'%(row)).style = "ali_white"
            # except:
            #     ws2.cell('g%s'%(row)).style = "ali_white"
            #     next
            
   
    ws2.row_dimensions[5].height = 51
    ws2.column_dimensions['A'].width = 30 
    ws2.column_dimensions['b'].width = 20
    ws2.column_dimensions['C'].width = 30 
    #ws2.column_dimensions['c'].width = 20 
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_heat_statistic_'+str(electric_data_end)
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response 
    
    
def report_water_impulse_res_status(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]  

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Статистика опроса по воде(импульсные ПУ) на ' + str(electric_data_end)
    

    ws['A5'] = 'Объект'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Опрошено счётчиков'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Всего счётчиков'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Процент опроса'
    ws['D5'].style = "ali_grey"
    
    ws['e5'] = 'Не опрошено счётчиков'
    ws['e5'].style = "ali_grey"
    
  
#Запрашиваем данные для отчета
          
    dt_objects = common_sql.get_water_impulse_objects()
    
    dtAll_statistic=[]
    dtAll_no_data_meters=[]
    for obj in dt_objects:
        #print electric_data_end
        dt_statistic= common_sql.get_water_impulse_count(obj[0],  electric_data_end)
        dt_no_data_meters=common_sql.get_water_impulse_no_data(obj[0],  electric_data_end)
        dtAll_statistic.append(dt_statistic)
        if len(dt_no_data_meters)>0: 
            dt_no_data_meters=common_sql.ChangeNull(dt_no_data_meters, None)
            dtAll_no_data_meters.append(dt_no_data_meters)

    
# Заполняем отчет значениями
    for row in range(6, len(dtAll_statistic)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][2]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][3]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('d%s'%(row)).style = "ali_white"
        except:
            ws.cell('d%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][4]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('e%s'%(row)).style = "ali_white"
        except:
            ws.cell('e%s'%(row)).style = "ali_white"
            next
            
    ws.row_dimensions[5].height = 51
    ws.column_dimensions['A'].width = 30 
    ws.column_dimensions['b'].width = 20 
    ws.column_dimensions['c'].width = 20 
   
    #ws.column_dimensions['H'].width = 
    #___________________________________________________
    
    #print val[1], val[2], val[3]
    ws2 = wb.create_sheet(title="water_detail")
    ws2.merge_cells('A1:E1')
           
    ws2.merge_cells('A2:E2')
    ws2['A2'] = 'Не ответившие счётчики на ' + str(electric_data_end)
    
    ws2['A5'] = 'Объект'
    ws2['A5'].style = "ali_grey"
    
    ws2['B5'] = 'Абонент'
    ws2['B5'].style = "ali_grey"
    
    ws2['C5'] = 'Счётчик '
    ws2['C5'].style = "ali_grey"
    
    ws2['d5'] = 'Регистратор'
    ws2['d5'].style = "ali_grey"
    
    ws2['e5'] = 'Канал'
    ws2['e5'].style = "ali_grey"
    
    ws2['f5'] = 'Показания'
    ws2['f5'].style = "ali_grey"
    
    
        #print val
    row = 5
    for value in dtAll_no_data_meters:                 
        for val in value: 
            row+=1
            try:
                ws2.cell('A%s'%(row)).value = '%s' % (val[0])  # Абонент
                ws2.cell('A%s'%(row)).style = "ali_white"
            except:
                ws2.cell('A%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('b%s'%(row)).value = '%s' % (val[1])  # Абонент
                ws2.cell('b%s'%(row)).style = "ali_white"
            except:
                ws2.cell('b%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('c%s'%(row)).value = '%s' % (val[2])  # Абонент
                ws2.cell('c%s'%(row)).style = "ali_white"
            except:
                ws2.cell('c%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('d%s'%(row)).value = '%s' % (val[3])  # Абонент
                ws2.cell('d%s'%(row)).style = "ali_white"
            except:
                ws2.cell('d%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('e%s'%(row)).value = '%s' % (val[4])  # Абонент
                ws2.cell('e%s'%(row)).style = "ali_white"
            except:
                ws2.cell('e%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('f%s'%(row)).value = '%s' % (val[5])  # Абонент
                ws2.cell('f%s'%(row)).style = "ali_white"
            except:
                ws2.cell('f%s'%(row)).style = "ali_white"
                next
                            
   
    ws2.row_dimensions[5].height = 51
    ws2.column_dimensions['A'].width = 35 
    ws2.column_dimensions['b'].width = 20 
    ws2.column_dimensions['c'].width = 30 
    ws2.column_dimensions['d'].width = 30 
    #ws2.column_dimensions['c'].width = 20 
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_water_impulse_statistic_'+str(electric_data_end)
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
    
"""def report_heat_res_status(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]  

    #Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Статистика опроса по теплу на ' + str(electric_data_end)
    

    ws['A5'] = 'Объект'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Опрошено счётчиков'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Всего счётчиков'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Процент опроса'
    ws['D5'].style = "ali_grey"
    
    ws['e5'] = 'Не опрошено счётчиков'
    ws['e5'].style = "ali_grey"
    
  
#Запрашиваем данные для отчета
    
    obj_title           = request.session['obj_title']
    obj_parent_title    = request.session['obj_parent_title']       
    electric_data_end   = request.session['electric_data_end']  
               
    dt_objects = common_sql.get_res_objects('heat')
    #print 'print len(dt_objects) ', len(dt_objects)
    dtAll_statistic=[]
    dtAll_no_data_meters=[]
    for obj in dt_objects:
        #print obj[0]
        dt_statistic= common_sql.get_heat_count(obj[0],  electric_data_end)
        dt_no_data_meters=common_sql.get_heat_no_data(obj[0],  electric_data_end)
        dtAll_statistic.append(dt_statistic)
        if len(dt_no_data_meters)>0: 
            dt_no_data_meters=common_sql.ChangeNull(dt_no_data_meters, None)
            dtAll_no_data_meters.append(dt_no_data_meters)

    
# Заполняем отчет значениями
    for row in range(6, len(dtAll_statistic)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][2]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][3]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('d%s'%(row)).style = "ali_white"
        except:
            ws.cell('d%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][4]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('e%s'%(row)).style = "ali_white"
        except:
            ws.cell('e%s'%(row)).style = "ali_white"
            next
            
    ws.row_dimensions[5].height = 51
    ws.column_dimensions['A'].width = 30 
    ws.column_dimensions['b'].width = 20 
    ws.column_dimensions['c'].width = 20 
   
    #ws.column_dimensions['H'].width = 
    #___________________________________________________
    
    #print val[1], val[2], val[3]
    ws2 = wb.create_sheet(title="heat_detail")
    ws2.merge_cells('A1:E1')
           
    ws2.merge_cells('A2:E2')
    ws2['A2'] = 'Не ответившие счётчики на ' + str(electric_data_end)
    
    ws2['A5'] = 'Объект'
    ws2['A5'].style = "ali_grey"
    
    ws2['B5'] = 'Абонент'
    ws2['B5'].style = "ali_grey"
    
    ws2['C5'] = 'Счётчик '
    ws2['C5'].style = "ali_grey"
    
    ws2['d5'] = 'Энергия, Гкал'
    ws2['d5'].style = "ali_grey"
    
    ws2['e5'] = 'Объем, м3'
    ws2['e5'].style = "ali_grey"
    
    ws2['f5'] = 'Температура входа, С'
    ws2['f5'].style = "ali_grey"
    
    ws2['g5'] = 'Температура выхода, С'
    ws2['g5'].style = "ali_grey"
    
        #print val
    row = 5
    for value in dtAll_no_data_meters:                 
        for val in value: 
            row+=1
            try:
                ws2.cell('A%s'%(row)).value = '%s' % (val[0])  # Абонент
                ws2.cell('A%s'%(row)).style = "ali_white"
            except:
                ws2.cell('A%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('b%s'%(row)).value = '%s' % (val[1])  # Абонент
                ws2.cell('b%s'%(row)).style = "ali_white"
            except:
                ws2.cell('b%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('c%s'%(row)).value = '%s' % (val[2])  # Абонент
                ws2.cell('c%s'%(row)).style = "ali_white"
            except:
                ws2.cell('c%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('d%s'%(row)).value = '%s' % (val[3])  # Абонент
                ws2.cell('d%s'%(row)).style = "ali_white"
            except:
                ws2.cell('d%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('e%s'%(row)).value = '%s' % (val[4])  # Абонент
                ws2.cell('e%s'%(row)).style = "ali_white"
            except:
                ws2.cell('e%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('f%s'%(row)).value = '%s' % (val[5])  # Абонент
                ws2.cell('f%s'%(row)).style = "ali_white"
            except:
                ws2.cell('f%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('g%s'%(row)).value = '%s' % (val[6])  # Абонент
                ws2.cell('g%s'%(row)).style = "ali_white"
            except:
                ws2.cell('g%s'%(row)).style = "ali_white"
                next
            
   
    ws2.row_dimensions[5].height = 51
    ws2.column_dimensions['A'].width = 30 
    ws2.column_dimensions['b'].width = 20 
    #ws2.column_dimensions['c'].width = 20 
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_heat_statistic_'+str(electric_data_end)
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    """
    
def report_balance_period_water_impulse(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]  
    electric_data_start   = request.session["electric_data_start"] 

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Баланс по воде (импульсные приборы) c ' + str(electric_data_start) + ' по ' + str(electric_data_end)
    

    ws['A5'] = 'Дата'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Потребление за сутки, м3'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Подача за сутки, м3 (вводный приборы)'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Небаланс, м3'
    ws['D5'].style = "ali_grey"
    
    ws['e5'] = 'Процент, %'
    ws['e5'].style = "ali_grey"
    
    ws['f5'] = 'Опрошено счётчиков'
    ws['f5'].style = "ali_grey"
    
  
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')      
    obj_title           = request.session['obj_title']
    obj_key             = request.session['obj_key']
    obj_parent_title    = request.session['obj_parent_title']    
            
   
    if not(bool(is_abonent_level.search(obj_key))):
         data_table = common_sql.get_data_table_balance_water_impulse_perid(obj_parent_title, obj_title,electric_data_start, electric_data_end)
    if (len( data_table) >0):                
        data_table=common_sql.ChangeNull(data_table, None)

    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][3])  # тип
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][6])  # стояк
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][7])  # счётчик 
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][8])  # значения на начало
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][9])  # значения на конец
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
    ws.row_dimensions[5].height = 51
    ws.column_dimensions['A'].width = 20 
    ws.column_dimensions['b'].width = 20 
    ws.column_dimensions['c'].width = 20 
   
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_water_impulse_balance_'+str(electric_data_end)
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_forma_80040(request):
    import zipfile
    # response = io.StringIO()
    response = io.BytesIO()
    
    #Запрашиваем данные для отчета
    
    group_80040_name    = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']                        
   

    # Формируем список дат на основе начальной и конечной даты полученной от web-календарей
    end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
    start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
    list_of_dates = [x for x in common_sql.daterange(start_date,
                  end_date,
                  step=datetime.timedelta(days=1),
                  inclusive=True)]
    
    info_group_80040 = common_sql.get_info_group_80020(group_80040_name) # Всю информацию для макета 80040 берем из данных группы 80020. Данные идентичны. 
    inn_sender_from_base      = info_group_80040[0][0]
    name_sender_from_base     = info_group_80040[0][1]
    inn_postavshik_from_base  = info_group_80040[0][2]
    name_postavshik_from_base = info_group_80040[0][3]
    dogovor_number_from_base  = info_group_80040[0][4]
    
    #Узнаем GUID счётчиков, которые входят в группу 80040
    meters_guid_list = common_sql.get_meters_guid_list_by_group_name(group_80040_name)

   
    #Создаем архив
    zf = zipfile.ZipFile(response, mode='w', compression=zipfile.ZIP_DEFLATED)
    
    for dates in range(len(list_of_dates)):
    #Формируем файл xml по форме Мосэнергосбыт 80040 Часовые профили мощности
        
        # Создание корневого элемента message
        root = etree.Element('message', {'class': '80040', 'version': '2', 'number': '1' })
        
        # Добавление дочерних элементов - <datetime> <sender> <area> в <root>
        datetimeElt = etree.SubElement(root, 'datetime')
        senderElt = etree.SubElement(root, 'sender')
        areaElt = etree.SubElement(root, 'area')
        
        # Присваиваем значения в <day> <timestamp> <daylightsavingtime>
        day = etree.SubElement(datetimeElt, 'day')
        timestamp = etree.SubElement(datetimeElt, 'timestamp')
        daylightsavingtime = etree.SubElement(datetimeElt, 'daylightsavingtime')
        
        day.text = str(list_of_dates[dates].strftime('%Y%m%d'))
        timestamp.text = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
        daylightsavingtime.text = '0'
        
        # Присваиваем значения <name> <inn> для <sender>
        inn_sender = etree.SubElement(senderElt, 'inn')
        name_sender = etree.SubElement(senderElt, 'name')
        
        inn_sender.text  = inn_sender_from_base
        name_sender.text = name_sender_from_base
        
        # Присваиваем значения для <area>
        inn_area = etree.SubElement(areaElt, 'inn')
        name_abonent_area = etree.SubElement(areaElt, 'name_abonent')
        
        inn_area.text = inn_postavshik_from_base
        name_abonent_area.text = name_postavshik_from_base
        
        # Добавление дочерних элементов
        for x in range(len(meters_guid_list)):
            info_measuring_point = common_sql.get_info_measuring_point_in_group_80020(meters_guid_list[x])
            my_measure_code = str(info_measuring_point[0][0])
            my_measure_name = str(info_measuring_point[0][1])
            measurepointElt = etree.SubElement(areaElt, 'measuringpoint', code = my_measure_code , name = my_measure_name )
        
            list_of_taken_params = []
            #Получаем считываемые параметры по guid счётчика.
             #A+
            name_of_type_meters = common_sql.get_name_of_type_meter_by_guid(meters_guid_list[x])
            if name_of_type_meters[0][0] == 'Меркурий 230-УМ':
                guid_params = '922ad57c-8f5e-4f00-a78d-e3ba89ef859f'
            elif name_of_type_meters[0][0] == 'Меркурий 230':
                guid_params = '6af9ddce-437a-4e07-bd70-6cf9dcc10b31'
            else:
                pass
            result = common_sql.get_taken_param_by_guid_meters_and_guid_params(meters_guid_list[x], guid_params)
            result_list = []
            result_list.append(str(result[0][0]))
            result_list.append(str(result[0][1]))
            list_of_taken_params.append(result_list)
            #R+
            if name_of_type_meters[0][0] == 'Меркурий 230-УМ':
                 guid_params = '61101fa3-a96a-4934-9482-e32036c12829'
            elif name_of_type_meters[0][0] == 'Меркурий 230':
                 guid_params = '66e997c0-8128-40a7-ae65-7e8993fbea61'
            else:
                pass
            
            result = common_sql.get_taken_param_by_guid_meters_and_guid_params(meters_guid_list[x], guid_params)
            result_list = []
            result_list.append(str(result[0][0]))
            result_list.append(str(result[0][1]))
            list_of_taken_params.append(result_list)
            
            for y in range(len(list_of_taken_params)):
                my_dict_of_profil =           {0: ['0000', '0030', 0, 1],
                                               1: ['0030', '0100', 0, 1],
                                               2: ['0100', '0130', 0, 1],
                                               3: ['0130', '0200', 0, 1],
                                               4: ['0200', '0230', 0, 1],
                                               5: ['0230', '0300', 0, 1],
                                               6: ['0300', '0330', 0, 1],
                                               7: ['0330', '0400', 0, 1],
                                               8: ['0400', '0430', 0, 1],
                                               9: ['0430', '0500', 0, 1],
                                              10: ['0500', '0530', 0, 1],
                                              11: ['0530', '0600', 0, 1],
                                              12: ['0600', '0630', 0, 1],
                                              13: ['0630', '0700', 0, 1],
                                              14: ['0700', '0730', 0, 1],
                                              15: ['0730', '0800', 0, 1],
                                              16: ['0800', '0830', 0, 1],
                                              17: ['0830', '0900', 0, 1],
                                              18: ['0900', '0930', 0, 1],
                                              19: ['0930', '1000', 0, 1],
                                              20: ['1000', '1030', 0, 1],
                                              21: ['1030', '1100', 0, 1],
                                              22: ['1100', '1130', 0, 1],                                
                                              23: ['1130', '1200', 0, 1],
                                              24: ['1200', '1230', 0, 1],
                                              25: ['1230', '1300', 0, 1],
                                              26: ['1300', '1330', 0, 1],
                                              27: ['1330', '1400', 0, 1],
                                              28: ['1400', '1430', 0, 1],
                                              29: ['1430', '1500', 0, 1],
                                              30: ['1500', '1530', 0, 1],
                                              31: ['1530', '1600', 0, 1],
                                              32: ['1600', '1630', 0, 1],
                                              33: ['1630', '1700', 0, 1],
                                              34: ['1700', '1730', 0, 1],
                                              35: ['1730', '1800', 0, 1],
                                              36: ['1800', '1830', 0, 1],
                                              37: ['1830', '1900', 0, 1],
                                              38: ['1900', '1930', 0, 1],
                                              39: ['1930', '2000', 0, 1],
                                              40: ['2000', '2030', 0, 1],                                
                                              41: ['2030', '2100', 0, 1],
                                              42: ['2100', '2130', 0, 1],
                                              43: ['2130', '2200', 0, 1],
                                              44: ['2200', '2230', 0, 1],
                                              45: ['2230', '2300', 0, 1],
                                              46: ['2300', '2330', 0, 1],
                                              47: ['2330', '0000', 0, 1] }
                
                my_dict_of_profil_hour =  {0: ['0000', '0100', 0, 1],
                                           1: ['0100', '0200', 0, 1],
                                           2: ['0200', '0300', 0, 1],
                                           3: ['0300', '0400', 0, 1],
                                           4: ['0400', '0500', 0, 1],
                                           5: ['0500', '0600', 0, 1],
                                           6: ['0600', '0700', 0, 1],
                                           7: ['0700', '0800', 0, 1],
                                           8: ['0800', '0900', 0, 1],
                                           9: ['0900', '1000', 0, 1],
                                          10: ['1000', '1100', 0, 1],
                                          11: ['1100', '1200', 0, 1],
                                          12: ['1200', '1300', 0, 1],
                                          13: ['1300', '1400', 0, 1],
                                          14: ['1400', '1500', 0, 1],
                                          15: ['1500', '1600', 0, 1],
                                          16: ['1600', '1700', 0, 1],
                                          17: ['1700', '1800', 0, 1],
                                          18: ['1800', '1900', 0, 1],
                                          19: ['1900', '2000', 0, 1],
                                          20: ['2000', '2100', 0, 1],
                                          21: ['2100', '2200', 0, 1],
                                          22: ['2200', '2300', 0, 1],                                
                                          23: ['2300', '0000', 0, 1]}

                time_table =  ['00:00:00', '00:30:00',
                               '01:00:00', '01:30:00',
                               '02:00:00', '02:30:00',
                               '03:00:00', '03:30:00',
                               '04:00:00', '04:30:00',
                               '05:00:00', '05:30:00',
                               '06:00:00', '06:30:00',
                               '07:00:00', '07:30:00',
                               '08:00:00', '08:30:00',
                               '09:00:00', '09:30:00',
                               '10:00:00', '10:30:00',
                               '11:00:00', '11:30:00',
                               '12:00:00', '12:30:00',
                               '13:00:00', '13:30:00',
                               '14:00:00', '14:30:00',
                               '15:00:00', '15:30:00',
                               '16:00:00', '16:30:00',
                               '17:00:00', '17:30:00',
                               '18:00:00', '18:30:00',
                               '19:00:00', '19:30:00',
                               '20:00:00', '20:30:00',
                               '21:00:00', '21:30:00',
                               '22:00:00', '22:30:00',
                               '23:00:00', '23:30:00']
                
                if list_of_taken_params[y][1] == 'A+ Профиль':
                    my_measuring_channel_code = '01'
                    for time in range(len(time_table)):
                        result_30_min = common_sql.get_30_min_value_by_meters_number_param_names_and_datetime(list_of_taken_params[y][0], list_of_taken_params[y][1], list_of_dates[dates].strftime('%Y-%m-%d'), time_table[time])
                        if result_30_min:
                            my_dict_of_profil[time]=[my_dict_of_profil[time][0], my_dict_of_profil[time][1], float(result_30_min[0][4])*common_sql.get_k_t_t_by_factory_number_manual(list_of_taken_params[y][0]), 0] #Квт.ч
                            
                        else:
                            pass
                    
                elif list_of_taken_params[y][1] == 'R+ Профиль':
                    my_measuring_channel_code = '03'
                    for time in range(len(time_table)):
                        result_30_min = common_sql.get_30_min_value_by_meters_number_param_names_and_datetime(list_of_taken_params[y][0], list_of_taken_params[y][1], list_of_dates[dates].strftime('%Y-%m-%d'), time_table[time])
                        if result_30_min:
                            my_dict_of_profil[time]=[my_dict_of_profil[time][0], my_dict_of_profil[time][1], float(result_30_min[0][4])*common_sql.get_k_t_t_by_factory_number_manual(list_of_taken_params[y][0]), 0] # Квар.ч
                            
                        else:
                           pass
                else:
                    my_measuring_channel_code = ''
                       
                my_measuring_channel_desc = str(list_of_taken_params[y][0]) + ' ' + str(list_of_taken_params[y][1])            
                measuringchannelElt = etree.SubElement(measurepointElt, 'measuringchannel', code = my_measuring_channel_code, desc = my_measuring_channel_desc )
                
                #Берем словарь с получасовыми значениями и делаем часовой словарь, складывая значения получасовок
                hour_iterator = 0
                for half_hour_iterator in range(0, len(my_dict_of_profil), 2):
                    my_dict_of_profil_hour[hour_iterator][2] = my_dict_of_profil[half_hour_iterator][2] + my_dict_of_profil[half_hour_iterator+1][2]
                    my_dict_of_profil_hour[hour_iterator][3] = 0

                    hour_iterator=hour_iterator+1
                
                #Пишем часовой словарь с данными в структуру xml макета 80020                
                for z in range (len(my_dict_of_profil_hour)):
                    periodElt = etree.SubElement(measuringchannelElt, 'period', start = str(my_dict_of_profil_hour[z][0]), end = str(my_dict_of_profil_hour[z][1]))
                    value  = etree.SubElement(periodElt, 'value', status = str(my_dict_of_profil_hour[z][3]))
                    value.text = str(round(my_dict_of_profil_hour[z][2],2))
        
        # Создание и сохранение документа
        doc = etree.ElementTree(root) 
        #myxml_IO=io.StringIO()
        myxml_IO=io.BytesIO()   
        doc.write(myxml_IO, xml_declaration=True, encoding='UTF-8')
        # Формируем имя документа
        name_of_document = '80040'
        name_of_file_80040 =name_of_document + '_'+ inn_sender_from_base + '_' + str(list_of_dates[dates].strftime('%Y%m%d')) + '_1' + '.xml'
        zf.writestr(name_of_file_80040, myxml_IO.getvalue())
       
    zf.close()
    
    response=HttpResponse(response.getvalue())
    response['Content-Type'] = 'application/x-zip-compressed'
    
    output_name = '80040_' + str(dogovor_number_from_base) + '_' + start_date.strftime('%Y%m%d') + '-'+ end_date.strftime('%Y%m%d')
    file_ext = 'zip'
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
  
    return response

def report_electric_report_for_c300(request):
    response = io.StringIO()
    import unicodecsv
    
    #Шапка отчёта
    caption=['N',
            'AREA',
            ' FUNCTION',
            'CODE',
            'SUBNET',
            'MBUS',
            'SERIAL_NUM',
            'SENSOR_TYPE',
            'PREVIOUS',
            'CURRENT',
            'DELTA']

    
    #Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')

    obj_title           = request.session['obj_title']
    obj_parent_title    = request.session['obj_parent_title']       
    electric_data_start   = request.session['electric_data_start']  
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']

    data_table = []
    
    if (bool(is_abonent_level.search(obj_key))): 
        pass
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_electric_period_c300(obj_parent_title, obj_title ,electric_data_start, electric_data_end)
        
    i=0
    while i<len(data_table):
        #print data_table[i][6]
        data_table[i]=list(data_table[i])
        data_table[i+1]=list(data_table[i+1])
        data_table[i+2]=list(data_table[i+2])

        if data_table[i][8] == '-' or data_table[i+1][8] =='-' or data_table[i+2][8] == '-':
            data_table[i][8]     = '-'
            data_table[i+1][8]   = '-'
            data_table[i+2][8]   = '-'
            data_table[i][10]     = '-'
            data_table[i+1][10]   = '-'
            data_table[i+2][10]   = '-'
        
        if data_table[i][9] == '-' or data_table[i+1][9] =='-' or data_table[i+2][9] == '-':
            data_table[i][9]     = '-'
            data_table[i+1][9]   = '-'
            data_table[i+2][9]   = '-'
            data_table[i][10]     = '-'
            data_table[i+1][10]   = '-'
            data_table[i+2][10]   = '-'

        data_table[i]   = tuple(data_table[i])
        data_table[i+1] = tuple(data_table[i+1])
        data_table[i+2] = tuple(data_table[i+2])
        i+=3   

    #Запрашиваем данные для отчета - конец
    
    # Заполняем отчет значениями
    file_name='electric_'+translate(obj_title )+'_'+str(electric_data_start)+'-'+str(electric_data_end)
    response = HttpResponse(save_virtual_workbook(wb),content_type='text/csv')       
    file_ext = 'csv'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (file_name.replace('"', '\"'), file_ext)
    
    response.write('\ufeff'.encode('utf8'))
    writer = unicodecsv.writer(response, delimiter=str(';'))  
    # zagolovok   
    writer.writerow(caption)
    # dannie
    for row in data_table: 
        writer.writerow(row)     
        #writer.writerow([unicode(row[0]), unicode(row[1]),row[2],unicode(row[3]),unicode(row[4]),unicode(row[5]),unicode(row[6]),unicode(row[7]),unicode(row[8]),unicode(row[9]),unicode(row[10])] )

    return response

def report_water_impulse_report_for_c300(request):
    response = io.StringIO()
    import unicodecsv
    
    #Шапка отчёта
    caption=['N',
            'AREA',
            ' FUNCTION',
            'CODE',
            'SUBNET',
            'MBUS',
            'SERIAL_NUM',
            'SENSOR_TYPE',
            'PREVIOUS',
            'CURRENT',
            'DELTA']

    
    #Запрашиваем данные для отчета
    is_object_level_1 = re.compile(r'level')

    obj_title           = request.session['obj_title']
    obj_parent_title    = request.session['obj_parent_title']       
    electric_data_start   = request.session['electric_data_start']  
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']

    data_table = []
    
    if (bool(is_object_level_1.search(obj_key))) : 
         data_table = common_sql.get_data_table_water_period_c300(obj_parent_title, obj_title ,electric_data_start, electric_data_end)
    else:
        pass
    
    
    #Запрашиваем данные для отчета - конец
    
    # Заполняем отчет значениями
    file_name='water_'+translate(obj_title )+'_'+str(electric_data_start)+'-'+str(electric_data_end)
    response = HttpResponse(save_virtual_workbook(wb),content_type='text/csv')       
    file_ext = 'csv'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (file_name.replace('"', '\"'), file_ext)
    
    response.write('\ufeff'.encode('utf8'))
    writer = unicodecsv.writer(response, delimiter=str(';'))  
    # zagolovok   
    writer.writerow(caption)
    # dannie
    for row in data_table: 
        writer.writerow(row)     
        #writer.writerow([unicode(row[0]), unicode(row[1]),row[2],unicode(row[3]),unicode(row[4]),unicode(row[5]),unicode(row[6]),unicode(row[7]),unicode(row[8]),unicode(row[9]),unicode(row[10])] )

    return response

def report_water_digital_pulsar_res_status(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]  

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Статистика опроса по воде(цифровые ПУ) на ' + str(electric_data_end)
    

    ws['A5'] = 'Объект'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Опрошено счётчиков'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Всего счётчиков'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Процент опроса'
    ws['D5'].style = "ali_grey"
    
    ws['e5'] = 'Не опрошено счётчиков'
    ws['e5'].style = "ali_grey"
    
  
#Запрашиваем данные для отчета
          
    dt_objects = common_sql.get_water_digital_pulsar_objects()
    
    dtAll_statistic=[]
    dtAll_no_data_meters=[]
    for obj in dt_objects:
        #print electric_data_end
        dt_statistic= common_sql.get_water_digital_pulsar_count(obj[0],  electric_data_end)
        dt_no_data_meters=common_sql.get_water_digital_pulsar_no_data(obj[0],  electric_data_end)
        dtAll_statistic.append(dt_statistic)
        if len(dt_no_data_meters)>0: 
            dt_no_data_meters=common_sql.ChangeNull(dt_no_data_meters, None)
            dtAll_no_data_meters.append(dt_no_data_meters)

    
# Заполняем отчет значениями
    for row in range(6, len(dtAll_statistic)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][2]) # '%s' % (data_table[row-6][2])  # 
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][3]) # '%s' % (data_table[row-6][2])  # 
            ws.cell('d%s'%(row)).style = "ali_white"
        except:
            ws.cell('d%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][4]) # '%s' % (data_table[row-6][2])  # 
            ws.cell('e%s'%(row)).style = "ali_white"
        except:
            ws.cell('e%s'%(row)).style = "ali_white"
            next
            
    ws.row_dimensions[5].height = 51
    ws.column_dimensions['A'].width = 30 
    ws.column_dimensions['b'].width = 20 
    ws.column_dimensions['c'].width = 20 
   
    #ws.column_dimensions['H'].width = 
    #___________________________________________________
    
    #print val[1], val[2], val[3]
    ws2 = wb.create_sheet(title="water_detail")
    ws2.merge_cells('A1:E1')
           
    ws2.merge_cells('A2:E2')
    ws2['A2'] = 'Не ответившие счётчики на ' + str(electric_data_end)
    
    ws2['A5'] = 'Объект'
    ws2['A5'].style = "ali_grey"
    
    ws2['B5'] = 'Абонент'
    ws2['B5'].style = "ali_grey"
    
    ws2['C5'] = 'Счётчик '
    ws2['C5'].style = "ali_grey"
    
    ws2['d5'] = 'Показания'
    ws2['d5'].style = "ali_grey"
    
    ws2['e5'] = 'Тип счётчика'
    ws2['e5'].style = "ali_grey"
    
    
        #print val
    row = 5
    for value in dtAll_no_data_meters:                 
        for val in value: 
            row+=1
            try:
                ws2.cell('A%s'%(row)).value = '%s' % (val[0])  # Абонент
                ws2.cell('A%s'%(row)).style = "ali_white"
            except:
                ws2.cell('A%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('b%s'%(row)).value = '%s' % (val[1])  # Абонент
                ws2.cell('b%s'%(row)).style = "ali_white"
            except:
                ws2.cell('b%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('c%s'%(row)).value = '%s' % (val[2])  # Абонент
                ws2.cell('c%s'%(row)).style = "ali_white"
            except:
                ws2.cell('c%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('d%s'%(row)).value = '%s' % (val[3])  # Абонент
                ws2.cell('d%s'%(row)).style = "ali_white"
            except:
                ws2.cell('d%s'%(row)).style = "ali_white"
                next
                
            try:
                ws2.cell('e%s'%(row)).value = '%s' % (val[4])  # Абонент
                ws2.cell('e%s'%(row)).style = "ali_white"
            except:
                ws2.cell('e%s'%(row)).style = "ali_white"
                next
                
                          
   
    ws2.row_dimensions[5].height = 51
    ws2.column_dimensions['A'].width = 35 
    ws2.column_dimensions['b'].width = 20 
    ws2.column_dimensions['c'].width = 30 
    ws2.column_dimensions['d'].width = 30 
    ws2.column_dimensions['e'].width = 20 
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_water_digital_pulsar_statistic_'+str(electric_data_end)
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def electric_period_graphic_activ_reactiv_report(request):
    
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+'. Значения профиля показаний за период с' + ' '+electric_data_start +' по '+ electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:C5')
    ws['C4'] = 'Дата'
    ws['C4'].style = "ali_grey"
    ws['C5'].style = "ali_grey"
    
    # Сумма
    ws.merge_cells('D4:D5')
    ws['D4'] = 'Показания T0 A+ '
    ws['D4'].style = "ali_grey"
    ws['D5'].style = "ali_grey"
 
    # Дельта
    ws.merge_cells('E4:E5')
    ws['E4'] = 'Расход за прошедшие сутки T0'
    ws['E4'].style = "ali_grey"
    ws['E5'].style = "ali_grey"
    
        # Сумма
    ws.merge_cells('F4:F5')
    ws['F4'] = 'Показания T0 R+ '
    ws['F4'].style = "ali_grey"
    ws['F5'].style = "ali_grey"
 
    # Дельта
    ws.merge_cells('G4:G5')
    ws['G4'] = 'Расход за прошедшие сутки T0 R+'
    ws['G4'].style = "ali_grey"
    ws['G5'].style = "ali_grey"
    
        # ктт
    ws.merge_cells('H4:H5')
    ws['H4'] = 'КТТ'
    ws['H4'].style = "ali_grey"
    ws['H5'].style = "ali_grey"
 
    # ктн
    ws.merge_cells('I4:I5')
    ws['I4'] = 'КТН'
    ws['I4'].style = "ali_grey"
    ws['I5'].style = "ali_grey"
    
        # а
    ws.merge_cells('J4:J5')
    ws['J4'] = 'А'
    ws['J4'].style = "ali_grey"
    ws['J5'].style = "ali_grey"
    
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_electric_daily   = request.session['is_electric_daily']
    #obj__title          = request.session['obj_t_title']  
    obj_parent_title    = request.session['obj_parent_title']    
    obj_key             = request.session['obj_key']

    data_table = []
    if (is_electric_daily == '1') & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
        #print obj_title, obj_parent_title,electric_data_start, electric_data_end
        params=['T0 A+','T0 R+', 'Электричество']
        data_table= common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
    else:
        pass
            
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][3])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][4])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # дата
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % get_val(data_table[row-6][5])  # сумма-показания t0
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % get_val(data_table[row-6][10]) #str(data_table[row-6][12]).replace('.', separator)   # Расход за прошедшие сутки t0
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % get_val(data_table[row-6][6])   # сумма-показанияt1
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % get_val(data_table[row-6][11])   # Расход за прошедшие суткиt1
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('H%s'%(row)).value = '%s' % get_val(data_table[row-6][7])  # 
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % get_val(data_table[row-6][8])  # 
            ws.cell('I%s'%(row)).style = "ali_white"
        except:
            ws.cell('I%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('J%s'%(row)).value = '%s' % get_val(data_table[row-6][9])   # 
            ws.cell('J%s'%(row)).style = "ali_white"
        except:
            ws.cell('J%s'%(row)).style = "ali_white"
            next            
        

# Сохраняем в ecxel    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'potreblenie_electric_activ_reactiv_'+translate(obj_title)+'_' + electric_data_start + ' - ' +electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def electric_restored_activ_reactiv_daily_report(request):
    
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    #electric_data_start   = request.session['electric_data_start']
# Шапка отчета    
    ws.merge_cells('A2:H2')
    ws['A2'] = obj_title+'. Значения профиля показаний восстановленное посредством получасовых срезов без учёта Кт на ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Дата'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:C5')
    ws['C4'] = 'T0 A+, кВт*ч'
    ws['C4'].style = "ali_grey"
    ws['C5'].style = "ali_grey"

    ws.merge_cells('D4:D5')
    ws['D4'] = 'T0 R+, кВт*ч'
    ws['D4'].style = "ali_grey"
    ws['D5'].style = "ali_grey"
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    data_table=[]
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    obj_title           = request.GET['obj_title']
    obj_key             = request.GET['obj_key']
    obj_parent_title    = request.GET['obj_parent_title']            
    electric_data_end   = request.GET['electric_data_end']
        

    d= datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
    electric_data_start=datetime.date(d.year, d.month, 1)
    if (bool(is_abonent_level.search(obj_key))):   #             
        params=['T0 A+','T0 R+', 'Электричество']
        dt = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
        #print dt
        i=len(dt) - 1

        if not(dt[i][5] == 'Н/Д') and not(dt[i][6] == 'Н/Д'):
            data_table=[dt[i]] # на дату есть срез, просто выводим daily_value        
        
        else:
            for row in reversed(dt):
                #print row
                #print row[5],row[6]
                if (row[5] == 'Н/Д') or (row[6] == 'Н/Д'): #activ-5, reactiv-6                        
                    continue
                else: #начинаем суммировать получасовки
                    date = row[0]
                    activ = row[5]
                    reactiv = row[6]
                    date2 = d - datetime.timedelta(days=1)
                    data_table = common_sql.get_restored_activ_reactiv(obj_title, obj_parent_title, date, activ,reactiv,date2,electric_data_end)
                    break
            #если на 1 число нет суточных, то берем месячные
            if ((dt[0][5] == 'Н/Д') or (dt[0][6] == 'Н/Д')) and len(data_table)<1:
                #print 'monthly'
                dt_monthly = common_sql.get_dt_monthly_activ_reactiv(obj_title, obj_parent_title, electric_data_end)
                if len(dt_monthly)>0:
                    activ = dt_monthly[4]
                    reactiv = dt_monthly[5]
                    date2 = d - datetime.timedelta(days=1)
                    data_table = common_sql.get_restored_activ_reactiv(obj_title, obj_parent_title, date, activ,reactiv,date2,electric_data_end)

        if(datetime.datetime.now().date() < d.date()):
            data_table = []
    else:
        pass
            
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][4])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][5])  # дата
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next

        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][6])  # дата
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next

# Сохраняем в ecxel    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'vosstanovlen_activ_reactiv_'+' - ' +electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def electric_period_30_report(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
# Шапка отчета    
    ws.merge_cells('A2:H2')
    ws['A2'] = obj_title+'. Профиль мощности без учёта Кт с ' + electric_data_end + ' по ' + electric_data_start
    
    ws.merge_cells('A4:A5')
    ws['A4'] = ' № '
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:C5')
    ws['C4'] = 'Дата'
    ws['C4'].style = "ali_grey"
    ws['C5'].style = "ali_grey"
    
    # Сумма
    ws.merge_cells('D4:D5')
    ws['D4'] = 'Время'
    ws['D4'].style = "ali_grey"
    ws['D5'].style = "ali_grey"
 
    # Дельта
    ws.merge_cells('E4:E5')
    ws['E4'] = 'Показания T0 A+ '
    ws['E4'].style = "ali_grey"
    ws['E5'].style = "ali_grey"
    
        # Сумма
    ws.merge_cells('F4:F5')
    ws['F4'] = 'Показания T0 R+ '
    ws['F4'].style = "ali_grey"
    ws['F5'].style = "ali_grey"
 
    # Дельта
    ws.merge_cells('G4:G5')
    ws['G4'] = 'Интервал, мин.'
    ws['G4'].style = "ali_grey"
    ws['G5'].style = "ali_grey"
    
        # ктт
    ws.merge_cells('H4:H5')
    ws['H4'] = ' UTC, мс '
    ws['H4'].style = "ali_grey"
    ws['H5'].style = "ali_grey"
    
    ws.row_dimensions[5].height = 41
        
# Шапка отчета конец
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    electric_data_start   = request.session['electric_data_start']
    electric_data_end   = request.session['electric_data_end']
    obj_title          = request.session['obj_title']  
    obj_parent_title    = request.session['obj_parent_title']    
    obj_key             = request.session['obj_key']

    data_table = []
    if (bool(is_abonent_level.search(obj_key))):   #  Получасовки по абоненту
                params=['A+ Профиль','R+ Профиль']
                data_table= common_sql.get_electric_30_by_abonent_for_period(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
    else:
        pass
        
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][10])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # дата
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % get_val(data_table[row-6][4])  # сумма-показания t0
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % get_val(data_table[row-6][6]) #str(data_table[row-6][12]).replace('.', separator)   # Расход за прошедшие сутки t0
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % get_val(data_table[row-6][7])   # сумма-показанияt1
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % get_val(data_table[row-6][8])   # Расход за прошедшие суткиt1
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('H%s'%(row)).value = '%s' % get_val(data_table[row-6][9])  # 
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['H'].width = 25

# Сохраняем в ecxel    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = '30_activ_reactiv_'+translate(obj_title)
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def heat_danfoss_daily_report(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    
# Шапка отчета    
    ws.merge_cells('A2:H2')
    ws['A2'] = obj_title+'. Показания по цифровым счётчикам Danfoss на  ' + electric_data_end 
    
    ws.merge_cells('A4:A5')
    ws['A4'] = ' Абонент '
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = ' Счётчик '
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:C5')
    ws['C4'] = 'Энергия, Гкал'
    ws['C4'].style = "ali_grey"
    ws['C5'].style = "ali_grey"
    
    # Сумма
    ws.merge_cells('D4:D5')
    ws['D4'] = 'Объём, м3'
    ws['D4'].style = "ali_grey"
    ws['D5'].style = "ali_grey"
 
    # Дельта
    ws.merge_cells('E4:E5')
    ws['E4'] = 'Температура входа, С '
    ws['E4'].style = "ali_grey"
    ws['E5'].style = "ali_grey"
    
        # Сумма
    ws.merge_cells('F4:F5')
    ws['F4'] = 'Температура выхода, С '
    ws['F4'].style = "ali_grey"
    ws['F5'].style = "ali_grey"
         
# Шапка отчета конец
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')

    electric_data_end   = request.session['electric_data_end']
    obj_title          = request.session['obj_title']  
    obj_parent_title    = request.session['obj_parent_title']    
    obj_key             = request.session['obj_key']
    
    dc = 'daily'
        
    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_heat_danfos_daily(obj_parent_title, obj_title, electric_data_end, True,dc)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_heat_danfos_daily(obj_parent_title, obj_title, electric_data_end, False,dc)
        
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # энергия
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % get_val(data_table[row-6][4])  # объём
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % get_val(data_table[row-6][5]) #str(data_table[row-6][12]).replace('.', separator)   # t_vh
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % get_val(data_table[row-6][6])   # t_vih
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
                   
            
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25

# Сохраняем в ecxel    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'heat_danfos_'+translate(obj_title)
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def heat_danfoss_period_report(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
# Шапка отчета    
    ws.merge_cells('A2:H2')
    ws['A2'] = obj_title+'. Потребление по теплосчётчикам Danfoss за период с ' + electric_data_end + ' по ' + electric_data_start
    
    ws.merge_cells('A4:A5')
    ws['A4'] = ' Абонент '
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Счётчик'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:C5')
    ws['C4'] = 'Показания Энергии на ' + electric_data_start +', Гкал'
    ws['C4'].style = "ali_grey"
    ws['C5'].style = "ali_grey"
    
    # Сумма
    ws.merge_cells('D4:D5')
    ws['D4'] = 'Показания Энергии на ' + electric_data_end +', Гкал'
    ws['D4'].style = "ali_grey"
    ws['D5'].style = "ali_grey"
 
    # Дельта
    ws.merge_cells('E4:E5')
    ws['E4'] = 'Расход Энергии, Гкал '
    ws['E4'].style = "ali_grey"
    ws['E5'].style = "ali_grey"
    
        # Сумма
    ws.merge_cells('F4:F5')
    ws['F4'] = 'Показания Объёма на ' + electric_data_start + ', м3'
    ws['F4'].style = "ali_grey"
    ws['F5'].style = "ali_grey"
 
    # Дельта
    ws.merge_cells('G4:G5')
    ws['G4'] = 'Показания Объёма на ' + electric_data_end +', м3'
    ws['G4'].style = "ali_grey"
    ws['G5'].style = "ali_grey"
    
        # ктт
    ws.merge_cells('H4:H5')
    ws['H4'] = ' Расход Объёма, м3 '
    ws['H4'].style = "ali_grey"
    ws['H5'].style = "ali_grey"
    
    ws.row_dimensions[5].height = 41
        
# Шапка отчета конец
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')

    electric_data_start   = request.session['electric_data_start']
    electric_data_end   = request.session['electric_data_end']
    obj_title          = request.session['obj_title']  
    obj_parent_title    = request.session['obj_parent_title']    
    obj_key             = request.session['obj_key']

    dc='current'

    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_danfoss_period(obj_parent_title, obj_title, electric_data_start, electric_data_end, True,dc)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_danfoss_period(obj_parent_title, obj_title, electric_data_start, electric_data_end, False,dc)
        
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # дата
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % get_val(data_table[row-6][4])  # сумма-показания t0
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % get_val(data_table[row-6][5]) #str(data_table[row-6][12]).replace('.', separator)   # Расход за прошедшие сутки t0
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % get_val(data_table[row-6][6])   # сумма-показанияt1
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % get_val(data_table[row-6][7])   # Расход за прошедшие суткиt1
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('H%s'%(row)).value = '%s' % get_val(data_table[row-6][8])  # 
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25


# Сохраняем в ecxel    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'heat_danfos_period_'+translate(obj_title)
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def water_consumption_impuls_report(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    obj_title         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')
    electric_data_start   = request.GET.get('electric_data_start')
    
#Шапка
    ws.merge_cells('A2:G2')
    ws['A2'] = obj_title+'. Потребление по импульсным водосчётчикам в период с ' + electric_data_start + ' по ' +electric_data_end
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Марка счётчика'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Номер счётчика'
    ws['C5'].style = "ali_grey"
    
    ws['d5'] = 'Тип ресурса'
    ws['d5'].style = "ali_grey"

    ws['e5'] = 'Показания на '  + electric_data_start+', м3'
    ws['e5'].style = "ali_grey"
    
    ws['f5'] = 'Показания на '  + electric_data_end+', м3'
    ws['f5'].style = "ali_grey"
    
    ws['g5'] = 'Потребление, м3'
    ws['g5'].style = "ali_grey"
    
    ws['h5'] = 'Лицевой номер '
    ws['h5'].style = "ali_grey"    
  
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'level2')
    is_object_level_2 = re.compile(r'level1')    
    obj_parent_title         = request.GET['obj_parent_title']
    obj_title         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    electric_data_start   = request.GET['electric_data_start']            
    obj_key             = request.GET['obj_key']    
    data_table = []

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_end"]   = electric_data_start   = request.GET['electric_data_start']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']

    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_dt_water_impulse_consumption(obj_title, obj_parent_title,electric_data_start, electric_data_end, True)        
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_dt_water_impulse_consumption(obj_title, obj_parent_title,electric_data_start, electric_data_end, False)
        
    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Тип ресурса
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % (data_table[row-6][5])  # 
            ws.cell('d%s'%(row)).style = "ali_white"
        except:
            ws.cell('d%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % get_val(data_table[row-6][6])  # 
            ws.cell('e%s'%(row)).style = "ali_white"
        except:
            ws.cell('e%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('f%s'%(row)).value = '%s' % get_val(data_table[row-6][7])  # 
            ws.cell('f%s'%(row)).style = "ali_white"
        except:
            ws.cell('f%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('g%s'%(row)).value = '%s' % get_val(data_table[row-6][8])  
            ws.cell('g%s'%(row)).style = "ali_white"
        except:
            ws.cell('g%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('h%s'%(row)).value = '%s' % (data_table[row-6][11])  
            ws.cell('h%s'%(row)).style = "ali_white"
        except:
            ws.cell('h%s'%(row)).style = "ali_white"
            next


    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['F'].width = 18
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    
    output_name = 'consumption_water_impulse_'+translate(obj_title)+'_'+electric_data_start+'-'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_electric_3_zones(request):
    SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')
    ROUND_SIZE = getattr(settings, 'ROUND_SIZE', 3)
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    wb.guess_types = True
    
    obj_title         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')

# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+' .Срез показаний с коэффициентами на дату' + ' ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование абонента'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = "ali_grey"
    ws['D4'].style = "ali_grey"
    ws['E4'].style = "ali_grey"
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = "ali_grey"
    ws['D5'] = 'Ктт'
    ws['D5'].style = "ali_grey"
    ws['E5'] = 'А'
    ws['E5'].style = "ali_grey" 
    
    # Сумма
    ws.merge_cells('F4:G4')
    ws['F4'] = 'Сумма'
    ws['F4'].style = "ali_grey"
    ws['G4'].style = "ali_grey"
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['F5'] = 'Показания A+ на ' + electric_data_end
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'Энергия A+ на ' + electric_data_end
    ws['G5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('H4:I4')
    ws['H4'] = 'Тариф 1'
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['H5'] = 'Показания A+ на ' +  electric_data_end
    ws['H5'].style = "ali_grey"
    
    ws['I5'] = 'Энергия A+ на ' +  electric_data_end
    ws['I5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('J4:K4')
    ws['J4'] = 'Тариф 2'
    ws['J4'].style = "ali_grey"
    ws['K4'].style = "ali_grey"
    ws['J4'].style = "ali_grey"
    ws['K4'].style = "ali_grey"
    ws['J5'] = 'Показания A+ на ' +  electric_data_end
    ws['J5'].style = "ali_grey"
    
    ws['K5'] = 'Энергия A+ на ' +  electric_data_end
    ws['K5'].style = "ali_yellow"
    
    # Тариф 3
    ws.merge_cells('L4:M4')
    ws['L4'] = 'Тариф 3'
    ws['L4'].style = "ali_grey"
    ws['M4'].style = "ali_grey"
    ws['L4'].style = "ali_grey"
    ws['M4'].style = "ali_grey"
    ws['L5'] = 'Показания A+ на ' +  electric_data_end
    ws['L5'].style = "ali_grey"
    
    ws['M5'] = 'Энергия A+ на ' +  electric_data_end
    ws['M5'].style = "ali_yellow"

    if SHOW_LIC_NUM:
        ws.merge_cells('N4:N5')
        ws['N4'] = 'Лицевой номер абонента'
        ws['N4'].style = "ali_grey"
        ws['N5'].style = "ali_grey"
    
    ws.column_dimensions['N'].width = 17
    ws.row_dimensions[5].height = 43
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17 
        
# Шапка отчета конец
    
    #выборка данных из БД
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    
    obj_title           = request.GET['obj_title']
    obj_key             = request.GET['obj_key']
    obj_parent_title    = request.GET['obj_parent_title']
    is_electric_monthly = request.GET['is_electric_monthly']
    is_electric_daily   = request.GET['is_electric_daily']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']
    is_electric_period  = request.GET['is_electric_period']

    if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # месячные для абонента
        data_table = common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'monthly', True)
        
    elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # суточные для абонента
        data_table = common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'daily', True)

#*********************************************************************************************************************************************************************      
    elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # месячные для объекта
            data_table= common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'monthly', False)
            if not data_table:
                data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]        

#*********************************************************************************************************************************************************************
    elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for object
            data_table= common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'daily', False)
            if not data_table:
                data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]

    elif (is_electric_daily == '1') & (bool(is_group_level.search(obj_key))): # показания по баланскной группе                    
            data_table = common_sql.get_electric_by_date_balance(obj_parent_title, obj_title, electric_data_end, 'daily')
            
        
    elif (is_electric_monthly == '1') & (bool(is_group_level.search(obj_key))): # показания по баланскной группе месячные
            data_table = common_sql.get_electric_by_date_balance(obj_parent_title, obj_title, electric_data_end, 'monthly')

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:            
            ws.cell('C%s'%(row)).value = '%s' % str(data_table[row-6][9]).replace('.', separator)# Ктн
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % str(data_table[row-6][8]).replace('.', separator)  # Ктт
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % str(data_table[row-6][10]).replace('.', separator)  # Ка
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        #print(get_val_by_round(data_table[row-6][3], ROUND_SIZE, separator), data_table[row-6][3], ROUND_SIZE, separator)
        try:
            ws.cell('F%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][3], ROUND_SIZE, separator)  #str(val).replace('.', separator)
            ws.cell('F%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
    
        try:            
            ws.cell('G%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][3]*data_table[row-6][8]*data_table[row-6][9]),ROUND_SIZE, separator)  # "Энергия Сумма А+
            ws.cell('G%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('G%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][4]),ROUND_SIZE, separator)  # Тариф 1 А+
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            #val = round((data_table[row-6][4]*data_table[row-6][8]*data_table[row-6][9]),ROUND_SIZE)            
            ws.cell('I%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][4]*data_table[row-6][8]*data_table[row-6][9]),ROUND_SIZE, separator)  # "Энергия Тариф 1 А+
            ws.cell('I%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('I%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][5],ROUND_SIZE, separator)  # Тариф 2 А+
            ws.cell('J%s'%(row)).style = "ali_white"
        except:
            ws.cell('J%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][5]*data_table[row-6][8]*data_table[row-6][9]),ROUND_SIZE, separator) # "Энергия Тариф 2 А+
            ws.cell('K%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('K%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][6]),ROUND_SIZE, separator)  # Тариф 3 А+
            ws.cell('L%s'%(row)).style = "ali_white"
        except:
            ws.cell('L%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][6]*data_table[row-6][8]*data_table[row-6][9]),ROUND_SIZE, separator)  # "Энергия Тариф 3 А+
            ws.cell('M%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('M%s'%(row)).style = "ali_yellow"
            next
        
        if SHOW_LIC_NUM:
            try:                
                ws.cell('N%s'%(row)).value = '%s' % (data_table[row-6][15])   # Лицевой нмоер абонента
                ws.cell('N%s'%(row)).style = "ali_white"
            except:
                ws.cell('N%s'%(row)).style = "ali_white"
                next

# Сохраняем в ecxel  
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb), content_type="application/vnd.ms-excel")
        
    output_name = u'3_tariffa_'+translate(obj_title)+'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def report_electric_2_zones(request):
    SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')
    ROUND_SIZE = getattr(settings, 'ROUND_SIZE', 3)
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    wb.guess_types = True
    
    obj_title         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')

# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+' .Срез показаний с коэффициентами на дату' + ' ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование абонента'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = "ali_grey"
    ws['D4'].style = "ali_grey"
    ws['E4'].style = "ali_grey"
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = "ali_grey"
    ws['D5'] = 'Ктт'
    ws['D5'].style = "ali_grey"
    ws['E5'] = 'А'
    ws['E5'].style = "ali_grey" 
    
    # Сумма
    ws.merge_cells('F4:G4')
    ws['F4'] = 'Сумма'
    ws['F4'].style = "ali_grey"
    ws['G4'].style = "ali_grey"
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['F5'] = 'Показания A+ на ' + electric_data_end
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'Энергия A+ на ' + electric_data_end
    ws['G5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('H4:I4')
    ws['H4'] = 'Тариф 1'
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['H5'] = 'Показания A+ на ' +  electric_data_end
    ws['H5'].style = "ali_grey"
    
    ws['I5'] = 'Энергия A+ на ' +  electric_data_end
    ws['I5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('J4:K4')
    ws['J4'] = 'Тариф 2'
    ws['J4'].style = "ali_grey"
    ws['K4'].style = "ali_grey"
    ws['J4'].style = "ali_grey"
    ws['K4'].style = "ali_grey"
    ws['J5'] = 'Показания A+ на ' +  electric_data_end
    ws['J5'].style = "ali_grey"
    
    ws['K5'] = 'Энергия A+ на ' +  electric_data_end
    ws['K5'].style = "ali_yellow"
    
    if SHOW_LIC_NUM:
        ws.merge_cells('L4:L5')
        ws['L4'] = 'Лицевой номер абонента'
        ws['L4'].style = "ali_grey"
        ws['L5'].style = "ali_grey"
    
    ws.column_dimensions['L'].width = 17         
    ws.row_dimensions[5].height = 43
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
    #выборка данных из БД
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    
    obj_title           = request.GET['obj_title']
    obj_key             = request.GET['obj_key']
    obj_parent_title    = request.GET['obj_parent_title']
    is_electric_monthly = request.GET['is_electric_monthly']
    is_electric_daily   = request.GET['is_electric_daily']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']
    is_electric_period  = request.GET['is_electric_period']

    if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # месячные для абонента
        data_table = common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'monthly', True)
        
    elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # суточные для абонента
        data_table = common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'daily', True)

#*********************************************************************************************************************************************************************      
    elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # месячные для объекта
            data_table= common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'monthly', False)
            if not data_table:
                data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]        

#*********************************************************************************************************************************************************************
    elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for object
            data_table= common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'daily', False)
            if not data_table:
                data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]

    elif (is_electric_daily == '1') & (bool(is_group_level.search(obj_key))): # показания по баланскной группе                    
            data_table = common_sql.get_electric_by_date_balance(obj_parent_title, obj_title, electric_data_end, 'daily')
            
        
    elif (is_electric_monthly == '1') & (bool(is_group_level.search(obj_key))): # показания по баланскной группе месячные
            data_table = common_sql.get_electric_by_date_balance(obj_parent_title, obj_title, electric_data_end, 'monthly')

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:            
            ws.cell('C%s'%(row)).value = '%s' % str(data_table[row-6][9]).replace('.', separator)# Ктн
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % str(data_table[row-6][8]).replace('.', separator)  # Ктт
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % str(data_table[row-6][10]).replace('.', separator)  # Ка
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
                   
        try:
            #ws.cell('F%s'%(row)).number_format = 'Comma'
            #ws.cell('F%s'%(row)).value = '%s' % str(data_table[row-6][3]).replace('.',',')  # Сумма А+
            ws.cell('F%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][3]), ROUND_SIZE, separator)  # Сумма А+
            ws.cell('F%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
    
        try:            
            ws.cell('G%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][3]*data_table[row-6][8]*data_table[row-6][9]),ROUND_SIZE, separator)  # "Энергия Сумма А+
            ws.cell('G%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('G%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][4]),ROUND_SIZE, separator)  # Тариф 1 А+
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            #val = round((data_table[row-6][4]*data_table[row-6][8]*data_table[row-6][9]),ROUND_SIZE)            
            ws.cell('I%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][4]*data_table[row-6][8]*data_table[row-6][9]),ROUND_SIZE, separator)  # "Энергия Тариф 1 А+
            ws.cell('I%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('I%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][5]),ROUND_SIZE, separator)  # Тариф 2 А+
            ws.cell('J%s'%(row)).style = "ali_white"
        except:
            ws.cell('J%s'%(row)).style = "ali_white"
            next
            
        try:
            #val = round((data_table[row-6][5]*data_table[row-6][8]*data_table[row-6][9]),ROUND_SIZE)
            ws.cell('K%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][5]*data_table[row-6][8]*data_table[row-6][9]),ROUND_SIZE, separator) # "Энергия Тариф 2 А+
            ws.cell('K%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('K%s'%(row)).style = "ali_yellow"
            next
        if SHOW_LIC_NUM:
                    try:                
                        ws.cell('L%s'%(row)).value = '%s' % (data_table[row-6][15])   # Лицевой нмоер абонента
                        ws.cell('L%s'%(row)).style = "ali_white"
                    except:
                        ws.cell('L%s'%(row)).style = "ali_white"
                        next
# Сохраняем в ecxel  
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'2_tariffa_'+translate(obj_title)+'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def report_electric_1_zones(request):
    SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')
    ROUND_SIZE = getattr(settings, 'ROUND_SIZE', 'False')
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    wb.guess_types = True
    
    obj_title         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')

# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+' .Срез показаний с коэффициентами на дату' + ' ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование абонента'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = "ali_grey"
    ws['D4'].style = "ali_grey"
    ws['E4'].style = "ali_grey"
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = "ali_grey"
    ws['D5'] = 'Ктт'
    ws['D5'].style = "ali_grey"
    ws['E5'] = 'А'
    ws['E5'].style = "ali_grey" 
    
    # Сумма
    ws.merge_cells('F4:G4')
    ws['F4'] = 'Сумма'
    ws['F4'].style = "ali_grey"
    ws['G4'].style = "ali_grey"
    ws['F5'] = 'Показания A+ на ' + electric_data_end
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'Энергия A+ на ' + electric_data_end
    ws['G5'].style = "ali_yellow"

    if SHOW_LIC_NUM:
        ws.merge_cells('H4:H5')
        ws['H4'] = 'Лицевой номер абонента'
        ws['H4'].style = "ali_grey"
        ws['H5'].style = "ali_grey"
    
    ws.column_dimensions['H'].width = 17             
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
    #выборка данных из БД
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    
    obj_title           = request.GET['obj_title']
    obj_key             = request.GET['obj_key']
    obj_parent_title    = request.GET['obj_parent_title']
    is_electric_monthly = request.GET['is_electric_monthly']
    is_electric_daily   = request.GET['is_electric_daily']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']
    is_electric_period  = request.GET['is_electric_period']

    if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # месячные для абонента
        data_table = common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'monthly', True)
        
    elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # суточные для абонента
        data_table = common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'daily', True)

#*********************************************************************************************************************************************************************      
    elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # месячные для объекта
            data_table= common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'monthly', False)
            if not data_table:
                data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]        

#*********************************************************************************************************************************************************************
    elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for object
            data_table= common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'daily', False)
            if not data_table:
                data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]

    elif (is_electric_daily == '1') & (bool(is_group_level.search(obj_key))): # показания по баланскной группе                    
            data_table = common_sql.get_electric_by_date_balance(obj_parent_title, obj_title, electric_data_end, 'daily')
            
        
    elif (is_electric_monthly == '1') & (bool(is_group_level.search(obj_key))): # показания по баланскной группе месячные
            data_table = common_sql.get_electric_by_date_balance(obj_parent_title, obj_title, electric_data_end, 'monthly')

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:            
            ws.cell('C%s'%(row)).value = '%s' % str(data_table[row-6][9]).replace('.', separator)# Ктн
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % str(data_table[row-6][8]).replace('.', separator)  # Ктт
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % str(data_table[row-6][10]).replace('.', separator)  # Ка
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
                   
        try:
            ws.cell('F%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][3]), ROUND_SIZE, separator)  # Сумма А+
            ws.cell('F%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
    
        try:            
            ws.cell('G%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][3]*data_table[row-6][8]*data_table[row-6][9],ROUND_SIZE, separator)  # "Энергия Сумма А+
            ws.cell('G%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('G%s'%(row)).style = "ali_yellow"
            next
        
        if SHOW_LIC_NUM:
            try:                
                ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][15])   # Лицевой нмоер абонента
                ws.cell('H%s'%(row)).style = "ali_white"
            except:
                ws.cell('H%s'%(row)).style = "ali_white"
                next
                
# Сохраняем в ecxel  
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'1_tariff_'+translate(obj_title)+'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def report_electric_consumption_2_zones(request):
    SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')
    ROUND_SIZE = getattr(settings, 'ROUND_SIZE', 'False')
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+'. Потребление электроэнергии в период с ' + electric_data_start + ' по ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование абонента'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = "ali_grey"
    ws['D4'].style = "ali_grey"
    ws['E4'].style = "ali_grey"
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = "ali_grey"
    ws['D5'] = 'Ктт'
    ws['D5'].style = "ali_grey"
    ws['E5'] = 'А'
    ws['E5'].style = "ali_grey" 
    
    # Сумма
    ws.merge_cells('F3:I3')
    ws['F3'] = 'Сумма A+, кВт*ч'
    ws['F3'].style = "ali_grey"
    ws['G3'].style = "ali_grey"
    ws['H3'].style = "ali_grey"
    ws['I3'].style = "ali_grey"
    
    ws.merge_cells('F4:G4')
    ws['F4'] = 'На ' + str(request.session["electric_data_start"])
    ws['F4'].style = "ali_grey"
    
    ws.merge_cells('H4:I4')
    ws['H4'] = 'На ' + str(request.session["electric_data_end"])
    ws['H4'].style = "ali_grey"

    ws['F5'] = 'Показания'
    ws['F5'].style = "ali_grey"     
    ws['G5'] = 'Энергия'
    ws['G5'].style = "ali_yellow"
    
    ws['H5'] = 'Показания'
    ws['H5'].style = "ali_grey"     
    ws['I5'] = 'Энергия'
    ws['I5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('J3:M3')
    ws['J3'] = 'Тариф 1 A+, кВт*ч'
    ws['J3'].style = "ali_grey"
    ws['K3'].style = "ali_grey"
    ws['L3'].style = "ali_grey"
    ws['M3'].style = "ali_grey"
    
    ws.merge_cells('J4:K4')
    ws['J4'] = 'На ' + str(request.session["electric_data_start"])
    ws['J4'].style = "ali_grey"
    
    ws.merge_cells('L4:M4')
    ws['L4'] = 'На ' + str(request.session["electric_data_end"])
    ws['L4'].style = "ali_grey"

    ws['J5'] = 'Показания'
    ws['J5'].style = "ali_grey"     
    ws['K5'] = 'Энергия'
    ws['K5'].style = "ali_yellow"
    
    ws['L5'] = 'Показания'
    ws['L5'].style = "ali_grey"     
    ws['M5'] = 'Энергия'
    ws['M5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('N3:Q3')
    ws['N3'] = 'Тариф 2 A+, кВт*ч'
    ws['N3'].style = "ali_grey"
    ws['O3'].style = "ali_grey"
    ws['O3'].style = "ali_grey"
    ws['Q3'].style = "ali_grey"
    
    ws.merge_cells('N4:O4')
    ws['N4'] = 'На ' + str(request.session["electric_data_start"])
    ws['N4'].style = "ali_grey"
    
    ws.merge_cells('P4:Q4')
    ws['P4'] = 'На ' + str(request.session["electric_data_end"])
    ws['P4'].style = "ali_grey"

    ws['N5'] = 'Показания'
    ws['N5'].style = "ali_grey"     
    ws['O5'] = 'Энергия'
    ws['O5'].style = "ali_yellow"
    
    ws['P5'] = 'Показания'
    ws['P5'].style = "ali_grey"     
    ws['Q5'] = 'Энергия'
    ws['Q5'].style = "ali_yellow"
             
    # Расход
    ws.merge_cells('R3:W3')
    ws['R3'] = 'Расход А+, кВт*ч'
    ws['R3'].style = "ali_grey"
    ws['W3'].style = "ali_grey"
        # Расход Т0
    ws.merge_cells('R4:S4')
    ws['R4'] = 'Сумма'
    ws['R4'].style = "ali_grey"
    ws['R5'] = 'Показания'
    ws['R5'].style = "ali_grey"
    ws['S5'] = 'Энергия'
    ws['S5'].style = "ali_yellow"
        # Расход Т1
    ws.merge_cells('T4:U4')
    ws['t4'] = 'Tариф 1'
    ws['t4'].style = "ali_grey"
    ws['t5'] = 'Показания'
    ws['t5'].style = "ali_grey"
    ws['u5'] = 'Энергия'
    ws['u5'].style = "ali_yellow"
        # Расход Т2
    ws.merge_cells('V4:W4')
    ws['v4'] = 'Tариф 2'
    ws['v4'].style = "ali_grey"
    ws['v5'] = 'Показания'
    ws['v5'].style = "ali_grey"
    ws['w5'] = 'Энергия'
    ws['w5'].style = "ali_yellow"

    if SHOW_LIC_NUM:
            ws.merge_cells('X4:X5')
            ws['X4'] = 'Лицевой номер абонента'
            ws['X4'].style = "ali_grey"
            ws['X5'].style = "ali_grey"

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['AD'].width = 17     
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    
    obj_parent_title    = request.session['obj_parent_title']
    
    obj_key             = request.session['obj_key']
    is_electric_delta  = request.session['is_electric_delta']
    is_electric_monthly=request.session['is_electric_monthly']
    data_table = []
                            
    res='Электричество'
            
    if (is_electric_monthly=="1"):
        dm='monthly'
    else:
        dm='daily'
    if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
            isAbon=True                    
            data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                        
    elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
            isAbon=False
            data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
            
    #*********************************************************************************************************************************************************************
    elif (is_electric_delta == '1') &(bool(is_group_level.search(obj_key))):
            
            data_table=common_sql.get_data_table_electric_period_for_group(obj_title,obj_parent_title, electric_data_start, electric_data_end, res)
          
#Запрашиваем данные для отчета конец
                    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' %  str(round(float(data_table[row-6][23]), 1)).replace('.', separator)  # Ктн
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % str(round(float(data_table[row-6][20]), 1)).replace('.', separator)  # Ктт
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % str(round(float(data_table[row-6][24]), 1)).replace('.', separator)  # Ка
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
                   
        try:            
            ws.cell('H%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][7]), ROUND_SIZE, separator)  # Сумма А+ на конец интервала
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][7]), ROUND_SIZE, separator)  # Энергия Сумма А+ на конец интервала
            ws.cell('I%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('I%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s ' % get_val_by_round(float(data_table[row-6][2]), ROUND_SIZE, separator)  # '%s' % (data_table[row-6][2])  # Сумма А+ на начало интервала
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][2]), ROUND_SIZE, separator)  # Энергия Сумма А+ на начало интервала
            ws.cell('G%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('G%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][8]), ROUND_SIZE, separator)   # Тариф 1 А+ на конец интервала
            ws.cell('L%s'%(row)).style = "ali_white"
        except:
            ws.cell('L%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][8]), ROUND_SIZE, separator)  # Энергия Тариф 1 А+ на конец интервала
            ws.cell('M%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('M%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][3]), ROUND_SIZE, separator)  # Тариф 1 А+ на начало интервала
            ws.cell('J%s'%(row)).style = "ali_white"
        except:
            ws.cell('J%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][3]), ROUND_SIZE, separator)  # Энергия Тариф 1 А+ на начало интервала
            ws.cell('K%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('K%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][9]), ROUND_SIZE, separator)   # Тариф 2 А+ на конец интервала
            ws.cell('P%s'%(row)).style = "ali_white"
        except:
            ws.cell('P%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('Q%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][9]), ROUND_SIZE, separator)  # Энергия Тариф 2 А+ на конец интервала
            ws.cell('Q%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('Q%s'%(row)).style = "ali_yellow"
            next

        try:
            ws.cell('N%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][4]), ROUND_SIZE, separator)  # Тариф 2 А+ на начало интервала
            ws.cell('N%s'%(row)).style = "ali_white"
        except:
            ws.cell('N%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][4]), ROUND_SIZE, separator)  # Энергия Тариф 2 А+ на начало интервала
            ws.cell('O%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('O%s'%(row)).style = "ali_yellow"
            next

        # Расход
        try:
            ws.cell('R%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][12]), ROUND_SIZE, separator)  # Расход Сумма А+
            ws.cell('R%s'%(row)).style = "ali_white"
        except:
            ws.cell('R%s'%(row)).style = "ali_white"
            next
            
        try:            
            ws.cell('S%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][12]), ROUND_SIZE, separator)  # Расход Сумма Энергия А+
            ws.cell('S%s'%(row)).style = "ali_yellow"            
        except:
            ws.cell('S%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('T%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][13]), ROUND_SIZE, separator)    # Расход Тариф 1 А+
            ws.cell('T%s'%(row)).style = "ali_white"
        except:
            ws.cell('T%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('U%s'%(row)).value = '%s' % str(round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][13]), ROUND_SIZE)).replace('.', separator)    # Расход Тариф 1 Энергия А+
            ws.cell('U%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('U%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('V%s'%(row)).value = '%s' % str(round(float(data_table[row-6][14]), ROUND_SIZE)).replace('.', separator)  # Расход Тариф 2 А+
            ws.cell('V%s'%(row)).style = "ali_white"
        except:
            ws.cell('V%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('w%s'%(row)).value = '%s' % str(round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][14]), ROUND_SIZE)).replace('.', separator)  # Расход Тариф 2 Энергия А+
            ws.cell('w%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('w%s'%(row)).style = "ali_yellow"
            next
                   
        if SHOW_LIC_NUM:
            try:                
                ws.cell('x%s'%(row)).value = '%s' % (data_table[row-6][25])   # Лицевой нмоер абонента
                ws.cell('x%s'%(row)).style = "ali_white"
            except:
                ws.cell('x%s'%(row)).style = "ali_white"                
                next
# Конец наполнения отчёта
            
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb), content_type="application/vnd.ms-excel")
    
    output_name = u'rashod_2_zones_'+translate(obj_title)+'_' + str(electric_data_start) + u'-' + str(electric_data_end)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_electric_consumption_1_zone(request):
    SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')
    ROUND_SIZE = getattr(settings, 'ROUND_SIZE', 'False')
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+'. Потребление электроэнергии в период с ' + electric_data_start + ' по ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование абонента'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = "ali_grey"
    ws['D4'].style = "ali_grey"
    ws['E4'].style = "ali_grey"
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = "ali_grey"
    ws['D5'] = 'Ктт'
    ws['D5'].style = "ali_grey"
    ws['E5'] = 'А'
    ws['E5'].style = "ali_grey" 
    
    # Сумма
    ws.merge_cells('F3:I3')
    ws['F3'] = 'Сумма A+, кВт*ч'
    ws['F3'].style = "ali_grey"
    ws['G3'].style = "ali_grey"
    ws['H3'].style = "ali_grey"
    ws['I3'].style = "ali_grey"
    
    ws.merge_cells('F4:G4')
    ws['F4'] = 'На ' + str(request.session["electric_data_start"])
    ws['F4'].style = "ali_grey"
    
    ws.merge_cells('H4:I4')
    ws['H4'] = 'На ' + str(request.session["electric_data_end"])
    ws['H4'].style = "ali_grey"

    ws['F5'] = 'Показания'
    ws['F5'].style = "ali_grey"     
    ws['G5'] = 'Энергия'
    ws['G5'].style = "ali_yellow"
    
    ws['H5'] = 'Показания'
    ws['H5'].style = "ali_grey"     
    ws['I5'] = 'Энергия'
    ws['I5'].style = "ali_yellow"
    
             
    # Расход
    ws.merge_cells('J3:K3')
    ws['j3'] = 'Расход А+, кВт*ч'
    ws['j3'].style = "ali_grey"
    ws['k3'].style = "ali_grey"
        # Расход Т0
    ws.merge_cells('J4:K4')
    ws['j4'] = 'Сумма'
    ws['j4'].style = "ali_grey"
    ws['j5'] = 'Показания'
    ws['j5'].style = "ali_grey"
    ws['k5'] = 'Энергия'
    ws['k5'].style = "ali_yellow"

    if SHOW_LIC_NUM:
            ws.merge_cells('L4:L5')
            ws['l4'] = 'Лицевой номер абонента'
            ws['l4'].style = "ali_grey"
            ws['l5'].style = "ali_grey"

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['L'].width = 17     
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    
    obj_parent_title    = request.session['obj_parent_title']
    
    obj_key             = request.session['obj_key']
    is_electric_delta  = request.session['is_electric_delta']
    is_electric_monthly=request.session['is_electric_monthly']
    data_table = []
                            
    res='Электричество'
            
    if (is_electric_monthly=="1"):
        dm='monthly'
    else:
        dm='daily'
    if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
            isAbon=True                    
            data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                        
    elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
            isAbon=False
            data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
            
    #*********************************************************************************************************************************************************************
    elif (is_electric_delta == '1') &(bool(is_group_level.search(obj_key))):
            
            data_table=common_sql.get_data_table_electric_period_for_group(obj_title,obj_parent_title, electric_data_start, electric_data_end, res)
          
#Запрашиваем данные для отчета конец
                    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' %  str(round(float(data_table[row-6][23]), 1)).replace('.', separator)  # Ктн
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % str(round(float(data_table[row-6][20]), 1)).replace('.', separator)  # Ктт
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % str(round(float(data_table[row-6][24]), 1)).replace('.', separator)  # Ка
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
                   
        try:            
            ws.cell('H%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][7]), ROUND_SIZE, separator)  # Сумма А+ на конец интервала
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][7]), ROUND_SIZE, separator)  # Энергия Сумма А+ на конец интервала
            ws.cell('I%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('I%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s ' % get_val_by_round(float(data_table[row-6][2]), ROUND_SIZE, separator)  # '%s' % (data_table[row-6][2])  # Сумма А+ на начало интервала
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][2]), ROUND_SIZE, separator)  # Энергия Сумма А+ на начало интервала
            ws.cell('G%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('G%s'%(row)).style = "ali_yellow"
            next

        # Расход
        try:
            ws.cell('j%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][12]), ROUND_SIZE, separator)  # Расход Сумма А+
            ws.cell('j%s'%(row)).style = "ali_white"
        except:
            ws.cell('j%s'%(row)).style = "ali_white"
            next
            
        try:            
            ws.cell('k%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][12]), ROUND_SIZE, separator)  # Расход Сумма Энергия А+
            ws.cell('k%s'%(row)).style = "ali_yellow"            
        except:
            ws.cell('k%s'%(row)).style = "ali_yellow"
            next
       
        if SHOW_LIC_NUM:
            try:                
                ws.cell('L%s'%(row)).value = '%s' % (data_table[row-6][25])   # Лицевой нмоер абонента
                ws.cell('L%s'%(row)).style = "ali_white"
            except:
                ws.cell('L%s'%(row)).style = "ali_white"
                next
# Конец наполнения отчёта
            
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb), content_type="application/vnd.ms-excel")
    
    output_name = u'rashod_1_zone_'+translate(obj_title)+'_' + str(electric_data_start) + u'-' + str(electric_data_end)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_electric_consumption_podolsk(request):
    response =io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+'. Потребление электроэнергии в период с ' + electric_data_start + ' по ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование абонента'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = "ali_grey"
    ws['D4'].style = "ali_grey"
    ws['E4'].style = "ali_grey"
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = "ali_grey"
    ws['D5'] = 'Ктт'
    ws['D5'].style = "ali_grey"
    ws['E5'] = 'А'
    ws['E5'].style = "ali_grey" 
    
    # Сумма
    ws.merge_cells('F3:I3')
    ws['F3'] = 'Сумма A+, кВт*ч'
    ws['F3'].style = "ali_grey"
    ws['G3'].style = "ali_grey"
    ws['H3'].style = "ali_grey"
    ws['I3'].style = "ali_grey"
    
    ws.merge_cells('F4:G4')
    ws['F4'] = 'На ' + str(request.session["electric_data_start"])
    ws['F4'].style = "ali_grey"
    
    ws.merge_cells('H4:I4')
    ws['H4'] = 'На ' + str(request.session["electric_data_end"])
    ws['H4'].style = "ali_grey"

    ws['F5'] = 'Показания'
    ws['F5'].style = "ali_grey"     
    ws['G5'] = 'Энергия'
    ws['G5'].style = "ali_yellow"
    
    ws['H5'] = 'Показания'
    ws['H5'].style = "ali_grey"     
    ws['I5'] = 'Энергия'
    ws['I5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('J3:M3')
    ws['J3'] = 'Тариф 1 A+, кВт*ч'
    ws['J3'].style = "ali_grey"
    ws['K3'].style = "ali_grey"
    ws['L3'].style = "ali_grey"
    ws['M3'].style = "ali_grey"
    
    ws.merge_cells('J4:K4')
    ws['J4'] = 'На ' + str(request.session["electric_data_start"])
    ws['J4'].style = "ali_grey"
    
    ws.merge_cells('L4:M4')
    ws['L4'] = 'На ' + str(request.session["electric_data_end"])
    ws['L4'].style = "ali_grey"

    ws['J5'] = 'Показания'
    ws['J5'].style = "ali_grey"     
    ws['K5'] = 'Энергия'
    ws['K5'].style = "ali_yellow"
    
    ws['L5'] = 'Показания'
    ws['L5'].style = "ali_grey"     
    ws['M5'] = 'Энергия'
    ws['M5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('N3:Q3')
    ws['N3'] = 'Тариф 2 A+, кВт*ч'
    ws['N3'].style = "ali_grey"
    ws['O3'].style = "ali_grey"
    ws['O3'].style = "ali_grey"
    ws['Q3'].style = "ali_grey"
    
    ws.merge_cells('N4:O4')
    ws['N4'] = 'На ' + str(request.session["electric_data_start"])
    ws['N4'].style = "ali_grey"
    
    ws.merge_cells('P4:Q4')
    ws['P4'] = 'На ' + str(request.session["electric_data_end"])
    ws['P4'].style = "ali_grey"

    ws['N5'] = 'Показания'
    ws['N5'].style = "ali_grey"     
    ws['O5'] = 'Энергия'
    ws['O5'].style = "ali_yellow"
    
    ws['P5'] = 'Показания'
    ws['P5'].style = "ali_grey"     
    ws['Q5'] = 'Энергия'
    ws['Q5'].style = "ali_yellow"
             
    # Расход
    ws.merge_cells('R3:W3')
    ws['R3'] = 'Расход А+, кВт*ч'
    ws['R3'].style = "ali_grey"
    ws['W3'].style = "ali_grey"
        # Расход Т0
    ws.merge_cells('R4:S4')
    ws['R4'] = 'Сумма'
    ws['R4'].style = "ali_grey"
    ws['R5'] = 'Показания'
    ws['R5'].style = "ali_grey"
    ws['S5'] = 'Энергия'
    ws['S5'].style = "ali_yellow"
        # Расход Т1
    ws.merge_cells('T4:U4')
    ws['t4'] = 'Tариф 1'
    ws['t4'].style = "ali_grey"
    ws['t5'] = 'Показания'
    ws['t5'].style = "ali_grey"
    ws['u5'] = 'Энергия'
    ws['u5'].style = "ali_yellow"
        # Расход Т2
    ws.merge_cells('V4:W4')
    ws['v4'] = 'Tариф 2'
    ws['v4'].style = "ali_grey"
    ws['v5'] = 'Показания'
    ws['v5'].style = "ali_grey"
    ws['w5'] = 'Энергия'
    ws['w5'].style = "ali_yellow"

    
    SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')

    if SHOW_LIC_NUM:
            ws.merge_cells('X4:X5')
            ws['X4'] = 'Лицевой номер абонента'
            ws['X4'].style = "ali_grey"
            ws['X5'].style = "ali_grey"

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['AD'].width = 17     
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    
    obj_parent_title    = request.session['obj_parent_title']
    
    obj_key             = request.session['obj_key']
    is_electric_delta  = request.session['is_electric_delta']
    is_electric_monthly=request.session['is_electric_monthly']
    data_table = []
                            
    res='Электричество'
            
    if (is_electric_monthly=="1"):
        dm='monthly'
    else:
        dm='daily'
    if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
            isAbon=True                    
            data_table=common_sql.get_data_table_electric_period_podolsk(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
            request.session["data_table_export"] = data_table      
                
    elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
            isAbon=False
            data_table=common_sql.get_data_table_electric_period_podolsk(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
            request.session["data_table_export"] = data_table
          
#Запрашиваем данные для отчета конец
                    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' %  str(round(float(data_table[row-6][23]), 0)).replace('.', separator)  # Ктн
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % str(round(float(data_table[row-6][20]), 0)).replace('.', separator)  # Ктт
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % str(round(float(data_table[row-6][24]), 0)).replace('.', separator)  # Ка
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
                   
        try:            
            ws.cell('H%s'%(row)).value = '%s' % str(round(float(data_table[row-6][7]),0)).replace('.', separator)  # Сумма А+ на конец интервала
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            #val = float(data_table[row-6][23])*float(data_table[row-6][24])*float(data_table[row-6][7])
            ws.cell('I%s'%(row)).value = '%s' % str(round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][7]), 0)).replace('.', separator)  # Энергия Сумма А+ на конец интервала
            ws.cell('I%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('I%s'%(row)).style = "ali_yellow"
            next
            
        try:
            #val = format(data_table[row-6][2], '.3f')
            ws.cell('F%s'%(row)).value = '%s ' % str(round(float(data_table[row-6][2]), 0)).replace('.', separator)  # '%s' % (data_table[row-6][2])  # Сумма А+ на начало интервала
            ws.cell('F%s'%(row)).style = "ali_white"
            #ws.cell('F%s'%(row)).number_format = '0.000'
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            #val = format(data_table[row-6][2]*data_table[row-6][20]*data_table[row-6][23],'.3f')
            ws.cell('G%s'%(row)).value = '%s' % str(round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][2]), 0)).replace('.', separator)  # Энергия Сумма А+ на начало интервала
            ws.cell('G%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('G%s'%(row)).style = "ali_yellow"
            next
            
        try:
            #val = format(data_table[row-6][8],'.3f')
            ws.cell('L%s'%(row)).value = '%s' % str(round(float(data_table[row-6][8]), 3)).replace('.', separator)   # Тариф 1 А+ на конец интервала
            ws.cell('L%s'%(row)).style = "ali_white"
        except:
            ws.cell('L%s'%(row)).style = "ali_white"
            next
            
        try:
            #val = format(data_table[row-6][8]*data_table[row-6][20]*data_table[row-6][23],'.3f')
            ws.cell('M%s'%(row)).value = '%s' % str(round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][8]), 0)).replace('.', separator)  # Энергия Тариф 1 А+ на конец интервала
            ws.cell('M%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('M%s'%(row)).style = "ali_yellow"
            next
            
        try:
            #val = format(data_table[row-6][3],'.3f')
            ws.cell('J%s'%(row)).value = '%s' % str(round(float(data_table[row-6][3]), 3)).replace('.', separator)  # Тариф 1 А+ на начало интервала
            ws.cell('J%s'%(row)).style = "ali_white"
        except:
            ws.cell('J%s'%(row)).style = "ali_white"
            next
            
        try:
            #val = format(data_table[row-6][3]*data_table[row-6][20]*data_table[row-6][23],'.3f')
            ws.cell('K%s'%(row)).value = '%s' % str(round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][3]), 0)).replace('.', separator)  # Энергия Тариф 1 А+ на начало интервала
            ws.cell('K%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('K%s'%(row)).style = "ali_yellow"
            next
            
        try:
            #val = format(data_table[row-6][9],'.3f')
            ws.cell('P%s'%(row)).value = '%s' % str(round(float(data_table[row-6][9]), 0)).replace('.', separator)   # Тариф 2 А+ на конец интервала
            ws.cell('P%s'%(row)).style = "ali_white"
        except:
            ws.cell('P%s'%(row)).style = "ali_white"
            next
            
        try:
            #val = format(data_table[row-6][9]*data_table[row-6][20]*data_table[row-6][23],'.3f')
            ws.cell('Q%s'%(row)).value = '%s' % str(round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][9]), 0)).replace('.', separator)  # Энергия Тариф 2 А+ на конец интервала
            ws.cell('Q%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('Q%s'%(row)).style = "ali_yellow"
            next

        try:
            #val = format(data_table[row-6][4],'.3f')
            ws.cell('N%s'%(row)).value = '%s' % str(round(float(data_table[row-6][4]), 0)).replace('.', separator)  # Тариф 2 А+ на начало интервала
            ws.cell('N%s'%(row)).style = "ali_white"
        except:
            ws.cell('N%s'%(row)).style = "ali_white"
            next
            
        try:
            #val = format(data_table[row-6][4]*data_table[row-6][20]*data_table[row-6][23],'.3f')
            ws.cell('O%s'%(row)).value = '%s' % str(round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][4]), 0)).replace('.', separator)  # Энергия Тариф 2 А+ на начало интервала
            ws.cell('O%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('O%s'%(row)).style = "ali_yellow"
            next

        # Расход
        try:
            #val = format(data_table[row-6][12],'.3f')
            ws.cell('R%s'%(row)).value = '%s' % str(round(float(data_table[row-6][12]), 0)).replace('.', separator)  # Расход Сумма А+
            ws.cell('R%s'%(row)).style = "ali_white"
            #ws.cell('V%s'%(row)).number_format = '0.000'
        except:
            ws.cell('R%s'%(row)).style = "ali_white"
            next
            
        try:            
            ws.cell('S%s'%(row)).value = '%s' % str(round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][12]), 0)).replace('.', separator)  # Расход Сумма Энергия А+
            ws.cell('S%s'%(row)).style = "ali_yellow"            
        except:
            ws.cell('S%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('T%s'%(row)).value = '%s' % str(round(float(data_table[row-6][13]), 0)).replace('.', separator)    # Расход Тариф 1 А+
            ws.cell('T%s'%(row)).style = "ali_white"
        except:
            ws.cell('T%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('U%s'%(row)).value = '%s' % str(round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][13]), 0)).replace('.', separator)    # Расход Тариф 1 Энергия А+
            ws.cell('U%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('U%s'%(row)).style = "ali_yellow"
            next
            
        try:
            #val = format(data_table[row-6][14],'.3f')
            ws.cell('V%s'%(row)).value = '%s' % str(round(float(data_table[row-6][14]), 0)).replace('.', separator)  # Расход Тариф 2 А+
            ws.cell('V%s'%(row)).style = "ali_white"
        except:
            ws.cell('V%s'%(row)).style = "ali_white"
            next
            
        try:
            #val = format(data_table[row-6][14]*data_table[row-6][20]*data_table[row-6][23],'.3f')
            ws.cell('w%s'%(row)).value = '%s' % str(round(float(data_table[row-6][23])*float(data_table[row-6][20])*float(data_table[row-6][14]), 0)).replace('.', separator)  # Расход Тариф 2 Энергия А+
            ws.cell('w%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('w%s'%(row)).style = "ali_yellow"
            next
            
       

        if SHOW_LIC_NUM:
            try:                
                ws.cell('x%s'%(row)).value = '%s' % (data_table[row-6][25])   # Лицевой нмоер абонента
                ws.cell('x%s'%(row)).style = "ali_white"
            except:
                ws.cell('x%s'%(row)).style = "ali_white"
                next
# Конец наполнения отчёта
            
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb), content_type="application/vnd.ms-excel")
    
    output_name = u'rashod_2_zones_'+translate(obj_title)+'_' + str(electric_data_start) + u'-' + str(electric_data_end)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_electric_podolsk(request):
    SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    wb.guess_types = True
    
    obj_title         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')

# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+' .Срез показаний с коэффициентами на дату' + ' ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование абонента'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = "ali_grey"
    ws['D4'].style = "ali_grey"
    ws['E4'].style = "ali_grey"
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = "ali_grey"
    ws['D5'] = 'Ктт'
    ws['D5'].style = "ali_grey"
    ws['E5'] = 'А'
    ws['E5'].style = "ali_grey" 
    
    # Сумма
    ws.merge_cells('F4:G4')
    ws['F4'] = 'Сумма'
    ws['F4'].style = "ali_grey"
    ws['G4'].style = "ali_grey"
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['F5'] = 'Показания A+ на ' + electric_data_end
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'Энергия A+ на ' + electric_data_end
    ws['G5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('H4:I4')
    ws['H4'] = 'Тариф 1'
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['H4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['H5'] = 'Показания A+ на ' +  electric_data_end
    ws['H5'].style = "ali_grey"
    
    ws['I5'] = 'Энергия A+ на ' +  electric_data_end
    ws['I5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('J4:K4')
    ws['J4'] = 'Тариф 2'
    ws['J4'].style = "ali_grey"
    ws['K4'].style = "ali_grey"
    ws['J4'].style = "ali_grey"
    ws['K4'].style = "ali_grey"
    ws['J5'] = 'Показания A+ на ' +  electric_data_end
    ws['J5'].style = "ali_grey"
    
    ws['K5'] = 'Энергия A+ на ' +  electric_data_end
    ws['K5'].style = "ali_yellow"
    
    if SHOW_LIC_NUM:
        ws.merge_cells('L4:L5')
        ws['L4'] = 'Лицевой номер абонента'
        ws['L4'].style = "ali_grey"
        ws['L5'].style = "ali_grey"
    
    ws.column_dimensions['L'].width = 17         
    ws.row_dimensions[5].height = 43
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
    #выборка данных из БД
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    
    obj_title           = request.GET['obj_title']
    obj_key             = request.GET['obj_key']
    obj_parent_title    = request.GET['obj_parent_title']
    is_electric_monthly = request.GET['is_electric_monthly']
    is_electric_daily   = request.GET['is_electric_daily']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']
    is_electric_period  = request.GET['is_electric_period']

    if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # месячные для абонента
        data_table = common_sql.get_electric_by_date_podolsk(obj_parent_title, obj_title, electric_data_end, 'monthly', True)
        
    elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # суточные для абонента
        data_table = common_sql.get_electric_by_date_podolsk(obj_parent_title, obj_title, electric_data_end, 'daily', True)

#*********************************************************************************************************************************************************************      
    elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # месячные для объекта
            data_table= common_sql.get_electric_by_date_podolsk(obj_parent_title, obj_title, electric_data_end, 'monthly', False)
            if not data_table:
                data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]        

#*********************************************************************************************************************************************************************
    elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # суточные для объекта
            data_table= common_sql.get_electric_by_date_podolsk(obj_parent_title, obj_title, electric_data_end, 'daily', False)
            if not data_table:
                data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:            
            ws.cell('C%s'%(row)).value = '%s' % str(data_table[row-6][9]).replace('.', separator)# Ктн
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % str(data_table[row-6][8]).replace('.', separator)  # Ктт
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % str(data_table[row-6][10]).replace('.', separator)  # Ка
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
                   
        try:
            #ws.cell('F%s'%(row)).number_format = 'Comma'
            #ws.cell('F%s'%(row)).value = '%s' % str(data_table[row-6][3]).replace('.',',')  # Сумма А+
            ws.cell('F%s'%(row)).value = '%s' % str(round((data_table[row-6][3]), 3)).replace('.', separator)  # Сумма А+
            ws.cell('F%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
    
        try:            
            ws.cell('G%s'%(row)).value = '%s' % str(round((data_table[row-6][3]*data_table[row-6][8]*data_table[row-6][9]),3)).replace('.', separator)  # "Энергия Сумма А+
            ws.cell('G%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('G%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % str(round((data_table[row-6][4]),3)).replace('.', separator)  # Тариф 1 А+
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            val = round((data_table[row-6][4]*data_table[row-6][8]*data_table[row-6][9]),3)            
            ws.cell('I%s'%(row)).value = '%s' % str(val).replace('.', separator)  # "Энергия Тариф 1 А+
            ws.cell('I%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('I%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % str(round((data_table[row-6][5]),3)).replace('.', separator)  # Тариф 2 А+
            ws.cell('J%s'%(row)).style = "ali_white"
        except:
            ws.cell('J%s'%(row)).style = "ali_white"
            next
            
        try:
            val = round((data_table[row-6][5]*data_table[row-6][8]*data_table[row-6][9]),3)
            ws.cell('K%s'%(row)).value = '%s' % str(val).replace('.', separator) # "Энергия Тариф 2 А+
            ws.cell('K%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('K%s'%(row)).style = "ali_yellow"
            next
        if SHOW_LIC_NUM:
                    try:                
                        ws.cell('L%s'%(row)).value = '%s' % (data_table[row-6][15])   # Лицевой нмоер абонента
                        ws.cell('L%s'%(row)).style = "ali_white"
                    except:
                        ws.cell('L%s'%(row)).style = "ali_white"
                        next


# Сохраняем в ecxel  
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb), content_type="application/vnd.ms-excel")
    
    output_name = u'2_tariffa_'+translate(obj_title)+'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def report_water_tem104_consumption(request):
    #SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')
    ROUND_SIZE = getattr(settings, 'ROUND_SIZE', 'False')
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    obj_title         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')
    electric_data_start   = request.GET.get('electric_data_start')
    
#Шапка
    ws.merge_cells('A2:G2')
    ws['A2'] = obj_title+'. Потребление по водосчётчикам ТЭМ-104 в период с ' + electric_data_start + ' по ' +electric_data_end
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Номер счётчика'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Номер подсистемы'
    ws['C5'].style = "ali_grey"
    
    ws['d5'] = 'Показания на '  + electric_data_start+', м3'
    ws['d5'].style = "ali_grey"
    
    ws['e5'] = 'Показания на '  + electric_data_end+', м3'
    ws['e5'].style = "ali_grey"
    
    ws['f5'] = 'Потребление, м3'
    ws['f5'].style = "ali_grey"

    # if SHOW_LIC_NUM:
    #     ws['g5'] = 'Лицевой номер '
    #     ws['g5'].style = "ali_grey"    
  
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    obj_title           = request.GET['obj_title']
    obj_key             = request.GET['obj_key']
    obj_parent_title    = request.GET['obj_parent_title']    
    is_electric_daily   = request.GET['is_electric_daily']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']
    is_electric_period  = request.GET['is_electric_period']
    isWater = True
    if (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # суточные для абонента
        data_table = common_sql.get_tem104_consumption(obj_parent_title, obj_title,electric_data_start, electric_data_end, isWater, True)
        
    elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # суточные для объекта
        data_table= common_sql.get_tem104_consumption(obj_parent_title, obj_title, electric_data_start, electric_data_end, isWater, False)
        
    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][5])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # подсистема
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][16]),ROUND_SIZE, separator)
            ws.cell('d%s'%(row)).style = "ali_white"
        except:
            ws.cell('d%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][17]),ROUND_SIZE, separator)  # 
            ws.cell('e%s'%(row)).style = "ali_white"
        except:
            ws.cell('e%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('f%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][18]),ROUND_SIZE, separator)  # 
            ws.cell('f%s'%(row)).style = "ali_white"
        except:
            ws.cell('f%s'%(row)).style = "ali_white"
            next
        
        
        # try:
        #     ws.cell('g%s'%(row)).value = '%s' % get_val(data_table[row-6][2])  # лицевой,е сли понадобиться
        #     ws.cell('g%s'%(row)).style = "ali_white"
        # except:
        #     ws.cell('g%s'%(row)).style = "ali_white"
        #     next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    
    output_name = 'consumption_water_tem104_'+translate(obj_title)+'_'+electric_data_start+'-'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_water_tem104_daily(request):
    #SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')
    ROUND_SIZE = getattr(settings, 'ROUND_SIZE', 'False')
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    obj_title         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')
    electric_data_start   = request.GET.get('electric_data_start')
    
#Шапка
    ws.merge_cells('A2:G2')
    ws['A2'] = obj_title+'. Показания по водосчётчикам ТЭМ-104 на '+electric_data_end
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Номер счётчика'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Номер подсистемы'
    ws['C5'].style = "ali_grey"
            
    ws['d5'] = 'Показания V, м3'
    ws['d5'].style = "ali_grey"

    # if SHOW_LIC_NUM:
    #     ws['j5'] = 'Лицевой номер '
    #     ws['j5'].style = "ali_grey"    
  
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    obj_title           = request.GET['obj_title']
    obj_key             = request.GET['obj_key']
    obj_parent_title    = request.GET['obj_parent_title']    
    is_electric_daily   = request.GET['is_electric_daily']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']
    is_electric_period  = request.GET['is_electric_period']
    isWater = True    
    if (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # суточные для абонента
        data_table = common_sql.get_tem104_by_date(obj_parent_title, obj_title, electric_data_end, isWater, True)

    elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # суточные для объекта
        data_table= common_sql.get_tem104_by_date(obj_parent_title, obj_title, electric_data_end, isWater, False)

        
    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][5])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # подсистема
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][22]),ROUND_SIZE, separator)
            ws.cell('d%s'%(row)).style = "ali_white"
        except:
            ws.cell('d%s'%(row)).style = "ali_white"
            next
            
        
        # try:
        #     ws.cell('j%s'%(row)).value = '%s' % get_val(data_table[row-6][2])  # лицевой,е сли понадобиться
        #     ws.cell('j%s'%(row)).style = "ali_white"
        # except:
        #     ws.cell('g%s'%(row)).style = "ali_white"
        #     next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20

    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    
    output_name = 'daily_water_tem104_'+translate(obj_title)+'_'+electric_data_start+'-'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_heat_tem104_consumption(request):
    #SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')
    ROUND_SIZE = getattr(settings, 'ROUND_SIZE', 'False')
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    obj_title         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')
    electric_data_start   = request.GET.get('electric_data_start')
    
#Шапка
    ws.merge_cells('A2:G2')
    ws['A2'] = obj_title+'. Потребление по водосчётчикам ТЭМ-104 в период с ' + electric_data_start + ' по ' +electric_data_end
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Номер счётчика'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Номер подсистемы'
    ws['C5'].style = "ali_grey"
    
    ws['d5'] = 'Показания Q на '  + electric_data_start+', Гкал'
    ws['d5'].style = "ali_grey"
    
    ws['e5'] = 'Показания Q на '  + electric_data_end+', Гкал'
    ws['e5'].style = "ali_grey"
    
    ws['f5'] = 'Показания V на '  + electric_data_start+', м3'
    ws['f5'].style = "ali_grey"
    
    ws['g5'] = 'Показания V на '  + electric_data_end+', м3'
    ws['g5'].style = "ali_grey"
    
    ws['h5'] = 'Потребление Q, Гкал'
    ws['h5'].style = "ali_grey"

    ws['i5'] = 'Потребление V, м3'
    ws['i5'].style = "ali_grey"

    # if SHOW_LIC_NUM:
    #     ws['j5'] = 'Лицевой номер '
    #     ws['j5'].style = "ali_grey"    
  
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    obj_title           = request.GET['obj_title']
    obj_key             = request.GET['obj_key']
    obj_parent_title    = request.GET['obj_parent_title']    
    is_electric_daily   = request.GET['is_electric_daily']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']
    is_electric_period  = request.GET['is_electric_period']
    isWater = False
    if (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # суточные для абонента
        data_table = common_sql.get_tem104_consumption(obj_parent_title, obj_title,electric_data_start, electric_data_end, isWater, True)
        
    elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # суточные для объекта
        data_table= common_sql.get_tem104_consumption(obj_parent_title, obj_title, electric_data_start, electric_data_end, isWater, False)
        
    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][5])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # подсистема
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][7]),ROUND_SIZE, separator)
            ws.cell('d%s'%(row)).style = "ali_white"
        except:
            ws.cell('d%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][9]),ROUND_SIZE, separator)  # 
            ws.cell('e%s'%(row)).style = "ali_white"
        except:
            ws.cell('e%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('f%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][13]),ROUND_SIZE, separator)  # 
            ws.cell('f%s'%(row)).style = "ali_white"
        except:
            ws.cell('f%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('g%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][14]),ROUND_SIZE, separator)  # 
            ws.cell('g%s'%(row)).style = "ali_white"
        except:
            ws.cell('g%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('h%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][9]),ROUND_SIZE, separator)  # 
            ws.cell('h%s'%(row)).style = "ali_white"
        except:
            ws.cell('h%s'%(row)).style = "ali_white"
            next

        try:
            ws.cell('i%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][15]),ROUND_SIZE, separator)  # 
            ws.cell('i%s'%(row)).style = "ali_white"
        except:
            ws.cell('i%s'%(row)).style = "ali_white"
            next

        # try:
        #     ws.cell('j%s'%(row)).value = '%s' % get_val(data_table[row-6][2])  # лицевой,е сли понадобиться
        #     ws.cell('j%s'%(row)).style = "ali_white"
        # except:
        #     ws.cell('g%s'%(row)).style = "ali_white"
        #     next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20

    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    
    output_name = 'consumption_heat_tem104_'+translate(obj_title)+'_'+electric_data_start+'-'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_heat_tem104_daily(request):
    #SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')
    ROUND_SIZE = getattr(settings, 'ROUND_SIZE', 'False')
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    obj_title         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')
    electric_data_start   = request.GET.get('electric_data_start')
    
#Шапка
    ws.merge_cells('A2:G2')
    ws['A2'] = obj_title+'. Показания по теплосчётчикам ТЭМ-104 на '+electric_data_end
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Номер счётчика'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Номер подсистемы'
    ws['C5'].style = "ali_grey"
    
    ws['d5'] = 'Показания Рвх, МПа'
    ws['d5'].style = "ali_grey"    
    
    ws['e5'] = 'Показания Рвых, МПа'
    ws['E5'].style = "ali_grey"

    ws['f5'] = 'Показания Q, Гкал'
    ws['f5'].style = "ali_grey"    
    
    ws['g5'] = 'Показания tвх, С'
    ws['g5'].style = "ali_grey"

    ws['h5'] = 'Показания tвых, С'
    ws['h5'].style = "ali_grey"    
    
    ws['i5'] = 'Показания V, м3'
    ws['i5'].style = "ali_grey"

    # if SHOW_LIC_NUM:
    #     ws['j5'] = 'Лицевой номер '
    #     ws['j5'].style = "ali_grey"    
  
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    obj_title           = request.GET['obj_title']
    obj_key             = request.GET['obj_key']
    obj_parent_title    = request.GET['obj_parent_title']    
    is_electric_daily   = request.GET['is_electric_daily']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']
    is_electric_period  = request.GET['is_electric_period']
    isWater = False    
    if (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # суточные для абонента
        data_table = common_sql.get_tem104_by_date(obj_parent_title, obj_title, electric_data_end, isWater, True)

    elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # суточные для объекта
        data_table= common_sql.get_tem104_by_date(obj_parent_title, obj_title, electric_data_end, isWater, False)

        
    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][5])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # подсистема
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][9]),ROUND_SIZE, separator)
            ws.cell('d%s'%(row)).style = "ali_white"
        except:
            ws.cell('d%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][11]),ROUND_SIZE, separator)  # 
            ws.cell('e%s'%(row)).style = "ali_white"
        except:
            ws.cell('e%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('f%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][13]),ROUND_SIZE, separator)  # 
            ws.cell('f%s'%(row)).style = "ali_white"
        except:
            ws.cell('f%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('g%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][15]),ROUND_SIZE, separator)  # 
            ws.cell('g%s'%(row)).style = "ali_white"
        except:
            ws.cell('g%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('h%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][19]),ROUND_SIZE, separator)  # 
            ws.cell('h%s'%(row)).style = "ali_white"
        except:
            ws.cell('h%s'%(row)).style = "ali_white"
            next

        try:
            ws.cell('i%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][21]),ROUND_SIZE, separator)  # 
            ws.cell('i%s'%(row)).style = "ali_white"
        except:
            ws.cell('i%s'%(row)).style = "ali_white"
            next

        # try:
        #     ws.cell('j%s'%(row)).value = '%s' % get_val(data_table[row-6][2])  # лицевой,е сли понадобиться
        #     ws.cell('j%s'%(row)).style = "ali_white"
        # except:
        #     ws.cell('g%s'%(row)).style = "ali_white"
        #     next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20

    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    
    output_name = 'daily_heat_tem104_'+translate(obj_title)+'_'+electric_data_start+'-'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response


def report_electric_3_zones_v2(request):
    SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')
    ROUND_SIZE = getattr(settings, 'ROUND_SIZE', 3)
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    wb.guess_types = True
    
    obj_title         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')

# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+' . Срез показаний с коэффициентами на дату' + ' ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование объекта'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"

    ws.merge_cells('B4:B5')
    ws['B4'] = 'Наименование абонента'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"
    
    ws.merge_cells('C4:C5')
    ws['C4'] = 'Заводской номер'
    ws['C4'].style = "ali_grey"
    ws['C5'].style = "ali_grey"
    
    ws.merge_cells('D4:F4')
    ws['D4'] = 'Коэффициенты'
    ws['D4'].style = "ali_grey"
    ws['F4'].style = "ali_grey"
    ws['E4'].style = "ali_grey"
    
    ws['D5'] = 'Ктн'
    ws['D5'].style = "ali_grey"
    ws['E5'] = 'Ктт'
    ws['E5'].style = "ali_grey"
    ws['F5'] = 'А'
    ws['F5'].style = "ali_grey" 
    
    # Сумма
    ws.merge_cells('G4:H4')
    ws['G4'] = 'Сумма'
    ws['G4'].style = "ali_grey"
    ws['G4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['K4'].style = "ali_grey"
    ws['G5'] = 'Показания A+ на ' + electric_data_end
    ws['G5'].style = "ali_grey"
    
    ws['H5'] = 'Энергия A+ на ' + electric_data_end
    ws['H5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('I4:J4')
    ws['I4'] = 'Тариф 1'
    ws['I4'].style = "ali_grey"
    ws['J4'].style = "ali_grey"
    ws['I4'].style = "ali_grey"
    ws['J4'].style = "ali_grey"
    ws['I5'] = 'Показания A+ на ' +  electric_data_end
    ws['I5'].style = "ali_grey"
    
    ws['J5'] = 'Энергия A+ на ' +  electric_data_end
    ws['J5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('K4:L4')
    ws['K4'] = 'Тариф 2'
    ws['K4'].style = "ali_grey"
    #ws['K5'].style = "ali_grey"
    ws['L4'].style = "ali_grey"
    #ws['K4'].style = "ali_grey"
    ws['K5'] = 'Показания A+ на ' +  electric_data_end
    ws['K5'].style = "ali_grey"
    
    ws['L5'] = 'Энергия A+ на ' +  electric_data_end
    ws['L5'].style = "ali_yellow"
    
    # Тариф 3
    ws.merge_cells('M4:N4')
    ws['M4'] = 'Тариф 3'
    ws['M4'].style = "ali_grey"
    ws['N4'].style = "ali_grey"
    #ws['L4'].style = "ali_grey"
    #ws['M4'].style = "ali_grey"
    ws['M5'] = 'Показания A+ на ' +  electric_data_end
    ws['M5'].style = "ali_grey"
    
    ws['N5'] = 'Энергия A+ на ' +  electric_data_end
    ws['N5'].style = "ali_yellow"

    if SHOW_LIC_NUM:
        ws.merge_cells('O4:O5')
        ws['O4'] = 'Лицевой номер абонента'
        ws['O4'].style = "ali_grey"
        ws['O5'].style = "ali_grey"
    
    ws.column_dimensions['O'].width = 17
    ws.row_dimensions[5].height = 43
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 13
    ws.column_dimensions['I'].width = 13
    ws.column_dimensions['J'].width = 13
    ws.column_dimensions['K'].width = 13
    ws.column_dimensions['L'].width = 13
    ws.column_dimensions['M'].width = 13
    ws.column_dimensions['N'].width = 13
        
# Шапка отчета конец
    
    #выборка данных из БД
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    
    obj_title           = request.GET['obj_title']
    obj_key             = request.GET['obj_key']
    obj_parent_title    = request.GET['obj_parent_title']
    is_electric_monthly = request.GET['is_electric_monthly']
    is_electric_daily   = request.GET['is_electric_daily']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']
    is_electric_period  = request.GET['is_electric_period']

    if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # месячные для абонента
        data_table = common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'monthly', True)
        
    elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # суточные для абонента
        data_table = common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'daily', True)

#*********************************************************************************************************************************************************************      
    elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # месячные для объекта
            data_table = common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'monthly', False)
            if obj_key == 'level1-0':
                data_table = common_sql.get_electric_by_date_level2(obj_parent_title, obj_title, electric_data_end, 'monthly')
            if not data_table:
                data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]        

#*********************************************************************************************************************************************************************
    elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for object
            data_table= common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'daily', False)
            #print(obj_key, obj_key == 'level2-0')
            if obj_key == 'level1-0':
                data_table = common_sql.get_electric_by_date_level2(obj_parent_title, obj_title, electric_data_end, 'daily')
            if not data_table:
                data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]

    elif (is_electric_daily == '1') & (bool(is_group_level.search(obj_key))): # показания по баланскной группе                    
            data_table = common_sql.get_electric_by_date_balance(obj_parent_title, obj_title, electric_data_end, 'daily')
            
        
    elif (is_electric_monthly == '1') & (bool(is_group_level.search(obj_key))): # показания по баланскной группе месячные
            data_table = common_sql.get_electric_by_date_balance(obj_parent_title, obj_title, electric_data_end, 'monthly')

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][7])  # Наименование
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:            
            ws.cell('C%s'%(row)).value = '%s' % str(data_table[row-6][2]) #заводской номер #.replace('.', separator)# Ктн
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % str(data_table[row-6][9]).replace('.', separator)  # Ктн
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next

        try:
            ws.cell('E%s'%(row)).value = '%s' % str(data_table[row-6][8]).replace('.', separator)  # Ктт
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next

        try:
            ws.cell('F%s'%(row)).value = '%s' % str(data_table[row-6][10]).replace('.', separator)  # Ка
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next

        try:
            ws.cell('G%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][3], ROUND_SIZE, separator)  #str(val).replace('.', separator)
            ws.cell('G%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
    
        try:            
            ws.cell('H%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][3]*data_table[row-6][8]*data_table[row-6][9]),ROUND_SIZE, separator)  # "Энергия Сумма А+
            ws.cell('H%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('H%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][4]),ROUND_SIZE, separator)  # Тариф 1 А+
            ws.cell('I%s'%(row)).style = "ali_white"
        except:
            ws.cell('I%s'%(row)).style = "ali_white"
            next
            
        try:
            #val = round((data_table[row-6][4]*data_table[row-6][8]*data_table[row-6][9]),ROUND_SIZE)            
            ws.cell('J%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][4]*data_table[row-6][8]*data_table[row-6][9]),ROUND_SIZE, separator)  # "Энергия Тариф 1 А+
            ws.cell('J%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('J%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][5],ROUND_SIZE, separator)  # Тариф 2 А+
            ws.cell('K%s'%(row)).style = "ali_white"
        except:
            ws.cell('K%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][5]*data_table[row-6][8]*data_table[row-6][9]),ROUND_SIZE, separator) # "Энергия Тариф 2 А+
            ws.cell('L%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('L%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][6]),ROUND_SIZE, separator)  # Тариф 3 А+
            ws.cell('M%s'%(row)).style = "ali_white"
        except:
            ws.cell('M%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('N%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][6]*data_table[row-6][8]*data_table[row-6][9]),ROUND_SIZE, separator)  # "Энергия Тариф 3 А+
            ws.cell('N%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('N%s'%(row)).style = "ali_yellow"
            next
        
        if SHOW_LIC_NUM:
            try:                
                ws.cell('O%s'%(row)).value = '%s' % (data_table[row-6][15])   # Лицевой нмоер абонента
                ws.cell('O%s'%(row)).style = "ali_white"
            except:
                ws.cell('O%s'%(row)).style = "ali_white"
                next

# Сохраняем в ecxel  
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb), content_type="application/vnd.ms-excel")
        
    output_name = u'3_tariffa_'+translate(obj_title)+'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def report_stk_heat_daily(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Пульс СТК. Показания по теплу на ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Энергия, Гкал'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Энергия, кВт*ч'
    ws['D5'].style = "ali_grey"
    
    #ws['D5'] = 'Объем, м3'
    #ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Температура входа, С'
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 'Температура выхода, С'
    ws['F5'].style = "ali_grey"
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         =  request.session['obj_parent_title']
    obj_title         =  request.session['obj_title']
    electric_data_end   =  request.session['electric_data_end']            
    obj_key             =  request.session['obj_key']
    
    data_table = []
                     
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_daily_pulsar_teplo(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_daily_pulsar_teplo(obj_parent_title, obj_title, electric_data_end, False)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # тип
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % get_val(data_table[row-6][9])  # энергия Гкал
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % get_val(data_table[row-6][3])  # энергия Кват*ч
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % get_val(data_table[row-6][5])  # t вх
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next

        try:
            ws.cell('F%s'%(row)).value = '%s' % get_val(data_table[row-6][6])  # t вых
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next

    ws.row_dimensions[5].height = 63
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20  
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20 
    ws.column_dimensions['E'].width = 17 
    ws.column_dimensions['F'].width = 17


#------------
            
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'stk_heat_report_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_stk_heat_period(request):
    #SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')
    ROUND_SIZE = getattr(settings, 'ROUND_SIZE', 3)
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]
    electric_data_start   = request.session["electric_data_start"]

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Пульс СТК. Потребление тепла с ' +str(electric_data_start)+' по '+ str(electric_data_end)
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Показания Энергии на ' + str(electric_data_start)+', Гкал '
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания Энергии на ' + str(electric_data_end)+', Гкал '
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Потребление Энергии, Гкал'
    ws['E5'].style = "ali_grey"

    ws['F5'] = 'Показания Энергии на ' + str(electric_data_start)+', кВт*ч '
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'Показания Энергии на ' + str(electric_data_end)+', кВт*ч '
    ws['G5'].style = "ali_grey"
    
    ws['H5'] = 'Потребление Энергии, кВт*ч'
    ws['H5'].style = "ali_grey"
    
#Запрашиваем данные для отчета

    
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
          
    obj_key             = request.session['obj_key']
             
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title,electric_data_end, electric_data_start, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title, electric_data_end,electric_data_start, False)
               
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next

        try:
            ws.cell('C%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][8]),ROUND_SIZE, separator)  # '%s' % get_val(round(float(data_table[row-6][2]),7)) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][9]),ROUND_SIZE, separator)  # '%s' % get_val(round(float(data_table[row-6][3]),7)) # '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][10]),ROUND_SIZE, separator)  # '%s' % get_val(round(float(data_table[row-6][4]),7))  # '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next

            
        try:
            ws.cell('F%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][2]),ROUND_SIZE, separator)  # '%s' % get_val(round(float(data_table[row-6][2]),7)) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][3]),ROUND_SIZE, separator)  # '%s' % get_val(round(float(data_table[row-6][3]),7)) # '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][4]),ROUND_SIZE, separator)  # '%s' % get_val(round(float(data_table[row-6][4]),7))  # '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
     
    ws.row_dimensions[5].height = 54
    ws.column_dimensions['A'].width = 30 
    ws.column_dimensions['B'].width = 20 
    ws.column_dimensions['C'].width = 20 
    ws.column_dimensions['D'].width = 20 
    ws.column_dimensions['E'].width = 20 
    ws.column_dimensions['F'].width = 20 
    ws.column_dimensions['G'].width = 20 
    ws.column_dimensions['H'].width = 20 

        
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_heat_period_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+str(electric_data_start)+'-'+str(electric_data_end)
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_pulsar_frost_daily(request):
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Пульсар. Показания по холоду на ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Энергия, Гкал'
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Объем, м3'
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Температура входа, С'
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 'Температура выхода, С'
    ws['F5'].style = "ali_grey"
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         =  request.session['obj_parent_title']
    obj_title         =  request.session['obj_title']
    electric_data_end   =  request.session['electric_data_end']            
    obj_key             =  request.session['obj_key']
    
    data_table = []
                     
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_daily_pulsar_frost(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_daily_pulsar_frost(obj_parent_title, obj_title, electric_data_end, False)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # тип
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % get_val(data_table[row-6][3])  # стояк
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value =  '%s' % get_val(data_table[row-6][4])  # счётчик
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % get_val(data_table[row-6][5])  # показаня
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % get_val(data_table[row-6][6])  # показаня
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next

    ws.row_dimensions[5].height = 63
    ws.column_dimensions['A'].width = 30 
    ws.column_dimensions['C'].width = 23 
    ws.column_dimensions['E'].width = 17 
    ws.column_dimensions['F'].width = 17
    ws.column_dimensions['D'].width = 17

#------------
            
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'pulsar_frost_report_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pulsar_frost_period(request):
    #SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')
    ROUND_SIZE = getattr(settings, 'ROUND_SIZE', 3)
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]
    electric_data_start   = request.session["electric_data_start"]

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Пульсар. Потребление холода с ' +str(electric_data_start)+' по '+ str(electric_data_end)
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = "ali_grey"
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = "ali_grey"
    
    ws['C5'] = 'Показания Энергии на ' + str(electric_data_start)+', Гкал '
    ws['C5'].style = "ali_grey"
    
    ws['D5'] = 'Показания Энергии на ' + str(electric_data_end)+', Гкал '
    ws['D5'].style = "ali_grey"
    
    ws['E5'] = 'Потребление Энергии, Гкал'
    ws['E5'].style = "ali_grey"
    
    ws['F5'] = 'Показания Объёма на ' + str(electric_data_start)+', м3'
    ws['F5'].style = "ali_grey"
    
    ws['G5'] = 'Показания Объёма на ' + str(electric_data_end)+', м3'
    ws['G5'].style = "ali_grey"
    
    ws['H5'] = 'Потребление Объёма, м3'
    ws['H5'].style = "ali_grey"
#Запрашиваем данные для отчета

    
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
          
    obj_key             = request.session['obj_key']
             
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_frost_for_period(obj_parent_title, obj_title,electric_data_end, electric_data_start, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_frost_for_period(obj_parent_title, obj_title, electric_data_end,electric_data_start, False)
               
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][2]),ROUND_SIZE, separator)  # '%s' % get_val(round(float(data_table[row-6][2]),7)) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][3]),ROUND_SIZE, separator)  # '%s' % get_val(round(float(data_table[row-6][3]),7)) # '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][4]),ROUND_SIZE, separator)  # '%s' % get_val(round(float(data_table[row-6][4]),7))  # '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][5]),ROUND_SIZE, separator)  #  '%s' % get_val(round(float(data_table[row-6][5]),7)) # '%s' % (data_table[row-6][5])  # Время работы
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][6]),ROUND_SIZE, separator)  # '%s' % get_val(round(float(data_table[row-6][6]),7)) # '%s' % (data_table[row-6][6])  # Время работы
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('H%s'%(row)).value =  '%s' % get_val_by_round(float(data_table[row-6][7]),ROUND_SIZE, separator)  # '%s' % get_val(round(float(data_table[row-6][7]),7))  #'%s' % (data_table[row-6][7])  # Время работы
            ws.cell('H%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
     
    ws.row_dimensions[5].height = 54
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20    
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17 
    ws.column_dimensions['E'].width = 17 
    ws.column_dimensions['F'].width = 17
    ws.column_dimensions['G'].width = 17
    ws.column_dimensions['H'].width = 17
        
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_frost_pulsar_period_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+str(electric_data_start)+'-'+str(electric_data_end)
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def get_month_name(date): #по дате возвращает рускоязычное название месяца
    datem = datetime.datetime.strptime(date, "%Y-%m-%d")
    num = datem.month
    month = ['Январь','Февраль','Март','Апрель','Май','Июнь','Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']
    #print(str(num))
    #print(month[num-1])
    return str(month[num-1])

def report_electric_by_day_for_year(request):
    #SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')
    ROUND_SIZE = getattr(settings, 'ROUND_SIZE', 3)
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    wb.guess_types = True
    
    obj_parent_title         = request.GET.get('obj_parent_title')
    obj_title         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')

# Шапка отчета    
    ws.merge_cells('A1:E1')
    ws['A1'] = obj_parent_title+'. ' + obj_title + '.'
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Электрика. Срез показаний с коэффициентами за год по ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование абонента'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер счётчика'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"

    ws.merge_cells('C4:C5')
    ws['C4'] = 'Тип прибора'
    ws['C4'].style = "ali_grey"
    ws['C5'].style = "ali_grey"
    
    ws.merge_cells('D4:F4')
    ws['D4'] = 'Коэффициенты'
    ws['D4'].style = "ali_grey"
    ws['E4'].style = "ali_grey"
    ws['F4'].style = "ali_grey"
    
    ws['D5'] = 'Ктн'
    ws['D5'].style = "ali_grey"
    ws['E5'] = 'Ктт'
    ws['E5'].style = "ali_grey"
    ws['F5'] = 'А'
    ws['F5'].style = "ali_grey" 
    
    data_table_dates_list = common_sql.get_data_table_12month(electric_data_end) #список дат по месяцам, начиная от выбранной даты
    #month0
    ws.merge_cells('G3:N3')
    ws['G3'] = get_month_name(str(data_table_dates_list[0][0]))
    ws['G3'].style = "ali_blue"
    ws['N3'].style = "ali_blue"
    # Сумма
    ws.merge_cells('G4:H4')
    ws['G4'] = 'Сумма'
    ws['G4'].style = "ali_grey"
    ws['H4'].style = "ali_grey"
    ws['G5'] = 'Показания A+ на ' + str(data_table_dates_list[0][0])
    ws['G5'].style = "ali_grey"
    
    ws['H5'] = 'Энергия A+ на ' + str(data_table_dates_list[0][0])
    ws['H5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('I4:J4')
    ws['I4'] = 'Тариф 1'
    ws['I4'].style = "ali_grey"
    ws['J4'].style = "ali_grey"
    ws['I5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][0])
    ws['I5'].style = "ali_grey"
    
    ws['J5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][0])
    ws['J5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('K4:L4')
    ws['K4'] = 'Тариф 2'
    ws['K4'].style = "ali_grey"
    ws['L4'].style = "ali_grey"
    ws['K5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][0])
    ws['K5'].style = "ali_grey"
    
    ws['L5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][0])
    ws['L5'].style = "ali_yellow"
    
    # Тариф 3
    ws.merge_cells('M4:N4')
    ws['M4'] = 'Тариф 3'
    ws['M4'].style = "ali_grey"
    ws['N4'].style = "ali_grey"
    ws['M5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][0])
    ws['M5'].style = "ali_grey"
    
    ws['N5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][0])
    ws['N5'].style = "ali_yellow"

#month1
    ws.merge_cells('O3:V3')
    ws['O3'] = get_month_name(str(data_table_dates_list[0][1]))
    ws['O3'].style = "ali_blue"
    ws['V3'].style = "ali_blue"
        # Сумма
    ws.merge_cells('O4:P4')
    ws['O4'] = 'Сумма'
    ws['O4'].style = "ali_grey"
    ws['P4'].style = "ali_grey"
    ws['O5'] = 'Показания A+ на ' + str(data_table_dates_list[0][1])
    ws['O5'].style = "ali_grey"
    
    ws['P5'] = 'Энергия A+ на ' + str(data_table_dates_list[0][1])
    ws['P5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('Q4:R4')
    ws['Q4'] = 'Тариф 1'
    ws['Q4'].style = "ali_grey"
    ws['R4'].style = "ali_grey"
    ws['Q5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][1])
    ws['Q5'].style = "ali_grey"
    
    ws['R5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][1])
    ws['R5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('S4:T4')
    ws['S4'] = 'Тариф 2'
    ws['S4'].style = "ali_grey"
    ws['T4'].style = "ali_grey"
    ws['S5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][1])
    ws['S5'].style = "ali_grey"
    
    ws['T5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][1])
    ws['T5'].style = "ali_yellow"
    
    # Тариф 3
    ws.merge_cells('U4:V4')
    ws['U4'] = 'Тариф 3'
    ws['U4'].style = "ali_grey"
    ws['V4'].style = "ali_grey"
    ws['U5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][1])
    ws['U5'].style = "ali_grey"
    
    ws['V5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][1])
    ws['V5'].style = "ali_yellow"
        
#month2
    ws.merge_cells('W3:AD3')
    ws['W3'] = get_month_name(str(data_table_dates_list[0][2]))
    ws['W3'].style = "ali_blue"
    ws['AD3'].style = "ali_blue"
        # Сумма
    ws.merge_cells('W4:X4')
    ws['W4'] = 'Сумма'
    ws['W4'].style = "ali_grey"
    ws['X4'].style = "ali_grey"
    ws['W5'] = 'Показания A+ на ' + str(data_table_dates_list[0][2])
    ws['W5'].style = "ali_grey"
    
    ws['X5'] = 'Энергия A+ на ' + str(data_table_dates_list[0][2])
    ws['X5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('Y4:Z4')
    ws['Y4'] = 'Тариф 1'
    ws['Y4'].style = "ali_grey"
    ws['Z4'].style = "ali_grey"
    ws['Y5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][2])
    ws['Y5'].style = "ali_grey"
    
    ws['Z5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][2])
    ws['Z5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('AA4:AB4')
    ws['AA4'] = 'Тариф 2'
    ws['AA4'].style = "ali_grey"
    ws['AB4'].style = "ali_grey"
    ws['AA5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][2])
    ws['AA5'].style = "ali_grey"
    
    ws['AB5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][2])
    ws['AB5'].style = "ali_yellow"
    
    # Тариф 3
    ws.merge_cells('AC4:AD4')
    ws['AC4'] = 'Тариф 3'
    ws['AC4'].style = "ali_grey"
    ws['AD4'].style = "ali_grey"
    ws['AC5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][2])
    ws['AC5'].style = "ali_grey"
    
    ws['AD5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][2])
    ws['AD5'].style = "ali_yellow"

#month3
    ws.merge_cells('AE3:AL3')
    ws['AE3'] = get_month_name(str(data_table_dates_list[0][3]))
    ws['AE3'].style = "ali_blue"
    ws['AL3'].style = "ali_blue"
        # Сумма
    ws.merge_cells('AE4:AF4')
    ws['AE4'] = 'Сумма'
    ws['AE4'].style = "ali_grey"
    ws['AF4'].style = "ali_grey"
    ws['AE5'] = 'Показания A+ на ' + str(data_table_dates_list[0][3])
    ws['AE5'].style = "ali_grey"
    
    ws['AF5'] = 'Энергия A+ на ' + str(data_table_dates_list[0][3])
    ws['AF5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('AG4:AH4')
    ws['AG4'] = 'Тариф 1'
    ws['AG4'].style = "ali_grey"
    ws['AH4'].style = "ali_grey"
    ws['AG5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][3])
    ws['AG5'].style = "ali_grey"
    
    ws['AH5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][3])
    ws['AH5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('AI4:AJ4')
    ws['AI4'] = 'Тариф 2'
    ws['AI4'].style = "ali_grey"
    ws['AJ4'].style = "ali_grey"
    ws['AI5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][3])
    ws['AI5'].style = "ali_grey"
    
    ws['AJ5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][3])
    ws['AJ5'].style = "ali_yellow"
    
    # Тариф 3
    ws.merge_cells('AK4:AL4')
    ws['AK4'] = 'Тариф 3'
    ws['AK4'].style = "ali_grey"
    ws['AK4'].style = "ali_grey"
    ws['AK5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][3])
    ws['AK5'].style = "ali_grey"
    
    ws['AL5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][3])
    ws['AL5'].style = "ali_yellow"

#month4
    ws.merge_cells('AM3:AT3')
    ws['AM3'] = get_month_name(str(data_table_dates_list[0][4]))
    ws['AM3'].style = "ali_blue"
    ws['AT3'].style = "ali_blue"
        # Сумма
    ws.merge_cells('AM4:AN4')
    ws['AM4'] = 'Сумма'
    ws['AM4'].style = "ali_grey"
    ws['AN4'].style = "ali_grey"
    ws['AM5'] = 'Показания A+ на ' + str(data_table_dates_list[0][4])
    ws['AM5'].style = "ali_grey"
    
    ws['AN5'] = 'Энергия A+ на ' + str(data_table_dates_list[0][4])
    ws['AN5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('AO4:AP4')
    ws['AO4'] = 'Тариф 1'
    ws['AO4'].style = "ali_grey"
    ws['AP4'].style = "ali_grey"
    ws['AO5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][4])
    ws['AO5'].style = "ali_grey"
    
    ws['AP5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][4])
    ws['AP5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('AQ4:AR4')
    ws['AQ4'] = 'Тариф 2'
    ws['AQ4'].style = "ali_grey"
    ws['AR4'].style = "ali_grey"
    ws['AQ5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][4])
    ws['AQ5'].style = "ali_grey"
    
    ws['AR5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][4])
    ws['AR5'].style = "ali_yellow"
    
    # Тариф 3
    ws.merge_cells('AS4:AT4')
    ws['AS4'] = 'Тариф 3'
    ws['AS4'].style = "ali_grey"
    ws['AT4'].style = "ali_grey"
    ws['AS5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][4])
    ws['AS5'].style = "ali_grey"
    
    ws['AT5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][4])
    ws['AT5'].style = "ali_yellow"

#month5
    ws.merge_cells('AU3:BB3')
    ws['AU3'] = get_month_name(str(data_table_dates_list[0][5]))
    ws['AU3'].style = "ali_blue"
    ws['BB3'].style = "ali_blue"
        # Сумма
    ws.merge_cells('AU4:AV4')
    ws['AU4'] = 'Сумма'
    ws['AU4'].style = "ali_grey"
    ws['AV4'].style = "ali_grey"
    ws['AU5'] = 'Показания A+ на ' + str(data_table_dates_list[0][5])
    ws['AU5'].style = "ali_grey"
    
    ws['AV5'] = 'Энергия A+ на ' + str(data_table_dates_list[0][5])
    ws['AV5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('AW4:AX4')
    ws['AW4'] = 'Тариф 1'
    ws['AW4'].style = "ali_grey"
    ws['AX4'].style = "ali_grey"
    ws['AW5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][5])
    ws['AW5'].style = "ali_grey"
    
    ws['AX5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][5])
    ws['AX5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('AY4:AZ4')
    ws['AY4'] = 'Тариф 2'
    ws['AY4'].style = "ali_grey"
    ws['AZ4'].style = "ali_grey"
    ws['AY5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][5])
    ws['AY5'].style = "ali_grey"
    
    ws['AZ5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][5])
    ws['AZ5'].style = "ali_yellow"
    
    # Тариф 3
    ws.merge_cells('BA4:BB4')
    ws['BA4'] = 'Тариф 3'
    ws['BA4'].style = "ali_grey"
    ws['BB4'].style = "ali_grey"
    ws['BA5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][5])
    ws['BA5'].style = "ali_grey"
    
    ws['BB5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][5])
    ws['BB5'].style = "ali_yellow"

#month6
    ws.merge_cells('BC3:BJ3')
    ws['BC3'] = get_month_name(str(data_table_dates_list[0][6]))
    ws['BC3'].style = "ali_blue"
    ws['BJ3'].style = "ali_blue"

     # Сумма
    ws.merge_cells('BC4:BD4')
    ws['BC4'] = 'Сумма'
    ws['BC4'].style = "ali_grey"
    ws['BD4'].style = "ali_grey"
    ws['BC5'] = 'Показания A+ на ' + str(data_table_dates_list[0][6])
    ws['BC5'].style = "ali_grey"
    
    ws['BD5'] = 'Энергия A+ на ' + str(data_table_dates_list[0][6])
    ws['BD5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('BE4:BF4')
    ws['BE4'] = 'Тариф 1'
    ws['BE4'].style = "ali_grey"
    ws['BF4'].style = "ali_grey"
    ws['BE5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][6])
    ws['BE5'].style = "ali_grey"
    
    ws['BF5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][6])
    ws['BF5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('BG4:BH4')
    ws['BG4'] = 'Тариф 2'
    ws['BG4'].style = "ali_grey"
    ws['BH4'].style = "ali_grey"
    ws['BG5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][6])
    ws['BG5'].style = "ali_grey"
    
    ws['BH5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][6])
    ws['BH5'].style = "ali_yellow"
    
    # Тариф 3
    ws.merge_cells('BI4:BJ4')
    ws['BI4'] = 'Тариф 3'
    ws['BI4'].style = "ali_grey"
    ws['BJ4'].style = "ali_grey"
    ws['BI5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][6])
    ws['BI5'].style = "ali_grey"
    
    ws['BJ5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][6])
    ws['BJ5'].style = "ali_yellow"

#month7
    ws.merge_cells('BK3:BR3')
    ws['BK3'] = get_month_name(str(data_table_dates_list[0][7]))
    ws['BK3'].style = "ali_blue"
    ws['BR3'].style = "ali_blue"

            # Сумма
    ws.merge_cells('BK4:BL4')
    ws['BK4'] = 'Сумма'
    ws['BK4'].style = "ali_grey"
    ws['BL4'].style = "ali_grey"
    ws['BK5'] = 'Показания A+ на ' + str(data_table_dates_list[0][7])
    ws['BK5'].style = "ali_grey"
    
    ws['BL5'] = 'Энергия A+ на ' + str(data_table_dates_list[0][7])
    ws['BL5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('BM4:BN4')
    ws['BM4'] = 'Тариф 1'
    ws['BM4'].style = "ali_grey"
    ws['BN4'].style = "ali_grey"
    ws['BM5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][7])
    ws['BM5'].style = "ali_grey"
    
    ws['BN5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][7])
    ws['BN5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('BO4:BP4')
    ws['BO4'] = 'Тариф 2'
    ws['BO4'].style = "ali_grey"
    ws['BP4'].style = "ali_grey"
    ws['BO5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][7])
    ws['BO5'].style = "ali_grey"
    
    ws['BP5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][7])
    ws['BP5'].style = "ali_yellow"
    
    # Тариф 3
    ws.merge_cells('BQ4:BR4')
    ws['BQ4'] = 'Тариф 3'
    ws['BQ4'].style = "ali_grey"
    ws['BR4'].style = "ali_grey"
    ws['BQ5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][7])
    ws['BQ5'].style = "ali_grey"
    
    ws['BR5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][7])
    ws['BR5'].style = "ali_yellow"

#month8
    ws.merge_cells('BS3:BZ3')
    ws['BS3'] = get_month_name(str(data_table_dates_list[0][8]))
    ws['BS3'].style = "ali_blue"
    ws['BZ3'].style = "ali_blue"

        # Сумма
    ws.merge_cells('BS4:BT4')
    ws['BS4'] = 'Сумма'
    ws['BS4'].style = "ali_grey"
    ws['BT4'].style = "ali_grey"
    ws['BS5'] = 'Показания A+ на ' + str(data_table_dates_list[0][8])
    ws['BS5'].style = "ali_grey"
    
    ws['BT5'] = 'Энергия A+ на ' + str(data_table_dates_list[0][8])
    ws['BT5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('BU4:BV4')
    ws['BU4'] = 'Тариф 1'
    ws['BU4'].style = "ali_grey"
    ws['BV4'].style = "ali_grey"
    ws['BU5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][8])
    ws['BU5'].style = "ali_grey"
    
    ws['BV5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][8])
    ws['BV5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('BW4:BX4')
    ws['BW4'] = 'Тариф 2'
    ws['BW4'].style = "ali_grey"
    ws['BX4'].style = "ali_grey"
    ws['BW5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][8])
    ws['BW5'].style = "ali_grey"
    
    ws['BX5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][8])
    ws['BX5'].style = "ali_yellow"
    
    # Тариф 3
    ws.merge_cells('BY4:BZ4')
    ws['BY4'] = 'Тариф 3'
    ws['BY4'].style = "ali_grey"
    ws['BZ4'].style = "ali_grey"
    ws['BY5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][8])
    ws['BY5'].style = "ali_grey"
    
    ws['BZ5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][8])
    ws['BZ5'].style = "ali_yellow"

    #month9
    ws.merge_cells('CA3:CH3')
    ws['CA3'] = get_month_name(str(data_table_dates_list[0][9]))
    ws['CA3'].style = "ali_blue"
    ws['CH3'].style = "ali_blue"

            # Сумма
    ws.merge_cells('CA4:CB4')
    ws['CA4'] = 'Сумма'
    ws['CA4'].style = "ali_grey"
    ws['CB4'].style = "ali_grey"
    ws['CA5'] = 'Показания A+ на ' + str(data_table_dates_list[0][9])
    ws['CA5'].style = "ali_grey"
    
    ws['CB5'] = 'Энергия A+ на ' + str(data_table_dates_list[0][9])
    ws['CB5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('CC4:CD4')
    ws['CC4'] = 'Тариф 1'
    ws['CC4'].style = "ali_grey"
    ws['CD4'].style = "ali_grey"
    ws['CC5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][9])
    ws['CC5'].style = "ali_grey"
    
    ws['CD5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][9])
    ws['CD5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('CE4:CF4')
    ws['CE4'] = 'Тариф 2'
    ws['CE4'].style = "ali_grey"
    ws['CF4'].style = "ali_grey"
    ws['CE5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][9])
    ws['CE5'].style = "ali_grey"
    
    ws['CF5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][9])
    ws['CF5'].style = "ali_yellow"
    
    # Тариф 3
    ws.merge_cells('CG4:CH4')
    ws['CG4'] = 'Тариф 3'
    ws['CG4'].style = "ali_grey"
    ws['CH4'].style = "ali_grey"
    ws['CG5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][9])
    ws['CG5'].style = "ali_grey"
    
    ws['CH5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][9])
    ws['CH5'].style = "ali_yellow"

#month10
    ws.merge_cells('Ci3:CP3')
    ws['CI3'] = get_month_name(str(data_table_dates_list[0][10]))
    ws['CI3'].style = "ali_blue"
    ws['CP3'].style = "ali_blue"

        # Сумма
    ws.merge_cells('CI4:CJ4')
    ws['CI4'] = 'Сумма'
    ws['CI4'].style = "ali_grey"
    ws['CJ4'].style = "ali_grey"
    ws['CI5'] = 'Показания A+ на ' + str(data_table_dates_list[0][10])
    ws['CI5'].style = "ali_grey"
    
    ws['CJ5'] = 'Энергия A+ на ' + str(data_table_dates_list[0][10])
    ws['CJ5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('CK4:CL4')
    ws['CK4'] = 'Тариф 1'
    ws['CK4'].style = "ali_grey"
    ws['CL4'].style = "ali_grey"
    ws['CK5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][10])
    ws['CK5'].style = "ali_grey"
    
    ws['CL5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][10])
    ws['CL5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('CM4:CN4')
    ws['CM4'] = 'Тариф 2'
    ws['CM4'].style = "ali_grey"
    ws['CN4'].style = "ali_grey"
    ws['CM5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][10])
    ws['CM5'].style = "ali_grey"
    
    ws['CN5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][10])
    ws['CN5'].style = "ali_yellow"
    
    # Тариф 3
    ws.merge_cells('CO4:CP4')
    ws['CO4'] = 'Тариф 3'
    ws['CO4'].style = "ali_grey"
    ws['CP4'].style = "ali_grey"
    ws['CO5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][10])
    ws['CO5'].style = "ali_grey"
    
    ws['CP5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][10])
    ws['CP5'].style = "ali_yellow"

    #month11
    ws.merge_cells('CQ3:CX3')
    ws['CQ3'] = get_month_name(str(data_table_dates_list[0][11]))
    ws['CQ3'].style = "ali_blue"
    ws['CX3'].style = "ali_blue"

            # Сумма
    ws.merge_cells('CQ4:CR4')
    ws['CQ4'] = 'Сумма'
    ws['CQ4'].style = "ali_grey"
    ws['CR4'].style = "ali_grey"
    ws['CQ5'] = 'Показания A+ на ' + str(data_table_dates_list[0][11])
    ws['CQ5'].style = "ali_grey"
    
    ws['CR5'] = 'Энергия A+ на ' + str(data_table_dates_list[0][11])
    ws['CR5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('CS4:CT4')
    ws['CS4'] = 'Тариф 1'
    ws['CS4'].style = "ali_grey"
    ws['CT4'].style = "ali_grey"
    ws['CS5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][11])
    ws['CS5'].style = "ali_grey"
    
    ws['CT5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][11])
    ws['CT5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('CU4:CV4')
    ws['CU4'] = 'Тариф 2'
    ws['CU4'].style = "ali_grey"
    ws['CV4'].style = "ali_grey"
    ws['CU5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][11])
    ws['CU5'].style = "ali_grey"
    
    ws['CV5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][11])
    ws['CV5'].style = "ali_yellow"
    
    # Тариф 3
    ws.merge_cells('CW4:CX4')
    ws['CW4'] = 'Тариф 3'
    ws['CW4'].style = "ali_grey"
    ws['CX4'].style = "ali_grey"
    ws['CW5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][11])
    ws['CW5'].style = "ali_grey"
    
    ws['CX5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][11])
    ws['CX5'].style = "ali_yellow"

    #month12
    ws.merge_cells('CY3:DF3')
    ws['CY3'] = get_month_name(str(data_table_dates_list[0][12]))
    ws['CY3'].style = "ali_blue"
    ws['DF3'].style = "ali_blue"

            # Сумма
    ws.merge_cells('CY4:CZ4')
    ws['CY4'] = 'Сумма'
    ws['CY4'].style = "ali_grey"
    ws['CZ4'].style = "ali_grey"
    ws['CY5'] = 'Показания A+ на ' + str(data_table_dates_list[0][12])
    ws['CY5'].style = "ali_grey"
    
    ws['CZ5'] = 'Энергия A+ на ' + str(data_table_dates_list[0][12])
    ws['CZ5'].style = "ali_yellow"
    
    # Тариф 1
    ws.merge_cells('DA4:DB4')
    ws['DA4'] = 'Тариф 1'
    ws['DA4'].style = "ali_grey"
    ws['DB4'].style = "ali_grey"
    ws['DA5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][12])
    ws['DA5'].style = "ali_grey"
    
    ws['DB5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][12])
    ws['DB5'].style = "ali_yellow"
    
    # Тариф 2
    ws.merge_cells('DC4:DD4')
    ws['DC4'] = 'Тариф 2'
    ws['DC4'].style = "ali_grey"
    ws['DD4'].style = "ali_grey"
    ws['DC5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][12])
    ws['DC5'].style = "ali_grey"
    
    ws['DD5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][12])
    ws['DD5'].style = "ali_yellow"
    
    # Тариф 3
    ws.merge_cells('DE4:DF4')
    ws['DE4'] = 'Тариф 3'
    ws['DE4'].style = "ali_grey"
    ws['DF4'].style = "ali_grey"
    ws['DE5'] = 'Показания A+ на ' +  str(data_table_dates_list[0][12])
    ws['DE5'].style = "ali_grey"
    
    ws['DF5'] = 'Энергия A+ на ' +  str(data_table_dates_list[0][12])
    ws['DF5'].style = "ali_yellow"

# Шапка отчета конец
    
    #выборка данных из БД
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    
    obj_title           = request.GET['obj_title']
    obj_key             = request.GET['obj_key']
    obj_parent_title    = request.GET['obj_parent_title']
    is_electric_monthly = request.GET['is_electric_monthly']
    is_electric_daily   = request.GET['is_electric_daily']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']
    is_electric_period  = request.GET['is_electric_period']
    
    #электрика
    data_table = common_sql.get_data_table_electric_for_year_by_day(obj_parent_title, obj_title, electric_data_end)

    #data_table_water = common_sql.get_data_table_water_for_year_by_day(obj_parent_title, obj_title, electric_data_end)
    #data_table_heat = common_sql.get_data_table_heat_for_year_by_day(obj_parent_title, obj_title, electric_data_end)

#Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование фбонента
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:            
            ws.cell('C%s'%(row)).value = '%s' % str(data_table[row-6][3])# тип прибора
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next

        try:            
            ws.cell('D%s'%(row)).value = '%s' % str(data_table[row-6][4]).replace('.', separator)# Ктн
            ws.cell('D%s'%(row)).style = "ali_white"
        except:
            ws.cell('D%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % str(data_table[row-6][5]).replace('.', separator)  # Ктт
            ws.cell('E%s'%(row)).style = "ali_white"
        except:
            ws.cell('E%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s' % str(data_table[row-6][6]).replace('.', separator)  # Ка
            ws.cell('F%s'%(row)).style = "ali_white"
        except:
            ws.cell('F%s'%(row)).style = "ali_white"
            next
        
        #print(get_val_by_round(data_table[row-6][3], ROUND_SIZE, separator), data_table[row-6][3], ROUND_SIZE, separator)
        try:
            ws.cell('G%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][7], ROUND_SIZE, separator)  #str(val).replace('.', separator)
            ws.cell('G%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
    
        try:            
            ws.cell('H%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][7]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Сумма А+
            ws.cell('H%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('H%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][8]),ROUND_SIZE, separator)  # Тариф 1 А+
            ws.cell('I%s'%(row)).style = "ali_white"
        except:
            ws.cell('I%s'%(row)).style = "ali_white"
            next
            
        try:
            #val = round((data_table[row-6][4]*data_table[row-6][8]*data_table[row-6][9]),ROUND_SIZE)            
            ws.cell('J%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][8]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 1 А+
            ws.cell('J%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('J%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][9],ROUND_SIZE, separator)  # Тариф 2 А+
            ws.cell('K%s'%(row)).style = "ali_white"
        except:
            ws.cell('K%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][9]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator) # "Энергия Тариф 2 А+
            ws.cell('L%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('L%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][10]),ROUND_SIZE, separator)  # Тариф 3 А+
            ws.cell('M%s'%(row)).style = "ali_white"
        except:
            ws.cell('M%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('N%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][10]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 3 А+
            ws.cell('N%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('N%s'%(row)).style = "ali_yellow"
            next

        #month1
        try:
            ws.cell('O%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][11], ROUND_SIZE, separator)  #str(val).replace('.', separator)
            ws.cell('O%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('O%s'%(row)).style = "ali_white"
            next
    
        try:            
            ws.cell('P%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][11]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Сумма А+
            ws.cell('P%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('p%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('Q%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][12]),ROUND_SIZE, separator)  # Тариф 1 А+
            ws.cell('Q%s'%(row)).style = "ali_white"
        except:
            ws.cell('Q%s'%(row)).style = "ali_white"
            next
            
        try:           
            ws.cell('R%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][12]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 1 А+
            ws.cell('R%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('R%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('S%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][13],ROUND_SIZE, separator)  # Тариф 2 А+
            ws.cell('S%s'%(row)).style = "ali_white"
        except:
            ws.cell('S%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('T%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][13]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator) # "Энергия Тариф 2 А+
            ws.cell('T%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('T%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('U%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][14]),ROUND_SIZE, separator)  # Тариф 3 А+
            ws.cell('U%s'%(row)).style = "ali_white"
        except:
            ws.cell('U%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('V%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][14]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 3 А+
            ws.cell('V%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('V%s'%(row)).style = "ali_yellow"
            next

#month2
        try:
            ws.cell('W%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][15], ROUND_SIZE, separator)  #str(val).replace('.', separator)
            ws.cell('W%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('W%s'%(row)).style = "ali_white"
            next
    
        try:            
            ws.cell('X%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][15]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Сумма А+
            ws.cell('X%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('X%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('Y%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][16]),ROUND_SIZE, separator)  # Тариф 1 А+
            ws.cell('Y%s'%(row)).style = "ali_white"
        except:
            ws.cell('Y%s'%(row)).style = "ali_white"
            next
            
        try:           
            ws.cell('Z%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][16]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 1 А+
            ws.cell('Z%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('Z%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('AA%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][17],ROUND_SIZE, separator)  # Тариф 2 А+
            ws.cell('AA%s'%(row)).style = "ali_white"
        except:
            ws.cell('AA%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('AB%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][17]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator) # "Энергия Тариф 2 А+
            ws.cell('AB%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('AB%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('AC%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][18]),ROUND_SIZE, separator)  # Тариф 3 А+
            ws.cell('AC%s'%(row)).style = "ali_white"
        except:
            ws.cell('AC%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('AD%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][18]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 3 А+
            ws.cell('AD%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('AD%s'%(row)).style = "ali_yellow"
            next

            #MONTH3
        try:
            ws.cell('AE%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][19], ROUND_SIZE, separator)  #str(val).replace('.', separator)
            ws.cell('AE%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('AE%s'%(row)).style = "ali_white"
            next
    
        try:            
            ws.cell('AF%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][19]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Сумма А+
            ws.cell('AF%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('AF%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('AG%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][20]),ROUND_SIZE, separator)  # Тариф 1 А+
            ws.cell('AG%s'%(row)).style = "ali_white"
        except:
            ws.cell('AG%s'%(row)).style = "ali_white"
            next
            
        try:           
            ws.cell('AH%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][20]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 1 А+
            ws.cell('AH%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('AH%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('AI%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][21],ROUND_SIZE, separator)  # Тариф 2 А+
            ws.cell('AI%s'%(row)).style = "ali_white"
        except:
            ws.cell('AI%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('AJ%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][21]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator) # "Энергия Тариф 2 А+
            ws.cell('AJ%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('AJ%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('AK%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][22]),ROUND_SIZE, separator)  # Тариф 3 А+
            ws.cell('AK%s'%(row)).style = "ali_white"
        except:
            ws.cell('AK%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('AL%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][22]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 3 А+
            ws.cell('AL%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('AL%s'%(row)).style = "ali_yellow"
            next

            #MONTH4
        try:
            ws.cell('Am%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][23], ROUND_SIZE, separator)  #str(val).replace('.', separator)
            ws.cell('AM%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('AM%s'%(row)).style = "ali_white"
            next
    
        try:            
            ws.cell('AN%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][23]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Сумма А+
            ws.cell('AN%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('AN%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('AO%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][24]),ROUND_SIZE, separator)  # Тариф 1 А+
            ws.cell('AO%s'%(row)).style = "ali_white"
        except:
            ws.cell('AO%s'%(row)).style = "ali_white"
            next
            
        try:           
            ws.cell('AP%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][24]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 1 А+
            ws.cell('AP%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('AP%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('AQ%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][25],ROUND_SIZE, separator)  # Тариф 2 А+
            ws.cell('AQ%s'%(row)).style = "ali_white"
        except:
            ws.cell('AQ%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('AR%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][25]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator) # "Энергия Тариф 2 А+
            ws.cell('AR%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('AR%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('AS%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][26]),ROUND_SIZE, separator)  # Тариф 3 А+
            ws.cell('AS%s'%(row)).style = "ali_white"
        except:
            ws.cell('AS%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('AT%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][26]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 3 А+
            ws.cell('AT%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('AT%s'%(row)).style = "ali_yellow"
            next

            #MONTH5
        try:
            ws.cell('AU%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][27], ROUND_SIZE, separator)  #str(val).replace('.', separator)
            ws.cell('AU%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('AU%s'%(row)).style = "ali_white"
            next
    
        try:            
            ws.cell('AV%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][27]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Сумма А+
            ws.cell('AV%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('AV%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('AW%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][28]),ROUND_SIZE, separator)  # Тариф 1 А+
            ws.cell('AW%s'%(row)).style = "ali_white"
        except:
            ws.cell('AW%s'%(row)).style = "ali_white"
            next
            
        try:           
            ws.cell('AX%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][28]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 1 А+
            ws.cell('AX%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('AX%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('AY%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][29],ROUND_SIZE, separator)  # Тариф 2 А+
            ws.cell('AY%s'%(row)).style = "ali_white"
        except:
            ws.cell('AY%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('AZ%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][29]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator) # "Энергия Тариф 2 А+
            ws.cell('AZ%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('AZ%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('BA%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][30]),ROUND_SIZE, separator)  # Тариф 3 А+
            ws.cell('BA%s'%(row)).style = "ali_white"
        except:
            ws.cell('BA%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('BB%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][30]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 3 А+
            ws.cell('BB%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('BB%s'%(row)).style = "ali_yellow"
            next

            #MONTH6
        try:
            ws.cell('BC%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][31], ROUND_SIZE, separator)  #str(val).replace('.', separator)
            ws.cell('BC%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('BC%s'%(row)).style = "ali_white"
            next
    
        try:            
            ws.cell('BD%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][31]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Сумма А+
            ws.cell('BD%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('BD%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('BE%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][32]),ROUND_SIZE, separator)  # Тариф 1 А+
            ws.cell('BE%s'%(row)).style = "ali_white"
        except:
            ws.cell('BE%s'%(row)).style = "ali_white"
            next
            
        try:           
            ws.cell('BF%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][32]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 1 А+
            ws.cell('BF%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('BF%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('BG%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][33],ROUND_SIZE, separator)  # Тариф 2 А+
            ws.cell('BG%s'%(row)).style = "ali_white"
        except:
            ws.cell('BG%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('BH%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][33]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator) # "Энергия Тариф 2 А+
            ws.cell('BH%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('BH%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('BI%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][34]),ROUND_SIZE, separator)  # Тариф 3 А+
            ws.cell('BI%s'%(row)).style = "ali_white"
        except:
            ws.cell('BI%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('BJ%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][34]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 3 А+
            ws.cell('BJ%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('BJ%s'%(row)).style = "ali_yellow"
            next

            #MONTH7
        try:
            ws.cell('BK%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][35], ROUND_SIZE, separator)  #str(val).replace('.', separator)
            ws.cell('BK%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('BK%s'%(row)).style = "ali_white"
            next
    
        try:            
            ws.cell('BL%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][35]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Сумма А+
            ws.cell('BL%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('BL%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('BM%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][36]),ROUND_SIZE, separator)  # Тариф 1 А+
            ws.cell('BM%s'%(row)).style = "ali_white"
        except:
            ws.cell('BM%s'%(row)).style = "ali_white"
            next
            
        try:           
            ws.cell('BN%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][36]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 1 А+
            ws.cell('BN%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('BN%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('BO%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][37],ROUND_SIZE, separator)  # Тариф 2 А+
            ws.cell('BO%s'%(row)).style = "ali_white"
        except:
            ws.cell('BO%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('BP%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][37]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator) # "Энергия Тариф 2 А+
            ws.cell('BP%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('BP%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('BQ%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][38]),ROUND_SIZE, separator)  # Тариф 3 А+
            ws.cell('BQ%s'%(row)).style = "ali_white"
        except:
            ws.cell('BQ%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('BR%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][38]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 3 А+
            ws.cell('BR%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('BR%s'%(row)).style = "ali_yellow"
            next


            #MONTH8
        try:
            ws.cell('BS%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][39], ROUND_SIZE, separator)  #str(val).replace('.', separator)
            ws.cell('BS%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('BS%s'%(row)).style = "ali_white"
            next
    
        try:            
            ws.cell('BT%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][39]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Сумма А+
            ws.cell('BT%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('BT%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('BU%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][40]),ROUND_SIZE, separator)  # Тариф 1 А+
            ws.cell('BU%s'%(row)).style = "ali_white"
        except:
            ws.cell('BU%s'%(row)).style = "ali_white"
            next
            
        try:           
            ws.cell('BV%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][40]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 1 А+
            ws.cell('BV%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('BV%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('Bw%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][41],ROUND_SIZE, separator)  # Тариф 2 А+
            ws.cell('BW%s'%(row)).style = "ali_white"
        except:
            ws.cell('BW%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('BX%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][41]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator) # "Энергия Тариф 2 А+
            ws.cell('BX%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('BX%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('BY%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][42]),ROUND_SIZE, separator)  # Тариф 3 А+
            ws.cell('BY%s'%(row)).style = "ali_white"
        except:
            ws.cell('BY%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('BZ%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][42]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 3 А+
            ws.cell('BZ%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('BZ%s'%(row)).style = "ali_yellow"
            next


            #MONTH9
        try:
            ws.cell('CA%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][43], ROUND_SIZE, separator)  #str(val).replace('.', separator)
            ws.cell('CA%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('CA%s'%(row)).style = "ali_white"
            next
    
        try:            
            ws.cell('CB%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][43]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Сумма А+
            ws.cell('CB%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('CB%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('CC%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][44]),ROUND_SIZE, separator)  # Тариф 1 А+
            ws.cell('CC%s'%(row)).style = "ali_white"
        except:
            ws.cell('CC%s'%(row)).style = "ali_white"
            next
            
        try:           
            ws.cell('CD%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][44]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 1 А+
            ws.cell('CD%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('CD%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('CE%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][45],ROUND_SIZE, separator)  # Тариф 2 А+
            ws.cell('CE%s'%(row)).style = "ali_white"
        except:
            ws.cell('CE%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('CF%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][45]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator) # "Энергия Тариф 2 А+
            ws.cell('CF%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('CF%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('CG%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][46]),ROUND_SIZE, separator)  # Тариф 3 А+
            ws.cell('CG%s'%(row)).style = "ali_white"
        except:
            ws.cell('CG%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('CH%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][46]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 3 А+
            ws.cell('CH%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('CH%s'%(row)).style = "ali_yellow"
            next


            #MONTH10
        try:
            ws.cell('CI%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][47], ROUND_SIZE, separator)  #str(val).replace('.', separator)
            ws.cell('CI%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('CI%s'%(row)).style = "ali_white"
            next
    
        try:            
            ws.cell('CJ%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][47]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Сумма А+
            ws.cell('CJ%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('CJ%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('CK%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][48]),ROUND_SIZE, separator)  # Тариф 1 А+
            ws.cell('CK%s'%(row)).style = "ali_white"
        except:
            ws.cell('CK%s'%(row)).style = "ali_white"
            next
            
        try:           
            ws.cell('CL%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][48]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 1 А+
            ws.cell('CL%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('CL%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('CM%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][49],ROUND_SIZE, separator)  # Тариф 2 А+
            ws.cell('CM%s'%(row)).style = "ali_white"
        except:
            ws.cell('CM%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('CN%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][49]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator) # "Энергия Тариф 2 А+
            ws.cell('CN%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('CN%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('CO%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][50]),ROUND_SIZE, separator)  # Тариф 3 А+
            ws.cell('CO%s'%(row)).style = "ali_white"
        except:
            ws.cell('CO%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('CP%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][50]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 3 А+
            ws.cell('CP%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('CP%s'%(row)).style = "ali_yellow"
            next


            #MONTH11
        try:
            ws.cell('CQ%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][51], ROUND_SIZE, separator)  #str(val).replace('.', separator)
            ws.cell('CQ%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('CQ%s'%(row)).style = "ali_white"
            next
    
        try:            
            ws.cell('CR%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][51]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Сумма А+
            ws.cell('CR%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('CR%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('CS%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][52]),ROUND_SIZE, separator)  # Тариф 1 А+
            ws.cell('CS%s'%(row)).style = "ali_white"
        except:
            ws.cell('CS%s'%(row)).style = "ali_white"
            next
            
        try:           
            ws.cell('CT%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][52]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 1 А+
            ws.cell('CT%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('CT%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('CU%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][53],ROUND_SIZE, separator)  # Тариф 2 А+
            ws.cell('CU%s'%(row)).style = "ali_white"
        except:
            ws.cell('CU%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('CV%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][53]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator) # "Энергия Тариф 2 А+
            ws.cell('CV%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('CV%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('CW%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][54]),ROUND_SIZE, separator)  # Тариф 3 А+
            ws.cell('CW%s'%(row)).style = "ali_white"
        except:
            ws.cell('CW%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('CX%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][54]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 3 А+
            ws.cell('CX%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('CX%s'%(row)).style = "ali_yellow"
            next


            #MONTH12
        try:
            ws.cell('CY%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][55], ROUND_SIZE, separator)  #str(val).replace('.', separator)
            ws.cell('CY%s'%(row)).style = "ali_white"
            
        except:
            ws.cell('CY%s'%(row)).style = "ali_white"
            next
    
        try:            
            ws.cell('CZ%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][55]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Сумма А+
            ws.cell('CZ%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('CZ%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('DA%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][56]),ROUND_SIZE, separator)  # Тариф 1 А+
            ws.cell('DA%s'%(row)).style = "ali_white"
        except:
            ws.cell('DA%s'%(row)).style = "ali_white"
            next
            
        try:           
            ws.cell('DB%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][56]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 1 А+
            ws.cell('DB%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('DB%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('DC%s'%(row)).value = '%s' % get_val_by_round(data_table[row-6][57],ROUND_SIZE, separator)  # Тариф 2 А+
            ws.cell('DC%s'%(row)).style = "ali_white"
        except:
            ws.cell('DC%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('DD%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][57]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator) # "Энергия Тариф 2 А+
            ws.cell('DD%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('DD%s'%(row)).style = "ali_yellow"
            next
            
        try:
            ws.cell('DE%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][58]),ROUND_SIZE, separator)  # Тариф 3 А+
            ws.cell('DE%s'%(row)).style = "ali_white"
        except:
            ws.cell('DE%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('DF%s'%(row)).value = '%s' % get_val_by_round((data_table[row-6][58]*data_table[row-6][4]*data_table[row-6][5]),ROUND_SIZE, separator)  # "Энергия Тариф 3 А+
            ws.cell('DF%s'%(row)).style = "ali_yellow"
        except:
            ws.cell('DF%s'%(row)).style = "ali_yellow"
            next
# Сохраняем в ecxel  

    ws.row_dimensions[5].height = 60
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb), content_type="application/vnd.ms-excel")
        
    output_name = u'electric_year_'+translate(obj_title)+'_' + electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def report_water_by_day_for_year(request):
    ROUND_SIZE = getattr(settings, 'ROUND_SIZE', 'False')
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    obj_parent_title         = request.GET.get('obj_parent_title')
    obj_title         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')
    
#Шапка
    ws.merge_cells('A1:E1')
    ws['A1'] = obj_parent_title+'. ' + obj_title + '.'
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Вода. Срез показаний за год по ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование абонента'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер счётчика'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"

    ws.merge_cells('C4:C5')
    ws['C4'] = 'Тип прибора'
    ws['C4'].style = "ali_grey"
    ws['C5'].style = "ali_grey"
 
    data_table_dates_list = common_sql.get_data_table_12month(electric_data_end) #список дат по месяцам, начиная от выбранной даты
    #month0
    ws['d4'] = get_month_name(str(data_table_dates_list[0][0]))
    ws['d4'].style = "ali_blue"
    ws['d5'] = 'Объём на '  + str(data_table_dates_list[0][0]) + ', м3'
    ws['d5'].style = "ali_grey"
    
    #month1
    ws['e4'] = get_month_name(str(data_table_dates_list[0][1]))
    ws['E4'].style = "ali_blue"
    ws['E5'] = 'Объём на '  + str(data_table_dates_list[0][1]) + ', м3'
    ws['E5'].style = "ali_grey"

        #month2
    ws['F4'] = get_month_name(str(data_table_dates_list[0][2]))
    ws['F4'].style = "ali_blue"
    ws['F5'] = 'Объём на '  + str(data_table_dates_list[0][2]) + ', м3'
    ws['F5'].style = "ali_grey"
    
        #month3
    ws['G4'] = get_month_name(str(data_table_dates_list[0][3]))
    ws['G4'].style = "ali_blue"
    ws['G5'] = 'Объём на '  + str(data_table_dates_list[0][3]) + ', м3'
    ws['G5'].style = "ali_grey"

        #month4
    ws['H4'] = get_month_name(str(data_table_dates_list[0][4]))
    ws['H4'].style = "ali_blue"
    ws['H5'] = 'Объём на '  + str(data_table_dates_list[0][4]) + ', м3'
    ws['H5'].style = "ali_grey"
        #month5
    ws['I4'] = get_month_name(str(data_table_dates_list[0][5]))
    ws['I4'].style = "ali_blue"
    ws['I5'] = 'Объём на '  + str(data_table_dates_list[0][5]) + ', м3'
    ws['I5'].style = "ali_grey"
        #month6
    ws['J4'] = get_month_name(str(data_table_dates_list[0][6]))
    ws['J4'].style = "ali_blue"
    ws['J5'] = 'Объём на '  + str(data_table_dates_list[0][6]) + ', м3'
    ws['J5'].style = "ali_grey"
        #month7
    ws['K4'] = get_month_name(str(data_table_dates_list[0][7]))
    ws['K4'].style = "ali_blue"
    ws['K5'] = 'Объём на '  + str(data_table_dates_list[0][7]) + ', м3'
    ws['K5'].style = "ali_grey"
        #month8
    ws['L4'] = get_month_name(str(data_table_dates_list[0][8]))
    ws['L4'].style = "ali_blue"
    ws['L5'] = 'Объём на '  + str(data_table_dates_list[0][8]) + ', м3'
    ws['L5'].style = "ali_grey"
        #month9
    ws['M4'] = get_month_name(str(data_table_dates_list[0][9]))
    ws['M4'].style = "ali_blue"
    ws['M5'] = 'Объём на '  + str(data_table_dates_list[0][9]) + ', м3'
    ws['M5'].style = "ali_grey"
        #month10
    ws['N4'] = get_month_name(str(data_table_dates_list[0][10]))
    ws['N4'].style = "ali_blue"
    ws['N5'] = 'Объём на '  + str(data_table_dates_list[0][10]) + ', м3'
    ws['N5'].style = "ali_grey"
        #month11
    ws['O4'] = get_month_name(str(data_table_dates_list[0][11]))
    ws['O4'].style = "ali_blue"
    ws['O5'] = 'Объём на '  + str(data_table_dates_list[0][11]) + ', м3'
    ws['O5'].style = "ali_grey"
        #month12
    ws['P4'] = get_month_name(str(data_table_dates_list[0][12]))
    ws['P4'].style = "ali_blue"
    ws['P5'] = 'Объём на '  + str(data_table_dates_list[0][12]) + ', м3'
    ws['P5'].style = "ali_grey"

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    obj_title           = request.GET['obj_title']
    obj_key             = request.GET['obj_key']
    obj_parent_title    = request.GET['obj_parent_title']    
    is_electric_daily   = request.GET['is_electric_daily']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']
    is_electric_period  = request.GET['is_electric_period']
    isWater = True
    data_table = []
    data_table = common_sql.get_data_table_water_for_year_by_day(obj_parent_title, obj_title, electric_data_end)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][1])  # тип прибора
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][3]),ROUND_SIZE, separator)
            ws.cell('d%s'%(row)).style = "ali_white"
        except:
            ws.cell('d%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][4]),ROUND_SIZE, separator)  # 
            ws.cell('e%s'%(row)).style = "ali_white"
        except:
            ws.cell('e%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('f%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][5]),ROUND_SIZE, separator)  # 
            ws.cell('f%s'%(row)).style = "ali_white"
        except:
            ws.cell('f%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][6]),ROUND_SIZE, separator)  # 
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][7]),ROUND_SIZE, separator)  # 
            ws.cell('%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][8]),ROUND_SIZE, separator)  # 
            ws.cell('I%s'%(row)).style = "ali_white"
        except:
            ws.cell('I%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][9]),ROUND_SIZE, separator)  # 
            ws.cell('J%s'%(row)).style = "ali_white"
        except:
            ws.cell('J%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][10]),ROUND_SIZE, separator)  # 
            ws.cell('K%s'%(row)).style = "ali_white"
        except:
            ws.cell('K%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][11]),ROUND_SIZE, separator)  # 
            ws.cell('L%s'%(row)).style = "ali_white"
        except:
            ws.cell('L%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][12]),ROUND_SIZE, separator)  # 
            ws.cell('M%s'%(row)).style = "ali_white"
        except:
            ws.cell('M%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('N%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][13]),ROUND_SIZE, separator)  # 
            ws.cell('N%s'%(row)).style = "ali_white"
        except:
            ws.cell('N%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][14]),ROUND_SIZE, separator)  # 
            ws.cell('%s'%(row)).style = "ali_white"
        except:
            ws.cell('O%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][15]),ROUND_SIZE, separator)  # 
            ws.cell('P%s'%(row)).style = "ali_white"
        except:
            ws.cell('P%s'%(row)).style = "ali_white"
            next
        

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['N'].width = 12
    ws.column_dimensions['O'].width = 12
    ws.column_dimensions['P'].width = 12

    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    
    output_name = 'water_year_' + translate(obj_title) + '_' + electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_heat_by_day_for_year(request):
    ROUND_SIZE = getattr(settings, 'ROUND_SIZE', 'False')
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws = wb.active
    
    obj_parent_title         = request.GET.get('obj_parent_title')
    obj_title         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')
    
#Шапка
    ws.merge_cells('A1:E1')
    ws['A1'] = obj_parent_title+'. ' + obj_title + '.'
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Тепло. Срез показаний за год по ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование абонента'
    ws['A4'].style = "ali_grey"
    ws['A5'].style = "ali_grey"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер счётчика'
    ws['B4'].style = "ali_grey"
    ws['B5'].style = "ali_grey"

    ws.merge_cells('C4:C5')
    ws['C4'] = 'Тип прибора'
    ws['C4'].style = "ali_grey"
    ws['C5'].style = "ali_grey"
 
    data_table_dates_list = common_sql.get_data_table_12month(electric_data_end) #список дат по месяцам, начиная от выбранной даты
    #month0
    ws.merge_cells('d4:e4')
    ws['d4'] = get_month_name(str(data_table_dates_list[0][0]))
    ws['d4'].style = "ali_blue"
    ws['e4'].style = "ali_blue"
    ws['d5'] = 'Энергия на '  + str(data_table_dates_list[0][0]) + ', Гкал'
    ws['d5'].style = "ali_grey"
    ws['E5'] = 'Объём на '  + str(data_table_dates_list[0][0]) + ', м3'
    ws['E5'].style = "ali_grey"

    #month1
    ws.merge_cells('F4:G4')
    ws['F4'] = get_month_name(str(data_table_dates_list[0][1]))
    ws['F4'].style = "ali_blue"
    ws['G4'].style = "ali_blue"
    ws['F5'] = 'Энергия на '  + str(data_table_dates_list[0][1]) + ', Гкал'
    ws['F5'].style = "ali_grey"
    ws['G5'] = 'Объём на '  + str(data_table_dates_list[0][1]) + ', м3'
    ws['G5'].style = "ali_grey"

    #month2
    ws.merge_cells('H4:I4')
    ws['H4'] = get_month_name(str(data_table_dates_list[0][2]))
    ws['H4'].style = "ali_blue"
    ws['I4'].style = "ali_blue"
    ws['H5'] = 'Энергия на '  + str(data_table_dates_list[0][2]) + ', Гкал'
    ws['H5'].style = "ali_grey"
    ws['I5'] = 'Объём на '  + str(data_table_dates_list[0][2]) + ', м3'
    ws['I5'].style = "ali_grey"

    #month0
    ws.merge_cells('J4:K4')
    ws['J4'] = get_month_name(str(data_table_dates_list[0][3]))
    ws['J4'].style = "ali_blue"
    ws['K4'].style = "ali_blue"
    ws['J5'] = 'Энергия на '  + str(data_table_dates_list[0][3]) + ', Гкал'
    ws['J5'].style = "ali_grey"
    ws['K5'] = 'Объём на '  + str(data_table_dates_list[0][3]) + ', м3'
    ws['K5'].style = "ali_grey"

    #month0
    ws.merge_cells('L4:M4')
    ws['L4'] = get_month_name(str(data_table_dates_list[0][4]))
    ws['L4'].style = "ali_blue"
    ws['M4'].style = "ali_blue"
    ws['L5'] = 'Энергия на '  + str(data_table_dates_list[0][4]) + ', Гкал'
    ws['L5'].style = "ali_grey"
    ws['M5'] = 'Объём на '  + str(data_table_dates_list[0][4]) + ', м3'
    ws['M5'].style = "ali_grey"

    #month0
    ws.merge_cells('N4:O4')
    ws['N4'] = get_month_name(str(data_table_dates_list[0][5]))
    ws['N4'].style = "ali_blue"
    ws['O4'].style = "ali_blue"
    ws['N5'] = 'Энергия на '  + str(data_table_dates_list[0][5]) + ', Гкал'
    ws['N5'].style = "ali_grey"
    ws['O5'] = 'Объём на '  + str(data_table_dates_list[0][5]) + ', м3'
    ws['O5'].style = "ali_grey"

    #month0
    ws.merge_cells('P4:Q4')
    ws['P4'] = get_month_name(str(data_table_dates_list[0][6]))
    ws['P4'].style = "ali_blue"
    ws['Q4'].style = "ali_blue"
    ws['P5'] = 'Энергия на '  + str(data_table_dates_list[0][6]) + ', Гкал'
    ws['P5'].style = "ali_grey"
    ws['Q5'] = 'Объём на '  + str(data_table_dates_list[0][6]) + ', м3'
    ws['Q5'].style = "ali_grey"

    #month0
    ws.merge_cells('R4:S4')
    ws['R4'] = get_month_name(str(data_table_dates_list[0][7]))
    ws['R4'].style = "ali_blue"
    ws['S4'].style = "ali_blue"
    ws['R5'] = 'Энергия на '  + str(data_table_dates_list[0][7]) + ', Гкал'
    ws['R5'].style = "ali_grey"
    ws['S5'] = 'Объём на '  + str(data_table_dates_list[0][7]) + ', м3'
    ws['S5'].style = "ali_grey"

    #month0
    ws.merge_cells('T4:U4')
    ws['T4'] = get_month_name(str(data_table_dates_list[0][8]))
    ws['T4'].style = "ali_blue"
    ws['U4'].style = "ali_blue"
    ws['T5'] = 'Энергия на '  + str(data_table_dates_list[0][8]) + ', Гкал'
    ws['T5'].style = "ali_grey"
    ws['U5'] = 'Объём на '  + str(data_table_dates_list[0][8]) + ', м3'
    ws['U5'].style = "ali_grey"

    #month0
    ws.merge_cells('V4:W4')
    ws['V4'] = get_month_name(str(data_table_dates_list[0][9]))
    ws['V4'].style = "ali_blue"
    ws['W4'].style = "ali_blue"
    ws['V5'] = 'Энергия на '  + str(data_table_dates_list[0][9]) + ', Гкал'
    ws['V5'].style = "ali_grey"
    ws['W5'] = 'Объём на '  + str(data_table_dates_list[0][9]) + ', м3'
    ws['W5'].style = "ali_grey"

    #month0
    ws.merge_cells('X4:Y4')
    ws['X4'] = get_month_name(str(data_table_dates_list[0][10]))
    ws['X4'].style = "ali_blue"
    ws['Y4'].style = "ali_blue"
    ws['X5'] = 'Энергия на '  + str(data_table_dates_list[0][10]) + ', Гкал'
    ws['X5'].style = "ali_grey"
    ws['Y5'] = 'Объём на '  + str(data_table_dates_list[0][10]) + ', м3'
    ws['Y5'].style = "ali_grey"

    #month0
    ws.merge_cells('Z4:AA4')
    ws['Z4'] = get_month_name(str(data_table_dates_list[0][11]))
    ws['Z4'].style = "ali_blue"
    ws['AA4'].style = "ali_blue"
    ws['Z5'] = 'Энергия на '  + str(data_table_dates_list[0][11]) + ', Гкал'
    ws['Z5'].style = "ali_grey"
    ws['AA5'] = 'Объём на '  + str(data_table_dates_list[0][11]) + ', м3'
    ws['AA5'].style = "ali_grey"

    #month0
    ws.merge_cells('AB4:AC4')
    ws['AB4'] = get_month_name(str(data_table_dates_list[0][12]))
    ws['AB4'].style = "ali_blue"
    ws['AC4'].style = "ali_blue"
    ws['AB5'] = 'Энергия на '  + str(data_table_dates_list[0][12]) + ', Гкал'
    ws['AB5'].style = "ali_grey"
    ws['AC5'] = 'Объём на '  + str(data_table_dates_list[0][12]) + ', м3'
    ws['AC5'].style = "ali_grey"

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    obj_title           = request.GET['obj_title']
    obj_key             = request.GET['obj_key']
    obj_parent_title    = request.GET['obj_parent_title']    
    is_electric_daily   = request.GET['is_electric_daily']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']
    is_electric_period  = request.GET['is_electric_period']
    isWater = True
    data_table = []
    data_table = common_sql.get_data_table_heat_for_year_by_day(obj_parent_title, obj_title, electric_data_end)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = "ali_white"
        except:
            ws.cell('A%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = "ali_white"
        except:
            ws.cell('B%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # тип прибора
            ws.cell('C%s'%(row)).style = "ali_white"
        except:
            ws.cell('C%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][4]),ROUND_SIZE, separator)
            ws.cell('d%s'%(row)).style = "ali_white"
        except:
            ws.cell('d%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][5]),ROUND_SIZE, separator)  # 
            ws.cell('e%s'%(row)).style = "ali_white"
        except:
            ws.cell('e%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('f%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][6]),ROUND_SIZE, separator)  # 
            ws.cell('f%s'%(row)).style = "ali_white"
        except:
            ws.cell('f%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][7]),ROUND_SIZE, separator)  # 
            ws.cell('G%s'%(row)).style = "ali_white"
        except:
            ws.cell('G%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][8]),ROUND_SIZE, separator)  # 
            ws.cell('%s'%(row)).style = "ali_white"
        except:
            ws.cell('H%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][8]),ROUND_SIZE, separator)  # 
            ws.cell('I%s'%(row)).style = "ali_white"
        except:
            ws.cell('I%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][10]),ROUND_SIZE, separator)  # 
            ws.cell('J%s'%(row)).style = "ali_white"
        except:
            ws.cell('J%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][11]),ROUND_SIZE, separator)  # 
            ws.cell('K%s'%(row)).style = "ali_white"
        except:
            ws.cell('K%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][12]),ROUND_SIZE, separator)  # 
            ws.cell('L%s'%(row)).style = "ali_white"
        except:
            ws.cell('L%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][13]),ROUND_SIZE, separator)  # 
            ws.cell('M%s'%(row)).style = "ali_white"
        except:
            ws.cell('M%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('N%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][14]),ROUND_SIZE, separator)  # 
            ws.cell('N%s'%(row)).style = "ali_white"
        except:
            ws.cell('N%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][15]),ROUND_SIZE, separator)  # 
            ws.cell('%s'%(row)).style = "ali_white"
        except:
            ws.cell('O%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][16]),ROUND_SIZE, separator)  # 
            ws.cell('P%s'%(row)).style = "ali_white"
        except:
            ws.cell('P%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('q%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][17]),ROUND_SIZE, separator)  # 
            ws.cell('q%s'%(row)).style = "ali_white"
        except:
            ws.cell('q%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('r%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][18]),ROUND_SIZE, separator)  # 
            ws.cell('r%s'%(row)).style = "ali_white"
        except:
            ws.cell('r%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('s%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][19]),ROUND_SIZE, separator)  # 
            ws.cell('s%s'%(row)).style = "ali_white"
        except:
            ws.cell('s%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('t%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][20]),ROUND_SIZE, separator)  # 
            ws.cell('t%s'%(row)).style = "ali_white"
        except:
            ws.cell('t%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('u%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][21]),ROUND_SIZE, separator)  # 
            ws.cell('u%s'%(row)).style = "ali_white"
        except:
            ws.cell('u%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('v%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][22]),ROUND_SIZE, separator)  # 
            ws.cell('v%s'%(row)).style = "ali_white"
        except:
            ws.cell('v%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('w%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][23]),ROUND_SIZE, separator)  # 
            ws.cell('w%s'%(row)).style = "ali_white"
        except:
            ws.cell('w%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('x%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][24]),ROUND_SIZE, separator)  # 
            ws.cell('x%s'%(row)).style = "ali_white"
        except:
            ws.cell('x%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('y%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][25]),ROUND_SIZE, separator)  # 
            ws.cell('y%s'%(row)).style = "ali_white"
        except:
            ws.cell('y%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('z%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][26]),ROUND_SIZE, separator)  # 
            ws.cell('z%s'%(row)).style = "ali_white"
        except:
            ws.cell('z%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('aa%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][27]),ROUND_SIZE, separator)  # 
            ws.cell('aa%s'%(row)).style = "ali_white"
        except:
            ws.cell('aa%s'%(row)).style = "ali_white"
            next
            
        try:
            ws.cell('ab%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][28]),ROUND_SIZE, separator)  # 
            ws.cell('ab%s'%(row)).style = "ali_white"
        except:
            ws.cell('ab%s'%(row)).style = "ali_white"
            next
        
        try:
            ws.cell('ac%s'%(row)).value = '%s' % get_val_by_round(float(data_table[row-6][29]),ROUND_SIZE, separator)  # 
            ws.cell('ac%s'%(row)).style = "ali_white"
        except:
            ws.cell('ac%s'%(row)).style = "ali_white"
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 30 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['N'].width = 12
    ws.column_dimensions['O'].width = 12
    ws.column_dimensions['P'].width = 12    
    ws.column_dimensions['q'].width = 12    
    ws.column_dimensions['r'].width = 12    
    ws.column_dimensions['s'].width = 12    
    ws.column_dimensions['t'].width = 12    
    ws.column_dimensions['u'].width = 12    
    ws.column_dimensions['v'].width = 12    
    ws.column_dimensions['w'].width = 12    
    ws.column_dimensions['x'].width = 12    
    ws.column_dimensions['y'].width = 12    
    ws.column_dimensions['z'].width = 12
    ws.column_dimensions['aa'].width = 12    
    ws.column_dimensions['ab'].width = 12    
    ws.column_dimensions['ac'].width = 12

    
    #wb.save(response)
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    
    output_name = 'heat_year_' + translate(obj_title) + '_' + electric_data_end
    file_ext = 'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def electric_integral_month_hours(request):
    #SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')
    ROUND_SIZE = 0 #getattr(settings, 'ROUND_SIZE', 3)
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)
    ws_1 = wb.create_sheet('Приложение №5', 0)
    del wb['Sheet']

    # Шрифты
    font_1 = Font(name='Times New Roman',
                    size=12,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

    font_2 = Font(name='Times New Roman',
                    size=10,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

    font_3 = Font(name='Times New Roman',
                    size=9,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

    font_4 = Font(name='Times New Roman',
                    size=11,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

    # Форматирование
    alignment_1 = Alignment(horizontal='center',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0)

    alignment_1_2 = Alignment(horizontal='center',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0)

    alignment_2 = Alignment(horizontal='left',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0)

    alignment_3 = Alignment(horizontal='general',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0)

    alignment_4 = Alignment(horizontal='right',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0)

    alignment_5 = Alignment(horizontal='center',
                        vertical='center',
                        text_rotation=90,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0)

    alignment_6 = Alignment(horizontal='center',
                        vertical='bottom',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0)

    alignment_7 = Alignment(horizontal='left',
                        vertical='bottom',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0)

    alignment_8 = Alignment(horizontal='right',
                        vertical='bottom',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0)

    # Размерность строк и столбцов
    ws_1.row_dimensions[11].height = 27.75
    ws_1.row_dimensions[12].height = 71.25
    ws_1.row_dimensions[13].height = 21
    ws_1.row_dimensions[14].height = 32.25
    ws_1.row_dimensions[15].height = 33
    ws_1.row_dimensions[16].height = 39
    ws_1.row_dimensions[19].height = 16.5
    ws_1.row_dimensions[22].height = 17.22
    ws_1.row_dimensions[25].height = 16.5

    ws_1.column_dimensions['A'].width = 4.71
    ws_1.column_dimensions['B'].width = 12.29
    ws_1.column_dimensions['C'].width = 24.86
    ws_1.column_dimensions['D'].width = 12.43
    ws_1.column_dimensions['E'].width = 12
    ws_1.column_dimensions['F'].width = 12.86
    ws_1.column_dimensions['G'].width = 11.71
    ws_1.column_dimensions['H'].width = 7.71
    ws_1.column_dimensions['I'].width = 6.43
    ws_1.column_dimensions['J'].width = 6.43
    ws_1.column_dimensions['K'].width = 9.43
    ws_1.column_dimensions['L'].width = 16.14
    ws_1.column_dimensions['M'].width = 11

    # Границы
    brdr_1 = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thin'))

    brdr_2 = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thick'))

    brdr_3 = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

    brdr_4 = Border(bottom=Side(style='thin'))

    # Отчёт
    ws_1['G3'] = 'АКТ'
    ws_1['G3'].font = font_1
    ws_1['G3'].alignment = alignment_1_2
    ws_1['G4'] = 'снятия показаний приборов учета электрической энергии (интегральная форма первичного учета)'
    ws_1['G4'].font = font_1
    ws_1['G4'].alignment = alignment_1_2
    ws_1.merge_cells('A7:B7')
    ws_1['A7'] = 'Договор:  '
    ws_1['A7'].font = font_1
    ws_1['A7'].alignment = alignment_2
    ws_1['C7'] = '№2016-Э/Дх-МЦ-5002'
    ws_1['C7'].font = font_1
    ws_1['C7'].alignment = alignment_3
    ws_1.merge_cells('A9:B9')
    ws_1['A9'] = 'Потребитель:'
    ws_1['A9'].font = font_1
    ws_1['A9'].alignment = alignment_2
    ws_1['C9'] = 'ООО "ПромТраст"'
    ws_1['C9'].font = font_1
    ws_1['C9'].alignment = alignment_3
    ws_1['J9'] = 'месяц:'
    ws_1['J9'].font = font_1
    ws_1['J9'].alignment = alignment_4
    ws_1['K9'] = '-'
    ws_1['K9'].font = font_1
    ws_1['K9'].alignment = alignment_3
    ws_1.merge_cells('A11:A12')
    ws_1.merge_cells('B11:C12')
    ws_1['B11'] = 'Место установки (наименование присоединения)'
    ws_1['B11'].font = font_2
    ws_1['B11'].alignment = alignment_1
    ws_1.merge_cells('D11:D12')
    ws_1['D11'] = '№ счетчика'
    ws_1['D11'].font = font_2
    ws_1['D11'].alignment = alignment_1
    ws_1.merge_cells('E11:F11')
    ws_1['E11'] = 'показания счетчика'
    ws_1['E11'].font = font_2
    ws_1['E11'].alignment = alignment_1
    ws_1['E12'] = 'на начало месяца'
    ws_1['E12'].font = font_2
    ws_1['E12'].alignment = alignment_1
    ws_1['F12'] = 'на конец месяца'
    ws_1['F12'].font = font_2
    ws_1['F12'].alignment = alignment_1
    ws_1.merge_cells('G11:G12')
    ws_1['G11'] = 'Разница показаний'
    ws_1['G11'].font = font_2
    ws_1['G11'].alignment = alignment_1
    ws_1.merge_cells('H11:H12')
    ws_1['H11'] = 'Коэффициент ИК'
    ws_1['H11'].font = font_2
    ws_1['H11'].alignment = alignment_5
    ws_1.merge_cells('I11:J11')
    ws_1['I11'] = 'Потери до границы'
    ws_1['I11'].font = font_2
    ws_1['I11'].alignment = alignment_1
    ws_1['I12'] = '%'
    ws_1['I12'].font = font_2
    ws_1['I12'].alignment = alignment_1
    ws_1['J12'] = 'кВт ч'
    ws_1['J12'].font = font_2
    ws_1['J12'].alignment = alignment_1
    ws_1.merge_cells('K11:K12')
    ws_1['K11'] = 'Потери в ТП, кВт ч'
    ws_1['K11'].font = font_2
    ws_1['K11'].alignment = alignment_5
    ws_1.merge_cells('L11:L12')
    ws_1['L11'] = 'Итого расход электроэнергии, кВт ч'
    ws_1['L11'].font = font_2
    ws_1['L11'].alignment = alignment_1
    ws_1.merge_cells('M11:M12')
    ws_1['M11'] = 'Мощность, кВт'
    ws_1['M11'].font = font_2
    ws_1['M11'].alignment = alignment_5
    ws_1['A13'] = '1.'
    ws_1['A13'].font = font_3
    ws_1['A13'].alignment = alignment_1
    ws_1['B13'] = '(Наименование объекта)'
    ws_1['B13'].font = font_3
    ws_1['B13'].alignment = alignment_3
    ws_1['A14'] = '1'
    ws_1['A14'].font = font_4
    ws_1['A14'].alignment = alignment_6
    ws_1.merge_cells('B14:C14')
    ws_1['B14'] = ' РУ 0,4 кВ, ул. Дубининская, д.31А'
    ws_1['B14'].font = font_4
    ws_1['B14'].alignment = alignment_7
    ws_1['D14'] = ''#'0803136386'
    ws_1['D14'].font = font_4
    ws_1['D14'].alignment = alignment_8
    ws_1['E14'] = ''#'75097,0297'
    ws_1['E14'].font = font_4
    ws_1['E14'].alignment = alignment_8
    ws_1['F14'] = '' #'75830,6113'
    ws_1['F14'].font = font_4
    ws_1['F14'].alignment = alignment_8
    ws_1['G14'] = ''#'=F14-E14'
    ws_1['G14'].font = font_4
    ws_1['G14'].alignment = alignment_8
    ws_1['H14'] = ''#'500'
    ws_1['H14'].font = font_4
    ws_1['H14'].alignment = alignment_8
    ws_1['I14'] = '2,51'
    ws_1['I14'].font = font_4
    ws_1['I14'].alignment = alignment_8
    ws_1['J14'].font = font_4
    ws_1['J14'].alignment = alignment_8
    ws_1['K14'] = ''#'=G14*H14*0.0251'
    ws_1['K14'].font = font_4
    ws_1['K14'].alignment = alignment_8
    ws_1['L14'] = ''#'=G14*H14+K14'
    ws_1['L14'].font = font_4
    ws_1['L14'].alignment = alignment_8
    ws_1['M14'].font = font_4
    ws_1['M14'].alignment = alignment_8
    ws_1['A15'] = '2'
    ws_1['A15'].font = font_4
    ws_1['A15'].alignment = alignment_6
    ws_1.merge_cells('B15:C15')
    ws_1['B15'] = ' РУ 0,4 кВ, ул. Дубининская, д.31А'
    ws_1['B15'].font = font_4
    ws_1['B15'].alignment = alignment_7
    ws_1['D15'] = ''#'0803136290'
    ws_1['D15'].font = font_4
    ws_1['D15'].alignment = alignment_8
    ws_1['E15'] = ''#'59495,3285'
    ws_1['E15'].font = font_4
    ws_1['E15'].alignment = alignment_8
    ws_1['F15'] = ''#'59988,3705'
    ws_1['F15'].font = font_4
    ws_1['F15'].alignment = alignment_8
    ws_1['G15'] = ''#'=F15-E15'
    ws_1['G15'].font = font_4
    ws_1['G15'].alignment = alignment_8
    ws_1['H15'] = ''#'500'
    ws_1['H15'].font = font_4
    ws_1['H15'].alignment = alignment_8
    ws_1['I15'] = '2,51'
    ws_1['I15'].font = font_4
    ws_1['I15'].alignment = alignment_8
    ws_1['J15'].font = font_4
    ws_1['J15'].alignment = alignment_8
    ws_1['K15'] = ''#'=G15*H15*0.0251'
    ws_1['K15'].font = font_4
    ws_1['K15'].alignment = alignment_8
    ws_1['L15'] = ''#'=G15*H15+K15'
    ws_1['L15'].font = font_4
    ws_1['L15'].alignment = alignment_8
    ws_1.merge_cells('B16:C16')
    ws_1['B16'] = 'Расход абонента, кВт ч'
    ws_1['B16'].font = font_4
    ws_1['B16'].alignment = alignment_7
    ws_1['L16'] = '=SUM(L14:L15)'
    ws_1['L16'].font = font_4
    ws_1['L16'].alignment = alignment_8
    ws_1['A19'] = 'Представитель Потребителя'
    ws_1['A19'].font = font_4
    ws_1['A19'].alignment = alignment_2
    ws_1['F19'] = '    /  Мальцев С.Ю.'
    ws_1['F19'].font = font_4
    ws_1['F19'].alignment = alignment_2
    ws_1['I19'] = '(дата)'
    ws_1['I19'].font = font_4
    ws_1['I19'].alignment = alignment_1
    ws_1['A22'] = 'Представитель  Сетевой Организации'
    ws_1['A22'].font = font_4
    ws_1['A22'].alignment = alignment_2
    ws_1['F22'] = '    /'
    ws_1['F22'].font = font_4
    ws_1['F22'].alignment = alignment_2
    ws_1['I22'] = '(дата)'
    ws_1['I22'].font = font_4
    ws_1['I22'].alignment = alignment_1
    ws_1['A25'] = 'Представитель  Продавца'
    ws_1['A25'].font = font_4
    ws_1['A25'].alignment = alignment_2
    ws_1['F25'] = '    /'
    ws_1['F25'].font = font_4
    ws_1['F25'].alignment = alignment_2
    ws_1['I25'] = '(дата)'
    ws_1['I25'].font = font_4
    ws_1['I25'].alignment = alignment_1

    border_range_1 = ws_1['A11':'M11']
    border_range_2 = ws_1['A12':'M12']
    border_range_3 = ws_1['A14':'M14']
    border_range_4 = ws_1['A15':'M15']
    border_range_5 = ws_1['A16':'M16']
    border_range_6 = ws_1['D19':'J19']
    border_range_7 = ws_1['D22':'J22']
    border_range_8 = ws_1['D25':'J25']

    for cell in border_range_1:
        for x in cell:
            x.border = brdr_1

    for cell in border_range_2:
        for x in cell:
            x.border = brdr_2

    for cell in border_range_3:
        for x in cell:
            x.border = brdr_1

    for cell in border_range_4:
        for x in cell:
            x.border = brdr_3

    for cell in border_range_5:
        for x in cell:
            x.border = brdr_2

    for cell in border_range_6:
        for x in cell:
            x.border = brdr_4

    for cell in border_range_7:
        for x in cell:
            x.border = brdr_4

    for cell in border_range_8:
        for x in cell:
            x.border = brdr_4

    ws_1['M11'].border = Border(left=Side(style='thin'), 
                                right=Side(style='thick'), 
                                top=Side(style='thick'), 
                                bottom=Side(style='thin'))

    ws_1['M12'].border = Border(left=Side(style='thin'), 
                                right=Side(style='thick'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thick'))

    ws_1['M13'].border = Border(left=Side(style=None), 
                                right=Side(style='thick'), 
                                top=Side(style='thick'), 
                                bottom=Side(style='thick'))

    ws_1['M14'].border = Border(left=Side(style='thin'), 
                                right=Side(style='thick'), 
                                top=Side(style='thick'), 
                                bottom=Side(style='thin'))

    ws_1['M15'].border = Border(left=Side(style='thin'), 
                                right=Side(style='thick'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin'))

    ws_1['M16'].border = Border(left=Side(style='thin'), 
                                right=Side(style='thick'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thick'))

#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2') 
    obj_title           = request.GET['obj_title']
    obj_key             = request.GET['obj_key']
    obj_parent_title    = request.GET['obj_parent_title']    
    electric_data_end   = request.GET['electric_data_end']
    is_electric_monthly   = request.GET['is_electric_monthly']  
          
    res='Электричество'
    if (is_electric_monthly=="1"):
        dm='monthly'
    else:
        dm='daily'

    dt_date=common_sql.get_date_month_range_by_date(electric_data_end)

    common_sql.del_double_30_by_dates(dt_date[0][0],dt_date[-1][0])
    
    data_table=[]
    if (bool(is_abonent_level.search(obj_key))): #выбран абонент
        pass
        
    elif (bool(is_object_level_2.search(obj_key))):#выбран объект
        data_table = common_sql.get_data_integral_dubi(obj_parent_title, obj_title, dt_date[0][0],dt_date[-1][0], dm, res)


    my_date_list = electric_data_end.split(".")
    year = my_date_list[2]
    #print(year)
    month_name = ""
    month_name = common_sql.get_month_rus(dt_date[0][0])

    if len(data_table)>0:
    # Заполняем отчет значениями
        ws_1['K9'] = month_name + ' ' + year +'г.'

        try:
            ws_1['D14'] = data_table[0][1]#'0803136386'
        except:
            next

        try:
            ws_1['E14'] = get_val_by_round(float(data_table[0][2]), 4, ',').replace('.',',')  #data_table[0][2]#'75097,0297'
        except:
            next
    
        try:
            ws_1['F14'] = get_val_by_round(float(data_table[0][7]), 4, ',').replace('.',',')  #data_table[0][7] #'75830,6113'
        except:
            next

        try:
            ws_1['G14'] = get_val_by_round(float(data_table[0][12]), 4, ',').replace('.',',')  #data_table[0][12] #'=F14-E14'
        except:
            next

        try:
            ws_1['H14'] = data_table[0][20]#'500'
        except:
            next

        try:
            ws_1['K14'] = get_val_by_round(float(data_table[0][26]), 4, ',').replace('.',',')  #data_table[0][26]#'=G14*H14*0.0251'
        except:
            next

        try:
            ws_1['L14'] = get_val_by_round(float(data_table[0][27]), 4, ',').replace('.',',')  #data_table[0][27]#'=G14*H14+K14'
        except:
            next
        try:
            ws_1['D15'] = data_table[1][1]#'0803136290'
        except:
            next
        try:
            ws_1['E15'] = get_val_by_round(float(data_table[1][2]), 4, ',').replace('.',',')  #data_table[1][2]#'59495,3285'
        except:
            next
        try:
            ws_1['F15'] = get_val_by_round(float(data_table[1][7]), 4, ',').replace('.',',')  #data_table[1][7]#'59988,3705'
        except:
            next
        try:
            ws_1['G15'] = get_val_by_round(float(data_table[1][12]), 4, ',').replace('.',',')  #data_table[1][12]#'=F15-E15'
        except:
            next
        try:
            ws_1['H15'] = data_table[1][20]#'500'
        except:
            next
        try:
            ws_1['K15'] = get_val_by_round(float(data_table[1][26]), 4, ',').replace('.',',')  #data_table[1][26]#'=G15*H15*0.0251'
        except:
            next
        try:
            ws_1['L15'] = get_val_by_round(float(data_table[1][27]), 4, ',').replace('.',',')  #'=G15*H15+K15'
        except:
            next
        try:
            ws_1['L16'] = get_val_by_round(float(data_table[1][27]) + float(data_table[0][27]) ,4,',').replace('.',',')
        except:
            next

    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_integral_' + translate(obj_parent_title) + '_' + translate(obj_title) + '_' + common_sql.get_month_eng(dt_date[0][0])
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response



def electric_interval_month_hours(request):
    #SHOW_LIC_NUM = getattr(settings, 'SHOW_LIC_NUM', 'False')
    ROUND_SIZE = 0 #getattr(settings, 'ROUND_SIZE', 3)
    response = io.StringIO()
    wb = Workbook()
    wb.add_named_style(ali_grey)
    wb.add_named_style(ali_white)
    wb.add_named_style(ali_yellow)
    wb.add_named_style(ali_pink)
    wb.add_named_style(ali_blue)


#ШАпка

    # Лист
    ws_1 = wb.create_sheet('Сетевой номер прибора', 0)
    del wb['Sheet']

    # Шрифты
    font_1 = Font(name='Arial Cyr',
                    size=12,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

    font_2 = Font(name='Arial Cyr',
                    size=10,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

    font_3 = Font(name='Arial Cyr',
                    size=8,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

    font_4 = Font(name='Times New Roman',
                    size=11,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

    # Форматирование
    alignment_1 = Alignment(horizontal='center',
                        vertical='bottom',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0)

    alignment_1_2 = Alignment(horizontal='center',
                        vertical='bottom',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0)

    alignment_2 = Alignment(horizontal='left',
                        vertical='bottom',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0)

    alignment_3 = Alignment(horizontal='center',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0)

    alignment_4 = Alignment(horizontal='general',
                        vertical='bottom',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0)

    # Размерность строк и столбцов
    ws_1.column_dimensions['A'].width = 9.29

    #Границы
    brdr_1 = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

    brdr_2 = Border(bottom=Side(style='thin'))

    #Отчёт
    ws_1.merge_cells('A4:Y4')
    ws_1['A4'] = 'Интервальный акт снятия показаний приборов учета'
    ws_1['A4'].font = font_1
    ws_1['A4'].alignment = alignment_1
    ws_1['A6'] = 'Наименование потребителя:'
    ws_1['A6'].font = font_1
    ws_1['A6'].alignment = alignment_2
    ws_1['E6'] = 'ООО "ПромТраст"'
    ws_1['E6'].font = font_1
    ws_1['E6'].alignment = alignment_2
    ws_1['A7'] = '№ Договора энергоснабжения:'
    ws_1['A7'].font = font_1
    ws_1['A7'].alignment = alignment_2
    ws_1['E7'] = '№2016-Э/Дх-МЦ-5002'
    ws_1['E7'].font = font_1
    ws_1['E7'].alignment = alignment_2
    ws_1['A8'] = 'Наименование  Сетевой организации:'
    ws_1['A8'].font = font_1
    ws_1['A8'].alignment = alignment_2
    ws_1['F8'] = 'ОАО "МОЭСК" 1 р-н УКС ЦО МКС'
    ws_1['F8'].font = font_1
    ws_1['F8'].alignment = alignment_2
    ws_1['A10'] = 'Уровень напряжения:'
    ws_1['A10'].font = font_1
    ws_1['A10'].alignment = alignment_2
    ws_1['E10'] = 'СН-2'
    ws_1['E10'].font = font_1
    ws_1['E10'].alignment = alignment_2
    ws_1['A11'] = 'Наименование точки поставки:'
    ws_1['A11'].font = font_1
    ws_1['A11'].alignment = alignment_2
    ws_1['E11'] = 'Источник питания'
    ws_1['E11'].font = font_1
    ws_1['E11'].alignment = alignment_2
    ws_1.merge_cells('H11:L11')
    ws_1['H11'] = 'РТП-18197\РТП-16072\РУ-10кВ\КЛ-10кВ,'
    ws_1['H11'].font = font_1
    ws_1['H11'].alignment = alignment_2
    ws_1['M11'] = 'Наименование присоединения'
    ws_1['M11'].font = font_1
    ws_1['M11'].alignment = alignment_2
    ws_1['Q11'] = 'ТП-22969\РУ-0,4кВ'
    ws_1['Q11'].font = font_1
    ws_1['Q11'].alignment = alignment_2
    ws_1['A12'] = 'Номер прибора учета:'
    ws_1['A12'].font = font_1
    ws_1['A12'].alignment = alignment_2
    ws_1['E12'] = ''
    ws_1['E12'].font = font_1
    ws_1['E12'].alignment = alignment_2
    ws_1['A14'] = 'расчетный период:'
    ws_1['A14'].font = font_1
    ws_1['A14'].alignment = alignment_2
    ws_1.merge_cells('E14:G14')
    ws_1['E14'] = ''
    ws_1['E14'].font = font_1
    ws_1['E14'].alignment = alignment_2
    ws_1.merge_cells('W14:Y14')
    ws_1['W14'] = 'время московское'
    ws_1['W14'].font = font_1
    ws_1['W14'].alignment = alignment_2
    ws_1.merge_cells('A16:A17')
    ws_1['A16'] = 'Дата'
    ws_1['A16'].font = font_2
    ws_1['A16'].alignment = alignment_3
    ws_1.merge_cells('B16:Y16')
    ws_1['B16'] = 'Почасовые объемы потребления электрической энергии, кВт.ч.'
    ws_1['B16'].font = font_2
    ws_1['B16'].alignment = alignment_3
    ws_1['B51'] = 'Потребитель:'
    ws_1['B51'].font = font_1
    ws_1['B51'].alignment = alignment_4
    ws_1['J51'] = 'Сетевая организация:'
    ws_1['J51'].font = font_1
    ws_1['J51'].alignment = alignment_4
    ws_1['R51'] = 'Продавец:'
    ws_1['R51'].font = font_1
    ws_1['R51'].alignment = alignment_4
    ws_1['B53'] = '__________________________/   Мальцев С.Ю.'
    ws_1['B53'].font = font_1
    ws_1['B53'].alignment = alignment_4
    ws_1['J53'] = '__________________________________'
    ws_1['J53'].font = font_1
    ws_1['J53'].alignment = alignment_4
    ws_1['R53'] = '__________________________________'
    ws_1['R53'].font = font_1
    ws_1['R53'].alignment = alignment_4

    dates_range = ws_1['A18':'A48']
    time_range = ws_1['B17':'Y17']
    values_range = ws_1['B18':'Y48']
    border_range_1 = ws_1['A16':'Y48']
    border_range_2 = ws_1['E6':'G7']
    border_range_3 = ws_1['F8':'G8']
    border_range_4 = ws_1['E10':'G12']
    border_range_5 = ws_1['E14':'G14']

    #Заполнение таблицы
    def daterange(start_date, end_date):
        '''Генератор дат'''
        for n in range(int((end_date - start_date).days)):
            yield start_date + timedelta(n)

    def datetime_range(start, end, delta):
        '''Генератор интервалов времени'''
        current = start
        while current < end:
            yield current
            current += delta

    start_date = date(2023, 3, 1)
    end_date = date(2023, 4, 1)

    dates = []
    for single_date in daterange(start_date, end_date):
        dates.append(single_date.strftime("%d.%m.%Y"))

    intervals = [dt.strftime('%H:%M') for dt in 
        datetime_range(datetime(2016, 9, 1, 0), datetime(2016, 9, 1, 23), 
        timedelta(hours=1))]

    intervals.extend(('23:00', '24:00'))

    new_intervals = []
    b_index = 0
    e_index = 1

    for el in intervals:
        new_intervals.append(f'{intervals[b_index]}-{intervals[e_index]}')
        b_index += 1
        e_index += 1
        if b_index == 24:
            break

    # d_i = 0
    # for cell in dates_range:
    #     for x in cell:
    #         x.font = font_2
    #         x.alignment = alignment_1_2
    #         x.value = dates[d_i]
    #         d_i += 1

    t_i = 0
    for cell in time_range:
        for x in cell:
            x.font = font_3
            x.alignment = alignment_1
            x.value = new_intervals[t_i]
            t_i += 1



    for cell in border_range_1:
        for x in cell:
            x.border = brdr_1

    for cell in border_range_2:
        for x in cell:
            x.border = brdr_2

    for cell in border_range_3:
        for x in cell:
            x.border = brdr_2

    for cell in border_range_4:
        for x in cell:
            x.border = brdr_2

    for cell in border_range_5:
        for x in cell:
            x.border = brdr_2

#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2') 
    obj_title           = request.GET['obj_title']
    obj_key             = request.GET['obj_key']
    obj_parent_title    = request.GET['obj_parent_title']    
    electric_data_end   = request.GET['electric_data_end']

    dt_date=common_sql.get_date_month_range_by_date(electric_data_end)

    common_sql.del_double_30_by_dates(dt_date[0][0],dt_date[-1][0])

    dt_All_el = []
    factory_number_manual = "-"
    year = ""
    row_num = 0
    for dd in dt_date:
        data_table=[]
        if (bool(is_abonent_level.search(obj_key))): #выбран абонент
            data_table = common_sql.get_data_hours_various_by_date(obj_parent_title, obj_title, dd[0])
            if len(data_table) > 0:
                factory_number_manual = data_table[0][25]
            if len(data_table) == 0:
                data_table = (('','','','','','','','','','','','','','','','','','','','','','','','','','-'),)
                
            #заполняем данными
            # for cell in values_range:

            c_j = 1
            for x in values_range[row_num]:
                x.font = font_4
                x.alignment = alignment_1
                try:
                    x.value = data_table[0][c_j]
                except:
                    next
                c_j +=1
            row_num +=1

                

        elif (bool(is_object_level_2.search(obj_key))):#выбран объект
            pass
    
    my_date_list = electric_data_end.split(".")
    year = my_date_list[2]
    #print(year)
    month_name = ""
    month_name = common_sql.get_month_rus(dt_date[0][0])


# Заполняем отчет значениями
    try:
        ws_1.title = factory_number_manual
        ws_1['E12'] = factory_number_manual
        ws_1['E14'] = month_name + ' ' + year +'г.'
    except:
            next

    d_i = 0
    date_f = ""
    for cell in dates_range:
        if d_i == len(dt_date):
                break
        for x in cell:
            x.font = font_2
            x.alignment = alignment_1_2
            #print(dt_date[d_i], len(dt_date), d_i)
            try:
                date_f = ''.join(dt_date[d_i]).split('-')
                x.value = date_f[2] + '.' + date_f[1]+ '.' + date_f[0]
            except:
                next            
            d_i += 1
            if d_i == len(dt_date):
                break
            

    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = 'report_interval_' + translate(obj_parent_title) + '_' + translate(obj_title) + '_' + common_sql.get_month_eng(dd[0])
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response