# coding -*- coding: utf-8 -*-


from django.shortcuts import render
from django.shortcuts import HttpResponse
from django.http import HttpResponseRedirect
from django.shortcuts import redirect
from django import forms
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Font
from openpyxl.writer.excel import save_virtual_workbook
import os
from django.db import connection
#from general.models import Objects, Abonents, TypesAbonents, Meters, MonthlyValues, DailyValues, CurrentValues, VariousValues, TypesParams, Params, TakenParams, LinkAbonentsTakenParams, Resources, TypesMeters, Measurement, NamesParams, BalanceGroups, LinkMetersComportSettings, LinkMetersTcpipSettings, ComportSettings, TcpipSettings, LinkBalanceGroupsMeters, Groups80020, LinkGroups80020Meters
from general.models import  Objects, Abonents, TcpipSettings, TypesAbonents, Meters, TypesMeters,LinkAbonentsTakenParams,LinkMetersComportSettings, LinkMetersTcpipSettings, ComportSettings,  TakenParams,Params, LinkAbonentsAuthUser, Groups80020, LinkGroups80020Meters
from django.db.models.signals import pre_save
from django.db.models.signals import post_save
from django.db.models import signals
import datetime
from django.db.models import Max 
import uuid
import io
import sys

import common_sql, AskueReports
from html.parser import HTMLParser
import psycopg2

from django.contrib.auth.models import User
from django.contrib.auth import authenticate
from django.core.exceptions import ObjectDoesNotExist
cfg_excel_name=""
cfg_sheet_name=""

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
# Create your views here.

isService=False
import logging

#logging.basicConfig(filename=u"C:\\Users\\Lena\\Desktop\\m_errors\\service_log.log", level=logging.INFO)
logger=logging.getLogger('service_log') # path in settings.py

from django.contrib.auth.decorators import user_passes_test

# Стили
ali_grey   = NamedStyle(name = "ali_grey", fill=PatternFill(fill_type='solid', start_color='DCDCDC'), border=Border(left=Side(border_style='thin',color='FF000000'), bottom=Side(border_style='thin',color='FF000000'), right=Side(border_style='thin',color='FF000000'), top=Side(border_style='thin',color='FF000000')), alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True))
ali_white  = NamedStyle(name = "ali_white", border=Border(left=Side(border_style='thin',color='FF000000'), bottom=Side(border_style='thin',color='FF000000'), right=Side(border_style='thin',color='FF000000'), top=Side(border_style='thin',color='FF000000')), alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True))
ali_blue   = NamedStyle(name = "ali_blue", fill=PatternFill(fill_type='solid', start_color='E6E6FA'), border=Border(left=Side(border_style='thin',color='FF000000'), bottom=Side(border_style='thin',color='FF000000'), right=Side(border_style='thin',color='FF000000'), top=Side(border_style='thin',color='FF000000')), alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True))
ali_pink   = NamedStyle(name = "ali_pink", fill=PatternFill(fill_type='solid', start_color='FFF0F5'), border=Border(left=Side(border_style='thin',color='FF000000'), bottom=Side(border_style='thin',color='FF000000'), right=Side(border_style='thin',color='FF000000'), top=Side(border_style='thin',color='FF000000')), alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True))

# Конец описания стилей

def isAdmin(user):
    return user.is_superuser

#@user_passes_test(isAdmin)
class UploadFileForm(forms.Form):
    #title = forms.CharField(max_length=150)
    path  = forms.FileField()

def MakeSheet(request):
    args={}
    fileName=""
    sheets = ""
    if request.is_ajax():
        if request.method == 'GET':
            fileName    = request.GET['choice_file']
            #print fileName
            directory=os.path.join(BASE_DIR,'static/cfg/')
            try:
                wb=load_workbook(directory+fileName)
                sheets=wb.sheetnames
            except: # catch *all* exceptions
                pass

    args['sheets']=sheets
    return render(request,"service/service_sheets_excel.html", args)


def writeToLog(msg):
    ##################################
    log_date=datetime.datetime.now().strftime("%d-%m-%Y %H:%M")
    #print msg
    logger.info('['+log_date+']: '+str(msg))

@login_required(login_url='/auth/login/') 
@user_passes_test(isAdmin, login_url='/auth/login/')
def choose_service(request):
    args={}
    # directory=os.path.join(BASE_DIR,'static\\cfg\\')
    
    # if  not(os.path.exists(directory)):
    #     os.mkdir(directory)
    # #print directory
    # files = os.listdir(directory) 
    # #print files
    # args['filesFF']= files
    return render(request,"choose_service.html", args)

def make_excel(request):
    args = {}
    directory=os.path.join(BASE_DIR,'static\\cfg\\')    
    if  not(os.path.exists(directory)):
        os.mkdir(directory)
    files = os.listdir(directory) 
    args['filesFF']= files
    #print(files)
    return render(request,"service/service_excel.html", args)

@csrf_exempt
def service_electric(request):
    args={}
    return render(request,"service/service_electric.html", args)

#@login_required(login_url='/auth/login/') 
@csrf_exempt
def service_file(request):
    args={}
    #args.update(csrf(request))    
    data_table=[]
    status='Файл не загружен'
    args['data_table'] = data_table
    args['status']=status
    
    directory=os.path.join(BASE_DIR,'static\\cfg\\')    
    if  not(os.path.exists(directory)):
        os.mkdir(directory)    
    files = os.listdir(directory)    
    args['filesFF']= files
    return render(request,"service/service_file.html", args)

@login_required(login_url='/auth/login/') 
@user_passes_test(isAdmin, login_url='/auth/login/')
def service_file_loading(request):
    args={}
    data_table=[]
    status='Файл не загружен'
    sPath=""
    if request.method == 'POST':        
        form = UploadFileForm(request.POST, request.FILES)
        #print form.as_table()
        #print form.is_valid()
        
        #print sPath
        if form.is_valid():
            sPath=os.path.join(BASE_DIR,'static/cfg/'+request.FILES['path'].name)
            handle_uploaded_file(request.FILES['path'])
            status= 'Файл загружен'
    else:
        form = UploadFileForm()
        
    args['data_table'] = data_table
    args['status']=status
    args['sPath']=sPath

    directory=os.path.join(BASE_DIR,'static\\cfg\\')    
    if  not(os.path.exists(directory)):
        os.mkdir(directory)    
    files = os.listdir(directory)    
    args['filesFF']= files
    #print status
    return render(request,"choose_service.html", args)


@login_required(login_url='/auth/login/') 
@user_passes_test(isAdmin, login_url='/auth/login/')
def service_electric_load(request):
    args={}
    data_table=[]
    status='Файл не загружен'

    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():
            handle_uploaded_file(request.FILES['path'])
            status='Файл загружен'
    else:
        form = UploadFileForm()
        
    args['data_table'] = data_table
    args['status']=status
    directory=os.path.join(BASE_DIR,'static\\cfg\\')    
    if  not(os.path.exists(directory)):
        os.mkdir(directory)    
    files = os.listdir(directory)    
    args['filesFF']= files
    return render(request,"service/service_electric.html", args)
    #return render(request,"service/service_electric_load.html", args)

#@login_required(login_url='/auth/login/') 
#@user_passes_test(isAdmin, login_url='/auth/login/')
def handle_uploaded_file(f):
    destination = open(os.path.join(BASE_DIR,'static/cfg/'+f.name), 'wb+')
    for chunk in f.chunks():
        destination.write(chunk)
    #print 'file load'
    destination.close()


def load_port(request):
    args={}
    #print '!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!'
    fileName=""
    sheet    = ""
    tcp_ip_status    = ""
    object_status    = ""
    counter_status    = ""
    result=""
    try:
        if request.is_ajax():
            if request.method == 'GET':
                request.session["choice_file"]    = fileName    = request.GET['choice_file']
                request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']
                request.session["tcp_ip_status"]    = tcp_ip_status    = request.GET['tcp_ip_status']
                request.session["object_status"]    = object_status    = request.GET['object_status']
                request.session["counter_status"]    = counter_status    = request.GET['counter_status']
                
                directory=os.path.join(BASE_DIR,'static/cfg/')
                sPath=directory+fileName
                
                result=load_tcp_ip_or_com_ports_from_excel(sPath, sheet)
                if result:
                    result="Порт/ы был успешно добавлен"
                else:
                    result="Порт не был загружен, он уже существует в БД"
    except: # catch *all* exceptions
        e = sys.exc_info()[0]
        result = ( "Ошибка: %s" % e )

    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["tcp_ip_status"]=result
    args["object_status"]=object_status
    args["counter_status"]=counter_status
    return render(request,"service/service_electric.html", args)


def checkPortIsExist(ip_adr,ip_port):
    dt_ports=[]
    cursor = connection.cursor()
    sQuery="""
    SELECT guid, ip_address, ip_port, write_timeout, read_timeout, attempts, 
       delay_between_sending
  FROM tcpip_settings
  where ip_address='%s' and  ip_port='%s'"""%(str(ip_adr).rstrip(),str(ip_port).rstrip())
    #print sQuery
    cursor.execute(sQuery)
    dt_ports = cursor.fetchall()
    #print dt_ports
    if len(dt_ports):  
        return False
    else: 
        return True


def load_tcp_ip_or_com_ports_from_excel(sPath, sSheet):
    #Добавление tcp_ip портов
    global cfg_excel_name
    cfg_excel_name=sPath
    global cfg_sheet_name
    cfg_sheet_name=sSheet
    wb = load_workbook(filename = sPath)
    sheet_ranges = wb[sSheet]
    row = 2
    result=""
    IsAdded=False
    portType=sheet_ranges['L1'].value
    while (bool(sheet_ranges['G%s'%(row)].value)):
        if sheet_ranges['G%s'%(row)].value is not None:
            #writeToLog(u'Обрабатываем строку ' + str(u'G%s '%(row)) + str(sheet_ranges[u'G%s'%(row)].value))
            ip_adr=sheet_ranges['K%s'%(row)].value
            ip_port=sheet_ranges['L%s'%(row)].value
            com_port=sheet_ranges['L%s'%(row)].value
            #print ip_adr, ip_port
            if portType=='Com-port': #добавление com-порта
                writeToLog(com_port)
                if not com_port or com_port==None: 
                    result+="Отсутствует значение для com-порта в строке"+str(row)+". Заполните все ячейки excel таблицы."
                    break
                if not (SimpleCheckIfExist('comport_settings','name', com_port, "", "", "")):
                    add_port=ComportSettings(name=str(com_port).rstrip(),baudrate=9600,data_bits=8,parity=0,stop_bits=1, write_timeout=100, read_timeout=100, attempts=2, delay_between_sending=100)
                    add_port.save()
                    result+="Новый com-порт добавлен"
                    IsAdded=True
                else: result= 'Порт '+str(com_port)+" уже существует"
            else:
                # проверка есть ли уже такой порт, запрос в БД с адресом и портом, если ответ пустой-добавляем, в противном случае continue
                if not ip_adr or not ip_port or ip_adr==None or ip_port==None: 
                    result+="Отсутствует значение/я для tcp/ip-порта в строке"+str(row)+". Заполните все ячейки excel таблицы."
                    break
                else:
                    if (checkPortIsExist(ip_adr,ip_port)):
                        add_port=TcpipSettings(ip_address = str(ip_adr).rstrip(), ip_port =int(ip_port), write_timeout =300 , read_timeout =700 , attempts =3 , delay_between_sending =400)
                        add_port.save()
                        result ='Новый tcp/ip порт добавлен'
                        IsAdded=True
    #                add_meter = Meters(name = unicode(sheet_ranges[u'F%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'E%s'%(row)].value), address = unicode(sheet_ranges[u'E%s'%(row)].value),  factory_number_manual = unicode(sheet_ranges[u'E%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"7cd88751-d232-410c-a0ef-6354a79112f1") )
    #                add_meter.save()
                    else: 
                        newRes = 'Порт '+str(ip_adr)+": "+str(ip_port)+" уже существует "
                        # print result
                        # print newRes
                        # print result.find(newRes)
                        if bool(result.find(newRes) == -1):
                            result+= newRes
                            
        writeToLog( result)
        row+=1
    return IsAdded


def SimpleCheckIfExist(table1,fieldName1, value1, table2, fieldName2, value2):
    dt=[]
    cursor = connection.cursor()
    if len(table2)==0: #проверка для одной таблицы
        sQuery="""
        Select *
        from %s
        where %s.%s='%s'"""%(table1, table1, fieldName1, value1)
    else:#проверка для двух сводных таблиц
        sQuery="""
        Select *
        from %s, %s
        where %s.guid_%s=%s.guid and
        %s.%s='%s' and
        %s.%s='%s'
        """%(table1,table2, table2, table1,table1, table1, fieldName1, value1,table2, fieldName2, value2)
    #print sQuery
    #print bool(dt)
    cursor.execute(sQuery)
    dt = cursor.fetchall()

    if not dt:  
        return False
    else: 
        return True
    
def GetSimpleTable(table,fieldName1,value1, fieldName2='', value2=''):
    dt=[]
    cursor = connection.cursor()
    if len(fieldName2)==0: #одна таблица с 1 полем на запрос
        sQuery="""
            Select *
            from %s
            where %s.%s='%s'"""%(table, table, fieldName1, value1)
    else:
        sQuery="""
            Select *
            from %s
            where %s.%s='%s'
            and %s.%s='%s' """%(table, table, fieldName1, value1, table, fieldName2, value2)
    #print(sQuery)
    cursor.execute(sQuery)
    dt = cursor.fetchall()
    return dt
    
def GetSimpleCrossTable(table1,fieldName1,value1,table2,fieldName2, value2):
    dt=[]
    cursor = connection.cursor()
    sQuery="""
        Select *
        from %s, %s
        where %s.guid_%s=%s.guid and
        %s.%s='%s' and
        %s.%s='%s'
        """%(table1,table2, table2, table1,table1, table1, fieldName1, value1,table2, fieldName2, value2)
    #print sQuery
    cursor.execute(sQuery)
    dt = cursor.fetchall()
    return dt

def GetCrossTwoTable(table1, table2, crossField1, crossField2, whereTable1, fieldName1, fieldValue1, whereTable2, fieldName2, fieldValue2):
    #возвращает результат пересечения 2 таблиц с условием WHERE или без, елси поля пустые: whereTable, fieldName, fieldValue
    dt=[]
    cursor = connection.cursor()
    if len(whereTable1)>0:
        sQuery="""
        SELECT *
        FROM 
          %s, 
          %s
        WHERE 
          %s.%s = %s.%s AND
          %s.%s = '%s'"""%(table1, table2, table1, crossField1, table2, crossField2, whereTable1, fieldName1, fieldValue1)
        if len(whereTable2)>0:
            sQuery += """
             AND %s.%s = '%s'
            """%(whereTable2, fieldName2, fieldValue2)
    else: sQuery="""
        SELECT *  
        FROM 
          %s, 
          %s
        WHERE 
          %s.%s = %s.%s"""%(table1, table2, table1, crossField1, table2, crossField2)  
    cursor.execute(sQuery)
    dt = cursor.fetchall()
    return dt

def GetTableFromExcel(sPath,sSheet):
    wb = load_workbook(filename = sPath)
    ws = wb[sSheet]
    #print sPath, sSheet
    row = 1
    dt=[]
    while (bool(ws['A%s'%(row)].value)):
        #print row
        A=ws['A%s'%(row)].value
        B=ws['b%s'%(row)].value
        C=ws['c%s'%(row)].value
        D=ws['d%s'%(row)].value
        E=ws['e%s'%(row)].value
        F=ws['f%s'%(row)].value
        G=ws['g%s'%(row)].value
        H=ws['h%s'%(row)].value
        I=ws['i%s'%(row)].value
        J=ws['j%s'%(row)].value
        K=ws['k%s'%(row)].value
        L=ws['l%s'%(row)].value
        M=ws['m%s'%(row)].value
        N=ws['n%s'%(row)].value
        O=ws['o%s'%(row)].value
        
        vals =[A,B,C,D,E,F,G,H,I,J,K,L,M,N,O]
        dt.append(vals)
        row+=1
    return dt


def LoadObjectsAndAbons(sPath, sSheet):
    #Добавление объектов
    global cfg_excel_name
    cfg_excel_name=sPath
    global cfg_sheet_name
    cfg_sheet_name=sSheet
    result="Объекты не загружены"

    dtAll=GetTableFromExcel(sPath,sSheet) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    
    for i in range(1,len(dtAll)):
        #print  dtAll[i][2],dtAll[i][3]
        # print u'Обрабатываем строку ' + unicode(dtAll[i][2])+' - ' + unicode(dtAll[i][3])
        obj_l0=str(dtAll[i][0]).strip()
        writeToLog( obj_l0)
        obj_l1=str(dtAll[i][1]).strip()
        writeToLog(obj_l1)
        obj_l2=str(dtAll[i][2]).strip()
        writeToLog(obj_l2)
        abon=str(dtAll[i][3]).strip()
        writeToLog(abon)
        account_1=str(dtAll[i][4])
        writeToLog(account_1)
        account_2=str(dtAll[i][5])
        writeToLog(account_2)
        isNewObj_l0=SimpleCheckIfExist('objects','name',obj_l0,"","","")
        isNewObj_l1=SimpleCheckIfExist('objects','name',obj_l1,"","","")
        isNewObj_l2=SimpleCheckIfExist('objects','name',obj_l2,"","","")
        isNewAbon=SimpleCheckIfExist('objects','name', obj_l2,'abonents', 'name', abon)
        kv=0
        if not (isNewObj_l0):
            writeToLog('create object '+obj_l0)
            add_parent_object = Objects( name=obj_l0, level=0)
            add_parent_object.save()
            writeToLog('create object '+obj_l1)
            #print add_parent_object
            add_object1=Objects(name=obj_l1, level=1, guid_parent = add_parent_object)
            add_object1.save()
            writeToLog('create object '+obj_l2)
            add_object2=Objects(name=obj_l2, level=2, guid_parent = add_object1)
            add_object2.save()
            
            writeToLog('create abonent '+abon)
            add_abonent = Abonents(name = abon, account_1 =str(account_1), account_2 =str(account_2), guid_objects =add_object2, guid_types_abonents = TypesAbonents.objects.get(guid= "e4d813ca-e264-4579-ae15-385cdbf5d28c"))
            add_abonent.save()
            result="Объекты созданы"
            continue
        if not (isNewObj_l1):
            writeToLog('create object '+obj_l1)
            dtParent=GetSimpleTable('objects','name',obj_l0)
            if dtParent: #родительский объект есть
                guid_parent=dtParent[0][0]
                add_object1=Objects(name=obj_l1, level=1, guid_parent = Objects.objects.get(guid=guid_parent))
                add_object1.save()                
                add_object2=Objects(name=obj_l2, level=2, guid_parent = add_object1)
                add_object2.save()
                writeToLog('create abonent '+abon)
                add_abonent = Abonents(name = abon, account_1 =str(account_1), account_2 =str(account_2), guid_objects =add_object2, guid_types_abonents = TypesAbonents.objects.get(guid= "e4d813ca-e264-4579-ae15-385cdbf5d28c"))
                add_abonent.save()
                result="Объекты созданы"
                continue
        if not (isNewObj_l2):
            writeToLog('create object '+obj_l2)
            dtParent=GetSimpleTable('objects','name',obj_l1)
            if dtParent: #родительский объект есть
                guid_parent=dtParent[0][0]                
                add_object = Objects(name=obj_l2, level=2, guid_parent = Objects.objects.get(guid=guid_parent))
                add_object.save()
                result="Объекты созданы"
        if not (isNewAbon):
            writeToLog('create abonent '+ abon)
            dtObj=GetSimpleTable('objects','name',obj_l2)
            if dtObj: #родительский объект есть
                guid_object=dtObj[0][0]
                add_abonent = Abonents(name = abon, account_1 =str(account_1), account_2 =str(account_2), guid_objects = Objects.objects.get(guid=guid_object), guid_types_abonents = TypesAbonents.objects.get(guid= "e4d813ca-e264-4579-ae15-385cdbf5d28c"))
                add_abonent.save()
                kv+=1

    result+=" Абоненты созданы"
    return result


def load_electric_objects(request):
    args={}
    fileName=""
    sheet    = ""
    tcp_ip_status    = ""
    counter_status    = ""
    result="Не загружено"
    #writeToLog('test1') 
    try:    
        if request.is_ajax():
            if request.method == 'GET':            
                request.session["choice_file"]       = fileName    = request.GET['choice_file']
                request.session["choice_sheet"]      = sheet    = request.GET['choice_sheet']
                request.session["tcp_ip_status"]     = tcp_ip_status    = request.GET['tcp_ip_status']
                request.session["object_status"]     = object_status    = request.GET['object_status']
                request.session["counter_status"]    = counter_status    = request.GET['counter_status']
                
                directory=os.path.join(BASE_DIR,'static/cfg/')
                sPath=directory+fileName
                writeToLog(sPath)
                            
                result=LoadObjectsAndAbons(sPath, sheet)
    except: # catch *all* exceptions
        e = sys.exc_info()[0]
        result = ( "Ошибка: %s" % e )
    
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["tcp_ip_status"]=tcp_ip_status
    args["object_status"]=result
    args["counter_status"]=counter_status
    return render(request,"service/service_electric.html", args)


def LoadElectricMeters(sPath, sSheet):
    global cfg_excel_name
    cfg_excel_name = sPath
    global cfg_sheet_name
    cfg_sheet_name = sSheet
    result = u"Счётчики не загружены"
    #print type(sPath), sPath, type(sSheet), sSheet
    dtAll=GetTableFromExcel(sPath,sSheet) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    met=0
    #print('load dt - ok')
    #print('range(1,len(dtAll)) %s'%str(range(1,len(dtAll))))
    for i in range(1,len(dtAll)):
        #writeToLog(u'Обрабатываем строку ' + unicode(dtAll[i][3])+' - '+unicode(dtAll[i][6]))
        #print((dtAll[i][3]), (dtAll[i][6]))
        obj_l2=str(dtAll[i][2]).strip() #корпус
        abon=str(dtAll[i][3]).strip()  #квартира
        meter=str(dtAll[i][6]).strip()  #номер счётчика
        adr=str(dtAll[i][7]).strip()  #номер в сети
        type_meter=str(dtAll[i][8]).strip()  #тип счётчика
        NumLic=str(dtAll[i][5]).strip()  #номер лицевого счёта, тут используется как пароль для м-230-ум
        Group=str(dtAll[i][12]).strip() 
        attr1=str(dtAll[i][13]).strip() 
        attr2=str(dtAll[i][14]).strip() 
        isNewMeter=SimpleCheckIfExist('meters','factory_number_manual',meter,"","","")
        isNewAbon=SimpleCheckIfExist('objects','name', obj_l2,'abonents', 'name', abon)
        isR = False
        isHalfs = False
        #print('attr1, attr2', meter, attr1, attr2)
        if (attr1 == '+'):
            isR = True    
        if (attr2 == '+'):
            isHalfs = True
        #print('attr1, attr2', meter, isR, isHalfs)
        #writeToLog( u'счётчик существует ', isNewMeter)
        if not (isNewAbon):
            #print('Need create struct!')
            return "Сначала создайте структуру объектов и абонентов"
        if not (isNewMeter):
            print(str(type_meter))
            if str(type_meter) == 'М-200':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), guid_types_meters = TypesMeters.objects.get(guid = "6224d20b-1781-4c39-8799-b1446b60774d") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'М-200')
            elif str(type_meter) == 'М-230':
                #print('m-230')
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), password = 111111 , factory_number_manual = str(meter), guid_types_meters = TypesMeters.objects.get(guid = "423b33a7-2d68-47b6-b4f6-5b470aedc4f4") )
                add_meter.save()
                #print('Device added' + ' --->   ' + 'М-230')
                
            elif str(type_meter) == 'М-230-УМ':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), password = str(NumLic) , factory_number_manual = str(meter), guid_types_meters = TypesMeters.objects.get(guid = "20e4767a-49e5-4f84-890c-25e311339c28") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'М-230-УМ')
                
            elif str(type_meter) == 'Эльф 1.08':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), guid_types_meters = TypesMeters.objects.get(guid = "1c5a8a80-1c51-4733-8332-4ed8d510a650"), attr1=str(attr1), attr2=str(attr2) )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Эльф 1.08')
            elif str(type_meter) == 'СПГ762-1':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), guid_types_meters = TypesMeters.objects.get(guid = "c3ec5c22-d184-41c5-b6bf-66fa30215a41") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'СПГ762-1')                
            elif str(type_meter) == 'СПГ762-2':
                add_meter = Meters(name=str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), guid_types_meters = TypesMeters.objects.get(guid = "5eb7dd59-faf9-4ead-8654-4f3de74de2b0") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'СПГ762-2')
            elif str(type_meter) == 'СПГ762-3':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), guid_types_meters = TypesMeters.objects.get(guid = "e4fb7950-a44f-41f0-a6ff-af5e30d9d562") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'СПГ762-3')
            elif str(type_meter) == 'Sayany':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), guid_types_meters = TypesMeters.objects.get(guid = "5429b439-233e-4944-b91b-4b521a10f77b") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Sayany')
            elif str(type_meter) == 'Tekon_hvs':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), password = str(Group), guid_types_meters = TypesMeters.objects.get(guid = "8398e7d6-39f7-45d2-9c45-a1c48e751b61") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Tekon_gvs')
            elif str(type_meter) == 'Tekon_hvs':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), password = str(Group), guid_types_meters = TypesMeters.objects.get(guid = "64f02a2c-41e1-48b2-bc72-7873ea9b6431") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Tekon_gvs')

            elif str(type_meter) == 'Tekon_heat':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), password = str(Group), guid_types_meters = TypesMeters.objects.get(guid = "b53173f2-2307-4b70-b84c-61b634521e87") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Tekon_heat')
            elif str(type_meter) == 'Пульсар ХВС':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), attr1 = str(attr1), guid_types_meters = TypesMeters.objects.get(guid = "f1789bb7-7fcd-4124-8432-40320559890f") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Пульсар ХВС')
            
            elif str(type_meter) == 'Пульсар ГВС':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), attr1 = str(attr1), guid_types_meters = TypesMeters.objects.get(guid = "a1a349ba-e070-4ec9-975d-9f39e61c34da") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Пульсар ГВС')

            elif str(type_meter) == 'Пульс СТК ГВС':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), attr1 = str(attr1), guid_types_meters = TypesMeters.objects.get(guid = "12c9874c-1dc4-4cb0-95e7-4ff6ca7ab17f") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Пульс СТК ГВС')

            elif str(type_meter) == 'Пульс СТК ХВС':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), attr1 = str(attr1), guid_types_meters = TypesMeters.objects.get(guid = "fbc9874c-1dc4-4cb0-95e7-4ff6ca7ab17f") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Пульс СТК ХВС')

            elif str(type_meter) == 'Пульсар Теплосчётчик':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), guid_types_meters = TypesMeters.objects.get(guid = "82b96b1c-31cf-4753-9d64-d22e2f4d036e") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Пульсар Теплосчётчик')

            elif str(type_meter) == 'Пульсар Холодосчётчик':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), guid_types_meters = TypesMeters.objects.get(guid = "c1ae0de6-f071-4e07-8452-09059eef187b") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Пульсар Холодосчётчик')

            elif str(type_meter) == 'Пульсар 3Ф4Т':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), guid_types_meters = TypesMeters.objects.get(guid = "5f9e013c-378d-4947-a1a7-33e6ebdc1cef") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Пульсар 3Ф4Т')

            elif str(type_meter) == 'Пульс СТК Теплосчётчик':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), guid_types_meters = TypesMeters.objects.get(guid = "bb111ede-e00b-4e1d-a8ba-1ef61dba1caa") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Пульс СТК Теплосчётчик')

            elif str(type_meter) == 'Карат 307':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), guid_types_meters = TypesMeters.objects.get(guid = "84fb7a85-ab91-4e93-9154-76ddee35a316") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Карат 307')
            elif str(type_meter) == 'Danfoss SonoSelect':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), guid_types_meters = TypesMeters.objects.get(guid = "aa491ede-e00b-4e1d-a8ba-1ef61dba1caa") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Danfoss SonoSelect')
            elif str(type_meter) == 'СЭТ-4ТМ.03М':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), guid_types_meters = TypesMeters.objects.get(guid = "66b7ce6a-f280-4e54-8c8d-f69f34aabdf9") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'СЭТ-4ТМ.03М')
            elif str(type_meter) == 'СЕ301':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), guid_types_meters = TypesMeters.objects.get(guid = "17d88dbc-23b9-490a-9895-58ad24fe459d") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'СЕ301')
            elif str(type_meter) == 'Sanext':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), guid_types_meters = TypesMeters.objects.get(guid = "e8fa5e00-e1b9-4ef3-bc39-b8439a44b540") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Sanext')
            elif str(type_meter) == 'Нартис СПОДЭС':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), guid_types_meters = TypesMeters.objects.get(guid = "8790eaeb-671b-4596-b80e-d6475d74382c") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Нартис СПОДЭС')
            elif str(type_meter) == 'Пульс Эл. ГВС':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), attr1 = str(attr1), guid_types_meters = TypesMeters.objects.get(guid = "fbe5620f-726d-4b2d-9f70-6d835bd5e3c6") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Пульс Эл. ГВС')
            elif str(type_meter) == 'Пульс Эл. ХВС':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), attr1 = str(attr1), guid_types_meters = TypesMeters.objects.get(guid = "c922dec6-f830-4ee1-bb6a-cd0a96c53c81") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Пульс Эл. ХВС')
            elif str(type_meter) == 'ЭкоНом ГВС':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), attr1 = str(attr1), guid_types_meters = TypesMeters.objects.get(guid = "e2e6c4c5-636a-432a-bdbf-6a5ab4b1fdee") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'ЭкоНом ГВС')
            elif str(type_meter) == 'ЭкоНом ХВС':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), attr1 = str(attr1), guid_types_meters = TypesMeters.objects.get(guid = "50098019-7418-4661-baa9-b913de3596da") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'ЭкоНом ХВС')
            elif str(type_meter) == 'Декаст ГВС':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), attr1 = str(attr1), guid_types_meters = TypesMeters.objects.get(guid = "36b6ea95-beb1-490d-a39f-06163bfcaae5") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Декаст ГВС')
            elif str(type_meter) == 'Декаст ХВС':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), attr1 = str(attr1), guid_types_meters = TypesMeters.objects.get(guid = "657d8ad0-bdba-4459-a07e-4d4eb72950d6") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Декаст ХВС')
            elif str(type_meter) == 'ЭкоНом Теплосчётчик':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), attr1 = str(attr1), guid_types_meters = TypesMeters.objects.get(guid = "aefa5648-2240-42b4-88cf-04b093a60187") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'ЭкоНом Теплосчётчик')

            elif str(type_meter) == 'Декаст Теплосчётчик':
                add_meter = Meters(name = str(type_meter) + ' ' + str(meter), address = str(adr), factory_number_manual = str(meter), attr1 = str(attr1), guid_types_meters = TypesMeters.objects.get(guid = "b95134db-af0c-4eea-bc8e-32b2bcfc7e1d") )
                add_meter.save()
                writeToLog('Device added' + ' --->   ' + 'Декаст Теплосчётчик')

            else:
                writeToLog('Не найдено совпадение с существующим типом прибора')
                met-=1
            
            #Если экземпляр был создан, то добавляем считываемые параметры
            try:
                add_taken_param_no_signals(instance = add_meter, isR = isR, isHalfs = isHalfs)
            except:
                e = sys.exc_info()[0]
                #print(e)
                return( "Ошибка: %s" % e )
            
            met+=1
            
    result=" Загружено счётчиков "+str(met)
    
    return result


def load_electric_counters(request):
    global isService
    isService=True
    OnOffSignals()
    args={}
    fileName=""
    sheet    = ""
    tcp_ip_status    = ""
    object_status    = ""
    counter_status = ""
    result=""
    try:
        if request.is_ajax():
            if request.method == 'GET':
                request.session["choice_file"]    = fileName    = request.GET['choice_file']
                request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']
                request.session["tcp_ip_status"]    = tcp_ip_status    = request.GET['tcp_ip_status']
                request.session["object_status"]    = object_status    = request.GET['object_status']
                request.session["counter_status"]    = counter_status    = request.GET['counter_status']
                directory=os.path.join(BASE_DIR,'static/cfg/')
                sPath=directory+fileName
                print(sPath)
                result=LoadElectricMeters(sPath, sheet)                
    except: # catch *all* exceptions
        e = sys.exc_info()[0]
        result = ( "Ошибка: %s" % e )
        
    #print fileName
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["tcp_ip_status"]=tcp_ip_status
    args["object_status"]=object_status
    args["counter_status"]=result
    isService=False
    OnOffSignals()
    return render(request,"service/service_electric.html", args)


@csrf_exempt
def service_water(request):
    args={}
    return render(request,"service/service_water.html", args)


def add_link_meter(sender, instance, created, **kwargs):
    #print u'Start add link port - meter'
    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    writeToLog( str(dtAll[1][1]))
    if (dtAll[1][1] == 'Объект'): #вода
        #print(u'Add impulse connect')
        add_link_meter_port_from_excel_cfg_water_v2(sender, instance, created, **kwargs)
    else:# электрика
        #print(u'Add digital connect')
        add_link_meter_port_from_excel_cfg_electric(sender, instance, created, **kwargs)


def add_link_meter_port_from_excel_cfg_water_v2(sender, instance, created, **kwargs):
    """Делаем привязку счётчика к порту по excel файлу ведомости"""
    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    #print u'test'
    for i in range(1,len(dtAll)):
        #print u'Обрабатываем строку ' + unicode(dtAll[i][6])+' - '+unicode(dtAll[i][7])
        #print dtAll[i]
        meter=dtAll[i][5] #счётчик
        #print meter
        #print instance.factory_number_manual
        #print dtAll[0][5], dtAll[0][4]       
        ip_adr=str(dtAll[i][7]).strip()
        ip_port=str(dtAll[i][8]).strip()
        # Привязка к tpc порту
        if meter is not None:
            if str(meter) == instance.factory_number_manual :
                 guid_ip_port_from_excel = connection.cursor()
                 sQuery="""SELECT 
                                      tcpip_settings.guid
                                    FROM 
                                      public.tcpip_settings
                                    WHERE 
                                      tcpip_settings.ip_address = '%s' AND 
                                      tcpip_settings.ip_port = '%s';"""%(str(ip_adr), str(ip_port))
    #print sQuery
                 guid_ip_port_from_excel.execute(sQuery)
                 guid_ip_port_from_excel = guid_ip_port_from_excel.fetchall()
                 print(guid_ip_port_from_excel)
                 
                 IsExistLink=SimpleCheckIfExist("Link_Meters_Tcpip_Settings","guid_meters",instance.guid,"","guid_tcpip_settings", guid_ip_port_from_excel)
                 #print IsExistLink
                 if IsExistLink: break
                 if guid_ip_port_from_excel:
                     guid_ip_port = TcpipSettings.objects.get(guid=guid_ip_port_from_excel[0][0])
                     add_ip_port_link = LinkMetersTcpipSettings(guid_meters = instance, guid_tcpip_settings = guid_ip_port)            
                     add_ip_port_link.save()
                     #print u'Связь добавлена ', meter, ip_adr, ip_port
                 else: pass #print(u'Не прогружен порт')
                 

def add_link_meter_port_from_excel_cfg_electric(sender, instance, created, **kwargs):
    """Делаем привязку счётчика к порту по excel файлу ведомости"""    
    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    
    for i in range(1,len(dtAll)):
        #print u'Обрабатываем строку ' + unicode(dtAll[i][6])+' - '+unicode(dtAll[i][7])
        meter=dtAll[i][6] #счётчик
        #print dtAll[0][11], dtAll[0][12]
        PortType=str(dtAll[0][11]) # com или tcp-ip
        #print 'i=',i,' len=', len(dtAll)
        ip_adr=str(dtAll[i][10]).strip()
        ip_port=str(dtAll[i][11]).strip()
        # Привязка к tpc порту
        if meter is not None:
            if str(meter) == instance.factory_number_manual :
                if PortType == 'Com-port':
                    #print 'dtAll[i][12]', dtAll[i][12]
                    guid_com_port_from_excel = connection.cursor()
                    guid_com_port_from_excel.execute("""SELECT 
                                                      comport_settings.guid
                                                    FROM 
                                                      public.comport_settings
                                                    WHERE 
                                                      comport_settings.name = '%s';"""%(str(dtAll[i][11])))
                    guid_com_port_from_excel = guid_com_port_from_excel.fetchall()
                    #print guid_com_port_from_excel
                    if (len(guid_com_port_from_excel)>0):
                        guid_com_port = ComportSettings.objects.get(guid=guid_com_port_from_excel[0][0])
                        add_com_port_link = LinkMetersComportSettings(guid_meters = instance, guid_comport_settings = guid_com_port)
                        add_com_port_link.save()
                    else: print(' Com-port does not exist. Create it before! ')
                
                else:
                    guid_ip_port_from_excel = connection.cursor()
                    sQuery="""SELECT tcpip_settings.guid
                                                    FROM 
                                                      public.tcpip_settings
                                                    WHERE 
                                                      tcpip_settings.ip_address = '%s' AND 
                                                      tcpip_settings.ip_port = '%s';"""%(ip_adr, ip_port)
                    #print sQuery
                    guid_ip_port_from_excel.execute(sQuery)
                    guid_ip_port_from_excel = guid_ip_port_from_excel.fetchall()
            
                    #print guid_ip_port_from_excel
                    if (len(guid_ip_port_from_excel)>0):
                        guid_ip_port = TcpipSettings.objects.get(guid=guid_ip_port_from_excel[0][0])
                        add_ip_port_link = LinkMetersTcpipSettings(guid_meters = instance, guid_tcpip_settings = guid_ip_port)            
                        add_ip_port_link.save()
                    else: pass #print u'Привязки по портам не добавлены'
            else:
                pass


def add_link_taken_params(sender, instance, created, **kwargs):
    #print 'link taken params'
    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    if (dtAll[1][1] == 'Объект'): #вода
        add_link_abonents_taken_params2(sender, instance, created, **kwargs)
    else:# электрика
        add_link_abonent_taken_params_from_excel_cfg_electric(sender, instance, created, **kwargs)


def add_taken_param(sender, instance, created, **kwargs): # Добавляем считываемые параметры при создании счётчика
    if instance.guid_types_meters.name == 'Меркурий 230':
        #Добавляем параметры для Меркурия 230
    # T0 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "bdcd1268-37f3-4579-a9d9-5becb2ba8aa3")) # A+ T0 месячные
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "99cd6002-f81c-4ad6-9cb0-53a92a498519")) # A+ T0 суточные
        add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e8c20ce7-bdb6-4ea6-8401-cee28049a7d7")) # A+ T0 текущие
        #add_param.save()
    # T0 R+
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"2ebc02e6-65c6-40ab-b717-0d98d66b5701")) # R+ T0 месячные
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"345a24a4-95b7-4f67-b004-716706ed2560")) # R+ T0 суточные
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4c93dd55-1ec2-48c7-9865-9ceab580b7b3")) # R+ T0 текущие
        #add_param.save()
        
    # T1 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "17789c36-4593-4ff2-94eb-1d0cebdb5366")) # A+ T1 месячные
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "d262c71a-6da4-4ec0-a9c3-b9ea659c246d")) # A+ T1 суточные
        add_param.save()
    # T2 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c31297be-220b-4971-8642-6b614aa7ecee")) # A+ T2 месячные
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "37011b85-c8af-4f6c-857d-4b93a95d31e1")) # A+ T2 суточные
        add_param.save()
    # T3 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "79741ba9-e8b8-4352-862e-17a9c4d928ce")) # A+ T3 месячные
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c3bb9033-ffcb-4a28-91e2-6b45924b8858")) # A+ T3 суточные
        add_param.save()
    
    #Код ошибки
        # add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"7175f6c7-b816-40f6-86f4-e08a309c08f6")) # код ошибки
        # add_param.save()

    # Ток
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"aee312b0-adb1-4be9-9879-b3a3598f9b29")) # Ia текущее
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"7ed0d364-e790-4325-a927-9ef86a685f00")) # Ib текущее
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"474b0809-482a-4851-9a96-4587f8c59152")) # Ic текущее
        #add_param.save()
    # Напряжение
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c06f7315-abc6-4889-97ad-201a936c8f2c")) # Ua текущее
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"632f76fb-4dd9-4e7d-86a0-a57a27fc648a")) # Ub текущее
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1a3ca6ca-8866-4aad-8712-d9df003fe692")) # Uc текущее
        #add_param.save()
    # Мощность
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"3077b3ac-fde2-4435-9e6f-17464310c090")) # P Активная мощность
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e7617c95-7e42-4cfa-9acd-5bc119261c6d")) # Q Реактивная мощность
        #add_param.save()
    #Получасовки
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "6af9ddce-437a-4e07-bd70-6cf9dcc10b31")) # A+ 30-мин. срез мощности
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "66e997c0-8128-40a7-ae65-7e8993fbea61")) # R+ 30-мин. срез мощности
        add_param.save()
    elif instance.guid_types_meters.name == 'Меркурий 233':
        #Добавляем параметры для Меркурия 233
        pass
    elif instance.guid_types_meters.name == 'Пульсар16':
        #Добавляем параметры для Пульсар16
    # Суточные
      # Канал 1
      # Канал 2
      # Канал 3
      # Канал 4
      # Канал 5
      # Канал 6
      # Канал 7
      # Канал 8
      # Канал 9
      # Канал 10
      # Канал 11
      # Канал 12
      # Канал 13
      # Канал 14
      # Канал 15
      # Канал 16
   
    # Текущие
      # Канал 1
      # Канал 2
      # Канал 3
      # Канал 4
      # Канал 5
      # Канал 6
      # Канал 7
      # Канал 8
      # Канал 9
      # Канал 10
      # Канал 11
      # Канал 12
      # Канал 13
      # Канал 14
      # Канал 15
      # Канал 16
       pass
    elif instance.guid_types_meters.name == 'Пульсар10':
        #Добавляем параметры для Пульсар10
        pass
    elif instance.guid_types_meters.name == 'Пульсар 16M':
        print('add_taken_params_16m')
    # Суточные
      # Канал 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "fc4a9568-4674-4a80-b497-e4f34399acd5"))
        add_param.save()
      # Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9e6e308f-abec-4b47-9b99-9cb590c55d0c"))
        add_param.save()
      # Канал 3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "e6815dd5-fbbc-480f-8b95-025d7f9a0403"))
        add_param.save()
      # Канал 4
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "612d2f20-d454-4e14-910b-1fd89bbb31dd"))
        add_param.save()
      # Канал 5
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "d82c7576-8e5e-4e93-ae10-58459b31e4a0"))
        add_param.save()
      # Канал 6
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "6ccc7efb-d9fe-4285-b343-8ed22d2d3625"))
        add_param.save()
      # Канал 7
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "72567365-9a40-4f97-ab25-0911585035bf"))
        add_param.save()
      # Канал 8
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9203f5ed-d5da-4462-91d1-5aea42e99124"))
        add_param.save()
      # Канал 9
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "e4068568-d8c4-42ab-9957-7292753e2891"))
        add_param.save()
      # Канал 10
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9b5ab67b-40aa-4536-8b7c-340a773ab31b"))
        add_param.save()
      # Канал 11
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "4fd440c4-9ec5-4ab9-a073-6c4d3a174777"))
        add_param.save()
      # Канал 12
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "00b7f7c5-37f3-494a-8ceb-5a62f9ebf4e3"))
        add_param.save()
      # Канал 13
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "169a79e0-da6f-4091-9fc7-ab81adc0d7e0"))
        add_param.save()
      # Канал 14
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "17e9c8fe-0d69-4466-b64e-185452c61555"))
        add_param.save()
      # Канал 15
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "25de493d-c680-4ca6-ac02-b778022ee151"))
        add_param.save()
      # Канал 16
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "908e88f0-f9a0-421d-bbe7-9bafdf5d2565"))
        add_param.save() 
  
    # Текущие
      # Канал 1
       # add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e3f1325e-3018-40ba-b94a-ab6d6ac093e9"))
       # add_param.save()
      # Канал 2
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5a6b0338-c15d-4224-a04f-a10fc73c5fc7"))
        #add_param.save()
      # Канал 3
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"48a42afe-d9ac-4180-a733-6dd5f9d9ca80"))
        #add_param.save()
      # Канал 4
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"01a5419c-c701-4185-95b6-457b8c9ca2d0"))
        #add_param.save()
      # Канал 5
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"85c4295e-bc6a-46ec-9866-0bf9f77c6904"))
        #add_param.save()
      # Канал 6
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"68270d0a-5043-4ea2-9b61-4adaa298abad"))
        #add_param.save()
      # Канал 7
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"cd489c4b-6e74-4c65-bfee-c0fa78a853bf"))
        #add_param.save()
      # Канал 8
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"f29062a4-ab60-4117-8f85-0cdec634c797"))
        #add_param.save()
      # Канал 9
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e8521cd7-2f38-4619-935d-8fe86234dbe7"))
        #add_param.save()
      # Канал 10
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1349b747-41ca-4ba8-a690-69c649129f44"))
        #add_param.save()
      # Канал 11
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"99ab1a30-fde8-4b81-9f9e-2f731516ce1b"))
        #add_param.save()
      # Канал 12
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c7f6a397-833d-4020-9d2b-38c19bec272c"))
        #add_param.save()
      # Канал 13
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4413bffb-1832-4900-9351-5ac3666dd8b0"))
        #add_param.save()
      # Канал 14
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6280490b-123d-4e27-bef9-19fd7dc2cf54"))
        #add_param.save()
      # Канал 15
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"93891c5a-1c8f-4906-b7f0-961dc8ad3c9f"))
        #add_param.save()
      # Канал 16
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"22dd3a17-a828-44e0-80d9-db075ba120ae"))
        #add_param.save()

    elif instance.guid_types_meters.name == 'Пульсар 10M':
        #Добавляем параметры для Пульсар10
    # Месячные
      # Канал 1
      # Канал 2
      # Канал 3
      # Канал 4
      # Канал 5
      # Канал 6
      # Канал 7
      # Канал 8
      # Канал 9
      # Канал 10
    
    # Суточные
      # Канал 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "325ec164-9428-4a57-867c-33d4eaf8cc2a"))
        add_param.save()
      # Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "99a99024-65b4-44dd-99fc-6a5cf1d4aaee"))
        add_param.save()
      # Канал 3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "f897f0ca-4e35-4f0d-b345-3379668aa36f"))
        add_param.save()
      # Канал 4
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "034374bd-2dfb-4568-aa16-84255df33c88"))
        add_param.save()
      # Канал 5
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b6bdfae8-4f27-4056-af79-d746b44038ee"))
        add_param.save()
      # Канал 6
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "2c2f7176-8b77-44f4-9678-4773e95e67ce"))
        add_param.save()
      # Канал 7
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "91bb7c43-f802-4ebd-a8fe-75f833acedeb"))
        add_param.save()
      # Канал 8
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "cf24b669-1c5b-4db7-936a-5f9d5c8be928"))
        add_param.save()
      # Канал 9
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "96035c7c-ee7c-41f6-9723-8a75dd9ed573"))
        add_param.save()
      # Канал 10
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "253475ea-614d-4aad-93a8-e81e4c9028e9"))
        add_param.save()   

    # Текущие
      # Канал 1
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"32dad392-ca1e-4110-8f2c-a86b02e26fb3"))
        #add_param.save()
      # Канал 2
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"3e13694b-7cb5-4417-a091-af8a7db34dc7"))
        #add_param.save()
      # Канал 3
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1023b35b-3cbf-4519-aac3-3bf1ebae07c1"))
        #add_param.save()
      # Канал 4
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"eea27ade-44cd-4e66-8298-00a4a6ad915a"))
        #add_param.save()
      # Канал 5
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"25e09d4d-3a48-4381-ad5d-b783c03c4d35"))
        #add_param.save()
      # Канал 6
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"069898ea-9d74-4571-b719-e8e6f1513c12"))
        #add_param.save()
      # Канал 7
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"084aa5f4-75d5-41f6-b0d6-9f2403eacd2c"))
        #add_param.save()
      # Канал 8
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"786ed8b8-aed1-478c-ae75-99caf1358cf0"))
        #add_param.save()
      # Канал 9
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6fc4c39c-9a43-4cb7-a066-c40fd2ca47e5"))
        #add_param.save()
      # Канал 10
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"8b2aa40a-cd91-4e22-b9d1-596e49e5f839"))
        #add_param.save()  

    elif instance.guid_types_meters.name == 'Пульсар 2M':
        #Добавляем параметры для Пульсар10
    
    # Месячные
      # Канал 1
      # Канал 2
    
    # Суточные
      # Канал 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "0239dffb-de88-45e5-b6f6-18bf39f92525"))
        add_param.save()
      # Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "a1cb319d-ac09-466d-894b-91d90aba4239"))
        add_param.save()   
    
    # Текущие
      # Канал 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "fcc28118-66c0-4cdf-aeba-5da1171aae48"))
        add_param.save()
      # Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "1faeb517-bd1f-4ba0-96a5-67f00764822f"))
        add_param.save()

    elif instance.guid_types_meters.name == 'ПСЧ-3ТА.04':
        #Добавляем параметры для ПСЧ-3ТА.04
        pass
    elif instance.guid_types_meters.name == 'ТЭМ-104':
        #Добавляем параметры для ТЭМ-104
        pass
    elif instance.guid_types_meters.name == 'СЭТ-4ТМ.03М':
        #Добавляем параметры для СЭТ-4ТМ.03М
       
    # T0 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "e7624c25-9852-4ffd-8777-b2bfd16c29a8")) # A+ T0 месячные
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "aa83b499-6a9e-40e1-b68b-dc84fec8490b")) # A+ T0 суточные
        add_param.save()

    # T0 R+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "3d365f91-8bd3-476e-b07e-3f79134f6853")) # R+ T0 месячные
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "087b785e-5d59-4956-9cd3-57706f9557e6")) # R+ T0 суточные
        add_param.save()
           
    #Получасовки
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "4f505e17-7d71-4cf8-9880-c6ce33612e6e")) # A+ 30-мин. срез мощности
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "55abd40d-fb3c-4100-88f2-46d79be7733a")) # R+ 30-��ин. срез мощности
        add_param.save()

    elif instance.guid_types_meters.name == 'Меркурий 200':
        #Добавляем параметры для Меркурий 200

    # Значения суточные (текущие)
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9cbc001d-a262-481f-a1aa-47d02bf18af1")) #T0
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b65d4227-69a5-487b-9999-5539ca3fc004")) #T1
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5e312de9-34cd-4ba7-a744-c9b94a77d98b")) #T2
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "4260ea05-78f8-4c5c-9172-fa161fa96068")) #T3
        add_param.save()
    # Значения на начало месяца
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "86cd925b-48c2-40b8-b211-f116e0e6dbea")) #T0
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "62a3796a-eaae-445d-9166-2ad517186b78")) #T1
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5f6e1e3d-4128-4cfe-94cf-57ac84a7694a")) #T2
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "0c28c135-58f2-4dff-a222-9f3d9f3c742b")) #T3
        add_param.save()
    # Значения на начало суток
        #Не поддерживается прибором, но текущие переделаны на суточные
        

    elif instance.guid_types_meters.name == 'Эльф 1.08':
        #Добавляем параметры для счётчика тепла Elf 108
    
        #-------------Текущие
        # "Энергия"
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "f2bbf267-456e-477a-95d2-abb94c78ba43"))
        add_param.save()
        # "Объем"       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "dad6e2eb-e978-46f4-b7ec-442834b04e7a"))
        add_param.save()
        # "ElfTon"  Время работы прбора     
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "d3c9563d-51ed-4ca7-922f-ac3731065ead"))
        add_param.save()
        # "ElfErr"  Время работы прибора с ошибкой
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "dade3324-b9b0-41c8-bc76-70f617573e43"))
        add_param.save()
        # "Ti"      Температура входа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "acca627e-f21a-4f8b-be7e-038f534b5d11"))
        add_param.save()
        # "To"      Температура выхода
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "01487323-a28f-419e-9589-2563d785ab2a"))
        add_param.save()
        # "Канал 1"      Импульсный вход 1 текущий
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6e7f0d37-df5c-4850-991e-b5d7cb793924"))
        #add_param.save()
        #"Канал 1"      Импульсный вход 1 суточный
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9af27a62-d6c8-4b67-bd36-da7103e0b1f1"))
        add_param.save()
        #"Канал 2"      Импульсный вход 2 суточный
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "86acc33d-7bea-4977-a5b5-c5858ce9a09d"))
        add_param.save()
        # "Канал 2"      Импульсный вход 2 текущий
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"de7bfdfd-c17f-4a7c-942d-b28e85db33cb"))
        #add_param.save()
        #-------------Архивные
        # "Энергия"
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "ae439e1f-5c4b-494c-8a53-a61b85c804a0"))
        add_param.save()
        # "Объем"
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b02153a4-00c0-4800-a55a-c7f9dfbb14e7"))
        add_param.save()
        # "ElfTon" Время работы прибора
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "aa611d48-f1fe-462a-8b0a-0a7596792b69"))
        add_param.save()
        # "ElfErr" Время работы прибора с ошибкой
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "af047098-bd45-4579-a60c-b75bed376bbe"))
        add_param.save()
        
        
    elif instance.guid_types_meters.name == 'СПГ762-1':
        #Добавляем параметры для счётчика газа СПГ762 Подсистема 1
    
        #-------------Часовые
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "edfc6dc0-1628-4a7e-bd04-71107882039a"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "1e27e72c-79d6-4c68-bc04-1be84d061622"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "0fa5d9ef-4c6c-4f78-bc64-9d9b34002344"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "433d7025-15f3-4ab0-9d73-39bd0e425566"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "446a6bb6-c17f-4478-b1d8-252c7eb454d3"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "003c2fb2-0092-4d7d-a513-3dcc50a255da"))
        add_param.save()
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9c99da7b-1a73-48b0-a3f5-54438a3ea824"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "832374da-5834-4fe0-abe3-07d48d447af2"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "4920e70a-452e-4ecf-918e-14ca288c7a1f"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "825dfd73-82fa-4f1a-9635-77bdcb244997"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "daa4a90e-7993-493c-9ac0-c03241b2ab2c"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "72b5f4ba-d179-419c-8795-e2f86f5ee2ff"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "41774f43-9c9c-4867-bf21-e3d1df4fd2f8"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "f610969e-cb2d-436f-9357-e63da72d162e"))
        add_param.save()
        
        #-------------Суточные
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "ac9fe54a-6f51-4ee9-a849-448f0f10a4b6"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "2eefa93a-be60-4f23-9b09-8d1c6bad0a15"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "78c30686-a5d0-436f-9c56-57b076769774"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "d7900bac-b85a-4b83-ad67-b822b470a698"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "068f4de0-f041-4608-b09a-81dbc8f319ff"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "dad040cd-1047-4060-a7ac-e9a6e7f30fb4"))
        add_param.save()        
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "3be53eb7-c931-4728-a442-44d14c9da44f"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "1dff6784-dd8d-4898-ae8b-f1b60fbdc1af"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "af26b128-b634-47a4-96ba-42de6f039fdb"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5bb8af68-588e-470c-9b64-373482f71468"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c32e7e01-106a-49c6-b8b0-6490448548ad"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "8fc09300-4b19-4478-aa44-d2fb1cf792d5"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "53693586-e5b5-4204-922a-a0b0153298ea"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "bf06727c-9cab-4efe-a0f2-6242bb320372"))
        add_param.save()
        
        # Масса газа нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5399a54f-0d2c-47e8-8ffb-882f5dddc239"))
        add_param.save()
        # Объем газа при ст.условиях нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c14eaa34-3264-4fe4-98ab-8da6618fc431"))
        add_param.save()
        # Объем газа при раб.условиях нарастающим итогом              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "4ed5ba8d-9ead-4a60-96e4-726f38432d9a"))
        add_param.save()
        
    elif instance.guid_types_meters.name == 'СПГ762-2':
        #Добавляем параметры для счётчика газа СПГ762 Подсистема 1
    
        #-------------Часовые
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9a17ea4f-a1f8-4fb9-a21e-62f43978535a"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "49728e75-1d62-4b9e-8633-762cb7117b52"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c44af390-ed2a-49d8-9c67-25f543db9935"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "24bd767d-f667-444e-acfa-a935fb8f4699"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "1e675e58-2cc9-4103-8410-2d37704a2bcf"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "d14fa3fa-5bd5-4dd9-b740-61049d38e694"))
        add_param.save()
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "6de742eb-a391-473a-bb4b-ab780a4642b8"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "04b43217-af06-4f09-8d95-0e2d3dbd0905"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b62c6838-7578-41c9-a94c-d06788cc2d41"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "aad9d0b4-1c12-4165-a1b1-4fac9de00c38"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "205a8c8c-de26-44e5-ab72-efb7fe72040c"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c84c6130-ede7-487e-a414-b384964eb81e"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5907cc5f-1386-4fbc-9e5c-7d3f77dba6d6"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c5638410-44b6-41d7-b501-6e5c0a002f48"))
        add_param.save()
        
        #-------------Суточные
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c1643ab8-1707-4b73-9610-0226b1fb6860"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "dbacff0c-2b3d-40c5-aa03-8d51a64919dd"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "279bdfd5-7b22-4d7e-900c-21e4077506dd"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b74e6743-3996-4af7-8024-da3912d14b45"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c7220fc7-5c01-4bcc-ac2a-7c851276af4d"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "381ccc3f-a9d9-4dcf-a9aa-2e5bd0e4efc8"))
        add_param.save()        
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "7f0f0e09-3bd0-4595-84dd-754f4c21bc5e"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "ee687ef8-36de-4de8-9a05-4ac841c9c144"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "8ee128d0-5c21-4faa-a2fe-7432ff9be684"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c61c87eb-5a4f-4095-ac50-4324e7899340"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "bae3e866-f057-4be5-99a0-7474f6c7cbc1"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "8ce68aef-85fa-4ca6-8f9f-dfa1f9e71cdd"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "ea094191-7ce5-4c42-ad5e-e886d02e73e0"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "e4d751b4-ef6c-45ca-b31a-f107f47a97aa"))
        add_param.save()
        
        # Масса газа нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "add68490-64f4-47a2-a801-1fafa48c09a2"))
        add_param.save()
        # Объем газа при ст.условиях нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "525f9439-ce13-43be-a4f2-67f590f4842b"))
        add_param.save()
        # Объем газа при раб.условиях нарастающим итогом              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "a760c696-fda0-47d4-8fb5-899d742957f1"))
        add_param.save()
        
    elif instance.guid_types_meters.name == 'СПГ762-3':
        #Добавляем параметры для счётчика газа СПГ762 Подсистема 1
    
        #-------------Часовые
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "187b787c-1693-4c90-b6df-d868effef692"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "49c86cc3-57c2-4bdf-b4e3-b07f64673d37"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "1e92a9a8-1cd9-4252-b9c7-b33357bafce7"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "06cd5dad-7ea9-438e-abbf-043e8918eb3e"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "8064aad9-778a-4902-b0c0-75b23289469a"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b747eeb5-8c69-443e-b74b-2bb89af64206"))
        add_param.save()
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "38c0b41b-0883-4990-bb0c-8b532caed34c"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "7bf3d68d-4344-49f3-8169-370f6142351a"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "fa188255-c1cc-4c2d-844c-3b40a3a7559e"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "06c97066-1e35-4bb1-a96f-fe3c0056cf39"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "e530ab26-92e8-4edc-8e6d-5cd6184bfbe7"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5d315c1e-b237-46d6-9273-be4e597ad1c2"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "54b86222-3d1c-440a-b02f-bedbef0e9e28"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "acb12185-dd88-4449-8f03-76b6fd148958"))
        add_param.save()
        
        #-------------Суточные
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "d1f0258f-42ba-4e4f-a66c-74aed4d512ce"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "7ec1a0fb-88f0-497a-8917-01c0b731b88a"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "3465cb7c-57ea-4ad8-afde-74fb2814ddeb"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "a3014fec-73df-4fd4-a68c-9c3ff737d140"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "986f24b0-df76-4c36-9b7f-fbbe05a10c94"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "2a5ec8f1-b6fe-4eff-b91f-42a7712dd663"))
        add_param.save()        
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "cbfc543d-3a19-46c3-8075-ff59492d2620"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "8a24d34e-aee6-4865-bf21-56d9c07dcd1e"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "85de87d5-8e0a-4088-8248-8a64367db47e"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b2bd2c95-ee85-4156-9ef6-7fc25d29a244"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "8674a61c-af88-46c5-b553-fecc9a7d0837"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5880bd0f-699d-407c-a3f0-6cea0ebde423"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9057618b-445c-4581-86a6-4715469db938"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "4991e5c0-0827-4467-b9d2-7613d1b6dd09"))
        add_param.save()
        
        # Масса газа нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "dcd5ed6a-7bd0-41ba-8850-5b88a9831c04"))
        add_param.save()
        # Объем газа при ст.условиях нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "cbeaa4fa-d1fb-4bf5-9688-7084b57fbfe4"))
        add_param.save()
        # Объем газа при раб.условиях нарастающим итогом              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "d16b31ea-87d1-409a-bbf8-4a743b678dbb"))
        add_param.save()

    elif instance.guid_types_meters.name == 'Sayany':
        #Добавляем параметры для счётчика Sayany
    
        #-------------Суточные
        # "Q" Тепловая энергия. Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "e7f2ffba-9a40-43e1-80f3-ddd22596cdb8"))
        add_param.save()    
        # "Q" Тепловая энергия. Канал2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "6f9cd79e-ca34-447e-8ad1-d54531389fe1"))
        add_param.save() 
        # "M" Расход воды. Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b05de8e2-6176-4fc0-bc44-79ceb4229c80"))
        add_param.save() 
        # "M" ТРасход воды. Канал2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5f256e9b-1cb3-4f27-a53a-d08b446dda58"))
        add_param.save() 
        # "T" Температура. Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "75474616-f3db-4903-91d5-1f22f6593394"))
        add_param.save() 
        # "T" Температура. Канал2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "f3210c5b-afde-4c9a-b201-9c7c403c4cf2"))
        add_param.save() 
        # "T" Температура. Канал3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b12762a0-0a06-49a4-b842-8ad3378f4602"))
        add_param.save() 
        # "T" Температура. Канал4
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "472ba2fd-cc06-4147-a1e7-c1bb66096536"))
        add_param.save() 
        
    elif instance.guid_types_meters.name == 'Tekon_hvs':
        #Добавляем параметры для счётчика Tekon. Читаем один тэк с opcretranslator
    
        #-------------Суточные
        # "Показание". Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "d76796ea-ea63-4317-982f-ffbcde2074dc"))
        add_param.save()  
        
    elif instance.guid_types_meters.name == 'Tekon_gvs':
        #Добавляем параметры для счётчика Tekon. Читаем один тэк с opcretranslator
    
        #-------------Суточные
        # "Показание". Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "e71a7206-7a30-45d9-981d-0b7592b96337"))
        add_param.save() 
        
    elif instance.guid_types_meters.name == 'Tekon_heat':
        #Добавляем параметры для счётчика Tekon. Читаем один тэк с opcretranslator
    
        #-------------Суточные
        # "Показание". Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "1dca7dab-371a-4429-afa1-8b4877b38b5b"))
        add_param.save()
        
    elif instance.guid_types_meters.name == 'Меркурий 230-УМ':
        #Добавляем параметры для счётчика Меркурий на УСПД УМ-RTU.    
        
        #-------------Суточные
        # "Показание". T0 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b6e89205-3814-463d-86d1-f52cec7d8962"))
        add_param.save()
        # "Показание". T1 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "7f3c42e6-4000-4373-a0e6-37e66ce819a9"))
        add_param.save() 
        # "Показание". T2 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c6512649-56ea-4214-aa33-84516bfe8dc1"))
        add_param.save() 
        # "Показание". T3 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "4e20bda9-6e75-4b0f-a99a-0e4c1cd07d3b"))
        add_param.save()
        
        #-------------Мощность        
        #А+ Профиль
#        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"922ad57c-8f5e-4f00-a78d-e3ba89ef859f")) # A+ 30-мин. срез мощности
#        add_param.save()        
#        #R+ Профиль
#        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"61101fa3-a96a-4934-9482-e32036c12829")) # R+ 30-мин. срез мощности
#        add_param.save()
        
    elif instance.guid_types_meters.name == 'Пульсар Теплосчётчик':
        #Добавляем параметры для Теплосчётчика Пульсар.
        #------------Суточные
        # "Показание Энергии" Q, Гкал
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "24ae9f51-40a4-4758-a826-a5f8286e1a2e"))
        add_param.save()
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "a3da78fb-b07b-4d53-a980-54b51e26819a"))
        add_param.save()
        # "Показание Температура подачи" Ti, C0
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "de66ecd2-b243-467c-8d1a-cfcb42377300"))
        add_param.save()
        # "Показание Температура выхода" To, C0
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "d3433b80-cb8c-4038-a682-947e6d05955e"))
        add_param.save()

        
    elif instance.guid_types_meters.name == 'Пульсар ХВС':
        #Добавляем параметры для водосчётчика Пульсар ХВС.
        #------------Суточные
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "209894a8-8d19-4e4d-bad8-1767eec4fedf"))
        add_param.save()

    
    elif instance.guid_types_meters.name == 'Пульсар ГВС':
        #Добавляем параметры для водосчётчика Пульсар ГВС.
        #------------Суточные
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5fc2ff3b-999e-4154-ba49-84d3971369b0"))
        add_param.save()
        
    elif instance.guid_types_meters.name == 'Пульс СТК ХВС':
        #Добавляем параметры для водосчётчика Пульс СТК ХВС.
        #------------Суточные
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "31bc817a-2ccd-4021-a8a1-7d63d97dae2c"))
        add_param.save()
        #------------Месячные
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c5b9362b-5c59-47bb-bc61-b3b556b24dc3"))
        add_param.save()

    elif instance.guid_types_meters.name == 'Пульс СТК ГВС':
        #Добавляем параметры для водосчётчика Пульс СТК ГВС.
        #------------Суточные
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "1ae0a516-2975-4a6e-95e3-23412e0f2e67"))
        add_param.save()
        #------------Месячные
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "49f1197f-c6ae-4081-afbc-587ac614a3c3"))
        add_param.save()


    elif instance.guid_types_meters.name == 'Карат 307':
        #print u'Добавляем параметры для счётчика Карат 307'
        #Суточные 
        #Объём     
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "3024fd72-d1e8-4476-a876-4bc09553dde9"))
        add_param.save()
        #Тепло
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "46a63ef5-5761-4e16-a854-1979ddc9668f"))
        add_param.save()
        #Tout
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "6dd6ea63-20dc-46d0-b56e-6890a2b83f48"))
        add_param.save()
        #Tin
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "8a5f5921-5b70-410d-83de-8403ec2a4d87"))
        add_param.save()
        #Ton наработка в минутах
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9c86e183-dd53-4c7f-b728-ffe75a55c633"))
        add_param.save()
        #Terr время работы в ошибке
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "abd41546-02f6-4e2c-8bd2-a60ab80ffe66"))
        add_param.save()
        #Масса 
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "eb617f04-14a3-403c-90e8-286412872232"))
        add_param.save()
    elif instance.guid_types_meters.name == 'Danfoss SonoSelect':
        #print u'Добавляем параметры для счётчика Danfoss SonoSelect'
        #Суточные 
        #Объём     
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "83ba885f-1881-45db-9d63-52195e67cf64"))
        add_param.save()
        #Энергия  
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "510ba9e6-c18b-4982-9763-2ad86c8a8245"))
        add_param.save()
        #Т вход
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "a8940b63-2002-4a28-b671-a455419c1229"))
        add_param.save()
        #Т выход    
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "db1ed365-e85e-4547-aef6-89fed020898f"))
        add_param.save()

        #Месячные
        #Объём     
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "3c763ebc-6ad4-4193-967f-f352bfae92c5"))
        add_param.save()
        #Энергия  
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "7a5d6dd3-3a34-40e4-90af-0c00252b978d"))
        add_param.save()
        #Т вход
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "caf11a15-27ff-4c0d-9b18-d55028e4840b"))
        add_param.save()
        #Т выход    
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "f76371c9-5ecd-44b2-835e-5fc4cdce7141"))
        add_param.save()

        #Текущие 
        #Объём     
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "097122a4-b7d0-4700-add2-bc99a58347d0"))
        add_param.save()
        #Энергия  
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "af254ed7-d4ab-4293-8aaf-8bb13c81efb7"))
        add_param.save()
        #Т вход
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "caf11a15-27ff-4c0d-9b18-d55028e4840b"))
        add_param.save()
        #Т выход    
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "1d246b01-7fd8-441e-af1b-9851acf5f104"))
        add_param.save()

    elif instance.guid_types_meters.name == 'СЕ301':
        #Добавляем параметры для счётчика Энергомера СЕ301.    
        
        #-------------Суточные
        # "Показание". T0 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "dc142b91-20a3-4048-bbfb-571bd969fd66"))
        add_param.save()
        # "Показание". T1 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "63594901-393e-4fd9-bb1c-1da237d1264d"))
        add_param.save() 
        # "Показание". T2 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "ff7837a5-5552-4fc0-a66e-a554f92c1f02"))
        add_param.save() 
        # "Показание". T3 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "15e387b9-45ef-44fc-8110-a8d823af9140"))
        add_param.save()
        
        #-------------Мощность        
        #А+ Профиль
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"9a02431c6-7daf-4f0e-b35d-b19916c5a940")) # A+ 30-мин. срез мощности
        add_param.save()        
#        #R+ Профиль
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"610a6bf6-d00a-4fd4-a41e-0a0d9ffcba2f")) # R+ 30-мин. срез мощности
        add_param.save()

    else:
        pass
        #print u'Тип счётчика не определен'
        
#signals.post_save.disconnect(add_taken_param, sender=Meters)
signals.post_save.disconnect(add_link_meter, sender=Meters) 
signals.post_save.disconnect(add_link_taken_params, sender=TakenParams)  
        
if (isService):
    #print 'signals ON'
    signals.post_save.connect(add_link_taken_params, sender=TakenParams)
    signals.post_save.connect(add_link_meter, sender=Meters)
    #signals.post_save.connect(add_taken_param, sender=Meters)
else:
    signals.post_save.disconnect(add_link_meter, sender=Meters)
    #signals.post_save.disconnect(add_taken_param, sender=Meters)
    signals.post_save.disconnect(add_link_taken_params, sender=TakenParams)



def add_taken_param_no_signals(instance, isR, isHalfs): # Добавляем считываемые параметры при создании счётчика
    #print(instance.guid_types_meters.name, isR, isHalfs)
    if instance.guid_types_meters.name == 'Меркурий 230':
        #Добавляем параметры для Меркурия 230
        #print(instance, isR, isHalfs)
    # T0 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "bdcd1268-37f3-4579-a9d9-5becb2ba8aa3")) # A+ T0 месячные
        #print(TakenParams.objects.aggregate(Max('id'))['id__max']+1, add_param)
        add_param.save()
        #print('1111')
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "99cd6002-f81c-4ad6-9cb0-53a92a498519")) # A+ T0 суточные
        add_param.save()
        #print('22222')
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e8c20ce7-bdb6-4ea6-8401-cee28049a7d7")) # A+ T0 текущие
        #add_param.save()
    # T0 R+
        if isR:
            add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "2ebc02e6-65c6-40ab-b717-0d98d66b5701")) # R+ T0 месячные
            add_param.save()
            add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "345a24a4-95b7-4f67-b004-716706ed2560")) # R+ T0 суточные
            add_param.save()
            # add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4c93dd55-1ec2-48c7-9865-9ceab580b7b3")) # R+ T0 текущие
            # add_param.save()
        
    # T1 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "17789c36-4593-4ff2-94eb-1d0cebdb5366")) # A+ T1 месячные
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "d262c71a-6da4-4ec0-a9c3-b9ea659c246d")) # A+ T1 суточные
        add_param.save()
    # T2 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c31297be-220b-4971-8642-6b614aa7ecee")) # A+ T2 месячные
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "37011b85-c8af-4f6c-857d-4b93a95d31e1")) # A+ T2 суточные
        add_param.save()
    # T3 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "79741ba9-e8b8-4352-862e-17a9c4d928ce")) # A+ T3 месячные
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c3bb9033-ffcb-4a28-91e2-6b45924b8858")) # A+ T3 суточные
        add_param.save()
    
    #Код ошибки
        # add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"7175f6c7-b816-40f6-86f4-e08a309c08f6")) # код ошибки
        # add_param.save()

    # Ток
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"aee312b0-adb1-4be9-9879-b3a3598f9b29")) # Ia текущее
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"7ed0d364-e790-4325-a927-9ef86a685f00")) # Ib текущее
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"474b0809-482a-4851-9a96-4587f8c59152")) # Ic текущее
        #add_param.save()
    # Напряжение
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c06f7315-abc6-4889-97ad-201a936c8f2c")) # Ua текущее
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"632f76fb-4dd9-4e7d-86a0-a57a27fc648a")) # Ub текущее
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1a3ca6ca-8866-4aad-8712-d9df003fe692")) # Uc текущее
        #add_param.save()
    # Мощность
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"3077b3ac-fde2-4435-9e6f-17464310c090")) # P Активная мощность
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e7617c95-7e42-4cfa-9acd-5bc119261c6d")) # Q Реактивная мощность
        #add_param.save()
    #Получасовки
        if isHalfs:
            add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "6af9ddce-437a-4e07-bd70-6cf9dcc10b31")) # A+ 30-мин. срез мощности
            add_param.save()
            add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "66e997c0-8128-40a7-ae65-7e8993fbea61")) # R+ 30-мин. срез мощности
            add_param.save()
    
    elif instance.guid_types_meters.name == 'Пульсар 16M':
    # Суточные
      # Канал 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "fc4a9568-4674-4a80-b497-e4f34399acd5"))
        add_param.save()
      # Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9e6e308f-abec-4b47-9b99-9cb590c55d0c"))
        add_param.save()
      # Канал 3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "e6815dd5-fbbc-480f-8b95-025d7f9a0403"))
        add_param.save()
      # Канал 4
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "612d2f20-d454-4e14-910b-1fd89bbb31dd"))
        add_param.save()
      # Канал 5
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "d82c7576-8e5e-4e93-ae10-58459b31e4a0"))
        add_param.save()
      # Канал 6
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "6ccc7efb-d9fe-4285-b343-8ed22d2d3625"))
        add_param.save()
      # Канал 7
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "72567365-9a40-4f97-ab25-0911585035bf"))
        add_param.save()
      # Канал 8
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9203f5ed-d5da-4462-91d1-5aea42e99124"))
        add_param.save()
      # Канал 9
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "e4068568-d8c4-42ab-9957-7292753e2891"))
        add_param.save()
      # Канал 10
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9b5ab67b-40aa-4536-8b7c-340a773ab31b"))
        add_param.save()
      # Канал 11
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "4fd440c4-9ec5-4ab9-a073-6c4d3a174777"))
        add_param.save()
      # Канал 12
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "00b7f7c5-37f3-494a-8ceb-5a62f9ebf4e3"))
        add_param.save()
      # Канал 13
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "169a79e0-da6f-4091-9fc7-ab81adc0d7e0"))
        add_param.save()
      # Канал 14
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "17e9c8fe-0d69-4466-b64e-185452c61555"))
        add_param.save()
      # Канал 15
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "25de493d-c680-4ca6-ac02-b778022ee151"))
        add_param.save()
      # Канал 16
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "908e88f0-f9a0-421d-bbe7-9bafdf5d2565"))
        add_param.save() 
  
    # Текущие
      # Канал 1
       # add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e3f1325e-3018-40ba-b94a-ab6d6ac093e9"))
       # add_param.save()
      # Канал 2
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5a6b0338-c15d-4224-a04f-a10fc73c5fc7"))
        #add_param.save()
      # Канал 3
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"48a42afe-d9ac-4180-a733-6dd5f9d9ca80"))
        #add_param.save()
      # Канал 4
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"01a5419c-c701-4185-95b6-457b8c9ca2d0"))
        #add_param.save()
      # Канал 5
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"85c4295e-bc6a-46ec-9866-0bf9f77c6904"))
        #add_param.save()
      # Канал 6
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"68270d0a-5043-4ea2-9b61-4adaa298abad"))
        #add_param.save()
      # Канал 7
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"cd489c4b-6e74-4c65-bfee-c0fa78a853bf"))
        #add_param.save()
      # Канал 8
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"f29062a4-ab60-4117-8f85-0cdec634c797"))
        #add_param.save()
      # Канал 9
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e8521cd7-2f38-4619-935d-8fe86234dbe7"))
        #add_param.save()
      # Канал 10
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1349b747-41ca-4ba8-a690-69c649129f44"))
        #add_param.save()
      # Канал 11
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"99ab1a30-fde8-4b81-9f9e-2f731516ce1b"))
        #add_param.save()
      # Канал 12
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c7f6a397-833d-4020-9d2b-38c19bec272c"))
        #add_param.save()
      # Канал 13
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4413bffb-1832-4900-9351-5ac3666dd8b0"))
        #add_param.save()
      # Канал 14
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6280490b-123d-4e27-bef9-19fd7dc2cf54"))
        #add_param.save()
      # Канал 15
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"93891c5a-1c8f-4906-b7f0-961dc8ad3c9f"))
        #add_param.save()
      # Канал 16
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"22dd3a17-a828-44e0-80d9-db075ba120ae"))
        #add_param.save()

    elif instance.guid_types_meters.name == 'Пульсар 10M':
        #Добавляем параметры для Пульсар10
    # Суточные
      # Канал 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "325ec164-9428-4a57-867c-33d4eaf8cc2a"))
        add_param.save()
      # Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "99a99024-65b4-44dd-99fc-6a5cf1d4aaee"))
        add_param.save()
      # Канал 3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "f897f0ca-4e35-4f0d-b345-3379668aa36f"))
        add_param.save()
      # Канал 4
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "034374bd-2dfb-4568-aa16-84255df33c88"))
        add_param.save()
      # Канал 5
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b6bdfae8-4f27-4056-af79-d746b44038ee"))
        add_param.save()
      # Канал 6
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "2c2f7176-8b77-44f4-9678-4773e95e67ce"))
        add_param.save()
      # Канал 7
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "91bb7c43-f802-4ebd-a8fe-75f833acedeb"))
        add_param.save()
      # Канал 8
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "cf24b669-1c5b-4db7-936a-5f9d5c8be928"))
        add_param.save()
      # Канал 9
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "96035c7c-ee7c-41f6-9723-8a75dd9ed573"))
        add_param.save()
      # Канал 10
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "253475ea-614d-4aad-93a8-e81e4c9028e9"))
        add_param.save()   

    # Текущие
      # Канал 1
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"32dad392-ca1e-4110-8f2c-a86b02e26fb3"))
        #add_param.save()
      # Канал 2
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"3e13694b-7cb5-4417-a091-af8a7db34dc7"))
        #add_param.save()
      # Канал 3
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1023b35b-3cbf-4519-aac3-3bf1ebae07c1"))
        #add_param.save()
      # Канал 4
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"eea27ade-44cd-4e66-8298-00a4a6ad915a"))
        #add_param.save()
      # Канал 5
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"25e09d4d-3a48-4381-ad5d-b783c03c4d35"))
        #add_param.save()
      # Канал 6
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"069898ea-9d74-4571-b719-e8e6f1513c12"))
        #add_param.save()
      # Канал 7
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"084aa5f4-75d5-41f6-b0d6-9f2403eacd2c"))
        #add_param.save()
      # Канал 8
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"786ed8b8-aed1-478c-ae75-99caf1358cf0"))
        #add_param.save()
      # Канал 9
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6fc4c39c-9a43-4cb7-a066-c40fd2ca47e5"))
        #add_param.save()
      # Канал 10
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"8b2aa40a-cd91-4e22-b9d1-596e49e5f839"))
        #add_param.save()  

    elif instance.guid_types_meters.name == 'Пульсар 2M':
        #Добавляем параметры для Пульсар10

    # Суточные
      # Канал 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "0239dffb-de88-45e5-b6f6-18bf39f92525"))
        add_param.save()
      # Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "a1cb319d-ac09-466d-894b-91d90aba4239"))
        add_param.save()   
    
    # Текущие
    #   # Канал 1
    #     add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"fcc28118-66c0-4cdf-aeba-5da1171aae48"))
    #     add_param.save()
    #   # Канал 2
    #     add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1faeb517-bd1f-4ba0-96a5-67f00764822f"))
    #     add_param.save()

    elif instance.guid_types_meters.name == 'ПСЧ-3ТА.04':
        #Добавляем параметры для ПСЧ-3ТА.04
        pass
    elif instance.guid_types_meters.name == 'ТЭМ-104':
        #Добавляем параметры для ТЭМ-104
        pass
    elif instance.guid_types_meters.name == 'СЭТ-4ТМ.03М':
        #Добавляем параметры для СЭТ-4ТМ.03М
       
    # T0 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "e7624c25-9852-4ffd-8777-b2bfd16c29a8")) # A+ T0 месячные
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "aa83b499-6a9e-40e1-b68b-dc84fec8490b")) # A+ T0 суточные
        add_param.save()

    # T0 R+
        if isR:
            add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "3d365f91-8bd3-476e-b07e-3f79134f6853")) # R+ T0 месячные
            add_param.save()
            add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "087b785e-5d59-4956-9cd3-57706f9557e6")) # R+ T0 суточные
            add_param.save()
           
    #Получасовки
        if isHalfs:
            add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "4f505e17-7d71-4cf8-9880-c6ce33612e6e")) # A+ 30-мин. срез мощности
            add_param.save()
            add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "55abd40d-fb3c-4100-88f2-46d79be7733a")) # R+ 30-мин. срез мощности
            add_param.save()

    elif instance.guid_types_meters.name == 'Меркурий 200':
        #Добавляем параметры для Меркурий 200

    # Значения суточные (текущие)
    #Не поддерживается прибором, но текущие переделаны на суточные
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9cbc001d-a262-481f-a1aa-47d02bf18af1")) #T0
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b65d4227-69a5-487b-9999-5539ca3fc004")) #T1
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5e312de9-34cd-4ba7-a744-c9b94a77d98b")) #T2
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "4260ea05-78f8-4c5c-9172-fa161fa96068")) #T3
        add_param.save()
    # Значения на начало месяца
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "86cd925b-48c2-40b8-b211-f116e0e6dbea")) #T0
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "62a3796a-eaae-445d-9166-2ad517186b78")) #T1
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5f6e1e3d-4128-4cfe-94cf-57ac84a7694a")) #T2
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "0c28c135-58f2-4dff-a222-9f3d9f3c742b")) #T3
        add_param.save()

    elif instance.guid_types_meters.name == 'Эльф 1.08':
        #Добавляем параметры для счётчика тепла Elf 108
    
        #-------------Текущие
        # "Энергия"
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "f2bbf267-456e-477a-95d2-abb94c78ba43"))
        add_param.save()
        # "Объем"       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "dad6e2eb-e978-46f4-b7ec-442834b04e7a"))
        add_param.save()
        # "ElfTon"  Время работы прбора     
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "d3c9563d-51ed-4ca7-922f-ac3731065ead"))
        add_param.save()
        # "ElfErr"  Время работы прибора с ошибкой
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "dade3324-b9b0-41c8-bc76-70f617573e43"))
        add_param.save()
        # "Ti"      Температура входа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "acca627e-f21a-4f8b-be7e-038f534b5d11"))
        add_param.save()
        # "To"      Температура выхода
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "01487323-a28f-419e-9589-2563d785ab2a"))
        add_param.save()
        # "Канал 1"      Импульсный вход 1 текущий
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6e7f0d37-df5c-4850-991e-b5d7cb793924"))
        #add_param.save()
        #"Канал 1"      Импульсный вход 1 суточный
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9af27a62-d6c8-4b67-bd36-da7103e0b1f1"))
        add_param.save()
        #"Канал 2"      Импульсный вход 2 суточный
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "86acc33d-7bea-4977-a5b5-c5858ce9a09d"))
        add_param.save()
        # "Канал 2"      Импульсный вход 2 текущий
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"de7bfdfd-c17f-4a7c-942d-b28e85db33cb"))
        #add_param.save()
        #-------------Архивные
        # "Энергия"
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "ae439e1f-5c4b-494c-8a53-a61b85c804a0"))
        add_param.save()
        # "Объем"
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b02153a4-00c0-4800-a55a-c7f9dfbb14e7"))
        add_param.save()
        # "ElfTon" Время работы прибора
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "aa611d48-f1fe-462a-8b0a-0a7596792b69"))
        add_param.save()
        # "ElfErr" Время работы прибора с ошибкой
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "af047098-bd45-4579-a60c-b75bed376bbe"))
        add_param.save()
        
        
    elif instance.guid_types_meters.name == 'СПГ762-1':
        #Добавляем параметры для счётчика газа СПГ762 Подсистема 1
    
        #-------------Часовые
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "edfc6dc0-1628-4a7e-bd04-71107882039a"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "1e27e72c-79d6-4c68-bc04-1be84d061622"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "0fa5d9ef-4c6c-4f78-bc64-9d9b34002344"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "433d7025-15f3-4ab0-9d73-39bd0e425566"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "446a6bb6-c17f-4478-b1d8-252c7eb454d3"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "003c2fb2-0092-4d7d-a513-3dcc50a255da"))
        add_param.save()
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9c99da7b-1a73-48b0-a3f5-54438a3ea824"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "832374da-5834-4fe0-abe3-07d48d447af2"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "4920e70a-452e-4ecf-918e-14ca288c7a1f"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "825dfd73-82fa-4f1a-9635-77bdcb244997"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "daa4a90e-7993-493c-9ac0-c03241b2ab2c"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "72b5f4ba-d179-419c-8795-e2f86f5ee2ff"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "41774f43-9c9c-4867-bf21-e3d1df4fd2f8"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "f610969e-cb2d-436f-9357-e63da72d162e"))
        add_param.save()
        
        #-------------Суточные
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "ac9fe54a-6f51-4ee9-a849-448f0f10a4b6"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "2eefa93a-be60-4f23-9b09-8d1c6bad0a15"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "78c30686-a5d0-436f-9c56-57b076769774"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "d7900bac-b85a-4b83-ad67-b822b470a698"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "068f4de0-f041-4608-b09a-81dbc8f319ff"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "dad040cd-1047-4060-a7ac-e9a6e7f30fb4"))
        add_param.save()        
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "3be53eb7-c931-4728-a442-44d14c9da44f"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "1dff6784-dd8d-4898-ae8b-f1b60fbdc1af"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "af26b128-b634-47a4-96ba-42de6f039fdb"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5bb8af68-588e-470c-9b64-373482f71468"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c32e7e01-106a-49c6-b8b0-6490448548ad"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "8fc09300-4b19-4478-aa44-d2fb1cf792d5"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "53693586-e5b5-4204-922a-a0b0153298ea"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "bf06727c-9cab-4efe-a0f2-6242bb320372"))
        add_param.save()
        
        # Масса газа нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5399a54f-0d2c-47e8-8ffb-882f5dddc239"))
        add_param.save()
        # Объем газа при ст.условиях нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c14eaa34-3264-4fe4-98ab-8da6618fc431"))
        add_param.save()
        # Объем газа при раб.условиях нарастающим итогом              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "4ed5ba8d-9ead-4a60-96e4-726f38432d9a"))
        add_param.save()
        
    elif instance.guid_types_meters.name == 'СПГ762-2':
        #Добавляем параметры для счётчика газа СПГ762 Подсистема 1
    
        #-------------Часовые
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9a17ea4f-a1f8-4fb9-a21e-62f43978535a"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "49728e75-1d62-4b9e-8633-762cb7117b52"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c44af390-ed2a-49d8-9c67-25f543db9935"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "24bd767d-f667-444e-acfa-a935fb8f4699"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "1e675e58-2cc9-4103-8410-2d37704a2bcf"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "d14fa3fa-5bd5-4dd9-b740-61049d38e694"))
        add_param.save()
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "6de742eb-a391-473a-bb4b-ab780a4642b8"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "04b43217-af06-4f09-8d95-0e2d3dbd0905"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b62c6838-7578-41c9-a94c-d06788cc2d41"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "aad9d0b4-1c12-4165-a1b1-4fac9de00c38"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "205a8c8c-de26-44e5-ab72-efb7fe72040c"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c84c6130-ede7-487e-a414-b384964eb81e"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5907cc5f-1386-4fbc-9e5c-7d3f77dba6d6"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c5638410-44b6-41d7-b501-6e5c0a002f48"))
        add_param.save()
        
        #-------------Суточные
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c1643ab8-1707-4b73-9610-0226b1fb6860"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "dbacff0c-2b3d-40c5-aa03-8d51a64919dd"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "279bdfd5-7b22-4d7e-900c-21e4077506dd"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b74e6743-3996-4af7-8024-da3912d14b45"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c7220fc7-5c01-4bcc-ac2a-7c851276af4d"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "381ccc3f-a9d9-4dcf-a9aa-2e5bd0e4efc8"))
        add_param.save()        
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "7f0f0e09-3bd0-4595-84dd-754f4c21bc5e"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "ee687ef8-36de-4de8-9a05-4ac841c9c144"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "8ee128d0-5c21-4faa-a2fe-7432ff9be684"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c61c87eb-5a4f-4095-ac50-4324e7899340"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "bae3e866-f057-4be5-99a0-7474f6c7cbc1"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "8ce68aef-85fa-4ca6-8f9f-dfa1f9e71cdd"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "ea094191-7ce5-4c42-ad5e-e886d02e73e0"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "e4d751b4-ef6c-45ca-b31a-f107f47a97aa"))
        add_param.save()
        
        # Масса газа нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "add68490-64f4-47a2-a801-1fafa48c09a2"))
        add_param.save()
        # Объем газа при ст.условиях нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "525f9439-ce13-43be-a4f2-67f590f4842b"))
        add_param.save()
        # Объем газа при раб.условиях нарастающим итогом              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "a760c696-fda0-47d4-8fb5-899d742957f1"))
        add_param.save()
        
    elif instance.guid_types_meters.name == 'СПГ762-3':
        #Добавляем параметры для счётчика газа СПГ762 Подсистема 1
    
        #-------------Часовые
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "187b787c-1693-4c90-b6df-d868effef692"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "49c86cc3-57c2-4bdf-b4e3-b07f64673d37"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "1e92a9a8-1cd9-4252-b9c7-b33357bafce7"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "06cd5dad-7ea9-438e-abbf-043e8918eb3e"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "8064aad9-778a-4902-b0c0-75b23289469a"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b747eeb5-8c69-443e-b74b-2bb89af64206"))
        add_param.save()
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "38c0b41b-0883-4990-bb0c-8b532caed34c"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "7bf3d68d-4344-49f3-8169-370f6142351a"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "fa188255-c1cc-4c2d-844c-3b40a3a7559e"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "06c97066-1e35-4bb1-a96f-fe3c0056cf39"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "e530ab26-92e8-4edc-8e6d-5cd6184bfbe7"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5d315c1e-b237-46d6-9273-be4e597ad1c2"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "54b86222-3d1c-440a-b02f-bedbef0e9e28"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "acb12185-dd88-4449-8f03-76b6fd148958"))
        add_param.save()
        
        #-------------Суточные
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "d1f0258f-42ba-4e4f-a66c-74aed4d512ce"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "7ec1a0fb-88f0-497a-8917-01c0b731b88a"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "3465cb7c-57ea-4ad8-afde-74fb2814ddeb"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "a3014fec-73df-4fd4-a68c-9c3ff737d140"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "986f24b0-df76-4c36-9b7f-fbbe05a10c94"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "2a5ec8f1-b6fe-4eff-b91f-42a7712dd663"))
        add_param.save()        
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "cbfc543d-3a19-46c3-8075-ff59492d2620"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "8a24d34e-aee6-4865-bf21-56d9c07dcd1e"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "85de87d5-8e0a-4088-8248-8a64367db47e"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b2bd2c95-ee85-4156-9ef6-7fc25d29a244"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "8674a61c-af88-46c5-b553-fecc9a7d0837"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5880bd0f-699d-407c-a3f0-6cea0ebde423"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9057618b-445c-4581-86a6-4715469db938"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "4991e5c0-0827-4467-b9d2-7613d1b6dd09"))
        add_param.save()
        
        # Масса газа нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "dcd5ed6a-7bd0-41ba-8850-5b88a9831c04"))
        add_param.save()
        # Объем газа при ст.условиях нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "cbeaa4fa-d1fb-4bf5-9688-7084b57fbfe4"))
        add_param.save()
        # Объем газа при раб.условиях нарастающим итогом              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "d16b31ea-87d1-409a-bbf8-4a743b678dbb"))
        add_param.save()

    elif instance.guid_types_meters.name == 'Sayany':
        #Добавляем параметры для счётчика Sayany
    
        #-------------Суточные
        # "Q" Тепловая энергия. Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "e7f2ffba-9a40-43e1-80f3-ddd22596cdb8"))
        add_param.save()    
        # "Q" Тепловая энергия. Канал2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "6f9cd79e-ca34-447e-8ad1-d54531389fe1"))
        add_param.save() 
        # "M" Расход воды. Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b05de8e2-6176-4fc0-bc44-79ceb4229c80"))
        add_param.save() 
        # "M" ТРасход воды. Канал2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5f256e9b-1cb3-4f27-a53a-d08b446dda58"))
        add_param.save() 
        # "T" Температура. Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "75474616-f3db-4903-91d5-1f22f6593394"))
        add_param.save() 
        # "T" Температура. Канал2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "f3210c5b-afde-4c9a-b201-9c7c403c4cf2"))
        add_param.save() 
        # "T" Температура. Канал3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b12762a0-0a06-49a4-b842-8ad3378f4602"))
        add_param.save() 
        # "T" Температура. Канал4
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "472ba2fd-cc06-4147-a1e7-c1bb66096536"))
        add_param.save() 
        
    elif instance.guid_types_meters.name == 'Tekon_hvs':
        #Добавляем параметры для счётчика Tekon. Читаем один тэк с opcretranslator
    
        #-------------Суточные
        # "Показание". Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "d76796ea-ea63-4317-982f-ffbcde2074dc"))
        add_param.save()  
        
    elif instance.guid_types_meters.name == 'Tekon_gvs':
        #Добавляем параметры для счётчика Tekon. Читаем один тэк с opcretranslator
    
        #-------------Суточные
        # "Показание". Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "e71a7206-7a30-45d9-981d-0b7592b96337"))
        add_param.save() 
        
    elif instance.guid_types_meters.name == 'Tekon_heat':
        #Добавляем параметры для счётчика Tekon. Читаем один тэк с opcretranslator
    
        #-------------Суточные
        # "Показание". Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "1dca7dab-371a-4429-afa1-8b4877b38b5b"))
        add_param.save()
        
    elif instance.guid_types_meters.name == 'Меркурий 230-УМ':
        #Добавляем параметры для счётчика Меркурий на УСПД УМ-RTU.    
        
        #-------------Суточные
        # "Показание". T0 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b6e89205-3814-463d-86d1-f52cec7d8962"))
        add_param.save()
        # "Показание". T1 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "7f3c42e6-4000-4373-a0e6-37e66ce819a9"))
        add_param.save() 
        # "Показание". T2 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c6512649-56ea-4214-aa33-84516bfe8dc1"))
        add_param.save() 
        # "Показание". T3 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "4e20bda9-6e75-4b0f-a99a-0e4c1cd07d3b"))
        add_param.save()
        
        #-------------Мощность        
        #А+ Профиль
#        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"922ad57c-8f5e-4f00-a78d-e3ba89ef859f")) # A+ 30-мин. срез мощности
#        add_param.save()        
#        #R+ Профиль
#        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"61101fa3-a96a-4934-9482-e32036c12829")) # R+ 30-мин. срез мощности
#        add_param.save()
        
    elif instance.guid_types_meters.name == 'Пульсар Теплосчётчик':
        #Добавляем параметры для Теплосчётчика Пульсар.
        #------------Суточные
        # "Показание Энергии" Q, Гкал
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "24ae9f51-40a4-4758-a826-a5f8286e1a2e"))
        add_param.save()
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "a3da78fb-b07b-4d53-a980-54b51e26819a"))
        add_param.save()
        # "Показание Температура подачи" Ti, C0
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "de66ecd2-b243-467c-8d1a-cfcb42377300"))
        add_param.save()
        # "Показание Температура выхода" To, C0
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "d3433b80-cb8c-4038-a682-947e6d05955e"))
        add_param.save()
        # "Код ошибки" 
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "dce68a94-2fb1-4856-acd6-2a1b13d5ec99"))
        add_param.save()
        #Канал 1. Импульсный
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "af842772-2e25-45d0-9c40-3b21f30fe808"))
        add_param.save()
        #Канал 2. Импульсный    
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "318bd700-815c-46fe-aa7c-1e5265bab53e"))
        add_param.save()

    elif instance.guid_types_meters.name == 'Пульсар Холодосчётчик':
        #Добавляем параметры для Холодосчётчика Пульсар.
        #------------Суточные
        # "Показание Энергии" Q, Гкал
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "ba344ec3-d390-48e0-811c-991b25d9734d"))
        add_param.save()
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "8fd07daa-4ad0-4124-acbe-aca6da2301d0"))
        add_param.save()
        # "Показание Температура подачи" Ti, C0
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "6a50c871-8101-4a44-82c6-6214bd7a5ddd"))
        add_param.save()
        # "Показание Температура выхода" To, C0
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "244de93e-4dcd-41e2-bcc4-db1a113ead0a"))
        add_param.save()


    elif instance.guid_types_meters.name == 'Пульсар 3Ф4Т':
        #Добавляем параметры для электросчётчика Пульсар 3Ф4Т

    # Значения суточные
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "2161a1c4-6d66-4e9b-8dbb-9f7ab1bab67d")) #T0
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5f11bcf5-b058-4062-b259-6b96c80367b7")) #T1
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "e3b0a6c1-20d7-4753-b165-e29c78a1d9e9")) #T2
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "3aff9cc3-110b-4467-add7-c19d319a01cc")) #T3
        add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "7f273543-3985-43b0-a027-21311961ecb7")) #T4
        #add_param.save()
        
    elif instance.guid_types_meters.name == 'Пульсар ХВС':
        #Добавляем параметры для водосчётчика Пульсар ХВС.
        #------------Суточные
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "209894a8-8d19-4e4d-bad8-1767eec4fedf"))
        add_param.save()

    elif instance.guid_types_meters.name == 'Пульсар ГВС':
        #Добавляем параметры для водосчётчика Пульсар ГВС.
        #------------Суточные
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5fc2ff3b-999e-4154-ba49-84d3971369b0"))
        add_param.save()
        
    elif instance.guid_types_meters.name == 'Пульс СТК ХВС':
        #Добавляем параметры для водосчётчика Пульс СТК ХВС.
        #------------Суточные
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "31bc817a-2ccd-4021-a8a1-7d63d97dae2c"))
        add_param.save()
        #------------Месячные
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c5b9362b-5c59-47bb-bc61-b3b556b24dc3"))
        add_param.save()

    elif instance.guid_types_meters.name == 'Пульс СТК ГВС':
        #Добавляем параметры для водосчётчика Пульс СТК ГВС.
        #------------Суточные
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "1ae0a516-2975-4a6e-95e3-23412e0f2e67"))
        add_param.save()
        #------------Месячные
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "49f1197f-c6ae-4081-afbc-587ac614a3c3"))
        add_param.save()

    elif instance.guid_types_meters.name == 'Пульс СТК Теплосчётчик':
        #Добавляем параметры для водосчётчика Пульс СТК Теплосчётчик.
        #------------Суточные
        # "Показание Энергии" Q, Гкал
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "14275dc3-eebb-4b95-aaf1-066ee4edab4e"))
        add_param.save()
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9a97d8b8-992f-4e43-a6e0-9f1dc89d2dec"))
        add_param.save()
        # "Показание Температура подачи" Ti, C0
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "27b8f3a1-b10d-4327-9505-31f730a3b62b"))
        add_param.save()
        # "Показание Температура выхода" To, C0
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "1944aeec-58a6-48d4-bbcb-24d0ce3a3e1a"))
        add_param.save()

    elif instance.guid_types_meters.name == 'Карат 307':
        #print u'Добавляем параметры для счётчика Карат 307'
        #Суточные 
        #Объём     
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "3024fd72-d1e8-4476-a876-4bc09553dde9"))
        add_param.save()
        #Тепло
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "46a63ef5-5761-4e16-a854-1979ddc9668f"))
        add_param.save()
        #Tout
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "6dd6ea63-20dc-46d0-b56e-6890a2b83f48"))
        add_param.save()
        #Tin
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "8a5f5921-5b70-410d-83de-8403ec2a4d87"))
        add_param.save()
        #Ton наработка в минутах
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9c86e183-dd53-4c7f-b728-ffe75a55c633"))
        add_param.save()
        #Terr время работы в ошибке
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "abd41546-02f6-4e2c-8bd2-a60ab80ffe66"))
        add_param.save()
        #Масса 
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "eb617f04-14a3-403c-90e8-286412872232"))
        add_param.save()
    elif instance.guid_types_meters.name == 'Danfoss SonoSelect':
        #print u'Добавляем параметры для счётчика Danfoss SonoSelect'
        #Суточные 
        #Объём     
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "83ba885f-1881-45db-9d63-52195e67cf64"))
        add_param.save()
        #Энергия  
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "510ba9e6-c18b-4982-9763-2ad86c8a8245"))
        add_param.save()
        #Т вход
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "a8940b63-2002-4a28-b671-a455419c1229"))
        add_param.save()
        #Т выход    
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "db1ed365-e85e-4547-aef6-89fed020898f"))
        add_param.save()
        #Канал 1. Импульсный
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "0166c378-5696-489e-92fb-a1d360fc2921"))
        add_param.save()
        #Канал 2. Импульсный    
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b18bc638-eb07-429e-9c59-d859604be48f"))
        add_param.save()

        # #Месячные
        # #Объём     
        # add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "3c763ebc-6ad4-4193-967f-f352bfae92c5"))
        # add_param.save()
        # #Энергия  
        # add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "7a5d6dd3-3a34-40e4-90af-0c00252b978d"))
        # add_param.save()
        # #Т вход
        # add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "caf11a15-27ff-4c0d-9b18-d55028e4840b"))
        # add_param.save()
        # #Т выход    
        # add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "f76371c9-5ecd-44b2-835e-5fc4cdce7141"))
        # add_param.save()

        # #Текущие 
        # #Объём     
        # add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "097122a4-b7d0-4700-add2-bc99a58347d0"))
        # add_param.save()
        # #Энергия  
        # add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "af254ed7-d4ab-4293-8aaf-8bb13c81efb7"))
        # add_param.save()
        # #Т вход
        # add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "caf11a15-27ff-4c0d-9b18-d55028e4840b"))
        # add_param.save()
        # #Т выход    
        # add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "1d246b01-7fd8-441e-af1b-9851acf5f104"))
        # add_param.save()

    elif instance.guid_types_meters.name == 'Энергомера СЕ301':
        #Добавляем параметры для счётчика Энергомера СЕ301.    
        #print('CE301 add taken params')
        #-------------Суточные
        # "Показание". T0 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "dc142b91-20a3-4048-bbfb-571bd969fd66"))
        add_param.save()
        # "Показание". T1 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "63594901-393e-4fd9-bb1c-1da237d1264d"))
        add_param.save() 
        # "Показание". T2 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "ff7837a5-5552-4fc0-a66e-a554f92c1f02"))
        add_param.save() 
        # "Показание". T3 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "15e387b9-45ef-44fc-8110-a8d823af9140"))
        add_param.save()
        
        #-------------Мощность        
        #А+ Профиль
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"a02431c6-7daf-4f0e-b35d-b19916c5a940")) # A+ 30-мин. срез мощности
        add_param.save()        
#        #R+ Профиль
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"610a6bf6-d00a-4fd4-a41e-0a0d9ffcba2f")) # R+ 30-мин. срез мощности
        add_param.save()

    elif instance.guid_types_meters.name == 'Sanext':
        #Добавляем параметры для Теплосчётчика Sanext.
        #------------Суточные
        # "Показание Энергии" Q, Гкал
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "dd95cc37-023c-4d15-861d-8a7363f36d9c"))
        add_param.save()
        # "Показание Расход теплоносителя" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "f83d191d-1252-4257-ab85-5d0bba9a04c2"))
        add_param.save()
        # "Показание Температура подачи" Ti, C0
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "ae423b31-9b3d-4ab3-a6cf-8b745faf48f0"))
        add_param.save()
        # "Показание Температура выхода" To, C0
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "e2baef4b-3cfe-4f19-aec7-0d77f5c8a822"))
        add_param.save()

    elif instance.guid_types_meters.name == 'Valtec 16M':
        print('add_taken_params_valtec_16m')
    # Суточные
      # Канал 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "a11b0730-3975-40c1-bd2f-ed0c1a3132dc"))
        add_param.save()
      # Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9d2832c8-f116-434d-ab48-0f28cbfc03ad"))
        add_param.save()
      # Канал 3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "4fe447ea-de07-4d97-ad4f-20ded5503ddb"))
        add_param.save()
      # Канал 4
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "f859f4ed-166e-4b44-9df3-f3a60d778c35"))
        add_param.save()
      # Канал 5
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c1b97126-d5f8-4d8c-acc4-08eaa017d5fa"))
        add_param.save()
      # Канал 6
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "ca29cbf2-ab61-4cb0-b9fd-99121cb45b1f"))
        add_param.save()
      # Канал 7
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "28c0bd41-1281-4272-b33e-234669294644"))
        add_param.save()
      # Канал 8
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "616608c0-2343-4d22-b065-d8d5f9769731"))
        add_param.save()
      # Канал 9
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "22ff1bd3-506b-40db-9137-2f01a923698d"))
        add_param.save()
      # Канал 10
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "48a5e8d8-d065-423c-a7ac-f126048687f7"))
        add_param.save()
      # Канал 11
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9b2cb997-1bb0-4230-92f7-3cd01032e5b9"))
        add_param.save()
      # Канал 12
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "fe4e2925-d986-4727-b3d5-bd3f809009e5"))
        add_param.save()
      # Канал 13
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "a49323aa-135d-4a3f-93b8-1e32762e64e7"))
        add_param.save()
      # Канал 14
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "aae6e088-7dcb-49a6-aafc-df4834ceca11"))
        add_param.save()
      # Канал 15
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "1e1d2cef-11fc-4b81-9201-ce19a85cacab"))
        add_param.save()
      # Канал 16
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "310b935b-e7f9-47d3-8c87-a8ae13a3c81d"))
        add_param.save() 
  
    # Текущие
      # Канал 1
       # add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u""))
       # add_param.save()
      # Канал 2
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u""))
        #add_param.save()
      # Канал 3
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u""))
        #add_param.save()
      # Канал 4
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u""))
        #add_param.save()
      # Канал 5
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u""))
        #add_param.save()
      # Канал 6
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u""))
        #add_param.save()
      # Канал 7
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u""))
        #add_param.save()
      # Канал 8
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u""))
        #add_param.save()
      # Канал 9
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u""))
        #add_param.save()
      # Канал 10
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u""))
        #add_param.save()
      # Канал 11
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u""))
        #add_param.save()
      # Канал 12
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u""))
        #add_param.save()
      # Канал 13
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u""))
        #add_param.save()
      # Канал 14
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u""))
        #add_param.save()
      # Канал 15
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u""))
        #add_param.save()
      # Канал 16
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u""))
        #add_param.save()

    elif instance.guid_types_meters.name == 'Нартис СПОДЭС':
        #Добавляем параметры для счётчика Нартис СПОДЭС.    
        #print('Нартис СПОДЭС add taken params')
        #-------------Суточные
        # "Показание". T0 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "1b93aeca-461c-468a-a0b9-792bd89001ab"))
        add_param.save()
        # "Показание". T1 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "634d9f9a-a7f2-479e-9904-147cfa800ccc"))
        add_param.save() 
        # "Показание". T2 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "949b3f73-a0bb-4037-8f13-cb30c815c226"))
        add_param.save() 
        # "Показание". T3 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "14ba0cf1-d4c6-40d6-a70c-9ecdd2c73050"))
        add_param.save()
        
        #-------------Мощность        
        #А+ Профиль
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e9e1c90f-7b4a-4ba0-a51e-e8b7a0cfbb41")) # A+ 30-мин. срез мощности
        add_param.save()        
#        #R+ Профиль
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4d8a4b75-695a-43e1-9a40-87c0e9f57e24")) # R+ 30-мин. срез мощности
        add_param.save()

    elif instance.guid_types_meters.name == 'Пульс Эл. ХВС':
        #Добавляем параметры для водосчётчика Пульс СТК ХВС.
        #------------Суточные
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "c5cc2d99-9e62-43ce-8b59-efc4d99956cf"))
        add_param.save()


    elif instance.guid_types_meters.name == 'Пульс Эл. ГВС':
        #Добавляем параметры для водосчётчика Пульс Эл. ГВС.
        #------------Суточные
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "f4facbc6-6b26-426b-939f-b67dbdd5631a"))
        add_param.save()

    elif instance.guid_types_meters.name == 'ЭкоНом ХВС':
        #Добавляем параметры для водосчётчика ЭкоНом ХВС.
        #------------Суточные
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "4ff80736-bcc7-47e6-9314-e505b837191d"))
        add_param.save()


    elif instance.guid_types_meters.name == 'ЭкоНом ГВС':
        #Добавляем параметры для водосчётчика ЭкоНом ГВС.
        #------------Суточные
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9f280064-b8c7-4037-8a3c-617301221427"))
        add_param.save()

    elif instance.guid_types_meters.name == 'ЭкоНом Теплосчётчик':
        #Добавляем параметры для Теплосчётчика ЭкоНом.
        #------------Суточные
        # "Показание Энергии" Q, Гкал
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "b22a28d0-9cb1-45eb-baba-5bb696ceb50d"))
        add_param.save()
        # "Показание Расход теплоносителя" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "4400f32c-9074-4838-93cf-7c7066f088f5"))
        add_param.save()
        # "Показание Температура подачи" Ti, C0
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5d9407e9-8ff7-4d8b-937e-d47edfe27e31"))
        add_param.save()
        # "Показание Температура выхода" To, C0
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "3c4bf80c-1a62-41ef-8bdd-9e9593ee0c56"))
        add_param.save()

    elif instance.guid_types_meters.name == 'Декаст Теплосчётчик':
        #Добавляем параметры для Декаст Теплосчётчик.
        #------------Суточные
        # "Показание Энергии" Q, Гкал
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "7b109795-7076-444c-ba75-0f6494e523c5"))
        add_param.save()
        # "Показание Расход теплоносителя" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "99fd51ba-2b1a-4cd3-adde-e0e744250d96"))
        add_param.save()
        # "Показание Температура подачи" Ti, C0
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "47c98a3e-0db8-4c51-8f2e-a4eed31039f9"))
        add_param.save()
        # "Показание Температура выхода" To, C0
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "9a0c606e-6986-4a36-93a4-0769a2f851e9"))
        add_param.save()

    elif instance.guid_types_meters.name == 'МЗТА':
        print('add_taken_params_mzta')
    # Суточные
      # Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "82ce9adf-0a2e-4ceb-8d95-89fc4d9df289"))
        add_param.save()
      # Канал 3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "4fd62d3d-880f-4f00-81bd-41b380c012da"))
        add_param.save()
      # Канал 4
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "ec10fc08-9e1b-416b-ab8e-b6f2233c8196"))
        add_param.save()
      # Канал 5
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "12ea6b3b-4dcf-40e3-bdf1-9edb80b9585a"))
        add_param.save()
      # Канал 6
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "6352996a-dc82-4939-8e0d-8679af9a3cb4"))
        add_param.save()
      # Канал 7
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "99c31bc4-0764-472a-91f9-774c7eacfaed"))
        add_param.save()
      # Канал 8
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "f394d963-c94f-42ac-9a15-d83dd0a97e78"))
        add_param.save()
      # Канал 9
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "8f16a338-23db-4fe0-99e1-3aad34897fda"))
        add_param.save()
      # Канал 10
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "11136ad3-a88e-40e5-8d19-cbf43032d17c"))
        add_param.save()
      # Канал 11
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "41f7302d-4fe7-4b55-8cc4-5d12269671e1"))
        add_param.save()
      # Канал 12
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "6472c355-66cd-4f99-8c17-a01ad155449e"))
        add_param.save()
      # Канал 13
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "5a4f437f-8b6d-4a59-9ec3-ad5f47cefccf"))
        add_param.save()
      # Канал 14
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "447cf2bb-bc43-4a50-8cd0-b1cc0407d973"))
        add_param.save()
      # Канал 15
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "98bcdfdd-8965-42f3-84fe-978d367be9eb"))
        add_param.save()
      # Канал 16
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "8663465e-64d1-49d9-915b-b0cd7d71a503"))
        add_param.save()
      # Канал 17
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "f7af4d4f-e542-406d-9ad2-44a5b8e9d506"))
        add_param.save() 
      # Канал 18
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "ccd0c7f4-babd-4953-99c8-b6adf0cf1b12"))
        add_param.save()
      # Канал 19
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "7633d6a2-9cba-49d4-b456-a51e137b45f1"))
        add_param.save()
      # Канал 20
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "a556773c-888a-4bb7-8ac4-e083630f0f28"))
        add_param.save()

    elif instance.guid_types_meters.name == 'Декаст ХВС':
        #Добавляем параметры для водосчётчика Декаст ХВС.
        #------------Суточные
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "2c129d66-9551-43af-b841-3beeb31974e1"))
        add_param.save()


    elif instance.guid_types_meters.name == 'Декаст ГВС':
        #Добавляем параметры для водосчётчика Декаст ГВС.
        #------------Суточные
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = "a2df1925-fa6a-4e41-9ded-01f17f8db55d"))
        add_param.save()

    else:
        pass
        #print('!!!!!!!!!!!!!!', instance.guid_types_meters.name)
        #print u'Тип счётчика не определен'
    return

#_____________________________________________________________________________________________________________
#При изменении в админке будут переименовыватсья все линки

def rename_taken_params(sender, instance, **kwargs):
    #print 'rename taken params'
    try:   
        #переименовываем taken_params
        guid_meter=instance.guid
        #print instance.values()
        new_val=instance.name
        old_val= Meters.objects.get(guid=guid_meter).name    
        common_sql.update_table_with_replace('taken_params', 'name', 'guid_meters', guid_meter, old_val, new_val)
    except Meters.DoesNotExist:
        return False

    try:
    #переименовываем link_abonents_taken_params
        for row in TakenParams.objects.filter(guid_meters=guid_meter):
            guid_taken_params= row.guid
            common_sql.update_table_with_replace('link_abonents_taken_params', 'name', 'guid_taken_params', guid_taken_params, old_val, new_val)
    except TakenParams.DoesNotExist:
        return False


def rename_link_abonents_taken_params(sender, instance, **kwargs): 
    
    try:
        guid_abon = instance.guid
        new_val = instance.name
        old_val = Abonents.objects.get(guid=guid_abon).name
        common_sql.update_table_with_replace('link_abonents_taken_params', 'name', 'guid_abonents', guid_abon, old_val, new_val)
    except Abonents.DoesNotExist:
        return False


signals.pre_save.connect(rename_link_abonents_taken_params, sender=Abonents)
signals.pre_save.connect(rename_taken_params, sender=Meters)

#____________________________________________________________________________________________
def OnOffSignals():
    if (isService):
        #print 'signals ON'
        signals.post_save.connect(add_link_taken_params, sender=TakenParams)
        signals.post_save.connect(add_link_meter, sender=Meters)
        #signals.post_save.connect(add_taken_param, sender=Meters)

        signals.pre_save.disconnect(rename_link_abonents_taken_params, sender=Abonents)
        signals.pre_save.disconnect(rename_taken_params, sender=Meters)
    else:
        #print 'signals Off'
        signals.post_save.disconnect(add_link_meter, sender=Meters)
        #signals.post_save.disconnect(add_taken_param, sender=Meters)
        signals.post_save.disconnect(add_link_taken_params, sender=TakenParams)

        signals.pre_save.connect(rename_link_abonents_taken_params, sender=Abonents)
        signals.pre_save.connect(rename_taken_params, sender=Meters)  


def add_link_abonents_taken_params(sender, instance, created, **kwargs):
    def get_taken_param_by_abonent_from_excel_cfg(input_taken_param):
        """Функция, которая читает excel файл. Составляет имя считываемого параметра типа "Пульсар 16M 33555 Пульсар 16M Канал 11". В случае совпадения должна привязать этот параметр к абоненту. Абоненты должны быть предварительно созданы."""    
        dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    
        def shrink_taken_param_name(taken_param_name):
            if taken_param_name.find('Текущий') != -1: # Ищем слово "Текущий"
                nn = taken_param_name.find('Текущий')  # Если нашли. то Записываем позицию где.        
            elif taken_param_name.find('Суточный') != -1:
                nn = taken_param_name.find('Суточный')
            else:
                pass
            return taken_param_name[:nn -1]

        for i in range(2,len(dtAll)):
            #taken_param = u'Пульсар' + u' ' + unicode(dtAll[i][3])[17:20] + u' ' + unicode(dtAll[i][3])[2:8] + u' ' + u'Пульсар' + u' ' + unicode(dtAll[i][3])[17:20] + u' ' + u'Канал' + u' ' + unicode(dtAll[i][4])
            taken_param = str(dtAll[i][6]) + ' ' + str(dtAll[i][5]) + ' '+ str(dtAll[i][6]) + ' ' + 'Канал' + ' ' + str(dtAll[i][4])
#            print taken_param
#            print shrink_taken_param_name(input_taken_param)
            if taken_param == shrink_taken_param_name(input_taken_param):
                try:
                    return str(dtAll[i][2])
                except:
                    return None
            else:
                pass
    
    writeToLog('--------')
    writeToLog(instance.name)
    writeToLog('==>', get_taken_param_by_abonent_from_excel_cfg(instance.name))
    if get_taken_param_by_abonent_from_excel_cfg(instance.name) is not None:
        writeToLog('Совпадение')
        try:
            common_sql.InsertInLinkAbonentsTakenParams(name = Abonents.objects.get(name= get_taken_param_by_abonent_from_excel_cfg(instance.name)).name + " " + instance.guid_params.guid_names_params.name + " " + instance.guid_params.guid_types_params.name ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1, guid_abonents = Abonents.objects.get(name= get_taken_param_by_abonent_from_excel_cfg(str(instance.name))) , guid_taken_params = instance )
            add_link_abonents_taken_param.save()
        except:
            pass
    else:
        pass
    

def add_link_abonents_taken_params2(sender, instance, created, **kwargs):
    print(f'Зашли в функцию add_link_abonents_taken_params2')
    writeToLog(instance.name)
    isExistTakenParam=SimpleCheckIfExist('taken_params','name',instance.name,"","","")
    if not isExistTakenParam:
        print('ERR: Param not exist!')
        return None
    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    for i in range(2,len(dtAll)):
        abon=str(dtAll[i][2])
        type_pulsar=str(dtAll[i][6])
        channel=str(dtAll[i][4])
        num_pulsar=str(dtAll[i][5])
        taken_param = type_pulsar+' '+num_pulsar+' '+type_pulsar+' Канал '+channel+' Суточный -- adress: '+channel+'  channel: 0'
        print(f'Пытаемся привязать: {taken_param}')
        if (taken_param==instance.name):
            isExistAbonent=SimpleCheckIfExist('abonents','name',abon,'','','')
            if isExistAbonent:
                #writeToLog('Совпадение')
                #"ХВС, №47622 Канал 4 Суточный"
                #guidAbon=GetSimpleTable('abonents','name',abon)[0][0]
                t = Abonents.objects.filter(name = abon)
                guidAbon = t[0].guid
                print(guidAbon)
                linkName=abon+' Канал '+channel+' Суточный'
                writeToLog(linkName)
                try:
                    common_sql.InsertInLinkAbonentsTakenParams(name = linkName,coefficient=1, coefficient_2 = 1,coefficient_3 = 1, guid_abonents = Abonents.objects.get(guid=guidAbon) , guid_taken_params = instance.guid )
                    add_link_abonents_taken_param.save()
                    writeToLog('Связь добавлена: '+abon+' -- '+taken_param)
                except:
                    writeToLog('ошибка')
                else:
                    pass
    
#    
#    
#    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
#    for i in range(2,len(dtAll)):
#            #taken_param = u'Пульсар' + u' ' + unicode(dtAll[i][3])[17:20] + u' ' + unicode(dtAll[i][3])[2:8] + u' ' + u'Пульсар' + u' ' + unicode(dtAll[i][3])[17:20] + u' ' + u'Канал' + u' ' + unicode(dtAll[i][4])
#            # "Пульсар 2M 062726 Пульсар 2M Канал 1 Суточный -- adress: 1  channel: 0"
#            # "Пульсар 10M 203677 Пульсар 10M Канал 7 Суточный -- adress: 7  channel: 0"
#        type_pulsar=unicode(dtAll[i][6])
#        channel=unicode(dtAll[i][4])
#        num_pulsar=unicode(dtAll[i][5])
#        taken_param = type_pulsar+u' '+num_pulsar+u' '+type_pulsar+u' Канал '+channel+u' Суточный -- adress: '+channel+u'  channel: 0'
#        print taken_param
#    
#    print u'--------'
#    print instance.name
#    print u'==>', get_taken_param_by_abonent_from_excel_cfg(instance.name)
#    if get_taken_param_by_abonent_from_excel_cfg(instance.name) is not None:
#        print u'Совпадение'
#        try:
#            common_sql.InsertInLinkAbonentsTakenParams(name = Abonents.objects.get(name= get_taken_param_by_abonent_from_excel_cfg(instance.name)).name + u" " + instance.guid_params.guid_names_params.name + u" " + instance.guid_params.guid_types_params.name ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1, guid_abonents = Abonents.objects.get(name= get_taken_param_by_abonent_from_excel_cfg(unicode(instance.name))) , guid_taken_params = instance )
#            add_link_abonents_taken_param.save()
#        except:
#            pass
#    else:
#        pass


def add_link_abonent_taken_params_from_excel_cfg_electric(sender, instance, created, **kwargs):
    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    #print dtAll[0][0]
    for i in range(1,len(dtAll)):
        meter=dtAll[i][6]
        abon=str(dtAll[i][3])
        obj=str(dtAll[i][2])
        if meter is not None:
            cursor = connection.cursor()
            sQuery="""SELECT abonents.guid FROM public.objects, public.abonents
                      WHERE objects.guid = abonents.guid_objects 
                      AND abonents.name = '%s' 
                      AND objects.name = '%s';"""%(abon,obj )
            #print sQuery
            cursor.execute(sQuery)
            guid_abonent_by_excel = cursor.fetchall()
            #print guid_abonent_by_excel

            if str(meter) == instance.guid_meters.factory_number_manual:
                writeToLog('Абонент найден' + ' ' + str(instance.name))
                #print guid_abonent_by_excel 
                dtAbon = GetSimpleCrossTable('objects', 'name', obj, 'abonents','name', abon)
                #dtAbon=GetSimpleTable('abonents','name', abon)
                guidAbon=dtAbon[0][4]
                #print instance.name, instance.guid
                common_sql.InsertInLinkAbonentsTakenParams(name = str(dtAll[i][3]) + ' - ' +  str(instance.guid_meters.name)  ,coefficient=str(dtAll[i][9]), coefficient_2 = 1, coefficient_3 = 1, guid_abonents = guidAbon, guid_taken_params = instance.guid)
                #add_link_abonents_taken_param.save()
            else:
                pass
    

def load_water_objects(request):
    args={}
    fileName = ""
    sheet    = ""
    tcp_ip_status    = ""
    counter_status    = ""
    result="Не загружено"
    try:
        if request.is_ajax():
            if request.method == 'GET':
                request.session["choice_file"]    = fileName    = request.GET['choice_file']
                request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']
                request.session["tcp_ip_status"]    = tcp_ip_status    = request.GET['tcp_ip_status']
                request.session["object_status"]    = object_status    = request.GET['object_status']
                request.session["counter_status"]    = counter_status    = request.GET['counter_status']
                
                directory=os.path.join(BASE_DIR,'static/cfg/')
                sPath=directory+fileName
                result=LoadObjectsAndAbons_water(sPath, sheet)
    except: # catch *all* exceptions
        e = sys.exc_info()[0]
        result = ( "Ошибка: %s" % e )
    

    directory=os.path.join(BASE_DIR,'static\\cfg\\')    
    if  not(os.path.exists(directory)):
        os.mkdir(directory)    
    files = os.listdir(directory)    
    args['filesFF']= files
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["port_status"]=tcp_ip_status
    args["object_status"]=result
    args["counter_status"]=counter_status
    return render(request,"service/service_water.html", args)


def CheckIfExistInObjects(name_parent, name_child):
    dt=[]
    cursor = connection.cursor()
    sQuery="""
    With obj as 
(Select guid as guid_child, objects.name as name_child, objects.level as level_child, guid_parent
 from objects)
Select guid as grand_parent, objects.name as name_parent, objects.level, objects.guid_parent, 
obj.guid_child,obj.name_child, obj.level_child, obj.guid_parent
FROM 
  public.objects, obj
where obj.guid_parent=objects.guid
and objects.name='%s' 
and obj.name_child='%s'
order by name_parent    """%(name_parent, name_child)
    cursor.execute(sQuery)
    dt = cursor.fetchall()

    if not dt:  
        return None
    else: 
        return dt[0][4]# возвращаем guid квариры
    

def LoadObjectsAndAbons_water(sPath, sheet):
    result="Объекты не загружены"
    dtAll=GetTableFromExcel(sPath,sheet) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    kv=0
    #print 'len(dtAll)', str(len(dtAll))
    for i in range(2,len(dtAll)):
        obj_l0='Вода' # всегда будет Вода как объект-родитель
        obj_l1=dtAll[i][0] #корпус
        obj_l2=dtAll[i][1] #квартира
        if not dtAll[i][1] or dtAll[i][1]==None:
            j=i
            while not obj_l2 or obj_l2==None:
                j-=1
                obj_l2=dtAll[j][1]
        abon=dtAll[i][2] #абонент он же счётчик по воде
#        chanel=dtAll[i][4] # канал пульсара
#        numPulsar=dtAll[i][5] #номер пульсара
#        typePulsar=dtAll[i][5] #тип пульсара
        isNewObj_l0=SimpleCheckIfExist('objects','name',obj_l0,"","","")#вода
        isNewObj_l1=SimpleCheckIfExist('objects','name',obj_l1,"","","")#корпус
        
        guid_obj2=CheckIfExistInObjects(obj_l1, obj_l2)#возвращает guid квартиры или None
        
        isNewAbon=SimpleCheckIfExist('objects','name', obj_l2,'abonents', 'name', abon)
        
        #print 'isNewObj_l0 ', not isNewObj_l0,'isNewObj_l1 ', not isNewObj_l1, 'guid_obj2 ', str(guid_obj2), ' IsNewAbon', not isNewAbon 
        #print i, obj_l1, obj_l2, abon
        if not (isNewObj_l0):
            writeToLog('Level 0 create object '+obj_l0)
            add_parent_object = Objects(name=obj_l0, level=0) 
            add_parent_object.save()
            writeToLog( " Ok")
            writeToLog('create object '+obj_l1)
            #print add_parent_object
            add_object1=Objects(name=obj_l1, level=1, guid_parent = add_parent_object)
            add_object1.save()
            writeToLog('create object '+obj_l2)
            add_object2=Objects(name=obj_l2, level=2, guid_parent = add_object1)
            add_object2.save()            
            writeToLog('create abonent '+abon)
            add_abonent = Abonents(name = abon, guid_objects =add_object2, guid_types_abonents = TypesAbonents.objects.get(guid= "e4d813ca-e264-4579-ae15-385cdbf5d28c"))
            
            add_abonent.save()
            kv+=1
            result="Объекты созданы"
            continue
        if not (isNewObj_l1):#новый корпус
            writeToLog('Level 1 create object '+obj_l1)
            dtParent=GetSimpleTable('objects','name',obj_l0)
            if dtParent: #родительский объект есть - корпус
                guid_parent=dtParent[0][0]
                add_object1=Objects(name=obj_l1, level=1, guid_parent = Objects.objects.get(guid=guid_parent))
                add_object1.save()                
                writeToLog('create object '+obj_l2)
                add_object2=Objects(name=obj_l2, level=2, guid_parent = add_object1)
                add_object2.save()
                writeToLog('create abonent '+abon)
                add_abonent = Abonents(name = abon, guid_objects =add_object2, guid_types_abonents = TypesAbonents.objects.get(guid= "e4d813ca-e264-4579-ae15-385cdbf5d28c"))
                add_abonent.save()
                kv+=1
                result="Объекты созданы"
                continue
            else: 
                writeToLog('Не удалось создать объект '+obj_l1)
                continue
            
        if bool(not guid_obj2): #новая квартира
            #переделать добавление на добавление по гуиду
            writeToLog('Level 2 create object '+obj_l2)
            dtParent=GetSimpleTable('objects','name',obj_l1)
            if dtParent: #родительский объект есть
                guid_parent=dtParent[0][0]
                add_object = Objects(name=obj_l2, level=2, guid_parent = Objects.objects.get(guid=guid_parent))
                add_object.save()
                result="Объекты созданы"
                add_abonent = Abonents(name = abon, guid_objects = add_object, guid_types_abonents = TypesAbonents.objects.get(guid= "e4d813ca-e264-4579-ae15-385cdbf5d28c"))
                add_abonent.save()
                kv+=1
        if not (isNewAbon):
            writeToLog('Just create abonent '+ abon)
            if bool(guid_obj2): #родительский объект есть
                add_abonent = Abonents(name = abon, guid_objects = Objects.objects.get(guid=guid_obj2), guid_types_abonents = TypesAbonents.objects.get(guid= "e4d813ca-e264-4579-ae15-385cdbf5d28c"))
                add_abonent.save()
                kv+=1            
#            else: 
#                print u'Не удалось создать объект '+abon
                continue

    result+=" Структура счётчиков создана "
    return result


def load_water_pulsar(request):
    global isService
    isService=True
    OnOffSignals()
    args={}
    result=""
    fileName = ""
    sheet = ""
    tcp_ip_status = ""
    object_status =""
    try:
        if request.is_ajax():
            if request.method == 'GET':
                request.session["choice_file"]    = fileName    = request.GET['choice_file']
                request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']
                request.session["tcp_ip_status"]    = tcp_ip_status    = request.GET['tcp_ip_status']
                request.session["object_status"]    = object_status    = request.GET['object_status']                
                directory=os.path.join(BASE_DIR,'static/cfg/')
                sPath=directory+fileName
                result=LoadWaterPulsar(sPath, sheet)
    except: # catch *all* exceptions
        e = sys.exc_info()[0]
        result = ( "Ошибка: %s" % e )
        
    #print fileName
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["tcp_ip_status"]=tcp_ip_status
    args["object_status"]=object_status
    args["counter_status"]=result
    isService=False
    OnOffSignals()
    return render(request,"service/service_water.html", args)


def LoadWaterPulsar(sPath, sSheet):
    global cfg_excel_name
    cfg_excel_name=sPath
    global cfg_sheet_name
    cfg_sheet_name=sSheet
    result=""
    dtAll=GetTableFromExcel(sPath,sSheet) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    met=0
    con=0
    #print 'str(len(dtAll))', str(len(dtAll))
    for i in range(2,len(dtAll)):
        obj_l0='Вода' # всегда будет Вода как объект-родитель
        obj_l1=str(dtAll[i][0]).strip() #корпус
        obj_l2=str(dtAll[i][1]).strip() #квартира
        if not dtAll[i][1] or dtAll[i][1]==None:
            j=i
            while not obj_l2 or obj_l2==None:
                j-=1
                obj_l2=dtAll[j][1]
        abon=str(dtAll[i][2]).strip() #абонент он же счётчик по воде
        numPulsar=str(dtAll[i][5]).strip() #номер пульсара
        typePulsar=str(dtAll[i][6]).strip() #тип пульсара
        
        isNewAbon=SimpleCheckIfExist('objects','name', obj_l2,'abonents', 'name', abon)
        isNewPulsar=SimpleCheckIfExist('meters','address', numPulsar,'','','')
        #print(u'пульсар существует ',str(isNewPulsar),typePulsar,numPulsar)
        if not (isNewAbon):
            return "Нет структуры объектов и счётчиков для "+ obj_l2 + " " +abon
        if not (isNewPulsar):
            #print (u'Обрабатываем строку ',obj_l2,numPulsar)
            if str(typePulsar) == 'Пульсар 10M':
                    signals.post_save.disconnect(add_link_taken_params, sender=TakenParams)  
                    add_meter = Meters(name = str(typePulsar) + ' ' + str(numPulsar), address = str(numPulsar), factory_number_manual = str(numPulsar), guid_types_meters = TypesMeters.objects.get(guid = "cae994a2-6ab9-4ffa-aac3-f21491a2de0b") )
                    add_meter.save()
                    print ('OK Device 10M added in DB')
                    #Если экземпляр был создан, то добавляем считываемые параметры
                    add_taken_param_no_signals(instance = add_meter, isR = False, isHalfs = False)
                    met+=1
                    
            elif str(typePulsar) == 'Пульсар 16M':
                   signals.post_save.disconnect(add_link_taken_params, sender=TakenParams)  
                   add_meter = Meters(name = str(str(typePulsar) + ' ' + str(numPulsar)), address = str(numPulsar),  factory_number_manual = str(numPulsar), guid_types_meters = TypesMeters.objects.get(guid = "7cd88751-d232-410c-a0ef-6354a79112f1") )
                   add_meter.save()
                   print ('OK Device 16M added in DB')
                   #Если экземпляр был создан, то добавляем считываемые параметры
                   add_taken_param_no_signals(instance = add_meter, isR = False, isHalfs = False)
            
                   met+=1
                   
            elif str(typePulsar) == 'Пульсар 2M':
                   signals.post_save.disconnect(add_link_taken_params, sender=TakenParams)  
                   add_meter = Meters(name = str(str(typePulsar) + ' ' + str(numPulsar)), address = str(numPulsar),  factory_number_manual = str(numPulsar), guid_types_meters = TypesMeters.objects.get(guid = "6599be9a-1f4d-4a6e-a3d9-fb054b8d44e8") )
                   add_meter.save()
                   print ('OK Device 2M added in DB')
                   #Если экземпляр был создан, то добавляем считываемые параметры
                   add_taken_param_no_signals(instance = add_meter, isR = False, isHalfs = False)
                   met+=1

            elif str(typePulsar) == 'Valtec 16M':
                   signals.post_save.disconnect(add_link_taken_params, sender=TakenParams)  
                   add_meter = Meters(name = str(str(typePulsar) + ' ' + str(numPulsar)), address = str(numPulsar),  factory_number_manual = str(numPulsar), guid_types_meters = TypesMeters.objects.get(guid = "d8613ccc-a5b4-406b-8a04-b70e08f7f7a8") )
                   add_meter.save()
                   print ('OK Device Valtec 16M added in DB')
                   #Если экземпляр был создан, то добавляем считываемые параметры
                   add_taken_param_no_signals(instance = add_meter, isR = False, isHalfs = False)
            
                   met+=1

            elif str(typePulsar) == 'МЗТА':
                   signals.post_save.disconnect(add_link_taken_params, sender=TakenParams)  
                   add_meter = Meters(name = str(str(typePulsar) + ' ' + str(numPulsar)), address = str(numPulsar),  factory_number_manual = str(numPulsar), guid_types_meters = TypesMeters.objects.get(guid = "295f91bd-3e05-435e-9eb8-bda7eddaf6a4") )
                   add_meter.save()
                   print ('OK Device МЗТА added in DB')
                   #Если экземпляр был создан, то добавляем считываемые параметры
                   add_taken_param_no_signals(instance = add_meter, isR = False, isHalfs = False)
            
                   met+=1
            else:
                print('Pulsar already exists or you incorrectly indicated the type of device in the loading list')        
        # надо проверить каналы и подсоединить их 
        #Пульсар 16M 029571 Пульсар 16M Канал 16 Суточный -- adress: 16  channel: 0
        chanel=str(dtAll[i][4])
        pulsarName=str(dtAll[i][6])
        abonent_name=str(dtAll[i][2])
        taken_param = pulsarName + ' ' + str(dtAll[i][5]) + ' '+ pulsarName + ' ' + 'Канал ' + chanel+ ' Суточный -- adress: ' +chanel+'  channel: 0'
        #print "chanel ", chanel
        print(taken_param)
        #Sravnenie(taken_param)
        dtTakenParam=GetSimpleTable('taken_params','name',taken_param)
        #writeToLog(bool(dtTakenParam))
        print(dtTakenParam)
        if dtTakenParam:                
            print(u'taken param найден')
            guid_taken_param=dtTakenParam[0][2]
            dtLink=GetSimpleTable('link_abonents_taken_params','guid_taken_params',guid_taken_param)
            #print dtLink
            if (dtLink):
                #print 'link is exist '+ chanel + '  '+pulsarName
                result+="\n Привязка канала "+chanel+" Пульсара "+pulsarName+" уже существует. Перезапись НЕ произведена для счётчика "+abonent_name
                continue
            else:
                #dtAbon= GetSimpleTable('abonents','name', abonent_name)
                t = Abonents.objects.filter(name = abonent_name)
                guidAbon = t[0].guid     #dtAbon[0][0]
                print(abonent_name, 'guidAbon', guidAbon)
                #"миномес ГВС, №68208 Канал 5 Суточный"
                #print abonent_name, guidAbon, guid_taken_param
                common_sql.InsertInLinkAbonentsTakenParams(name = abonent_name+' Канал '+chanel+' Суточный',coefficient=1, coefficient_2 = 1,coefficient_3 = 1, guid_abonents = guidAbon, guid_taken_params = guid_taken_param)
                #add_link_abonents_taken_param.save()
                print (u'Abonent connected with taken param')
                con+=1
    result+='Прогружено новых пульсаров '+str(met)
    if con>0:
        result+='Созданы новые связи '
    #print('1111111111')
    signals.post_save.connect(add_link_taken_params, sender=TakenParams)  
    return result

#def Sravnenie(takenParam):
#    str_bd='Пульсар 2М 062726 Пульсар 2M Канал 1 Суточный -- adress: 1 channel: 0'
#    i=0
#    print str_bd
#    while i!=len(takenParam):
#        if ord(takenParam[i])!=ord(str_bd[i]):
#            print i, takenParam[i]
#        i+=1


def load_water_port(request):
    args={}

    fileName=""
    sheet    = ""
    result=""
    try:
        if request.is_ajax():
            if request.method == 'GET':
                request.session["choice_file"]    = fileName    = request.GET['choice_file']
                request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']
                request.session["tcp_ip_status"]    = tcp_ip_status    = request.GET['tcp_ip_status']
                
                directory=os.path.join(BASE_DIR,'static/cfg/')
                sPath=directory+fileName
                #print sPath, sheet
                result=load_tcp_ip_water_ports_from_excel(sPath, sheet)
                if result:
                    result="Порт/ы был успешно добавлен"
                else:
                    result="Порт не был загружен, он уже существует в БД"
    except: # catch *all* exceptions
        e = sys.exc_info()[0]
        result = ( "Ошибка: %s" % e )

    directory=os.path.join(BASE_DIR,'static\\cfg\\')    
    if  not(os.path.exists(directory)):
        os.mkdir(directory)    
    files = os.listdir(directory)
    # print(files) 
    # print(fileName)   
    args['filesFF']= files
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["tcp_ip_status"]=result

    return render(request,"service/service_water.html", args)


def UpdateTable(table,whereFieled, whereValue,field1,value1,field2,value2,field3,value3):
    isOk=False
    dt=[]
    cursor = connection.cursor()
    if (field2==""):
        sQuery="""           
     UPDATE %s
     SET  %s='%s'       
     WHERE %s='%s'
     RETURNING * 
   """%(table, field1, value1, whereFieled, whereValue)
    elif (field3==""):
        sQuery="""           
     UPDATE %s
     SET  %s='%s', %s='%s'      
     WHERE %s='%s'
     RETURNING * 
   """%(table, field1, value1,field2,value2,whereFieled, whereValue)
    else:
       sQuery="""           
     UPDATE %s
     SET  %s='%s', %s='%s', %s='%s'       
     WHERE %s='%s'
     RETURNING * 
   """%(table, field1, value1,field2,value2,field3,value3,whereFieled, whereValue)
    #print sQuery
    cursor.execute(sQuery)
    dt = cursor.fetchall()
    if len(dt):
        isOk=True   
    return isOk


def load_tcp_ip_water_ports_from_excel(sPath, sheet):
    #Добавление tcp_ip портов
    
    wb = load_workbook(filename = sPath)
    sheet_ranges = wb[sheet]
    row = 3
    IsAdded=False
    result=""
    writeToLog('Load port')
    writeToLog('Загрузка портов')
    while (bool(sheet_ranges['H%s'%(row)].value)):
        if sheet_ranges['H%s'%(row)].value is not None:
            ip_adr=str(sheet_ranges['H%s'%(row)].value)
            ip_port=str(sheet_ranges['I%s'%(row)].value)
            #print ip_adr, ip_port
            writeToLog('Обрабатываем адрес ' +ip_adr +' '+ ip_port)
            
            # проверка есть ли уже такой порт, запрос в БД с адресом и портом, если ответ пустой-добавляем, в противном случае continue
            if not ip_adr or not ip_port or ip_adr==None or ip_port==None: 
                result+="Отсутствует значение/я для tcp/ip-порта в строке"+str(row)+". Заполните все ячейки excel таблицы. "
                break
            else:
                if (checkPortIsExist(ip_adr,ip_port)):
                    add_port=TcpipSettings(ip_address = ip_adr, ip_port =int(ip_port), write_timeout =300 , read_timeout =700 , attempts =3 , delay_between_sending =400)
                    add_port.save()
                    result ='Новый tcp/ip порт добавлен'
                    IsAdded=True
                else: 
                        newRes = 'Порт '+str(ip_adr)+": "+str(ip_port)+" уже существует "
                        # print result
                        # print newRes
                        # print result.find(newRes)
                        if bool(result.find(newRes) == -1):
                            result+= newRes
        writeToLog( result)
        row+=1
    return IsAdded


def makeLinkabonentTakenParamName(abName,typeMeter,new_meter):
    #"Квартира 0103 - М-230 21949676"   
#LinkAbonentsTakenParams (name = Abonents.objects.get(name= get_taken_param_by_abonent_from_excel_cfg(instance.name)).name + u" " + instance.guid_params.guid_names_params.name + u" " + instance.guid_params.guid_types_params.name 
    newLinkAbonentTakenParamName=abName+ ' - '+ typeMeter +' ' + str(new_meter)
    return newLinkAbonentTakenParamName


def makeNewTakenParamName(nameParam1, old_meter, new_meter, typeMeter):
    newName=''

        # "М-230 22633939 Меркурий 230 T0 A+ Суточный -- adress: 0  channel: 0"
        #"Саяны Комбик 4443 Саяны Комбик Q Система1 Суточный -- adress: 0  channel: 1"
    n=nameParam1.find(old_meter)
    s=nameParam1[n+len(old_meter):]
    newName= typeMeter + ' ' + str(new_meter) + s
    return newName
    
def get_electric_progruz_com(request):
    response = io.StringIO()
    wb = Workbook()    
    wb.add_named_style(ali_grey)
    ws = wb.active

    obj_title         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')
    electric_data_start   = request.GET.get('electric_data_start')
    
#Шапка
    ws['A1'] = 'Населенный пункт'
    ws['B1'] = 'Наименование улицы'
    ws['C1'] = 'Наименование дома'
    ws['D1'] = 'Квартира'
    ws['E1'] = 'Идентификатор АСКУЭ'
    ws['F1'] = 'Номер лицевого счета'
    ws['G1'] = 'Номер счётчика заводской'
    ws['H1'] = 'Номер в сети'
    ws['I1'] = 'Тип счётчика'
    ws['J1'] = 'ТТ'
    ws['K1'] = 'ip adress'
    ws['L1'] = 'ip port'
    
    ws['A1'].style = "ali_grey"
    ws['B1'].style = "ali_grey"
    ws['C1'].style = "ali_grey"
    ws['D1'].style = "ali_grey"
    ws['E1'].style = "ali_grey"
    ws['F1'].style = "ali_grey"
    ws['G1'].style = "ali_grey"
    ws['H1'].style = "ali_grey"
    ws['I1'].style = "ali_grey"
    ws['J1'].style = "ali_grey"
    ws['K1'].style = "ali_grey"
    ws['L1'].style = "ali_grey"

#Запрашиваем данные для отчета
    data_table = common_sql.get_electric_register_com()        
   
        
# Заполняем отчет значениями
    for row in range(2, len(data_table)+2):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-2][0])  
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-2][1]) 
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-2][2])  
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % (data_table[row-2][3])  # 
            ws.cell('d%s'%(row)).style = ali_white
        except:
            ws.cell('d%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % data_table[row-2][4]  # 
            ws.cell('e%s'%(row)).style = ali_white
        except:
            ws.cell('e%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('f%s'%(row)).value = '%s' % data_table[row-2][5]  # 
            ws.cell('f%s'%(row)).style = ali_white
        except:
            ws.cell('f%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('g%s'%(row)).value = '%s' % data_table[row-2][6]  
            ws.cell('g%s'%(row)).style = ali_white
        except:
            ws.cell('g%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('h%s'%(row)).value = '%s' % (data_table[row-2][7])  
            ws.cell('h%s'%(row)).style = ali_white
        except:
            ws.cell('h%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('i%s'%(row)).value = '%s' % (data_table[row-2][8])  
            ws.cell('i%s'%(row)).style = ali_white
        except:
            ws.cell('i%s'%(row)).style = ali_white
            next

        try:
            ws.cell('j%s'%(row)).value = '%s' % (data_table[row-2][9])  
            ws.cell('j%s'%(row)).style = ali_white
        except:
            ws.cell('j%s'%(row)).style = ali_white
            next

        # try:
        #     ws.cell('k%s'%(row)).value = '%s' % (data_table[row-2][10])  
        #     ws.cell('k%s'%(row)).style = ali_white
        # except:
        #     ws.cell('k%s'%(row)).style = ali_white
        #     next

        try:
            ws.cell('l%s'%(row)).value = '%s' % (data_table[row-2][10])  
            ws.cell('l%s'%(row)).style = ali_white
        except:
            ws.cell('l%s'%(row)).style = ali_white
            next

    #ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 15 
    ws.column_dimensions['B'].width = 30 
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 10
    
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel") 
    now = datetime.datetime.now()
    electric_data_end = now.strftime("%d-%m-%Y %H:%M")
    #response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")    
    output_name = 'electric_register-'+electric_data_end + '_com'
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response


def get_electric_progruz(request):
    response = io.StringIO()
    wb = Workbook()    
    wb.add_named_style(ali_grey)
    ws = wb.active
    
    obj_title         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')
    electric_data_start   = request.GET.get('electric_data_start')
    
#Шапка
    ws['A1'] = 'Населенный пункт'
    ws['B1'] = 'Наименование улицы'
    ws['C1'] = 'Наименование дома'
    ws['D1'] = 'Квартира'
    ws['E1'] = 'Идентификатор АСКУЭ'
    ws['F1'] = 'Номер лицевого счета'
    ws['G1'] = 'Номер счётчика заводской'
    ws['H1'] = 'Номер в сети'
    ws['I1'] = 'Тип счётчика'
    ws['J1'] = 'ТТ'
    ws['K1'] = 'ip adress'
    ws['L1'] = 'ip port'
    
    ws['A1'].style = "ali_grey"
    ws['B1'].style = "ali_grey"
    ws['C1'].style = "ali_grey"
    ws['D1'].style = "ali_grey"
    ws['E1'].style = "ali_grey"
    ws['F1'].style = "ali_grey"
    ws['G1'].style = "ali_grey"
    ws['H1'].style = "ali_grey"
    ws['I1'].style = "ali_grey"
    ws['J1'].style = "ali_grey"
    ws['K1'].style = "ali_grey"
    ws['L1'].style = "ali_grey"

#Запрашиваем данные для отчета
    data_table = common_sql.get_electric_register()        
   
        
# Заполняем отчет значениями
    for row in range(2, len(data_table)+2):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-2][0])  
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-2][1]) 
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-2][2])  
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % (data_table[row-2][3])  # 
            ws.cell('d%s'%(row)).style = ali_white
        except:
            ws.cell('d%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % data_table[row-2][4]  # 
            ws.cell('e%s'%(row)).style = ali_white
        except:
            ws.cell('e%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('f%s'%(row)).value = '%s' % data_table[row-2][5]  # 
            ws.cell('f%s'%(row)).style = ali_white
        except:
            ws.cell('f%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('g%s'%(row)).value = '%s' % data_table[row-2][6]  
            ws.cell('g%s'%(row)).style = ali_white
        except:
            ws.cell('g%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('h%s'%(row)).value = '%s' % (data_table[row-2][7])  
            ws.cell('h%s'%(row)).style = ali_white
        except:
            ws.cell('h%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('i%s'%(row)).value = '%s' % (data_table[row-2][8])  
            ws.cell('i%s'%(row)).style = ali_white
        except:
            ws.cell('i%s'%(row)).style = ali_white
            next

        try:
            ws.cell('j%s'%(row)).value = '%s' % (data_table[row-2][9])  
            ws.cell('j%s'%(row)).style = ali_white
        except:
            ws.cell('j%s'%(row)).style = ali_white
            next

        try:
            ws.cell('k%s'%(row)).value = '%s' % (data_table[row-2][10])  
            ws.cell('k%s'%(row)).style = ali_white
        except:
            ws.cell('k%s'%(row)).style = ali_white
            next

        try:
            ws.cell('l%s'%(row)).value = '%s' % (data_table[row-2][11])  
            ws.cell('l%s'%(row)).style = ali_white
        except:
            ws.cell('l%s'%(row)).style = ali_white
            next

    #ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 15 
    ws.column_dimensions['B'].width = 30 
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 10
    
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel") 
    now = datetime.datetime.now()
    electric_data_end = now.strftime("%d-%m-%Y %H:%M")
    #response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")    
    output_name = 'electric_register-'+electric_data_end
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def get_water_progruz(request):
    response = io.StringIO()
    wb = Workbook()    
    wb.add_named_style(ali_grey)
    ws = wb.active
    
    obj_title         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')
    electric_data_start   = request.GET.get('electric_data_start')
    
#Шапка
    ws['A1'] = 'Населенный пункт'
    ws['B1'] = 'Наименование улицы'
    ws['C1'] = 'Наименование дома'
    ws['D1'] = 'Квартира'
    ws['E1'] = 'Идентификатор АСКУЭ'
    ws['F1'] = 'Номер лицевого счета'
    ws['G1'] = 'Номер счётчика заводской'
    ws['H1'] = 'Номер в сети'
    ws['I1'] = 'Тип счётчика'
    ws['J1'] = 'ТТ'
    ws['K1'] = 'ip adress'
    ws['L1'] = 'ip port'
    
    ws['A1'].style = "ali_grey"
    ws['B1'].style = "ali_grey"
    ws['C1'].style = "ali_grey"
    ws['D1'].style = "ali_grey"
    ws['E1'].style = "ali_grey"
    ws['F1'].style = "ali_grey"
    ws['G1'].style = "ali_grey"
    ws['H1'].style = "ali_grey"
    ws['I1'].style = "ali_grey"
    ws['J1'].style = "ali_grey"
    ws['K1'].style = "ali_grey"
    ws['L1'].style = "ali_grey"

#Запрашиваем данные для отчета
    data_table = common_sql.get_water_register()        
   
        
# Заполняем отчет значениями
    for row in range(2, len(data_table)+2):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-2][0])  
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-2][1]) 
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-2][2])  
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % (data_table[row-2][3])  # 
            ws.cell('d%s'%(row)).style = ali_white
        except:
            ws.cell('d%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % data_table[row-2][4]  # 
            ws.cell('e%s'%(row)).style = ali_white
        except:
            ws.cell('e%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('f%s'%(row)).value = '%s' % data_table[row-2][5]  # 
            ws.cell('f%s'%(row)).style = ali_white
        except:
            ws.cell('f%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('g%s'%(row)).value = '%s' % data_table[row-2][6]  
            ws.cell('g%s'%(row)).style = ali_white
        except:
            ws.cell('g%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('h%s'%(row)).value = '%s' % (data_table[row-2][7])  
            ws.cell('h%s'%(row)).style = ali_white
        except:
            ws.cell('h%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('i%s'%(row)).value = '%s' % (data_table[row-2][8])  
            ws.cell('i%s'%(row)).style = ali_white
        except:
            ws.cell('i%s'%(row)).style = ali_white
            next

        try:
            ws.cell('j%s'%(row)).value = '%s' % (data_table[row-2][9])  
            ws.cell('j%s'%(row)).style = ali_white
        except:
            ws.cell('j%s'%(row)).style = ali_white
            next

        try:
            ws.cell('k%s'%(row)).value = '%s' % (data_table[row-2][10])  
            ws.cell('k%s'%(row)).style = ali_white
        except:
            ws.cell('k%s'%(row)).style = ali_white
            next

        try:
            ws.cell('l%s'%(row)).value = '%s' % (data_table[row-2][11])  
            ws.cell('l%s'%(row)).style = ali_white
        except:
            ws.cell('l%s'%(row)).style = ali_white
            next

    #ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 15 
    ws.column_dimensions['B'].width = 30 
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 10
    ws.column_dimensions['I'].width = 15
    
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel") 
    now = datetime.datetime.now()
    electric_data_end = now.strftime("%d-%m-%Y %H:%M")
    #response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")    
    output_name = 'water_register-'+electric_data_end
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def get_water_impulse_progruz(request):
    response = io.StringIO()
    wb = Workbook()    
    wb.add_named_style(ali_grey)
    ws = wb.active
    
    obj_title         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')
    electric_data_start   = request.GET.get('electric_data_start')
    
#Шапка
    ws['A1'] = 'Импульсные приборы'
    ws['A2'] = 'Наименование дома'
    ws['B2'] = 'Объект'
    ws['C2'] = 'Счётчик'
    ws['D2'] = 'Регистратор'
    ws['E2'] = 'Канал'
    ws['F2'] = 'Номер Пульсара'
    ws['G2'] = 'Тип Пульсара'
    ws['H2'] = 'IP'
    ws['I2'] = 'Порт'

    
    ws['A2'].style = "ali_grey"
    ws['B2'].style = "ali_grey"
    ws['C2'].style = "ali_grey"
    ws['D2'].style = "ali_grey"
    ws['E2'].style = "ali_grey"
    ws['F2'].style = "ali_grey"
    ws['G2'].style = "ali_grey"
    ws['H2'].style = "ali_grey"
    ws['I2'].style = "ali_grey"

#Запрашиваем данные для отчета
    data_table = common_sql.get_water_impulse_register()        
   
        
# Заполняем отчет значениями
    for row in range(3, len(data_table)+2):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-3][0])  
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-3][1]) 
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-3][2])  
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % (data_table[row-3][3])  # 
            ws.cell('d%s'%(row)).style = ali_white
        except:
            ws.cell('d%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % data_table[row-3][4]  # 
            ws.cell('e%s'%(row)).style = ali_white
        except:
            ws.cell('e%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('f%s'%(row)).value = '%s' % data_table[row-3][5]  # 
            ws.cell('f%s'%(row)).style = ali_white
        except:
            ws.cell('f%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('g%s'%(row)).value = '%s' % data_table[row-3][6]  
            ws.cell('g%s'%(row)).style = ali_white
        except:
            ws.cell('g%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('h%s'%(row)).value = '%s' % (data_table[row-3][7])  
            ws.cell('h%s'%(row)).style = ali_white
        except:
            ws.cell('h%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('i%s'%(row)).value = '%s' % (data_table[row-3][8])  
            ws.cell('i%s'%(row)).style = ali_white
        except:
            ws.cell('i%s'%(row)).style = ali_white
            next


    #ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 30 
    ws.column_dimensions['B'].width = 30 
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['H'].width = 20
    
    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel") 
    now = datetime.datetime.now()
    electric_data_end = now.strftime("%d-%m-%Y %H:%M")
    #response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")    
    output_name = 'water_impulse_register-'+electric_data_end
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def get_heat_progruz(request):
    response = io.StringIO()
    wb = Workbook()    
    wb.add_named_style(ali_grey)
    ws = wb.active
    
    obj_title         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')
    electric_data_start   = request.GET.get('electric_data_start')
    
#Шапка
    ws['A1'] = 'Населенный пункт'
    ws['B1'] = 'Наименование улицы'
    ws['C1'] = 'Наименование дома'
    ws['D1'] = 'Квартира'
    ws['E1'] = 'Идентификатор АСКУЭ'
    ws['F1'] = 'Номер лицевого счета'
    ws['G1'] = 'Номер счётчика заводской'
    ws['H1'] = 'Номер в сети'
    ws['I1'] = 'Тип счётчика'
    ws['J1'] = 'ТТ'
    ws['K1'] = 'ip adress'
    ws['L1'] = 'ip port'
    
    ws['A1'].style = "ali_grey"
    ws['B1'].style = "ali_grey"
    ws['C1'].style = "ali_grey"
    ws['D1'].style = "ali_grey"
    ws['E1'].style = "ali_grey"
    ws['F1'].style = "ali_grey"
    ws['G1'].style = "ali_grey"
    ws['H1'].style = "ali_grey"
    ws['I1'].style = "ali_grey"
    ws['J1'].style = "ali_grey"
    ws['K1'].style = "ali_grey"
    ws['L1'].style = "ali_grey"

#Запрашиваем данные для отчета
    data_table = common_sql.get_heat_register()        
   
        
# Заполняем отчет значениями
    for row in range(2, len(data_table)+2):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-2][0])  
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-2][1]) 
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-2][2])  
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % (data_table[row-2][3])  # 
            ws.cell('d%s'%(row)).style = ali_white
        except:
            ws.cell('d%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % data_table[row-2][4]  # 
            ws.cell('e%s'%(row)).style = ali_white
        except:
            ws.cell('e%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('f%s'%(row)).value = '%s' % data_table[row-2][5]  # 
            ws.cell('f%s'%(row)).style = ali_white
        except:
            ws.cell('f%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('g%s'%(row)).value = '%s' % data_table[row-2][6]  
            ws.cell('g%s'%(row)).style = ali_white
        except:
            ws.cell('g%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('h%s'%(row)).value = '%s' % (data_table[row-2][7])  
            ws.cell('h%s'%(row)).style = ali_white
        except:
            ws.cell('h%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('i%s'%(row)).value = '%s' % (data_table[row-2][8])  
            ws.cell('i%s'%(row)).style = ali_white
        except:
            ws.cell('i%s'%(row)).style = ali_white
            next

        try:
            ws.cell('j%s'%(row)).value = '%s' % (data_table[row-2][9])  
            ws.cell('j%s'%(row)).style = ali_white
        except:
            ws.cell('j%s'%(row)).style = ali_white
            next

        try:
            ws.cell('k%s'%(row)).value = '%s' % (data_table[row-2][10])  
            ws.cell('k%s'%(row)).style = ali_white
        except:
            ws.cell('k%s'%(row)).style = ali_white
            next

        try:
            ws.cell('l%s'%(row)).value = '%s' % (data_table[row-2][11])  
            ws.cell('l%s'%(row)).style = ali_white
        except:
            ws.cell('l%s'%(row)).style = ali_white
            next

    #ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 15 
    ws.column_dimensions['B'].width = 30 
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 10
    ws.column_dimensions['I'].width = 20

    response.seek(0)
    response = HttpResponse(save_virtual_workbook(wb),content_type="application/vnd.ms-excel") 
    now = datetime.datetime.now()
    electric_data_end = now.strftime("%d-%m-%Y %H:%M")
    #response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")    
    output_name = 'heat_register-'+electric_data_end
    file_ext = 'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def get_info(request):    
    args={}   
    args['pulsar16m_status'] = ''
    return render(request,"service/service_get_info.html", args)


def load_balance_group(request):
    args={}
    fileName=""
    sheet    = ""
    balance_status    = ""
    result="Не загружено"
    if request.is_ajax():
        if request.method == 'GET':
            request.session["choice_file"]    = fileName    = request.GET['choice_file']
            request.session["choice_sheet"]    = sheet    = request.GET.get('choice_sheet')
            request.session["balance_status"]    = balance_status    = request.GET['balance_status']            
            directory=os.path.join(BASE_DIR,'static/cfg/')
            sPath=directory+fileName
            result=LoadBalance(sPath, sheet)
    
    balance_status=result

    #print fileName
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["balance_status"]=balance_status

    return render(request,"service/service_balance_load.html", args)


def InsertIntoBalanceGroup(guid,name):
    result=''
    cursor = connection.cursor()
    sQuery="""
    INSERT INTO balance_groups(
            guid, name)
    VALUES ('%s', '%s');
    """%(guid,name)   
    
    cursor.execute(sQuery)
    cursor.close()
    connection.commit()
    result ='Создана балансная группа '+str(name)
    return result


def InsertIntoTypesAbonents(guid,name):
    result=''
    cursor = connection.cursor()
    sQuery="""
    INSERT INTO types_abonents(
            guid, name)
    VALUES ('%s', '%s');
    """%(guid,name)   
    
    cursor.execute(sQuery)
    cursor.close()
    connection.commit()
    result ='Создан тип '+str(name)
    return result


def GetGuidFromFirstTableCrossWithSecondTable(table1,table2,field1,val1,field2,val2):
    dt=[]
    cursor = connection.cursor()
    sQuery="""
        SELECT 
  %s.guid
FROM 
  public.%s, 
  public.%s
WHERE 
  %s.guid_%s = %s.guid AND
  %s.%s = '%s' AND 
  %s.%s = '%s'"""%(table1, table1, table2,  table1, table2,    table2,table1,field1,val1,table2,field2,val2)
    #print sQuery
    cursor.execute(sQuery)
    dt = cursor.fetchall()
    #print sQuery
    return dt
   

def UpdateSimpleTable(table,guid,field,val):
    result=False
    cursor = connection.cursor()
    sQuery="""
    UPDATE %s
    SET %s = '%s'
    WHERE guid = '%s'
    """%(table,field,val,guid)   
    
    cursor.execute(sQuery)
    cursor.close()
    connection.commit()
    result =True
    return result
    

def LoadImpulseWaterBalance(dtAll):
    result = "Баланс по водным импульсным счётчикам"
    count_new_link=0
    for i in range(1,len(dtAll)):        
        balance_group=str(dtAll[i][0])
        znak=str(dtAll[i][1])        
        meter=str(dtAll[i][4])
        type_abonent=str(dtAll[i][5])
        #print balance_group, znak, meter, type_abonent
        isNewBalanceGroup=not SimpleCheckIfExist('balance_groups','name',balance_group,"","","")
        isNewMeter=not SimpleCheckIfExist('meters','factory_number_manual',meter,"","","")
        isNewTypeAbonent=not SimpleCheckIfExist('types_abonents','name',type_abonent,"","","")
        #print u'isNewBalanceGroup: ', isNewBalanceGroup
        #print u'isNewTypeAbonent: ', isNewTypeAbonent
        #print u'isNewMeter: ', isNewMeter
        if isNewBalanceGroup: #если балансной группы ещё не существует, то создаём её
            balance_group_guid=uuid.uuid4()
            result += InsertIntoBalanceGroup(balance_group_guid, balance_group)
            #print u'Создана балансная группа '+balance_group
        if isNewTypeAbonent: #если такого типа абонента не существует, то создаём
            types_abonents_guid=uuid.uuid4()
            result += InsertIntoTypesAbonents(types_abonents_guid, type_abonent)
            #print u'Создан тип абонента ' + type_abonent       
        if isNewMeter:#ничего не создаём, добавляем сообщение, что абонента надо создать
           result += 'Счётчика '+meter+' (в таблице строка '+str(i+1)+') не существует. В балансную группу не добавлен!'
           continue
              
        guid_meters=GetSimpleTable('meters','factory_number_manual',meter)[0][0]        
        if not isNewBalanceGroup:
           balance_group_guid=GetSimpleTable('balance_groups','name',balance_group)[0][0]
       
        #проверяем нет ли такой связи уже
        dt_link=GetSimpleTable('link_balance_groups_meters',"guid_meters",guid_meters[0][0])
        isNewLink=True
        for j in range(1,len(dt_link)):
            #print dt_link[j][3]
            if dt_link[j][3] == balance_group_guid:
                isNewLink=False
                result+= 'Счётчик ' + meter + ' уже принадлежит балансной группе ' + balance_group
                break
        if isNewLink:
            #print balance_group, meter
            cursor = connection.cursor()
            isZnak=True        
            if znak=='0' or znak == 0:
                isZnak=False
            sQuery="""
                  INSERT INTO link_balance_groups_meters(
                  guid, type, guid_balance_groups, guid_meters)
                  VALUES ('%s', '%s', '%s', '%s');
                  """%(uuid.uuid4(),isZnak,balance_group_guid,guid_meters)      
            cursor.execute(sQuery)
            cursor.close()
            connection.commit()
            count_new_link+=1
    result+= '  В балансную группу добавлено счётчиков: '+ str(count_new_link) 
    return result       


def LoadBalance(sPath, sheet):
    result="Баланс по цифровым счётчикам"
    count_new_link=0
    dtAll=GetTableFromExcel(sPath, sheet) #получили из excel все строки до первой пустой строки (проверка по колонке А)    

    if len(dtAll)==0: return 'Таблица пуста!'
    
    if (len(str(dtAll[1][3])) < 0) or (dtAll[1][3] is None):
        result = LoadImpulseWaterBalance(dtAll)
    else:      
    
        for i in range(1,len(dtAll)):        
            balance_group=str(dtAll[i][0])
            znak=str(dtAll[i][1])
            object_name=str(dtAll[i][2])
            abonent_name=str(dtAll[i][3])
            meter=str(dtAll[i][4])
            type_abonent=str(dtAll[i][5])
            #print balance_group, znak, abonent_name,meter, type_abonent
            isNewBalanceGroup=not SimpleCheckIfExist('balance_groups','name',balance_group,"","","")
            isNewMeter=not SimpleCheckIfExist('meters','factory_number_manual',meter,"","","")
            isNewTypeAbonent=not SimpleCheckIfExist('types_abonents','name',type_abonent,"","","")
            #print u'isNewBalanceGroup: ', isNewBalanceGroup
            #print u'isNewTypeAbonent: ', isNewTypeAbonent
            #print u'isNewMeter: ', isNewMeter
            if isNewBalanceGroup: #если балансной группы ещё не существует, то создаём её
                balance_group_guid=uuid.uuid4()
                result += InsertIntoBalanceGroup(balance_group_guid, balance_group)
                #print u'Создана балансная группа '+balance_group
            if isNewTypeAbonent: #если такого типа абонента не существует, то создаём
                types_abonents_guid=uuid.uuid4()
                result += InsertIntoTypesAbonents(types_abonents_guid, type_abonent)
                #print u'Создан тип абонента ' + type_abonent       
            if isNewMeter:#ничего не создаём, добавляем сообщение, что абонента надо создать
               result += 'Счётчика '+meter+' (в таблице должен принадлежать абоненту '+abonent_name+') не существует. В балансную группу не добавлен!'
               continue
           
            types_abonents_guid=GetSimpleTable('types_abonents','name',type_abonent)[0][0]
            guid_abonent=GetGuidFromFirstTableCrossWithSecondTable('abonents','objects','name',abonent_name,'name',object_name)[0][0]
            isOk=UpdateSimpleTable('abonents', guid_abonent,'guid_types_abonents',types_abonents_guid)
            #print u'type of abonents changed: ', isOk 
            
            guid_meters=GetSimpleTable('meters','factory_number_manual',meter)[0][0]        
            if not isNewBalanceGroup:
               balance_group_guid=GetSimpleTable('balance_groups','name',balance_group)[0][0]
           
            #проверяем нет ли такой связи уже
            dt_link=GetSimpleTable('link_balance_groups_meters',"guid_meters",guid_meters)
            isNewLink=True
            for j in range(1,len(dt_link)):
                #print dt_link[j][3]
                if dt_link[j][3] == balance_group_guid:
                    isNewLink=False
                    result+= 'Счётчик ' + meter + ' уже принадлежит балансной группе ' + balance_group
                    break
            if isNewLink:
                #print balance_group, meter
                cursor = connection.cursor()
                isZnak=True        
                if znak=='0' or znak == 0:
                    isZnak=False
                sQuery="""
                      INSERT INTO link_balance_groups_meters(
                      guid, type, guid_balance_groups, guid_meters)
                      VALUES ('%s', '%s', '%s', '%s');
                      """%(uuid.uuid4(),isZnak,balance_group_guid,guid_meters)      
                cursor.execute(sQuery)
                cursor.close()
                connection.commit()
                count_new_link+=1
        result+= '  В балансную группу добавлено счётчиков: '+ str(count_new_link)    
         
    return result


def service_balance_load(request):    
    args={}    
    return render(request,"service/service_balance_load.html", args)


def add_current_taken_params_pulsar16m(request):   
    
    result = "Прогрузка прошла не успешно"
    args={}
    
    dt_pulsar16m=common_sql.get_meters_by_type( 'Пульсар 16M')
    count16m=0
    for puls in dt_pulsar16m:
        #print u'счётчик', puls[1]
        guid_meter = puls[0]
        dt_current_count = common_sql.get_count_current_params_by_meters_guid(guid_meter)
        if len(dt_current_count)>0: continue
        # if dt_current_count[0][1] == 'Пульсар 16M':
        #     if dt_current_count[0][0] >=16: continue
        
    # Текущие
      #Канал 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "e3f1325e-3018-40ba-b94a-ab6d6ac093e9"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 1)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 1 Текущий", coefficient=1, coefficient_2 = 1,coefficient_3 = 1, guid_abonents = dt_abonent[0][0], guid_taken_params = add_param.guid )               
        #common_sql.InsertInLinkAbonentsTakenParams(  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        #add_link_abonents_taken_param.save()

      #Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "5a6b0338-c15d-4224-a04f-a10fc73c5fc7"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 2)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 2 Текущий", coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "48a42afe-d9ac-4180-a733-6dd5f9d9ca80"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 3)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 3 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 4
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "01a5419c-c701-4185-95b6-457b8c9ca2d0"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 4)        
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 4 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 5
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "85c4295e-bc6a-46ec-9866-0bf9f77c6904"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 5)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 5 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 6
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "68270d0a-5043-4ea2-9b61-4adaa298abad"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 6)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 6 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 7
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "cd489c4b-6e74-4c65-bfee-c0fa78a853bf"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 7)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 7 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
       
      #Канал 8
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "f29062a4-ab60-4117-8f85-0cdec634c797"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 8)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 8 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 9
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "e8521cd7-2f38-4619-935d-8fe86234dbe7"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 9)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 9 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 10
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "1349b747-41ca-4ba8-a690-69c649129f44"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 10)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 10 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 11
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "99ab1a30-fde8-4b81-9f9e-2f731516ce1b"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 11)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 11 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 12
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "c7f6a397-833d-4020-9d2b-38c19bec272c"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 12)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 12 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 13
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "4413bffb-1832-4900-9351-5ac3666dd8b0"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 13)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 13 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 14
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "6280490b-123d-4e27-bef9-19fd7dc2cf54"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 14)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 14 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 15
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "93891c5a-1c8f-4906-b7f0-961dc8ad3c9f"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 15)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 15 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 16
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "22dd3a17-a828-44e0-80d9-db075ba120ae"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 16)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 16 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        

        count16m+=1

    dt_pulsar10m=common_sql.get_meters_by_type( 'Пульсар 10M')
    count10m=0
    for puls in dt_pulsar10m:        
        #Добавляем параметры для Пульсар10 
        #print u'счётчик', puls[1]
        guid_meter = puls[0]
        dt_current_count = common_sql.get_count_current_params_by_meters_guid(guid_meter)
        if len(dt_current_count)>0: continue
    # Текущие
      #Канал 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "32dad392-ca1e-4110-8f2c-a86b02e26fb3"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 1)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 1 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "3e13694b-7cb5-4417-a091-af8a7db34dc7"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 2)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 2 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "1023b35b-3cbf-4519-aac3-3bf1ebae07c1"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 3)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 3 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 4
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "eea27ade-44cd-4e66-8298-00a4a6ad915a"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 4)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 4 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 5
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "25e09d4d-3a48-4381-ad5d-b783c03c4d35"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 5)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 5 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 6
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "069898ea-9d74-4571-b719-e8e6f1513c12"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 6)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 6 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
         
      #Канал 7
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "084aa5f4-75d5-41f6-b0d6-9f2403eacd2c"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 7)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 7 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 8
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "786ed8b8-aed1-478c-ae75-99caf1358cf0"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 8)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 8 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 9
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "6fc4c39c-9a43-4cb7-a066-c40fd2ca47e5"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 9)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 9 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        
      #Канал 10
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = Meters.objects.get(guid=puls[0]), guid_params = Params.objects.get(guid = "8b2aa40a-cd91-4e22-b9d1-596e49e5f839"))
        add_param.save()
        dt_abonent = common_sql.get_abonent_by_meter_and_pulsar_chanel(guid_meter, 10)
        if len(dt_abonent)>0:
            common_sql.InsertInLinkAbonentsTakenParams(name = dt_abonent[0][1] + " Канал 10 Текущий"  ,coefficient=1, coefficient_2 = 1,coefficient_3 = 1,guid_abonents = dt_abonent[0][0] , guid_taken_params = add_param.guid )
        

        count10m+=1

    result = 'Добавлены параметры для ' + str(count16m) + ' ПУ Пульсар 16М и ' + str(count10m) + ' ПУ Пульcар 10М'  
    #print(result)
    args['pulsar16m_status'] = result
    return render(request,"service/service_get_info.html", args)


def change_meters_v2(request):
    args={}

    old_meter=' '
    new_meter=' '
    change_meter_status="Функция в разработке"
    if request.is_ajax():
        if request.method == 'GET':            
            request.session["old_meter"]    = old_meter    = request.GET.get('old_meter')
            request.session["new_meter"]    = new_meter   = request.GET.get('new_meter')
            if (not old_meter or old_meter==None or new_meter==None or not new_meter):
                change_meter_status="Заполните обе ячейки"
            else:
                change_meter_status=ChangeMeters_v2(old_meter, new_meter)
                
    #print 'old_meter, new_meter', old_meter, new_meter
    if old_meter is None or new_meter is None or change_meter_status.find('Счётчик заменён')>-1:
        old_meter=' '
        new_meter=' '
    args["change_meter_status"]=change_meter_status
    args["old_meter"] = old_meter
    args["new_meter"] = new_meter

    return render(request,"service/service_change_electric.html", args)


def isInt(s):
    try:
        int(s)
        return True
    except ValueError:
        return False

def rename_taken_params_by_guid(guid_meter, old_met, new_met):
    #print 'rename taken params'
    try:  
        common_sql.update_table_with_replace('taken_params', 'name', 'guid_meters', guid_meter, old_met, new_met)
    except Meters.DoesNotExist:
        return False
    try:
    #переименовываем link_abonents_taken_params
        for row in TakenParams.objects.filter(guid_meters=guid_meter):
            guid_taken_params= row.guid
            common_sql.update_table_with_replace('link_abonents_taken_params', 'name', 'guid_taken_params', guid_taken_params, old_met, new_met)
    except TakenParams.DoesNotExist:
        return False


def ChangeMeters_v2(old_meter, new_meter):
    result=""
    # Проверяем существуют ли такие счётчики, в норме первый должен быть, а второй нет
    isExistOldMeter=SimpleCheckIfExist('meters','factory_number_manual',old_meter,"","","")
    isExistNewMeter=SimpleCheckIfExist('meters','factory_number_manual',new_meter,"","","")
    if not isExistOldMeter:
        return "Замена невозможна. Номера старого счётчика нет в базе"
    if isExistNewMeter:
        return "Замена невозможна. Новый счётчик уже существует в базе"
    
    if not(isInt(old_meter)) or not(isInt(new_meter)):
        return 'Замена невозможна. Номера счётчиков должны быть числами'
    #print 'old_meter, new_meter', old_meter, new_meter
    old_met_obj=Meters.objects.filter(factory_number_manual=old_meter)    
    #Просто меняем meters 
    new_name = str(old_met_obj[0].name).replace(str(old_meter), str(new_meter))
    new_num = str(old_met_obj[0].factory_number_manual).replace(str(old_meter), str(new_meter))
    #делаем проверку,если сетевой равен заводскому, то меняем, иначе не трогаем
    new_address = old_met_obj[0].address
    #print new_address, 'new_address'
    #print old_met_obj[0].address, 'old_met_obj[0].address'
    #print old_met_obj[0].factory_number_manual, 'old_met_obj[0].factory_number_manual'
    if str(old_met_obj[0].address) == str(old_met_obj[0].factory_number_manual):
                
        new_address = str(old_met_obj[0].address).replace(str(old_meter), str(new_meter))
        #print new_address, 'new_address'     
    
    #print old_met_obj.values()
    #print 'old_met_obj.guid', old_met_obj[0].guid
    
    rename_taken_params_by_guid(old_met_obj[0].guid, old_meter, new_meter)
    old_met_obj.update(name=new_name, factory_number_manual = new_num, address = new_address)
    
    result = 'Счётчик заменён, если необходимо изменить сетевой адрес сделайте это через панель администратора'
    
    return result


def replace_electric_meters_v2(request):
    args={}

    meter1=''
    meter2=''
    change_meter_status=""
    replace_meter_status='НЕ удалось поменять счётчики местами'
    if request.is_ajax():
        if request.method == 'GET':                        
            request.session["meter1"]    = meter1    = request.GET.get('meter1')
            request.session["meter2"]    = meter2   = request.GET.get('meter2')
            
            if (not meter1 or meter1==None or meter2==None or not meter2):
                replace_meter_status="Заполните обе ячейки"
            else:                
                replace_meter_status=ReplaceMeters_v2(meter1, meter2)
    
    if meter1 is None or meter2 is None or replace_meter_status.find('успешно')>-1:
        meter1=' '
        meter2=' '
        
    args["change_meter_status"]=change_meter_status
    args["replace_meter_status"]=replace_meter_status
    args["meter1"]=meter1
    args["meter2"]=meter2
    return render(request,"service/service_change_electric.html", args)


def ReplaceMeters_v2(meter1, meter2):
    result=''
    
    isExistOldMeter=SimpleCheckIfExist('meters','factory_number_manual',meter1,"","","")
    isExistNewMeter=SimpleCheckIfExist('meters','factory_number_manual',meter2,"","","")
    if not isExistOldMeter:
        return "Замена невозможна. Номера первого счётчика нет в базе"
    if not isExistNewMeter:
        return "Замена невозможна. Номера второго счётчика нет в базе"
    
    if not(isInt(meter1)) or not(isInt(meter2)):
        return 'Замена невозможна. Номера счётчиков должны быть числами'
    
    obj1 = Meters.objects.filter(factory_number_manual = meter1)
    obj2 = Meters.objects.filter(factory_number_manual = meter2)
    guid_type1 = obj1[0].guid_types_meters
    guid_type2 = obj2[0].guid_types_meters
    if not(guid_type1 == guid_type2):
        return 'Замена невозможна. Счётчики должны быть одного типа'
    
    guid_meter1 = obj1[0].guid
    guid_meter2 = obj2[0].guid

    dt_ip1=GetSimpleTable('link_meters_tcpip_settings','guid_meters', guid_meter1)
    dt_ip2=GetSimpleTable('link_meters_tcpip_settings','guid_meters', guid_meter2)
    if len(dt_ip1)>0 and len(dt_ip2)>0:
        if dt_ip1[0][2]!=dt_ip2[0][2]:
            result = result + 'Счётчики принадлежат разным подсетям (!) '
    
    #сделать запрос на update с replace в link_abonents_taken_params
    # надо менять привязку к абоненту и имя привязки

    dt_link1 = common_sql.get_link_abonents_taken_params_by_meter_guid(guid_meter1)
    dt_link2 = common_sql.get_link_abonents_taken_params_by_meter_guid(guid_meter2)
    
    guid_abonent1 = dt_link1[0][0]
    guid_abonent2 = dt_link2[0][0]
    
    for row in dt_link1:
        #print 'row[2], meter1, meter2', row[2], meter1, meter2
        common_sql.update_table_with_replace('link_abonents_taken_params', 'name', 'guid', row[2], meter2, meter1)
        #print row[2], guid_abonent1, guid_abonent2
        common_sql.update_table_with_replace_guid('link_abonents_taken_params', 'guid_abonents', 'guid', row[2], guid_abonent1, guid_abonent2)
        #print 'taken_params', 'name', 'guid', row[6], meter1, meter2
        common_sql.update_table_with_replace('taken_params', 'name', 'guid', row[6], meter2, meter1)
       
    for row in dt_link2:
        common_sql.update_table_with_replace('link_abonents_taken_params', 'name', 'guid', row[2], meter1, meter2)
        common_sql.update_table_with_replace_guid('link_abonents_taken_params', 'guid_abonents', 'guid', row[2], guid_abonent2, guid_abonent1)
        common_sql.update_table_with_replace('taken_params', 'name', 'guid', row[6], meter1, meter2)
    
    result = result + ' Привязки счётчиков успешно изменены '
    return result


def get_file(name_file):
    from django.contrib.staticfiles import finders
    result_url = finders.find('%s'%(name_file))
    if result_url == None:
        response = HttpResponse('Образец не найден', content_type="text/plain")
        output_name = 'empty'
        file_ext = 'txt'    
        response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)  
        return response
    else:
        with open(result_url, 'rb') as f:
            response = HttpResponse(f.read(), content_type="application/vnd.ms-excel")
            output_name = name_file
            #file_ext = u'xlsx'    
            response['Content-Disposition'] = 'attachment;filename="%s"' % (output_name.replace('"', '\"'))
            return response
        f.close() 


def get_electric_template(request):
    return get_file('electric_template_for_load.xlsx')
    
def get_heat_template(request):
    return get_file('heat_template_for_load.xlsx')

def get_water_digital_template(request):
    return get_file('water_digital_template_for_load.xlsx')

def get_water_impulse_template(request):
    return get_file('water_impulse_template_for_load.xlsx')

def get_balance_template(request):
    return get_file('balance_template_for_load.xlsx')

def service_load30_page(request):    
    args={}    
    return render(request,"service/service_30.html", args)

def get_users_account_template(request):
    return get_file('users_account_template.xlsx')

#_____________________________________________________________________________________
#Прогрузка получасовок
#________________________________________________________
con = connection
setev = 0
zavod = 0
col=0
row=0
dt=[]

class NumberParser(HTMLParser):
    def __init__(self):
        HTMLParser.__init__(self)
        self.in_h2 = False

    def handle_starttag(self, tag, attrs):
        if tag == 'h2':
            self.in_h2 = True

    def handle_data(self, data):
        if self.in_h2:
            global setev 
            global zavod
            setev_start_pos = data.find('-') + 2
            setev_end_pos = data.find(',') 
            setev = str(data[setev_start_pos:setev_end_pos])
            zavod_start_pos = data.rfind('-') + 2
            zavod_end_pos = data.rfind(')') 
            zavod = str(data[zavod_start_pos:zavod_end_pos])
            #print u'setevoy: '+setev
            #print u'zavodskoy: '+zavod
            #print data

    def handle_endtag(self, tag):
        self.in_h2 = False

class TableParser(HTMLParser):
    def __init__(self):
        HTMLParser.__init__(self)
        self.in_td = False
     
    def handle_starttag(self, tag, attrs):
        if tag == 'td':
            self.in_td = True
     
    def handle_data(self, data):
        if self.in_td:
            global col
            global row
            global dt
            #
            # print unicode(row),unicode(col)
            if row > 1487 : return
            dt[row][col]=data
            if col==9:
                col = 0
                row+=1
            else:  col+=1
            
            #print data
     
    def handle_endtag(self, tag):
        self.in_td = False
 
def Insert(id_var, date,time, param,param_id):
    curs=con.cursor()
    sQueryInsert="""
    INSERT INTO various_values(
            id, date, "time", value, status, id_taken_params)
    VALUES (%s, '%s', '%s', %s, %s, %s);
        """%(id_var, date, time, param, True ,param_id)
    #print sQueryInsert
    curs.execute(sQueryInsert)
    con.commit()

def checkIsExist(date,time, taken_param,taken_param_id):
    isExist = False
    curs=con.cursor()
    sQuery = """
  SELECT id, date, "time", value, status, id_taken_params
  FROM various_values
  WHERE  date = '%s'
  AND  time = '%s'
  AND id_taken_params = '%s'
"""%(date, time, taken_param_id)
    #print sQuery
    curs.execute(sQuery)
    temp =  curs.fetchall()
    if len(temp)>0:
        isExist = True
    curs.close()
    return isExist

def checkIsSameNumbers(address, zav):
    isSame=True
    curs=con.cursor()
    sQuery ="""SELECT 
    meters.name, 
    meters.address, 
    meters.factory_number_manual
    FROM 
    public.meters
    WHERE 
    meters.address = '%s' AND 
    meters.factory_number_manual = '%s' """%(address, zav)
    
    curs.execute(sQuery)
    temp =  curs.fetchall()
    if len(temp) == 0:
        isSame=False
    curs.close()
    return isSame

def get_id_taken_param_for_meter_by_guid_params(zav, guid):
    result = 0
    curs=con.cursor()
    sQuery = """
    SELECT id
    FROM taken_params
    where name like '%%%s%%'
    and guid_params = '%s'
    """%(zav, guid)
    curs.execute(sQuery)
    temp =  curs.fetchall()
    result = temp[0][0]
    curs.close()
    return result

# @login_required(login_url='/auth/login/') 
# @user_passes_test(isAdmin, login_url='/auth/login/')
def load_30_in_db(f_path,file_list):
    result = ""
    for f_name in file_list:
        path = f_path+'\\'+f_name
        #print path
        with io.open(path, "r",  encoding="cp1251") as profil_file:
            data = profil_file.read().replace('\n', '')
        
        global setev
        global zavod
        global col
        global row
        global dt

        setev = 0
        zavod = 0
        col=0
        row=0
        dt=[]
        n = NumberParser()
        n.feed(data)
              
        # проверяем соответствие заводского и сетевого в базе и в прогружаемом файле
        if not checkIsSameNumbers(setev, zavod):
            result += ' Сетевой и заводской номер не соотвествуют номерам в БАЗЕ. Файл не загружен: ' + str(f_name)
            continue
        
        #инициируем пустой список из 1488 строк и 9 столбцов
        for i in range(1488):
            dt.append([0] * 10)

        p = TableParser()
        p.feed(data)
        
        #получаем taken_param activ&reactiv
        activ_id = get_id_taken_param_for_meter_by_guid_params(zavod, '6af9ddce-437a-4e07-bd70-6cf9dcc10b31')
        reactiv_id = get_id_taken_param_for_meter_by_guid_params(zavod, '66e997c0-8128-40a7-ae65-7e8993fbea61')
       
       #получаем последний id_varius_values, чтобы начать присваивать новые id после него
        curs=con.cursor()
        sQuery ="""SELECT id
        FROM various_values
        order by id DESC
        Limit 1"""
        curs.execute(sQuery)
        dt_var = curs.fetchall()
        curs.close()
        #print sQuery
        if dt_var == []:
            id_var=0
        else:
            id_var = int(dt_var[0][0])
        
        counter = 0
        for dt_row in dt:    
            activ = float(dt_row[1])*0.5
            reactiv = float(dt_row[3])*0.5
            time = dt_row[5]
            date = dt_row[6]
            problem = dt_row[8]
            if problem == '-':
                if not(checkIsExist(date,time, activ,activ_id)):
                    id_var +=1
                    Insert(id_var, date,time, activ,activ_id)
                    counter+=1

                if not(checkIsExist(date,time, reactiv,reactiv_id)):
                    id_var +=1
                    Insert(id_var, date,time, reactiv, reactiv_id)
                    counter+=1
                
            else: continue

        result += ' В таблицу various_values  загружено: '+str(counter) + ' новых строк для файла ' + str(f_name) + '.'
        curs.close()
    
    return result


def service_load30(request):
    args={}
    data_table=[]
    status='Не удаётся найти указанный путь'
        
    result=""
    file30 =""
    if request.is_ajax():
        if request.method == 'GET':            
            request.session["file30"]    = file30    = request.GET['file30']
            #print file30
            file_list=[]            
            isExistDir = os.path.exists(file30)
            if isExistDir:
                #'извлекаем файлы и загружаем в БД'
                status = 'Директория существует'
                #получаем список html-файлов в указанной директории
                for f in os.listdir(file30):
                    if f.endswith(".html"):
                        file_list.append(f)
                f_count = len(file_list)
                status += ', в ней находится '+ str(f_count) + ' html-файлов. '
                if f_count>0:
                    status += load_30_in_db(file30,file_list)
    args['data_table'] = data_table
    args['status']=status
    return render(request,"service/service_30.html", args)

def service_user_account(request):
    args={}
    return render(request,"service/service_users_account.html", args)

def load_user_account(request):
    args={}
    fileName=""
    sheet    = ""    
    result = ""
    try:    
        if request.is_ajax():
            if request.method == 'GET':            
                request.session["choice_file"]    = fileName    = request.GET['choice_file']
                request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']
                
                directory=os.path.join(BASE_DIR,'static/cfg/')
                sPath=directory+fileName
                result = "Не загружено"
                result = load_users_account(sPath, sheet)
    except: # catch *all* exceptions
        e = sys.exc_info()[0]
        result = ( "Ошибка: %s" % e )
    args["result"]    = result
    return render(request,"service/service_users_account.html", args)

def create_user(login, u_pass, u_mail, u_last_name, u_name):
    result = ''
    guid_user = ''
    user_new = True
    try:    
        #проверяем, существует ли такой пользователь, если нет- вернётся None        
        user = authenticate(username=login, password=u_pass)      
        if user is not None:            
            #пользователь новый или пара логин-пароль не совпадет
            user_new = False
        else:
            #создаём пользователя
            #print login, u_pass, u_mail, u_last_name, u_name
            user = User.objects.create_user(password = u_pass, username=login, first_name = u_name, last_name = u_last_name, email = u_mail)
            user.save()
            #print 'new user created:', user
    
    except: # catch *all* exceptions
        e = sys.exc_info()[0]
        result = ( "Ошибка создания пользователя %s: %s" %(u_name, e) )
    return result, user, user_new


def create_link_user_abonent(user, obj, abon):
    result_link = ''
    is_new_link = False
    try:
        #проверяем существует ли такой абонент
        #print 'check abonent', obj, abon
        dt=[]
        dt = GetSimpleCrossTable('objects', 'name', obj, 'abonents', 'name', abon)
        #print 'dt', dt, type(dt)
        if len(dt) > 0:
            #абонент существует
            #первые 4 поля (0,1,2,3) - это поля objects, далее по порядку идут поля abonents
            guid_abon = dt[0][4]
            id_user = user.id
            name = str(user.last_name) + ' - ' +str(abon)
            #print name, guid_abon, id_user
            #проверяем есть ли уже такая привязка:    
            dt_check_link = []        
            dt_check_link = LinkAbonentsAuthUser.objects.filter(guid_abonents = guid_abon).filter(guid_auth_user = id_user)
            #print dt_check_link, 'dt_check_link'
            if len(dt_check_link) < 1:
                link = LinkAbonentsAuthUser(name = name, guid_abonents = Abonents.objects.get(guid= guid_abon), guid_auth_user = user)
                link.save()
                is_new_link = True
            else: result_link += ' Связь ''%s'' уже существует. '%(name)
        else:
            #print u' Не существует: %s -> %s '%(obj, abon)
            result_link += ' Не существует: %s -> %s. '%(obj, abon)
    except: # catch *all* exceptions
        e = sys.exc_info()[0]
        result = ( "Ошибка: %s" % e )
        #print result_link
    return result_link, is_new_link


def load_users_account(sPath, sheet):
    result = ''
    dtAll = GetTableFromExcel(sPath,sheet) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    user_count = 0 #счётчик кол-ва добавленных пользователей
    link_count = 0
    for i in range(2,len(dtAll)):
        obj = dtAll[i][0] #корпус
        abon = dtAll[i][1] #квартира
        login = dtAll[i][2] # логин
        u_last_name = dtAll[i][4] #фамилия
        u_name = dtAll[i][5] #имя 
        u_pass = dtAll[i][3] #пароль
        u_mail = dtAll[i][6] #мейл
        if login is None:
            result += 'Строка %s не загружена: Отсутствует логин. '%(str(i+3))
            continue
        if u_pass is None:
            result += 'Строка %s не загружена: Отсутствует пароль. '%(str(i+3))
            continue
        if abon is None or obj is None:
            result += 'Строка %s не загружена: Отсутствует объект и/или абонент . '%(str(i+3))
            continue

        #если пользователь новый/существует, то вернёт oбъект user
        result_create = ''
        is_new_user = False
        result_create, user, is_new_user = create_user(login, u_pass, u_mail, u_last_name, u_name)
        if len(result_create)>0: 
            result +=result_create
            continue
        if is_new_user:
            user_count +=1
        
        #связываем абонента(картира или имп.счётчик) и пользователя
        result_link = ''
        result_link, is_new_link = create_link_user_abonent(user, obj, abon) 
        if len(result_link) > 0:
            result += result_link
            continue
        if is_new_link:
            link_count +=1
    result += 'Добавлено пользователей: %s. Добавлено привязок пользователей к ПУ: %s' %(str(user_count), str(link_count))
    return result

def service_del_meters(request):
    args ={}
    return render(request,"service/service_del_meters.html", args)

def get_80020_template(request):
    return get_file('80020_template_for_load.xlsx')


def load_80020_group(request):
    args={}
    fileName=""
    sheet    = ""
    result = []
    #writeToLog('test1') 
    try:    
        if request.is_ajax():
            if request.method == 'GET':            
                request.session["choice_file"]    = fileName    = request.GET['choice_file']
                request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']

                directory=os.path.join(BASE_DIR,'static/cfg/')
                sPath=directory+fileName
                writeToLog(sPath)
                            
                result = make_80020_report(sPath, sheet)
    except: # catch *all* exceptions
        e = sys.exc_info()[0]
        result.append( "Ошибка: %s" % e )

    #print(result)
    args["choice_file"]  = fileName
    args["choice_sheet"] = sheet
    args["80020_status"] = result
    return render(request,"service/service_electric.html", args)

def make_80020_report(sPath, sSheet):
    global cfg_excel_name
    cfg_excel_name=sPath
    global cfg_sheet_name
    cfg_sheet_name = sSheet
    result = []
    dtAll=GetTableFromExcel(sPath,sSheet) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    i = 0
    for row in dtAll:
        i+=1
        if i<3: continue
        group_name = str(row[0]).strip()
        contract_number = str(row[1]).strip()
        measuringpoint_code = str(row[2]).strip()
        measuringpoint_name = str(row[3]).strip()
        meter_number = str(row[4]).strip()
        inn_sender = str(row[5]).strip()
        name_sender = str(row[6]).strip()
        area_inn = str(row[7]).strip()
        abonent_name =  name_sender #str(row[8]).strip()
        
        guid_groups_80020 = ''
        dt = common_sql.get_80020_group_by_name(group_name)
        if len(dt) > 0:
            #проверяем данные по группе, если она уже существует
            #SELECT guid, name, name_sender, inn_sender, name_postavshik, inn_postavshik, dogovor_number
            guid_groups_80020 = dt[0][0]           
            if name_sender != str(dt[0][2]): result.append('Группа ''%s'' уже существует. Строка %s: Не совпадает название организации.'%(group_name,str(i)) )
            if inn_sender != str(dt[0][3]): result.append( ' Группа ''%s'' уже существует. Строка %s: Не совпадает ИНН организации.'%(group_name,str(i)) )          
            if area_inn != str(dt[0][5]): result.append(' Группа ''%s'' уже существует. Строка %s: Не совпадает Идентификатор предоставляемый АТС.'%(group_name,str(i))  )          
            if abonent_name!= str(dt[0][4]): result.append(' Группа ''%s'' уже существует. Строка %s: Не совпадает название организации-поставщика.'%(group_name,str(i)) )
            if contract_number.strip() != str(dt[0][6]).strip(): result.append( ' Группа ''%s'' уже существует. Строка %s: Не совпадает номер договора.'%(group_name,str(i)) )
            
        else:
            #создаём новую группу
            #Groups80020, LinkGroups80020Meters
            group = Groups80020(name = group_name, name_sender = name_sender, inn_sender=inn_sender, name_postavshik=abonent_name, inn_postavshik=area_inn, dogovor_number=contract_number)
            group.save()
            guid_groups_80020 = group.guid #Objects.objects.get(guid=guid_parent)
            result.append( 'Содана группа %s .'%(group_name) )
        #проверяем существует ли счётчик
        is_meter_exist = SimpleCheckIfExist('meters','factory_number_manual', meter_number, table2='', fieldName2='', value2='')
        if is_meter_exist:
            #проверяем есть ли уже связь счётчика с этой группой
            #dt = GetSimpleCrossTable('groups_80020',fieldName1,value1,table2,fieldName2, value2):
            guid_meter = Meters.objects.get(factory_number_manual = meter_number).guid
            filter_groups = LinkGroups80020Meters.objects.filter(guid_groups_80020 = guid_groups_80020).filter(guid_meters = guid_meter)
            if filter_groups.count() > 0:
                continue
            else:
                #создаём связь
                link = LinkGroups80020Meters(measuringpoint_code = measuringpoint_code, measuringpoint_name = measuringpoint_name, guid_groups_80020 = Groups80020.objects.get(guid= guid_groups_80020), guid_meters = Meters.objects.get(guid = guid_meter))
                link.save()
                result.append(' Счётчик %s добавлен к группе ''%s''. '%(meter_number, group_name) )
        else:
            result.append( ' Строка %s: Прибора с номером %s не существует в БД, он не был добавлен к группе.'%(str(i),meter_number) )
            
    return result

def del_meters(request):
    args={}
    fileName=""
    sheet    = ""
    result = []
    #writeToLog('test1') 
    try:    
        if request.is_ajax():
            if request.method == 'GET':            
                request.session["choice_file"]     = fileName    = request.GET['choice_file']
                request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']

                directory=os.path.join(BASE_DIR,'static/cfg/')
                sPath=directory+fileName
                result = delete_meters_by_excel(sPath, sheet)
    except: # catch *all* exceptions
        e = sys.exc_info()[0]
        result.append( "Ошибка: %s" % e )

    #print(result)
    args["choice_file"]  = fileName
    args["choice_sheet"] = sheet
    args["del_status"] = result
    return render(request,"service/service_del_meters.html", args)

def delete_meters_by_excel(sPath, sheet):
    global cfg_excel_name
    cfg_excel_name=sPath
    global cfg_sheet_name
    cfg_sheet_name = sheet
    result = []
    dtAll = GetTableFromExcel(sPath,sheet) #получили из excel все строки до первой пустой строки (проверка по колонке А)

    isImpulse = False
    #выясняем какая ведомость импульсная или цифровая, в зависимости от этого - из какой колонки брать номера счётчиков
    if dtAll[0][0] == 'Населенный пункт':
        isImpulse = False
        result.append('Удаление цифровых ПУ')
        m_col = 6
        c = 1
    if dtAll[1][0] == 'Наименование дома':
        isImpulse = True
        result.append('Удаление импульсных ПУ')
        m_col = 5
        c = 2
    i = 0
    for row in dtAll:
        if i<c: 
            i+=1
            continue
        meter = row[m_col]
        #print(meter, type(meter))
        try:
            del_meter = Meters.objects.get(factory_number_manual = str(meter))
            del_meter.delete()
            print('del - good')
            result.append('Удалён ПУ: {}'.format(meter))
        except ObjectDoesNotExist:
            result.append('НЕ найден: {}'.format(meter))
        i+=1
        #в цикле 
        #если существует счётчик, то удаляем каскадно
        #если прибора нет, то доабвляем запись об этом у result
    return result

def del_various30(request):
    args={}
    result = []
    date_del30 = ""
    meter = ""
   
    try:    
        meter       = request.GET.get('num_del30')          
        date_del30  = request.GET.get('date_del30')        
        
        # проверяем дату!
        print(date_del30)
        if len(date_del30)>0:
            #удаление
            if len(meter)>0:
                print(len(meter))
                common_sql.del_various_values_by_factory_number_by_date(meter, date_del30)
                result.append('Получасовые показания удалены для прибора: %s за %s'%(meter, date_del30))
            else:
                common_sql.del_various_values_by_date(date_del30)
                result.append('Получасовые показания удалены для всех приборов за %s'%(date_del30))
                #удаляем все получасовки на дату
        else:
            result.append('Удаление НЕ выполнено. Выберите дату.')

    except: # catch *all* exceptions
        e = sys.exc_info()[0]
        result.append( "Ошибка: %s" % e )

    #print(result)    
    args["del30_status"] = result[0]
    return render(request,"service/service_30.html", args)