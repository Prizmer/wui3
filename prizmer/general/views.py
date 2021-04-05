# coding -*- coding: utf-8 -*-

from django.contrib.auth.decorators import login_required
from django.contrib import auth
from django.shortcuts import  HttpResponse
from django.shortcuts import render
import simplejson as json
from django.db.models import Max
from django.db import connection
import re
#from excel_response import ExcelResponse
import datetime
import decimal

from django.db.models.signals import post_save
from django.db.models import signals
from django.db.models import Q

from django.contrib.auth.decorators import user_passes_test
from functools import wraps
import pytz
#---------
import calendar
import common_sql
from django.shortcuts import redirect

#---------

from general.models import Objects, Abonents, BalanceGroups, Meters, LinkBalanceGroupsMeters, Comments, Resources
from django import forms
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings

import time
import string



 
def dictfetchall(cursor):
#"Returns all rows from a cursor as a dict"
    desc = cursor.description
    return [
        dict(list(zip([col[0] for col in desc], row)))
        for row in cursor.fetchall()
    ]
    
#!!!!!!!!!!!!
def simple_query(): # Пример запроса в БД на чистом SQL
    simpleq = connection.cursor()
    simpleq.execute("""SELECT 
                                 monthly_values.id, 
                                 monthly_values.date
                               FROM 
                                 public.monthly_values;""")
    simpleq = simpleq.fetchall()
    return simpleq
#!!!!!!!!!!!!    


#------------------------------------------------------------------------------------------------------------------------



    

# Отчет по СПГ на начало суток
def get_data_table_by_date_spg(obj_title, obj_parent_title, electric_data):
    data_table = []
    
    my_parametr = "Время работы узла учёта"    
    data_table_time = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, "СПГ762__")
    
    my_parametr = "Время работы при ненулевом расходе"    
    data_table_time_rashod = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, "СПГ762__")
    
    my_parametr = "Атмосферное давление"    
    data_table_p_atm = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, "СПГ762__")

    my_parametr = "Температура наружного воздуха"    
    data_table_temp_air = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, "СПГ762__")
    
    my_parametr = "Значение времени интегрирования"    
    data_table_time_integr = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, "СПГ762__")
    
    my_parametr = "Среднее значение расхода газа"    
    data_table_sr_rashod = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, "СПГ762__")
    
    my_parametr = "Среднее значение температуры газа"    
    data_table_sr_temp_air = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, "СПГ762__")
    
    my_parametr = "Среднее значение абсолютного давления"    
    data_table_sr_abs_p = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, "СПГ762__")
    
    my_parametr = "Среднее значение с доп. датчика 1"    
    data_table_sr_dop_1 = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, "СПГ762__")
    
    my_parametr = "Среднее значение с доп. датчика 2"    
    data_table_sr_dop_2 = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, "СПГ762__")
    
    my_parametr = "Масса газа при стандартных условиях"    
    data_table_m_gas_standart = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, "СПГ762__")
    
    my_parametr = "Объем газа при стандартных условиях"    
    data_table_v_gas_standart = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, "СПГ762__")
    
    my_parametr = "Объем газа при рабочих условиях"    
    data_table_v_work = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, "СПГ762__")
    
    my_parametr = "Обобщённое сообщение о нештатных ситуациях"    
    data_table_err_common = common_sql.get_data_table_parametr_by_date_daily(obj_title, obj_parent_title, electric_data, my_parametr, "СПГ762__")
    
              
    for x in range(len(data_table_time)):
        data_table_temp = []
        try:
            data_table_temp.append(data_table_time[x][0]) # дата
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_time[x][2]) # имя абонента
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_time[x][3]) # заводской номер
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_time[x][4]) # Время работы узла учёта
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_time_rashod[x][4]) # Время работы при ненулевом расходе
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_p_atm[x][4]) # Атмосферное давление
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")            
        try:
            data_table_temp.append(data_table_temp_air[x][4]) # Температура наружного воздуха
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")            
        try:
            data_table_temp.append(data_table_time_integr[x][4]) # Значение времени интегрирования
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")            
        try:
            data_table_temp.append(data_table_sr_rashod[x][4]) # Среднее значение расхода газа
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_sr_temp_air[x][4]) # Среднее значение температуры газа
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_sr_abs_p[x][4]) # Среднее значение абсолютного давления
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")            
        try:
            data_table_temp.append(data_table_sr_dop_1[x][4]) # Среднее значение с доп. датчика 1
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")            
        try:
            data_table_temp.append(data_table_sr_dop_2[x][4]) # Среднее значение с доп. датчика 2
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_m_gas_standart[x][4]) # Масса газа при стандартных условиях
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_v_gas_standart[x][4]) # Объем газа при стандартных условиях
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")            
        try:
            data_table_temp.append(data_table_v_work[x][4]) # Объем газа при рабочих условиях
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")            
        try:
            data_table_temp.append(data_table_err_common[x][4]) # Обобщённое сообщение о нештатных ситуациях
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
            

        data_table.append(data_table_temp)
    return data_table


    

#------------    


#-------------------------------------------------------------------------------------------------------------------------
    
def get_data_table_by_date_monthly(obj_title, obj_parent_title, electric_data):
    data_table = []
    my_parametr = "T0 A+"
    cursor_t0_aplus = connection.cursor()
    cursor_t0_aplus.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t0_aplus = cursor_t0_aplus.fetchall()
    
    my_parametr = "T1 A+"            
    cursor_t1_aplus = connection.cursor()
    cursor_t1_aplus.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t1_aplus = cursor_t1_aplus.fetchall()

    my_parametr = "T2 A+"                
    cursor_t2_aplus = connection.cursor()
    cursor_t2_aplus.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t2_aplus = cursor_t2_aplus.fetchall()
                
    my_parametr = "T3 A+"
    cursor_t3_aplus = connection.cursor()
    cursor_t3_aplus.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t3_aplus = cursor_t3_aplus.fetchall()

    my_parametr = "T4 A+"                
    cursor_t4_aplus = connection.cursor()
    cursor_t4_aplus.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t4_aplus = cursor_t4_aplus.fetchall()
    
    my_parametr = "T0 R+"
    cursor_t0_rplus = connection.cursor()
    cursor_t0_rplus.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t0_rplus = cursor_t0_rplus.fetchall()
                
    for x in range(len(data_table_t0_aplus)):
        data_table_temp = []
        try:
            data_table_temp.append(data_table_t0_aplus[x][0])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_t0_aplus[x][2])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_t0_aplus[x][6])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_t0_aplus[x][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_t1_aplus[x][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_t2_aplus[x][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_t3_aplus[x][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_t4_aplus[x][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_t0_rplus[x][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        data_table.append(data_table_temp) 
    
    return data_table
#!!!!!!!!!!!!
    
def get_data_table_by_date_daily(obj_title, obj_parent_title, electric_data):
    data_table = []
    my_parametr = "T0 A+"
    cursor_t0_aplus = connection.cursor()
    cursor_t0_aplus.execute("""SELECT 
                                daily_values.date, 
                                daily_values.value, 
                                abonents.name, 
                                daily_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.daily_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = daily_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                daily_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t0_aplus = cursor_t0_aplus.fetchall()
    
    my_parametr = "T1 A+"            
    cursor_t1_aplus = connection.cursor()
    cursor_t1_aplus.execute("""SELECT 
                                daily_values.date, 
                                daily_values.value, 
                                abonents.name, 
                                daily_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.daily_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = daily_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                daily_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t1_aplus = cursor_t1_aplus.fetchall()

    my_parametr = "T2 A+"                
    cursor_t2_aplus = connection.cursor()
    cursor_t2_aplus.execute("""SELECT 
                                daily_values.date, 
                                daily_values.value, 
                                abonents.name, 
                                daily_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.daily_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = daily_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                daily_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t2_aplus = cursor_t2_aplus.fetchall()
                
    my_parametr = "T3 A+"
    cursor_t3_aplus = connection.cursor()
    cursor_t3_aplus.execute("""SELECT 
                                daily_values.date, 
                                daily_values.value, 
                                abonents.name, 
                                daily_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.daily_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = daily_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                daily_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t3_aplus = cursor_t3_aplus.fetchall()

    my_parametr = "T4 A+"                
    cursor_t4_aplus = connection.cursor()
    cursor_t4_aplus.execute("""SELECT 
                                daily_values.date, 
                                daily_values.value, 
                                abonents.name, 
                                daily_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.daily_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = daily_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                daily_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t4_aplus = cursor_t4_aplus.fetchall()
    
    my_parametr = "T0 R+"
    cursor_t0_rplus = connection.cursor()
    cursor_t0_rplus.execute("""SELECT 
                                daily_values.date, 
                                daily_values.value, 
                                abonents.name, 
                                daily_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.daily_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = daily_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = %s AND 
                                daily_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[obj_title, obj_parent_title, my_parametr, electric_data])
    data_table_t0_rplus = cursor_t0_rplus.fetchall()
                
    for x in range(len(data_table_t0_aplus)):
        data_table_temp = []
        try:
            data_table_temp.append(data_table_t0_aplus[x][0])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_t0_aplus[x][2])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_t0_aplus[x][6])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_t0_aplus[x][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_t1_aplus[x][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_t2_aplus[x][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_t3_aplus[x][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        try:
            data_table_temp.append(data_table_t4_aplus[x][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        
        try:
            data_table_temp.append(data_table_t0_rplus[x][1])
        except IndexError:
            data_table_temp.append("Н/Д")
        except TypeError:
            data_table_temp.append("Н/Д")
        data_table.append(data_table_temp)
    
    return data_table

#my_decorators
def isStaff(user):
    return user.is_staff




# Create your views here.
@csrf_protect
@user_passes_test(isStaff, login_url='/account/')
@login_required(login_url='/auth/login/')
def default(request):
    args={}
    #-------------- get data new tree
    max_level = Objects.objects.aggregate(Max('level'))['level__max'] # Max number of levels
    if max_level < 3:
        all_level_0 = Objects.objects.filter(level=0)
        tree_data = []
        for l0 in range(len(all_level_0)):
            filter_level_1 = Objects.objects.filter(level=1).filter(guid_parent = all_level_0[l0].guid).order_by('name')
            children_data_l1 = []
            for l1 in range(len(filter_level_1)):
                children_data_l2 = []
                filter_level_2 = Objects.objects.filter(level=2).filter(guid_parent = filter_level_1[l1].guid).order_by('name')
                for l2 in range(len(filter_level_2)):
                    abonents_data = []
                    list_of_level_2 = {"key":"level2-"+str(l2), "title": filter_level_2[l2].name, "children":abonents_data}
                    filter_level_abonents = Abonents.objects.filter(guid_objects = filter_level_2[l2].guid).order_by('name')
                    for l3 in range(len(filter_level_abonents)):
                        meters_data = []
                        cursor = connection.cursor()
                        cursor.execute("""SELECT 
                                      abonents.name, 
                                      meters.name, 
                                      meters.factory_number_manual
                                    FROM 
                                      public.abonents, 
                                      public.meters, 
                                      public.taken_params, 
                                      public.link_abonents_taken_params
                                    WHERE 
                                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                                      meters.guid = taken_params.guid_meters AND
                                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                      abonents.name = %s
                                    GROUP BY
                                      abonents.name,
                                      meters.name, 
                                      meters.factory_number_manual;""", [filter_level_abonents[l3].name])
                        filter_level_meters = dictfetchall(cursor)
                        for meter in range(len(filter_level_meters)):
                           list_of_level_meters = {"key":"meter-"+str(meter), "title": filter_level_meters[meter]['factory_number_manual']}
                           meters_data.append(list_of_level_meters)
                        list_of_level_abonents = {"key":"abonent-"+str(l2), "title": filter_level_abonents[l3].name, "children":meters_data}
                        abonents_data.append(list_of_level_abonents)
                    list_of_level_2 = {"key":"level2-"+str(l2), "title": filter_level_2[l2].name, "children":abonents_data}                     
                    children_data_l2.append(list_of_level_2)             
                list_of_level_1 = {"key":"level1-"+str(l1), "title": filter_level_1[l1].name, "children":children_data_l2, "folder":bool(children_data_l2)}
                children_data_l1.append(list_of_level_1)
            list_of_level_0 = {"key":"level0-"+str(l0), "title": all_level_0[l0].name, "children":children_data_l1, "folder":bool(children_data_l1)}
            tree_data.append(list_of_level_0)
        tree_data_json = json.dumps(tree_data, )
        args['tree_data_json'] = tree_data_json
    else:
        pass
    #-------------- get data new tree end
    args['ico_url_electric'] = "/static/images/electric-ico36.png"
    args['ico_url_water'] = "/static/images/water-ico36.png"
    args['ico_url_heat'] = "/static/images/heat-ico36.png"    
    args['ico_url_gas'] = "/static/images/gas-ico36.png"
    args['ico_url_economic'] = "/static/images/economic-ico36.png"
    return render(request,'base.html', args)

def go_out(request):
    auth.logout(request)

    return redirect(default)

@csrf_protect
@login_required(login_url='/auth/login/') 
def tree_data_json_v2(request):
    args={}

    pre_url = str(request.GET.get('preurl'))
    tree_data_json = '_________'
    #tree_data = '_________'
    if (pre_url.find('electric') > -1):       
        tree_data_json=tree_data_json_electric(request)
    elif (pre_url.find('heat') > -1):   
        tree_data_json=tree_data_json_heat(request)
    elif (pre_url.find('water') > -1):   
        tree_data_json=tree_data_json_water_v2(request)
    else:
        tree_data_json=tree_data_json_all(request)
    #print tree_data_json
    #tree_data_json = tree_data
    args['tree_data_json'] = tree_data_json
    return HttpResponse(tree_data_json)



def del_object_no_children(all_level_0):
    result=all_level_0
    for obj in range(len(all_level_0)):
        children_abons = Abonents.objects.filter(guid_objects=all_level_0[obj].guid)
        if len(children_abons) == 0:
            #print 'del ', all_level_0[obj].name
            result.exclude(guid=all_level_0[obj].guid)  

    return result

def tree_data_json_all(request):
    
    #-------------- get data new tree
    max_level = Objects.objects.aggregate(Max('level'))['level__max'] # Max number of levels
    if max_level < 3:
        all_level_0 = Objects.objects.filter(level=0)
        tree_data = []
        for l0 in range(len(all_level_0)):
            filter_level_1 = Objects.objects.filter(level=1).filter(guid_parent = all_level_0[l0].guid)
            children_data_l1 = []
            for l1 in range(len(filter_level_1)):
                children_data_l2 = []
                filter_level_2 = Objects.objects.filter(level=2).filter(guid_parent = filter_level_1[l1].guid)
                for l2 in range(len(filter_level_2)):
                    abonents_data = []
                    list_of_level_2 = {"key":"level2-"+str(l2), "title": filter_level_2[l2].name, "children":abonents_data}
                    filter_level_abonents = Abonents.objects.filter(guid_objects = filter_level_2[l2].guid).order_by('name')
                    for l3 in range(len(filter_level_abonents)):
                        meters_data = []
                        cursor = connection.cursor()
                        cursor.execute("""SELECT 
                                          abonents.name, 
                                          meters.name, 
                                          meters.factory_number_manual
                                        FROM 
                                          public.abonents, 
                                          public.meters, 
                                          public.taken_params, 
                                          public.link_abonents_taken_params, 
                                          public.objects
                                        WHERE 
                                          abonents.guid = link_abonents_taken_params.guid_abonents AND
                                          meters.guid = taken_params.guid_meters AND
                                          taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                          objects.guid = abonents.guid_objects AND
                                          abonents.name = %s AND 
                                          objects.name = %s
                                        GROUP BY
                                          abonents.name,
                                          meters.name, 
                                          meters.factory_number_manual;""", [filter_level_abonents[l3].name, filter_level_2[l2].name ])
                        filter_level_meters = dictfetchall(cursor)
                        for meter in range(len(filter_level_meters)):
                           list_of_level_meters = {"key":"meter-"+str(meter), "title": filter_level_meters[meter]['factory_number_manual']}
                           meters_data.append(list_of_level_meters)
                        list_of_level_abonents = {"key":"abonent-"+str(l2), "title": filter_level_abonents[l3].name, "children":meters_data}
                        abonents_data.append(list_of_level_abonents)
                    list_of_level_2 = {"key":"level2-"+str(l2), "title": filter_level_2[l2].name, "children":abonents_data}                     
                    children_data_l2.append(list_of_level_2)             
                list_of_level_1 = {"key":"level1-"+str(l1), "title": filter_level_1[l1].name, "children":children_data_l2, "folder":bool(children_data_l2)}
                children_data_l1.append(list_of_level_1)
            list_of_level_0 = {"key":"level0-"+str(l0), "title": all_level_0[l0].name, "children":children_data_l1, "folder":bool(children_data_l1)}
            tree_data.append(list_of_level_0)
            
        # Получаем информацию по балансным группам
            balance_groups_list = []
            simpleq = connection.cursor()
            simpleq.execute(""" SELECT 
                                  balance_groups.name
                                FROM 
                                  public.balance_groups;""")
            simpleq = simpleq.fetchall()
            for x in range (len(simpleq)):
                balance_groups_list.append({"key": "group-"+str(x), "title": simpleq[x][0]})
             
        # Получаем информацию по группам 80020
            groups_80020_list = []
            simpleq = connection.cursor()
            simpleq.execute(""" SELECT 
                                  groups_80020.name
                                FROM 
                                  public.groups_80020;
                                """)
            simpleq = simpleq.fetchall()
            for x in range (len(simpleq)):
                groups_80020_list.append({"key": "group80020-"+str(x), "title": simpleq[x][0]})
        
        tree_data.append({"key": "group" + str(1000), "title": "Группы 80020", "children":groups_80020_list , "folder":bool(True)})
        tree_data.append({"key": "group" + str(1000), "title": "Группы", "children":balance_groups_list , "folder":bool(True)})
        
        # Создаем json данные для дерева объектов
        tree_data_json = json.dumps(tree_data, )        
    else:
        pass
    return HttpResponse(tree_data_json)



def tree_data_json_electric(request):
    
    name_res = 'Электричество'
    #-------------- get data new tree
    max_level = Objects.objects.aggregate(Max('level'))['level__max'] # Max number of levels
    if max_level < 3:
        all_level_0 = Objects.objects.filter(level=0).filter(~Q(name = 'Вода'))      
        tree_data = []
        #all_level_0.exclude(guid='7fc2741d-0d39-4dea-b856-9d4b146181d1')
        #print u'Объекты уровня 0 \n', all_level_0
        for l0 in range(len(all_level_0)):
            #print all_level_0[l0].name
            filter_level_1 = Objects.objects.filter(level=1).filter(guid_parent = all_level_0[l0].guid)
            #print all_level_0[l0].name    #москва      
            children_data_l1 = []
            for l1 in range(len(filter_level_1)):
                """if (filter_level_1[l1].name).find('Вода')>-1:
                    print u"Вода!!!!"
                    continue"""            
                children_data_l2 = []              
                #filter_level_2 = Objects.objects.filter(level=2).filter(guid_parent = filter_level_1[l1].guid)
                #print filter_level_1[l1].name #ботсад
                filter_level_2 = Objects.objects.filter(level=2).filter(guid_parent = filter_level_1[l1].guid) 
                # uroven korpusov
                for l2 in range(len(filter_level_2)):
                    abonents_data = [] 
                    #filter_level_abonents=common_sql.get_electric_abons_by_object_guid(filter_level_2[l2].guid)
                    filter_level_abonents = common_sql.get_abons_by_object_guid_and_res(filter_level_2[l2].guid, name_res)
                    #print unicode(filter_level_2[l2])
                    #print u"###############################"
                    #print filter_level_abonents
                    for abon in range(len(filter_level_abonents)):
                        meters_data = []
                        #print filter_level_abonents[abon][0]
                        #print abon
                        list_of_meters = common_sql.get_meters_by_abons_guid_and_res(filter_level_abonents[abon][1], name_res)
                        for meter in range(len(list_of_meters)):
                            list_of_level_meters = {"key":"meter-"+str(meter), "title": list_of_meters[meter][2]}
                            meters_data.append(list_of_level_meters)
                        list_of_level_abonents = {"key":"abonent-"+str(l2), "title": filter_level_abonents[abon][0], "children":meters_data}
                        abonents_data.append(list_of_level_abonents)
                    if len(filter_level_abonents) > 0:              
                        list_of_level_2 = {"key":"level2-"+str(l1), "title": filter_level_2[l2].name, "children":abonents_data}
                        children_data_l2.append(list_of_level_2)

                list_of_level_1 = {"key":"level1-"+str(l1), "title": filter_level_1[l1].name, "children":children_data_l2, "folder":bool(children_data_l2)}
                children_data_l1.append(list_of_level_1)
            list_of_level_0 = {"key":"level0-"+str(l0), "title": all_level_0[l0].name, "children":children_data_l1, "folder":bool(children_data_l1)}
            tree_data.append(list_of_level_0)

        # Получаем информацию по балансным группам
            balance_groups_list = []
            dt_balanse_group = common_sql.get_balance_groups_by_res(name_res)
            for x in range (len(dt_balanse_group)):
                balance_groups_list.append({"key": "group-"+str(x), "title": dt_balanse_group[x][0]})
             
        # Получаем информацию по группам 80020
            groups_80020_list = []
            simpleq = connection.cursor()
            simpleq.execute(""" SELECT 
                                  groups_80020.name
                                FROM 
                                  public.groups_80020;
                                """)
            simpleq = simpleq.fetchall()
            for x in range (len(simpleq)):
                groups_80020_list.append({"key": "group80020-"+str(x), "title": simpleq[x][0]})
        
        tree_data.append({"key": "group" + str(1000), "title": "Группы 80020", "children":groups_80020_list , "folder":bool(True)})
        tree_data.append({"key": "group" + str(1000), "title": "Группы", "children":balance_groups_list , "folder":bool(True)})
        
        tree_data_json = json.dumps(tree_data,)
        #print tree_data_json
    #else:
    #    pass
    #print tree_data_json
    return tree_data_json 

def tree_data_json_heat(request):
    
    name_res = 'Тепло'
    #-------------- get data new tree
    max_level = Objects.objects.aggregate(Max('level'))['level__max'] # Max number of levels
    if max_level < 3:
        all_level_0 = Objects.objects.filter(level=0).filter(~Q(name = 'Вода'))      
        tree_data = []       
        #print u'Объекты уровня 0 \n', all_level_0
        for l0 in range(len(all_level_0)):
            #print all_level_0[l0].name
            filter_level_1 = Objects.objects.filter(level=1).filter(guid_parent = all_level_0[l0].guid)           
            children_data_l1 = []
            for l1 in range(len(filter_level_1)):                            
                children_data_l2 = []                             
                filter_level_2 = Objects.objects.filter(level=2).filter(guid_parent = filter_level_1[l1].guid)                 
                for l2 in range(len(filter_level_2)):
                    abonents_data = []                     
                    filter_level_abonents = common_sql.get_abons_by_object_guid_and_res(filter_level_2[l2].guid, name_res)
                    #print filter_level_2[l2].name, 'abonentov: ', str(len(filter_level_abonents))
                    
                    for abon in range(len(filter_level_abonents)):
                        meters_data = []                        
                        list_of_meters = common_sql.get_meters_by_abons_guid_and_res(filter_level_abonents[abon][1], name_res)
                        for meter in range(len(list_of_meters)):
                            list_of_level_meters = {"key":"meter-"+str(meter), "title": list_of_meters[meter][2]}
                            meters_data.append(list_of_level_meters)
                        list_of_level_abonents = {"key":"abonent-"+str(l2), "title": filter_level_abonents[abon][0], "children":meters_data}
                        abonents_data.append(list_of_level_abonents)              
                    if len(filter_level_abonents) > 0:
                        list_of_level_2 = {"key":"level2-"+str(l1), "title": filter_level_2[l2].name, "children":abonents_data}
                        children_data_l2.append(list_of_level_2)

                list_of_level_1 = {"key":"level1-"+str(l1), "title": filter_level_1[l1].name, "children":children_data_l2, "folder":bool(children_data_l2)}
                children_data_l1.append(list_of_level_1)
            list_of_level_0 = {"key":"level0-"+str(l0), "title": all_level_0[l0].name, "children":children_data_l1, "folder":bool(children_data_l1)}
            tree_data.append(list_of_level_0)

        # Получаем информацию по балансным группам
            balance_groups_list = []
            dt_balanse_group = common_sql.get_balance_groups_by_res(name_res)
            for x in range (len(dt_balanse_group)):
                balance_groups_list.append({"key": "group-"+str(x), "title": dt_balanse_group[x][0]})
                     
        #tree_data.append({"key": u"group" + str(1000), "title": u"Группы 80020", "children":groups_80020_list , "folder":bool(True)})
        tree_data.append({"key": "group" + str(1000), "title": "Группы", "children":balance_groups_list , "folder":bool(True)})
        
        tree_data_json = json.dumps(tree_data,)
        
    return tree_data_json


def tree_data_json_water_v2(request):    
    name_res = ['ХВС', 'ГВС', 'Импульс']
    #-------------- get data new tree
    max_level = Objects.objects.aggregate(Max('level'))['level__max'] # Max number of levels
    if max_level < 3:
        all_level_0 = Objects.objects.filter(level=0)#.filter(~Q(name = 'Вода'))      
        tree_data = []       
        #print u'Объекты уровня 0 \n', all_level_0
        for l0 in range(len(all_level_0)):
            #print all_level_0[l0].name
            filter_level_1 = Objects.objects.filter(level=1).filter(guid_parent = all_level_0[l0].guid)           
            children_data_l1 = []
            for l1 in range(len(filter_level_1)):                            
                children_data_l2 = []
                #print '-->', filter_level_1[l1].name                           
                filter_level_2 = Objects.objects.filter(level=2).filter(guid_parent = filter_level_1[l1].guid).order_by('name')
                for l2 in range(len(filter_level_2)):
                    abonents_data = []
                    #print 'test1', 'filter_level_1[l1].guid', filter_level_1[l1].guid                     
                    filter_level_abonents = common_sql.get_water_abonents_by_obj_guid(filter_level_2[l2].guid, name_res)                    
                    #print '--->',filter_level_2[l2].name, filter_level_2[l2].guid
                    #print len(filter_level_abonents)
                    for abon in range(len(filter_level_abonents)):
                        meters_data = []
                        #print filter_level_abonents[abon][1], filter_level_abonents[abon][2], filter_level_abonents[abon][0]
                        list_of_meters =  common_sql.get_meters_by_abons_guid_and_res(filter_level_abonents[abon][1], filter_level_abonents[abon][2])
                        
                        #list_of_meters = common_sql.get_meters_by_abons_guid_and_res(filter_level_abonents[abon][1], name_res)
                        for meter in range(len(list_of_meters)):
                            list_of_level_meters = {"key":"meter-"+str(meter), "title": list_of_meters[meter][2]}
                            meters_data.append(list_of_level_meters)
                        list_of_level_abonents = {"key":"abonent-"+str(l2), "title": filter_level_abonents[abon][0], "children":meters_data}
                        abonents_data.append(list_of_level_abonents)              
                    if len(filter_level_abonents) > 0:
                        list_of_level_2 = {"key":"level2-"+str(l1), "title": filter_level_2[l2].name, "children":abonents_data}
                        children_data_l2.append(list_of_level_2)

                list_of_level_1 = {"key":"level1-"+str(l1), "title": filter_level_1[l1].name, "children":children_data_l2, "folder":bool(children_data_l2)}
                children_data_l1.append(list_of_level_1)
            list_of_level_0 = {"key":"level0-"+str(l0), "title": all_level_0[l0].name, "children":children_data_l1, "folder":bool(children_data_l1)}
            tree_data.append(list_of_level_0)

        # Получаем информацию по балансным группам
            balance_groups_list = []
            dt_balanse_group = common_sql.get_balance_groups_by_res(name_res[2])
            for x in range (len(dt_balanse_group)):
                balance_groups_list.append({"key": "group-"+str(x), "title": dt_balanse_group[x][0]})
                     
        #tree_data.append({"key": u"group" + str(1000), "title": u"Группы 80020", "children":groups_80020_list , "folder":bool(True)})
        tree_data.append({"key": "group" + str(1000), "title": "Группы", "children":balance_groups_list , "folder":bool(True)})
        
        tree_data_json = json.dumps(tree_data,)
        
    return tree_data_json

def get_object_title(request):
    if request.is_ajax():
        if request.method == 'GET':
            object_title = request.GET['object_title']
        elif request.method == 'POST':
            object_title = 'Не выбран'
    else:
        object_title = 'Не выбран'
    return HttpResponse(object_title)

    
def get_object_key(request):
    if request.is_ajax():
        if request.method == 'GET':
            object_key = request.GET['object_key']
        elif request.method == 'POST':
            object_key = 'Не выбран'
    else:
        object_key = 'Не выбран'
    return HttpResponse(object_key)

@login_required(login_url='/auth/login/')     
def get_data_table(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = get_data_table_by_date_monthly(obj_title, obj_parent_title, electric_data_end)
#                request.session["data_table_export"] = data_table
                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = get_data_table_by_date_daily(obj_title, obj_parent_title, electric_data_end)

#                request.session["data_table_export"] = data_table  ! Check. Not Working

            elif (is_electric_current == "1") & (bool(is_abonent_level.search(obj_key))):
                pass
            elif (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                cursor_t0_aplus_delta_start = connection.cursor()
                cursor_t0_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t0_aplus_delta_start = cursor_t0_aplus_delta_start.fetchall()
                
                cursor_t1_aplus_delta_start = connection.cursor()
                cursor_t1_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t1_aplus_delta_start = cursor_t1_aplus_delta_start.fetchall()
                
                cursor_t2_aplus_delta_start = connection.cursor()
                cursor_t2_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t2_aplus_delta_start = cursor_t2_aplus_delta_start.fetchall()
                
                cursor_t3_aplus_delta_start = connection.cursor()
                cursor_t3_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t3_aplus_delta_start = cursor_t3_aplus_delta_start.fetchall() 
                
                cursor_t4_aplus_delta_start = connection.cursor()
                cursor_t4_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t4_aplus_delta_start = cursor_t4_aplus_delta_start.fetchall()

                cursor_t0_aplus_delta_end = connection.cursor()
                cursor_t0_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t0_aplus_delta_end = cursor_t0_aplus_delta_end.fetchall()
                
                cursor_t1_aplus_delta_end = connection.cursor()
                cursor_t1_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t1_aplus_delta_end = cursor_t1_aplus_delta_end.fetchall()
                
                cursor_t2_aplus_delta_end = connection.cursor()
                cursor_t2_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t2_aplus_delta_end = cursor_t2_aplus_delta_end.fetchall()
                
                cursor_t3_aplus_delta_end = connection.cursor()
                cursor_t3_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t3_aplus_delta_end = cursor_t3_aplus_delta_end.fetchall() 
                
                cursor_t4_aplus_delta_end = connection.cursor()
                cursor_t4_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t4_aplus_delta_end = cursor_t4_aplus_delta_end.fetchall()
#                data_table = []
                for x in range(len(data_table_t0_aplus_delta_end)):
                    data_table_temp = []

                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][2])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][6])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][1] - data_table_t0_aplus_delta_start[x][1] )
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end[x][1] - data_table_t1_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end[x][1] - data_table_t2_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end[x][1] - data_table_t3_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end[x][1] - data_table_t4_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table
                
            elif (is_electric_period == "1") & (is_electric_daily =="1") & (bool(is_abonent_level.search(obj_key))): # pokazaniya za period

                end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
                start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
                dates = [x for x in common_sql.daterange(start_date,
                              end_date,
                              step=datetime.timedelta(days=1),
                              inclusive=True)]
                '''for x in range(len(dates)):
                    data_table_temp = [dates[x], dates[x], datetime.datetime.strftime(dates[x], "%d.%m.%Y")]
                    data_table.append(data_table_temp)'''

                for x in range(len(dates)):
                    data_table_temp = get_data_table_by_date_daily(obj_title, obj_parent_title, datetime.datetime.strftime(dates[x], "%d.%m.%Y"))
                    data_table.extend(data_table_temp)
                #------------

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # monthly for abonents group
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                               ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    # query data for each abonent
                    cursor_t0_aplus_monthly_temp = connection.cursor()
                    cursor_t0_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T0 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_aplus_monthly_temp = cursor_t0_aplus_monthly_temp.fetchall()
                    
                    cursor_t1_aplus_monthly_temp = connection.cursor()
                    cursor_t1_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T1 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t1_aplus_monthly_temp = cursor_t1_aplus_monthly_temp.fetchall()
                    
                    cursor_t2_aplus_monthly_temp = connection.cursor()
                    cursor_t2_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date,
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T2 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY 
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t2_aplus_monthly_temp = cursor_t2_aplus_monthly_temp.fetchall()
                    
                    cursor_t3_aplus_monthly_temp = connection.cursor()
                    cursor_t3_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T3 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t3_aplus_monthly_temp = cursor_t3_aplus_monthly_temp.fetchall()
                
                    cursor_t4_aplus_monthly_temp = connection.cursor()
                    cursor_t4_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T4 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t4_aplus_monthly_temp = cursor_t4_aplus_monthly_temp.fetchall()
                    
                    cursor_t0_rplus_monthly_temp = connection.cursor()
                    cursor_t0_rplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T0 R+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_rplus_monthly_temp = cursor_t0_rplus_monthly_temp.fetchall()
                
                    data_table_temp = []
                    try:
                        data_table_temp.append(electric_data_end)
                    except IndexError:
                        data_table_temp.append("Н/Д")
                        
                    data_table_temp.append(abonents_list[x][0])
                    
                    try:
                        data_table_temp.append(data_table_t0_aplus_monthly_temp[0][6])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t0_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                        
                    try:    
                        data_table_temp.append(data_table_t1_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t2_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_rplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table
#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for abonents group
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                               ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()                              
#                data_table = []
                for x in range(len(abonents_list)):
                    cursor_t0_aplus_daily_temp = connection.cursor()
                    cursor_t0_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_aplus_daily_temp = cursor_t0_aplus_daily_temp.fetchall()
                
                    cursor_t1_aplus_daily_temp = connection.cursor()
                    cursor_t1_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t1_aplus_daily_temp = cursor_t1_aplus_daily_temp.fetchall()
                
                    cursor_t2_aplus_daily_temp = connection.cursor()
                    cursor_t2_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t2_aplus_daily_temp = cursor_t2_aplus_daily_temp.fetchall()
                
                    cursor_t3_aplus_daily_temp = connection.cursor()
                    cursor_t3_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t3_aplus_daily_temp = cursor_t3_aplus_daily_temp.fetchall() 
                
                    cursor_t4_aplus_daily_temp = connection.cursor()
                    cursor_t4_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t4_aplus_daily_temp = cursor_t4_aplus_daily_temp.fetchall()
                    
                    data_table_temp = []
                    try:
                        data_table_temp.append(electric_data_end)
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    data_table_temp.append(abonents_list[x][0])
                    try:
                        data_table_temp.append(data_table_t0_aplus_daily_temp[0][6])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table
            elif (is_electric_current == '1') & (bool(is_object_level.search(obj_key))):
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                              ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    data_table_temp = []
                    data_table_temp.append('Дата')
                    data_table_temp.append(abonents_list[x][0])
                    data_table_temp.append('Какой-то заводской номер')
                    data_table_temp.append(0)
                    data_table_temp.append(100)
                    data_table_temp.append(200)
                    data_table_temp.append(300)
                    data_table_temp.append(400)
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table
                   
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                              ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    # delta for groups abonents 'start date'
                    cursor_t0_aplus_delta_start_temp = connection.cursor()
                    cursor_t0_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t0_aplus_delta_start_temp = cursor_t0_aplus_delta_start_temp.fetchall()
                
                    cursor_t1_aplus_delta_start_temp = connection.cursor()
                    cursor_t1_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t1_aplus_delta_start_temp = cursor_t1_aplus_delta_start_temp.fetchall()
                
                    cursor_t2_aplus_delta_start_temp = connection.cursor()
                    cursor_t2_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t2_aplus_delta_start_temp = cursor_t2_aplus_delta_start_temp.fetchall()
                
                    cursor_t3_aplus_delta_start_temp = connection.cursor()
                    cursor_t3_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t3_aplus_delta_start_temp = cursor_t3_aplus_delta_start_temp.fetchall() 
                
                    cursor_t4_aplus_delta_start_temp = connection.cursor()
                    cursor_t4_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t4_aplus_delta_start_temp = cursor_t4_aplus_delta_start_temp.fetchall()
                    
                    # delta for groups abonents 'end date'
                    cursor_t0_aplus_delta_end_temp = connection.cursor()
                    cursor_t0_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_aplus_delta_end_temp = cursor_t0_aplus_delta_end_temp.fetchall()
                
                    cursor_t1_aplus_delta_end_temp = connection.cursor()
                    cursor_t1_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t1_aplus_delta_end_temp = cursor_t1_aplus_delta_end_temp.fetchall()
                
                    cursor_t2_aplus_delta_end_temp = connection.cursor()
                    cursor_t2_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t2_aplus_delta_end_temp = cursor_t2_aplus_delta_end_temp.fetchall()
                
                    cursor_t3_aplus_delta_end_temp = connection.cursor()
                    cursor_t3_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t3_aplus_delta_end_temp = cursor_t3_aplus_delta_end_temp.fetchall() 
                
                    cursor_t4_aplus_delta_end_temp = connection.cursor()
                    cursor_t4_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t4_aplus_delta_end_temp = cursor_t4_aplus_delta_end_temp.fetchall()
                    
                    data_table_temp = []
                    data_table_temp.append(abonents_list[x][0])
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][6])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1] - data_table_t0_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end_temp[0][1] - data_table_t1_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end_temp[0][1] - data_table_t2_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end_temp[0][1] - data_table_t3_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end_temp[0][1] - data_table_t4_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                           
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table                 
#*********************************************************************************************************************************************************************          
            else:
                pass

        else:
            obj_title = 'Не выбран'
            obj_parent_title = 'Не выбран'
            obj_key = 'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
            is_electric_delta = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render(request, "data_table.html", args)

# def export_excel_electric(request):
#     data_table = request.session["data_table_export"]
#     return ExcelResponse(data_table, 'report')
    


@login_required(login_url='/auth/login/')  
def electric(request):
    args={}
    args['ico_url_electric'] = "/static/images/electric-ico42.png"
    args['ico_url_water'] = "/static/images/water-ico36.png"
    args['ico_url_heat'] = "/static/images/heat-ico36.png"    
    args['ico_url_gas'] = "/static/images/gas-ico36.png"
    args['ico_url_economic'] = "/static/images/economic-ico36.png"
    return render(request, 'control.html', args)

@login_required(login_url='/auth/login/')
def economic(request):
    args={}
    args['ico_url_electric'] = "/static/images/electric-ico36.png"
    args['ico_url_water'] = "/static/images/water-ico36.png"
    args['ico_url_heat'] = "/static/images/heat-ico36.png"    
    args['ico_url_gas'] = "/static/images/gas-ico36.png"
    args['ico_url_economic'] = "/static/images/economic-ico42.png"
    return render(request, 'economic.html', args)

@login_required(login_url='/auth/login/')    
def water(request):
    args={}
    args['ico_url_electric'] = "/static/images/electric-ico36.png"
    args['ico_url_water'] = "/static/images/water-ico42.png"
    args['ico_url_heat'] = "/static/images/heat-ico36.png"    
    args['ico_url_gas'] = "/static/images/gas-ico36.png"
    args['ico_url_economic'] = "/static/images/economic-ico36.png"
    return render(request, 'water.html', args)
    
@login_required(login_url='/auth/login/')    
def heat(request):
    args={}
    args['ico_url_electric'] = "/static/images/electric-ico36.png"
    args['ico_url_water'] = "/static/images/water-ico36.png"
    args['ico_url_heat'] = "/static/images/heat-ico42.png"    
    args['ico_url_gas'] = "/static/images/gas-ico36.png"
    args['ico_url_economic'] = "/static/images/economic-ico36.png"
    return render(request, 'heat.html', args)
    
@login_required(login_url='/auth/login/')    
def gas(request):
    args={}
    args['ico_url_electric'] = "/static/images/electric-ico36.png"
    args['ico_url_water'] = "/static/images/water-ico36.png"
    args['ico_url_heat'] = "/static/images/heat-ico36.png"    
    args['ico_url_gas'] = "/static/images/gas-ico42.png"
    args['ico_url_economic'] = "/static/images/economic-ico36.png"
    return render(request, 'gas.html', args)
      
   
    
# образец выгрузги экселя -------------------------------------------------------------------------------------------    
def test_xlsx(request):
    import io
    response = io.StringIO()
    from openpyxl import Workbook
    from openpyxl.compat import range
    from openpyxl.cell import get_column_letter
    
    wb = Workbook()
    

    ws = wb.active
    
    ws.title = "range names"
    
    for col_idx in range(1,40):
        col = get_column_letter(col_idx)
        for row in range(1,600):
            ws.cell('%s%s'%(col,row)).value = '%s%s' % (col, row)
            
    ws = wb.create_sheet()
    
    ws.title = 'Pi'
    
    ws['F5'] = 3.14
    
    wb.save(response)

    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    response['Content-Disposition'] = "attachment; filename=test.xlsx"

    return response
#--------------------------------------------------------------------------------------------------------------------
def choose_report(request):
    return render(request, "choose_report.html")

def report_2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = get_data_table_by_date_monthly(obj_title, obj_parent_title, electric_data_end)

                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = get_data_table_by_date_daily(obj_title, obj_parent_title, electric_data_end)


            elif (is_electric_current == "1") & (bool(is_abonent_level.search(obj_key))):
                pass
                            
            elif (is_electric_period == "1") & (is_electric_daily =="1") & (bool(is_abonent_level.search(obj_key))): # pokazaniya za period

                end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
                start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
                dates = [x for x in common_sql.daterange(start_date,
                              end_date,
                              step=datetime.timedelta(days=1),
                              inclusive=True)]
                '''for x in range(len(dates)):
                    data_table_temp = [dates[x], dates[x], datetime.datetime.strftime(dates[x], "%d.%m.%Y")]
                    data_table.append(data_table_temp)'''

                for x in range(len(dates)):
                    data_table_temp = get_data_table_by_date_daily(obj_title, obj_parent_title, datetime.datetime.strftime(dates[x], "%d.%m.%Y"))
                    data_table.extend(data_table_temp)
                #------------

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # monthly for abonents group
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                               ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    # query data for each abonent
                    cursor_t0_aplus_monthly_temp = connection.cursor()
                    cursor_t0_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T0 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_aplus_monthly_temp = cursor_t0_aplus_monthly_temp.fetchall()
                    
                    cursor_t1_aplus_monthly_temp = connection.cursor()
                    cursor_t1_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T1 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t1_aplus_monthly_temp = cursor_t1_aplus_monthly_temp.fetchall()
                    
                    cursor_t2_aplus_monthly_temp = connection.cursor()
                    cursor_t2_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date,
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T2 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY 
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t2_aplus_monthly_temp = cursor_t2_aplus_monthly_temp.fetchall()
                    
                    cursor_t3_aplus_monthly_temp = connection.cursor()
                    cursor_t3_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T3 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t3_aplus_monthly_temp = cursor_t3_aplus_monthly_temp.fetchall()
                
                    cursor_t4_aplus_monthly_temp = connection.cursor()
                    cursor_t4_aplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T4 A+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t4_aplus_monthly_temp = cursor_t4_aplus_monthly_temp.fetchall()
                    
                    cursor_t0_rplus_monthly_temp = connection.cursor()
                    cursor_t0_rplus_monthly_temp.execute("""SELECT 
                                monthly_values.date, 
                                monthly_values.value, 
                                abonents.name, 
                                monthly_values.id_taken_params, 
                                objects.name, 
                                names_params.name, 
                                meters.factory_number_manual, 
                                resources.name
                                FROM 
                                public.monthly_values, 
                                public.link_abonents_taken_params, 
                                public.taken_params, 
                                public.abonents, 
                                public.objects, 
                                public.names_params, 
                                public.params, 
                                public.meters, 
                                public.resources
                                WHERE 
                                taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                taken_params.id = monthly_values.id_taken_params AND
                                taken_params.guid_params = params.guid AND
                                taken_params.guid_meters = meters.guid AND
                                abonents.guid = link_abonents_taken_params.guid_abonents AND
                                objects.guid = abonents.guid_objects AND
                                names_params.guid = params.guid_names_params AND
                                resources.guid = names_params.guid_resources AND
                                abonents.name = %s AND 
                                objects.name = %s AND 
                                names_params.name = 'T0 R+' AND 
                                monthly_values.date = %s AND 
                                resources.name = 'Электричество'
                                ORDER BY
                                objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_rplus_monthly_temp = cursor_t0_rplus_monthly_temp.fetchall()
                
                    data_table_temp = []
                    try:
                        data_table_temp.append(electric_data_end)
                    except IndexError:
                        data_table_temp.append("Н/Д")
                        
                    data_table_temp.append(abonents_list[x][0])
                    
                    try:
                        data_table_temp.append(data_table_t0_aplus_monthly_temp[0][6])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t0_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                        
                    try:    
                        data_table_temp.append(data_table_t1_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t2_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_rplus_monthly_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table
#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1'):# & (bool(is_object_level.search(obj_key))): # daily for abonents group
               
                    
                if (bool(is_object_level.search(obj_key))):
                    cursor_abonents_list = connection.cursor()
                    cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                               ORDER BY
                                               abonents.name ASC;""",[obj_title])
                    abonents_list = cursor_abonents_list.fetchall()

                    
                elif (bool(is_group_level.search(obj_key))):
                    cursor_abonents_list = connection.cursor()
                    cursor_abonents_list.execute("""
                                                SELECT 
                                                  meters.name,
                                                  link_balance_groups_meters.type
                                                FROM 
                                                  public.meters, 
                                                  public.link_balance_groups_meters, 
                                                  public.balance_groups
                                                WHERE 
                                                  link_balance_groups_meters.guid_balance_groups = balance_groups.guid AND
                                                  link_balance_groups_meters.guid_meters = meters.guid AND
                                                  balance_groups.name = %s
                                                ORDER BY
                                                  meters.name ASC;""",[obj_title])
                    abonents_list = cursor_abonents_list.fetchall()
                    obj_title='Завод'
                else:
                    abonents_list = [12345678]
                              

                for x in range(len(abonents_list)):
                    cursor_t0_aplus_daily_temp = connection.cursor()
                    cursor_t0_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_aplus_daily_temp = cursor_t0_aplus_daily_temp.fetchall()
                
                    cursor_t1_aplus_daily_temp = connection.cursor()
                    cursor_t1_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t1_aplus_daily_temp = cursor_t1_aplus_daily_temp.fetchall()
                
                    cursor_t2_aplus_daily_temp = connection.cursor()
                    cursor_t2_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t2_aplus_daily_temp = cursor_t2_aplus_daily_temp.fetchall()
                
                    cursor_t3_aplus_daily_temp = connection.cursor()
                    cursor_t3_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t3_aplus_daily_temp = cursor_t3_aplus_daily_temp.fetchall() 
                
                    cursor_t4_aplus_daily_temp = connection.cursor()
                    cursor_t4_aplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t4_aplus_daily_temp = cursor_t4_aplus_daily_temp.fetchall()
                    
                    cursor_t0_rplus_daily_temp = connection.cursor()
                    cursor_t0_rplus_daily_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_rplus_daily_temp = cursor_t0_rplus_daily_temp.fetchall()
                    
                    data_table_temp = []
                    try:
                        data_table_temp.append(electric_data_end)
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    data_table_temp.append(abonents_list[x][0])
                    try:
                        data_table_temp.append(data_table_t0_aplus_daily_temp[0][6])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    try:
                        if (bool(is_group_level.search(obj_key))):                           
                            if abonents_list[x][1]: # Если абонент входит в группу со знаком плюс, то показания как есть
                                data_table_temp.append(data_table_t0_aplus_daily_temp[0][1])
                            else:                   # Если абонент входит в группу со знаком минус, то показазния инвертируются
                                data_table_temp.append(-data_table_t0_aplus_daily_temp[0][1])
                        else:
                           data_table_temp.append(data_table_t0_aplus_daily_temp[0][1]) 
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    try:
                        if (bool(is_group_level.search(obj_key))):                                           
                            if abonents_list[x][1]: # Если абонент входит в группу со знаком плюс, то показания как есть
                                data_table_temp.append(data_table_t0_rplus_daily_temp[0][1])
                            else:
                                data_table_temp.append(-data_table_t0_rplus_daily_temp[0][1])
                        else:
                            data_table_temp.append(data_table_t0_rplus_daily_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    data_table.append(data_table_temp)
                if (bool(is_group_level.search(obj_key))):  # Если это группа добавляем еще одну строку с суммой показаний
                    sum_a_plus = 0
                    sum_r_plus = 0
                    for x in range(len(data_table)):
                        try:
                            sum_a_plus = sum_a_plus + data_table[x][3]
                            sum_r_plus = sum_r_plus + data_table[x][8]
                        except:
                            next
                    data_table.append([])
                    data_table.append([' ',' ','<strong>Сумма</strong>',sum_a_plus,'-','-','-','-',sum_r_plus])

                request.session["data_table_export"] = data_table
            elif (is_electric_current == '1') & (bool(is_object_level.search(obj_key))): # текущие для объекта учёта
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                              ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    data_table_temp = []
                    data_table_temp.append('Дата')
                    data_table_temp.append(abonents_list[x][0])
                    data_table_temp.append('Какой-то заводской номер')
                    data_table_temp.append(0)
                    data_table_temp.append(100)
                    data_table_temp.append(200)
                    data_table_temp.append(300)
                    data_table_temp.append(400)
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table
#*********************************************************************************************************************************************************************
            else:
                pass
        else:
            obj_title = 'Не выбран'
            obj_parent_title = 'Не выбран'
            obj_key = 'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render(request, "data_table/2.html", args)
    
def data_table_3_tarifa_k(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = '1'
    electric_data_start = ''
    electric_data_end = ''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                cursor_t0_aplus_delta_start = connection.cursor()
                cursor_t0_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t0_aplus_delta_start = cursor_t0_aplus_delta_start.fetchall()
                
                cursor_t1_aplus_delta_start = connection.cursor()
                cursor_t1_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t1_aplus_delta_start = cursor_t1_aplus_delta_start.fetchall()
                
                cursor_t2_aplus_delta_start = connection.cursor()
                cursor_t2_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t2_aplus_delta_start = cursor_t2_aplus_delta_start.fetchall()
                
                cursor_t3_aplus_delta_start = connection.cursor()
                cursor_t3_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t3_aplus_delta_start = cursor_t3_aplus_delta_start.fetchall() 
                
                cursor_t4_aplus_delta_start = connection.cursor()
                cursor_t4_aplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t4_aplus_delta_start = cursor_t4_aplus_delta_start.fetchall()

                cursor_t0_aplus_delta_end = connection.cursor()
                cursor_t0_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t0_aplus_delta_end = cursor_t0_aplus_delta_end.fetchall()
                
                cursor_t1_aplus_delta_end = connection.cursor()
                cursor_t1_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t1_aplus_delta_end = cursor_t1_aplus_delta_end.fetchall()
                
                cursor_t2_aplus_delta_end = connection.cursor()
                cursor_t2_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t2_aplus_delta_end = cursor_t2_aplus_delta_end.fetchall()
                
                cursor_t3_aplus_delta_end = connection.cursor()
                cursor_t3_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t3_aplus_delta_end = cursor_t3_aplus_delta_end.fetchall() 
                
                cursor_t4_aplus_delta_end = connection.cursor()
                cursor_t4_aplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t4_aplus_delta_end = cursor_t4_aplus_delta_end.fetchall()
                
                cursor_t0_rplus_delta_start = connection.cursor()
                cursor_t0_rplus_delta_start.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_start])
                data_table_t0_rplus_delta_start = cursor_t0_rplus_delta_start.fetchall()
                
                cursor_t0_rplus_delta_end = connection.cursor()
                cursor_t0_rplus_delta_end.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество';""",[obj_title, obj_parent_title, electric_data_end])
                data_table_t0_rplus_delta_end = cursor_t0_rplus_delta_end.fetchall()
                
                
#                data_table = []
                for x in range(len(data_table_t0_aplus_delta_end)):
                    data_table_temp = []

                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][2])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][6])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end[x][1] - data_table_t0_aplus_delta_start[x][1] )
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end[x][1] - data_table_t1_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end[x][1] - data_table_t2_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end[x][1] - data_table_t3_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end[x][1] - data_table_t4_aplus_delta_start[x][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")                    
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_start[x][1]) # Показания R+ начальные
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_end[x][1]) # Показания R+ конечные
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_end[x][1] - data_table_t0_rplus_delta_start[x][1]) # Показания R+ разница
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    
                    try:
                        data_table_temp.append(common_sql.get_k_t_t(obj_title)) # Коэффициент трансформации тока параметр 20
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                        
                    try:
                        data_table_temp.append(common_sql.get_k_t_n(obj_title)) # Коэффициент трансформации напряжения параметр 21
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_temp[20]*data_table_temp[21]*data_table_temp[12]) # Энергия А+ параметр 22
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")

                    try:
                        data_table_temp.append(data_table_temp[20]*data_table_temp[21]*data_table_temp[19]) # Энергия R+ параметр 23
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                           
                    data_table.append(data_table_temp)
                    
                    
                    
                request.session["data_table_export"] = data_table
                
                               
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                cursor_abonents_list = connection.cursor()
                cursor_abonents_list.execute("""
                                              SELECT 
                                               abonents.name
                                              FROM 
                                               public.objects, 
                                               public.abonents
                                              WHERE 
                                               objects.guid = abonents.guid_objects AND
                                               objects.name = %s
                                              ORDER BY
                                               abonents.name ASC;""",[obj_title])
                abonents_list = cursor_abonents_list.fetchall()
#                data_table = []
                for x in range(len(abonents_list)):
                    # delta for groups abonents 'start date'
                    cursor_t0_aplus_delta_start_temp = connection.cursor()
                    cursor_t0_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t0_aplus_delta_start_temp = cursor_t0_aplus_delta_start_temp.fetchall()
                
                    cursor_t1_aplus_delta_start_temp = connection.cursor()
                    cursor_t1_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t1_aplus_delta_start_temp = cursor_t1_aplus_delta_start_temp.fetchall()
                
                    cursor_t2_aplus_delta_start_temp = connection.cursor()
                    cursor_t2_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t2_aplus_delta_start_temp = cursor_t2_aplus_delta_start_temp.fetchall()
                
                    cursor_t3_aplus_delta_start_temp = connection.cursor()
                    cursor_t3_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t3_aplus_delta_start_temp = cursor_t3_aplus_delta_start_temp.fetchall() 
                
                    cursor_t4_aplus_delta_start_temp = connection.cursor()
                    cursor_t4_aplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t4_aplus_delta_start_temp = cursor_t4_aplus_delta_start_temp.fetchall()
                    
                    # delta for groups abonents 'end date'
                    cursor_t0_aplus_delta_end_temp = connection.cursor()
                    cursor_t0_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_aplus_delta_end_temp = cursor_t0_aplus_delta_end_temp.fetchall()
                
                    cursor_t1_aplus_delta_end_temp = connection.cursor()
                    cursor_t1_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T1 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t1_aplus_delta_end_temp = cursor_t1_aplus_delta_end_temp.fetchall()
                
                    cursor_t2_aplus_delta_end_temp = connection.cursor()
                    cursor_t2_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T2 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t2_aplus_delta_end_temp = cursor_t2_aplus_delta_end_temp.fetchall()
                
                    cursor_t3_aplus_delta_end_temp = connection.cursor()
                    cursor_t3_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T3 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t3_aplus_delta_end_temp = cursor_t3_aplus_delta_end_temp.fetchall() 
                
                    cursor_t4_aplus_delta_end_temp = connection.cursor()
                    cursor_t4_aplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T4 A+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t4_aplus_delta_end_temp = cursor_t4_aplus_delta_end_temp.fetchall()
                    
                    cursor_t0_rplus_delta_start_temp = connection.cursor()
                    cursor_t0_rplus_delta_start_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
                    data_table_t0_rplus_delta_start_temp = cursor_t0_rplus_delta_start_temp.fetchall()
                    
                    cursor_t0_rplus_delta_end_temp = connection.cursor()
                    cursor_t0_rplus_delta_end_temp.execute("""
                                SELECT 
                                  daily_values.date, 
                                  daily_values.value, 
                                  abonents.name, 
                                  daily_values.id_taken_params, 
                                  objects.name, 
                                  names_params.name, 
                                  meters.factory_number_manual, 
                                  resources.name
                                FROM 
                                  public.daily_values, 
                                  public.link_abonents_taken_params, 
                                  public.taken_params, 
                                  public.abonents, 
                                  public.objects, 
                                  public.names_params, 
                                  public.params, 
                                  public.meters, 
                                  public.resources
                                WHERE 
                                  taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                                  taken_params.id = daily_values.id_taken_params AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  abonents.guid = link_abonents_taken_params.guid_abonents AND
                                  objects.guid = abonents.guid_objects AND
                                  names_params.guid = params.guid_names_params AND
                                  resources.guid = names_params.guid_resources AND
                                  abonents.name = %s AND 
                                  objects.name = %s AND 
                                  names_params.name = 'T0 R+' AND 
                                  daily_values.date = %s AND 
                                  resources.name = 'Электричество'
                                  ORDER BY
                                  objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
                    data_table_t0_rplus_delta_end_temp = cursor_t0_rplus_delta_end_temp.fetchall()
                    
                    data_table_temp = []
                    data_table_temp.append(abonents_list[x][0])
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][6])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1] - data_table_t0_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t1_aplus_delta_end_temp[0][1] - data_table_t1_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t2_aplus_delta_end_temp[0][1] - data_table_t2_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t3_aplus_delta_end_temp[0][1] - data_table_t3_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t4_aplus_delta_end_temp[0][1] - data_table_t4_aplus_delta_start_temp[0][1])
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_start_temp[0][1]) # Показания R+ начальные
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_end_temp[0][1]) # Показания R+ конечные
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_t0_rplus_delta_end_temp[0][1] - data_table_t0_rplus_delta_start_temp[0][1]) # Показания R+ разница параметр 19
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                        
                    try:
                        data_table_temp.append(common_sql.get_k_t_t(abonents_list[x][0])) # Коэффициент трансформации тока параметр 20
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                        
                    try:
                        data_table_temp.append(common_sql.get_k_t_n(abonents_list[x][0])) # Коэффициент трансформации напряжения параметр 21
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                        
                    try:
                        data_table_temp.append(data_table_temp[20]*data_table_temp[21]*data_table_temp[12]) # Энергия А+ параметр 22
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")

                    try:
                        data_table_temp.append(data_table_temp[20]*data_table_temp[21]*data_table_temp[19]) # Энергия R+ параметр 23
                    except IndexError:
                        data_table_temp.append("Н/Д")
                    except TypeError:
                        data_table_temp.append("Н/Д")
                           
                    data_table.append(data_table_temp)
                request.session["data_table_export"] = data_table                 
#*********************************************************************************************************************************************************************
            else:
                pass
        else:
            obj_title = 'Не выбран'
            obj_parent_title = 'Не выбран'
            obj_key = 'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
            is_electric_delta = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render(request, "data_table/1.html", args)
    
def data_table_period_3_tarifa(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    data_table = []
    data_table_graph_a_plus = []
    data_table_graph_r_plus = []

    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    dates = None
    is_electric_period = '1'
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period                       
            if (is_electric_period == "1") & (is_electric_daily =="1") & (bool(is_abonent_level.search(obj_key))): # pokazaniya za period
                 
                end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
                start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
                # dates формирует список дат от начальной до конечной даты                
                dates = [x for x in common_sql.daterange(start_date,
                              end_date,
                              step=datetime.timedelta(days=1),
                              inclusive=True)]
                # делаем выборку показаний по каждой дате в диапазоне указанных
                for x in range(len(dates)):
                    data_table_temp = get_data_table_by_date_daily(obj_title, obj_parent_title, datetime.datetime.strftime(dates[x], "%d.%m.%Y"))
                    if x >0:
                        try:
                            data_table_temp[0].append(data_table_temp[0][3] - data_table[x-1][3]) # Считаем разницу показаний по A+ за предыдущие сутки
                            data_table_temp[0].append(data_table_temp[0][8] - data_table[x-1][8]) # Считаем разницу показаний по R+ за предыдущие сутки                            
                        except:
                            next
                    # Блок проверки показаний за отдельную дату в диапазоне. Если показаний нет, то вставляем Н/Д в соответствующие поля        
                    if data_table_temp:
                        data_table.extend(data_table_temp)
                    else:
                        data_table.append([datetime.datetime.strftime(dates[x], "%d.%m.%Y"),obj_title,common_sql.get_serial_number_by_meter_name(obj_title), 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д'])
                #------------
                        
#*********************************************************************************************************************************************************************
            else:
                pass
        else:
            obj_title = 'Не выбран'
            obj_parent_title = 'Не выбран'
            obj_key = 'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
            is_electric_delta = 0
               
    for x in range(len(data_table)):
        data_table_graph_a_plus_temp = []
        data_table_graph_r_plus_temp = []

        try:
            data_table_graph_a_plus_temp.append(data_table[x][0].strftime("%d.%m.%y"))
            data_table_graph_a_plus_temp.append(data_table[x][9])
            
            data_table_graph_r_plus_temp.append(data_table[x][0].strftime("%d.%m.%y"))
            data_table_graph_r_plus_temp.append(data_table[x][10])

            data_table_graph_a_plus.append(data_table_graph_a_plus_temp)
            data_table_graph_r_plus.append(data_table_graph_r_plus_temp)

        except:
            next
            
    # Сдвигаем дату на 1 число назад, потому что считаем энергию за прошедшие сутки            
    for x in range(len(data_table_graph_a_plus)):
        data_table_graph_a_plus[x][0] = (datetime.datetime.strptime(data_table_graph_a_plus[x][0],"%d.%m.%y")-datetime.timedelta(days=1)).strftime("%d.%m.%y")
        data_table_graph_a_plus[x][1] = data_table_graph_a_plus[x][1]*common_sql.get_k_t_n(obj_title)*common_sql.get_k_t_t(obj_title)
        
    # Сдвигаем дату на 1 число назад, потому что считаем энергию за прошедшие сутки            
    for x in range(len(data_table_graph_r_plus)):
        data_table_graph_r_plus[x][0] = (datetime.datetime.strptime(data_table_graph_r_plus[x][0],"%d.%m.%y")-datetime.timedelta(days=1)).strftime("%d.%m.%y")
        data_table_graph_r_plus[x][1] = data_table_graph_r_plus[x][1]*common_sql.get_k_t_n(obj_title)*common_sql.get_k_t_t(obj_title)
        

                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    args['data_table_graph_a_plus'] = data_table_graph_a_plus
    args['data_table_graph_r_plus'] = data_table_graph_r_plus
    

    

    return render(request, "data_table/3.html", args)
    


    
    
    
    
def profil_30_aplus(request):
    args = {}
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = meters_name           = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']                     
            
            a_plus = connection.cursor()
            a_plus.execute("""SELECT 
                                  various_values.date, 
                                  various_values."time", 
                                  various_values.value, 
                                  meters.name, 
                                  meters.address, 
                                  names_params.name
                                FROM 
                                  public.various_values, 
                                  public.meters, 
                                  public.params, 
                                  public.taken_params, 
                                  public.names_params
                                WHERE 
                                  params.guid_names_params = names_params.guid AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  taken_params.id = various_values.id_taken_params AND
                                  various_values.date = %s AND 
                                  meters.name = %s AND 
                                  names_params.name = 'A+ Профиль';""",[electric_data_end, meters_name])
            a_plus = a_plus.fetchall()
            val_table_a_plus = []
           
            for x in range(len(a_plus)):
                my_val_table = [] 
                my_val_table.append(float(calendar.timegm(datetime.datetime.combine(a_plus[x][0], a_plus[x][1]).timetuple())*1000))
                my_val_table.append(a_plus[x][2]*2*common_sql.get_k_t_n(meters_name)*common_sql.get_k_t_t(meters_name))
                val_table_a_plus.append(my_val_table)
                
            r_plus = connection.cursor()
            r_plus.execute("""SELECT 
                                  various_values.date, 
                                  various_values."time", 
                                  various_values.value, 
                                  meters.name, 
                                  meters.address, 
                                  names_params.name
                                FROM 
                                  public.various_values, 
                                  public.meters, 
                                  public.params, 
                                  public.taken_params, 
                                  public.names_params
                                WHERE 
                                  params.guid_names_params = names_params.guid AND
                                  taken_params.guid_params = params.guid AND
                                  taken_params.guid_meters = meters.guid AND
                                  taken_params.id = various_values.id_taken_params AND
                                  various_values.date = %s AND 
                                  meters.name = %s AND 
                                  names_params.name = 'R+ Профиль';""",[electric_data_end, meters_name])
            r_plus = r_plus.fetchall()
            val_table_r_plus = []
           
            for x in range(len(r_plus)):
                my_val_table = [] 
                my_val_table.append(float(calendar.timegm(datetime.datetime.combine(r_plus[x][0], r_plus[x][1]).timetuple())*1000))
                my_val_table.append(r_plus[x][2]*2*common_sql.get_k_t_n(meters_name)*common_sql.get_k_t_t(meters_name))
                val_table_r_plus.append(my_val_table)
                
            data_table = []
            for x in range(len(a_plus)):
                data_table_temp = []
                try:
                    data_table_temp.append(a_plus[x][0])
                except IndexError:
                    data_table_temp.append("Н/Д")
                except TypeError:
                    data_table_temp.append("Н/Д")
                try:
                    data_table_temp.append(a_plus[x][1])
                except IndexError:
                    data_table_temp.append("Н/Д")
                except TypeError:
                    data_table_temp.append("Н/Д")
                try:
                    data_table_temp.append(a_plus[x][3])
                except IndexError:
                    data_table_temp.append("Н/Д")
                except TypeError:
                    data_table_temp.append("Н/Д")
                try:
                    data_table_temp.append(a_plus[x][2]*2*common_sql.get_k_t_n(meters_name)*common_sql.get_k_t_t(meters_name))
                except IndexError:
                    data_table_temp.append("Н/Д")
                except TypeError:
                    data_table_temp.append("Н/Д")
                try:
                    data_table_temp.append(r_plus[x][2]*2*common_sql.get_k_t_n(meters_name)*common_sql.get_k_t_t(meters_name))
                except IndexError:
                    data_table_temp.append("Н/Д")
                except TypeError:
                    data_table_temp.append("Н/Д")
                data_table.append(data_table_temp)

        
     
            args['min30_a_plus'] = val_table_a_plus
            args['min30_r_plus'] = val_table_r_plus
                
            args['data_table'] = data_table
            args['k_t_n'] = common_sql.get_k_t_n(meters_name)
            args['k_t_t'] = common_sql.get_k_t_t(meters_name)
            args['meters_name'] = meters_name
            args['electric_data_end'] = electric_data_end
    
    

    
        
    return render(request, "data_table/4.html", args)
    



def hour_increment(request):
    args = {}
#    meters_name= u'Не выбран'
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']  
    time_list = ['00:00', '00:30','01:00', '01:30', '02:00', '02:30', '03:00', '03:30', '04:00', '04:30', '05:00', '05:30', '06:00', '06:30', '07:00', '07:30', '08:00', '08:30', '09:00', '09:30', '10:00', '10:30', '11:00', '11:30', '12:00', '12:30', '13:00', '13:30', '14:00', '14:30', '15:00', '15:30', '16:00', '16:30', '17:00', '17:30', '18:00', '18:30', '19:00', '19:30', '20:00', '20:30', '21:00', '21:30', '22:00', '22:30', '23:00', '23:30']
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = meters_name           = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
    
    serial_number = common_sql.get_serial_number_by_meter_name(meters_name)
        
    data_table = []
    if meters_name != 'Не выбран':
        # Добавляем первую строку в таблицу данных. Делаем запрос показаний на начало суток.
        data_table.append([electric_data_end,'00:00', meters_name, serial_number, common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 A+' ),common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 R+' ), '---', '---'])
        
        if common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 A+' ) != 'Нет данных':  # Если есть показания на начало суток выполняем почасовое приращение  
            for x in range(24):
                data_table_temp = []
                data_table_temp.append(electric_data_end)
                data_table_temp.append(time_list[(2*x)])
                data_table_temp.append(meters_name)
                data_table_temp.append(serial_number)
                data_table_temp.append(data_table[len(data_table)-1][4] + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'A+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'A+ Профиль'))
                data_table_temp.append(common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 R+' ) + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'R+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'R+ Профиль'))
                data_table_temp.append(common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'A+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'A+ Профиль'))
                data_table_temp.append(common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'R+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'R+ Профиль'))    
                if x == 0: # Убираем первую строку. Так как показания на 00:00 берем отдельным запросом
                    next
                else:
                    data_table.append(data_table_temp)    
    
    args['data_table'] = data_table
    args['meters_name'] = meters_name           
    args['electric_data_end'] = electric_data_end
    return render(request, "data_table/6.html", args)
    


    

def economic_electric(request):
    args = {}
    data_table = []
    
    electric_data_end   = request.session["electric_data_end"]
    electric_data_start   = request.session["electric_data_start"]



    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']    

    end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
    start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
    dates = [x for x in common_sql.daterange(start_date,
                  end_date,
                  step=datetime.timedelta(days=1),
                  inclusive=True)]
    #print dates
                  
    for x in range(len(dates)):
        try:
            data_table_temp = []
            delta_a_plus = 1
            delta_r_plus = 1

            try:
                delta_a_plus = common_sql.delta_sum_a_plus(dates[x+1])-common_sql.delta_sum_a_plus(dates[x])
                if delta_a_plus > 0:
                    delta_a_plus = delta_a_plus
                else:
                    delta_a_plus = 'Н/Д'
                delta_r_plus = common_sql.delta_sum_r_plus(dates[x+1])-common_sql.delta_sum_r_plus(dates[x])
                if delta_r_plus > 0:
                    delta_r_plus = delta_r_plus
                else:
                    delta_r_plus = 'Н/Д'

            except:
                delta_a_plus = 'Н/Д'
                delta_r_plus = 'Н/Д'

            data_table_temp.append(dates[x])
            data_table_temp.append(common_sql.product_sum(dates[x]))
            data_table_temp.append(delta_a_plus)
            data_table_temp.append(delta_a_plus/(common_sql.product_sum(dates[x])))
            data_table_temp.append(delta_r_plus)
            data_table_temp.append(delta_r_plus/(common_sql.product_sum(dates[x])))
        except:
            next
        data_table.append(data_table_temp)
    data_graph = []    
    for x in range(len(data_table)):
        data_graph_temp = []
        #data_graph_temp.append(x)
        try:
            data_graph_temp.append(data_table[x][0].strftime("%d.%m.%y"))
            data_graph_temp.append(data_table[x][3])
            data_graph.append(data_graph_temp)
        except:
            next
        
    
    args['economic_graph_data'] = json.dumps(data_graph)#[[u'Jan', 13], [datetime.datetime.now().strftime("%Y-%m-%d"), 17], [50, 9]])
    #args['economic_graph_data'] = data_graph
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    return render(request, "data_table/7.html", args)
    
def rejim_day(request):
    args = {}
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = meters_name           = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
    return render(request, "data_table/8.html", args)    

def load_balance_groups(request):
    #сделать потом воду, пока тут балансные группы!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    parent_name         = request.GET['obj_parent_title']
    ab_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
                # Добавляем привязку к балансной группе 
    cfg_excel_name = 'C:/work/mitino/prizmer/static/cfg/kB_balance_for_load.xlsx'
    cfg_sheet_name = 'ВРУ-1'
    is_electic_cfg = True
    is_water_cfg = False
    is_heat_cfg = False
    from django.db import connection
    from openpyxl import load_workbook
    wb = load_workbook(filename = cfg_excel_name)
    sheet_ranges = wb[cfg_sheet_name]
    row = 2
    dt=[]

    while (bool(sheet_ranges['A%s'%(row)].value)):
        guid_balance_groups_from_excel = connection.cursor()
        balance_group_name=[str(sheet_ranges['A%s'%(row)].value)]
        guid_balance_groups_from_excel.execute("""SELECT balance_groups.guid FROM public.balance_groups WHERE balance_groups.name = %s;""",balance_group_name)
        guid_balance_groups_from_excel = guid_balance_groups_from_excel.fetchall()
        if len(guid_balance_groups_from_excel)>0:
            guid_balance_groups = BalanceGroups.objects.get(guid=guid_balance_groups_from_excel[0][0])
        else: 
            #print u'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!'
            #print u'Надо создать балансную группу(вручную), прежде чем добавлять в неё что-то.'
            break
        guid_meters_from_excel = connection.cursor()
        meters_name=[str(sheet_ranges['E%s'%(row)].value)]
        znak=[bool(sheet_ranges['D%s'%(row)].value)]
        guid_meters_from_excel.execute("""SELECT meters.guid FROM public.meters WHERE meters.factory_number_manual = %s;""",meters_name)
        guid_meters_from_excel = guid_meters_from_excel.fetchall()
        if len(guid_meters_from_excel)>0:
            guid_meters = Meters.objects.get(guid=guid_meters_from_excel[0][0])
        else:
            #print u'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!'
            #print u'Такого счётчика не существует в БД, он не может быть добавлен в балансную группу.'
            continue
        add_link_meter_balance_group = LinkBalanceGroupsMeters(guid_balance_groups = guid_balance_groups, guid_meters = guid_meters, type=znak[0])
        add_link_meter_balance_group.save()
        
        dt.append([balance_group_name,meters_name,znak])
        #print unicode(sheet_ranges[u'A%s'%(row)].value), unicode(row), unicode([balance_group_name,meters_name,znak])
        row = row + 1
    
    args = {}

    args['data_table'] = dt
    args['electric_data_end'] = electric_data_end
    return render(request, "data_table/water/24.html", args)

def pokazaniya_water(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_channel(meters_name, electric_data_end)
    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        #print list_of_abonents_2
        #print common_sql.return_parent_guid_by_abonent_name(parent_name)
        #print meters_name
        #print common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []        
        for x in range(len(list_of_abonents_2)):
            data_table_temp = common_sql.get_daily_water_channel(list_of_abonents_2[x], electric_data_end)
            data_table.extend(data_table_temp)
            
    elif(bool(is_object_level_1.search(obj_key))):
        list_of_objects_2 = common_sql.list_of_objects(common_sql.return_parent_guid_by_abonent_name(meters_name)) #Список квартир для объекта с пульсарами
        data_table = []
        for x in range(len(list_of_objects_2)):
            data_table_temp = [(list_of_objects_2[x][0],)]
            list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(meters_name), list_of_objects_2[x][0])
            for y in range(len(list_of_abonents_2)):
                data_table_temp2 = common_sql.get_daily_water_channel(list_of_abonents_2[y], electric_data_end)

                data_table_temp.extend(data_table_temp2)                                
                      
            data_table.extend(data_table_temp)
              
    else:
        data_table = []
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end

    return render(request, "data_table/water/10.html", args)
    
def pokazaniya_water_identificators(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_channel(meters_name, electric_data_end)
    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []        
        for x in range(len(list_of_abonents_2)):
            data_table_temp = common_sql.get_daily_water_channel(list_of_abonents_2[x], electric_data_end)
            data_table.extend(data_table_temp)
    elif(bool(is_object_level_1.search(obj_key))):
        
        list_of_objects_2 = common_sql.list_of_objects(common_sql.return_parent_guid_by_abonent_name(meters_name)) #Список квартир для объекта с пульсарами
        data_table = []
        for x in range(len(list_of_objects_2)):
            data_table_temp = [(list_of_objects_2[x][0],)]
            list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(meters_name), list_of_objects_2[x][0])
            for y in range(len(list_of_abonents_2)):
                data_table_temp2 = common_sql.get_daily_water_channel(list_of_abonents_2[y], electric_data_end)

                data_table_temp.extend(data_table_temp2)                                
                      
            data_table.extend(data_table_temp)
              
    else:
        data_table = []
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name

    return render(request, "data_table/water/12.html", args)     

def pokazaniya_water_gvs_hvs_daily(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.GET['obj_parent_title']
    obj_title         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end, 'daily', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table=common_sql.get_daily_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end, 'daily', False)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = obj_title

    return render(request, "data_table/water/28.html", args)     
def pokazaniya_water_gvs_hvs_current(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.GET['obj_parent_title']
    obj_title         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_current_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end,  True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table_temp=common_sql.get_current_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end,  False)
        for row in data_table_temp:
            if row[4]=='Н/Д' and row[5]=='Н/Д':
                row2=common_sql.get_current_water_gvs_hvs(str(row[2]), str(row[6]) , electric_data_end, True)
                #print row2
                #print unicode(row[2]), unicode(row[6]), electric_data_end, True
                if len(row2)==0:
                    r=[str(electric_data_end), 'Н/Д', str(row[2]),str(row[3]), 'Н/Д', 'Н/Д']
                    data_table.append(r)
                else:
                    data_table.append(row2[0])
            else:
                data_table.append(row)

    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = obj_title

    return render(request, "data_table/water/26.html", args)     

def water_elf_hvs_by_date(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.GET['obj_parent_title']
    obj_title         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_elf(obj_title, obj_parent_title , electric_data_end, '1','attr1', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_daily_water_elf(obj_title, obj_parent_title , electric_data_end, '1', 'attr1',False)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['res'] = 'ХВС'
    args['obj_title'] = obj_title
    return render(request, "data_table/water/52.html", args)     

def water_elf_gvs_by_date(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.GET['obj_parent_title']
    obj_title         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_elf(obj_title, obj_parent_title , electric_data_end, '2','attr2', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_daily_water_elf(obj_title, obj_parent_title , electric_data_end, '2', 'attr2',False)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['res'] = 'ГВС'
    args['obj_title'] = obj_title

    return render(request, "data_table/water/52.html", args)

def water_elf_hvs_potreblenie(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.GET['obj_parent_title']
    obj_title         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    electric_data_start = request.GET['electric_data_start']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_elf_period(obj_title, obj_parent_title , electric_data_end, electric_data_start,'1','attr1', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_daily_water_elf_period(obj_title, obj_parent_title , electric_data_end, electric_data_start,'1','attr1', False)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    args['res'] = 'ХВС'
    args['obj_title'] = obj_title

    return render(request, "data_table/water/53.html", args)
    
def water_elf_gvs_potreblenie(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    electric_data_start = request.GET['electric_data_start']
    obj_parent_title         = request.GET['obj_parent_title']
    obj_title         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_elf_period(obj_title, obj_parent_title , electric_data_end,electric_data_start, '2','attr2', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_daily_water_elf_period(obj_title, obj_parent_title , electric_data_end,electric_data_start, '2','attr2', False)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['res'] = 'ГВС'
    args['electric_data_start'] = electric_data_start
    args['obj_title'] = obj_title

    return render(request, "data_table/water/53.html", args)

def potreblenie_water(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']                        
    obj_key             = request.GET['obj_key']

    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']                        
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table_start = common_sql.get_daily_water_channel(meters_name, electric_data_start) # Таблица с начальными значениями
        data_table_end = common_sql.get_daily_water_channel(meters_name, electric_data_end)     # Таблица с конечными значениями
        
        data_table = [[data_table_start[0][0],data_table_start[0][1],data_table_start[0][2],data_table_start[0][3],data_table_end[0][2],data_table_end[0][2]-data_table_start[0][2]]]

        
    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []        
        for x in range(len(list_of_abonents_2)):
            data_table_temp_start = common_sql.get_daily_water_channel(list_of_abonents_2[x], electric_data_start)
            data_table_temp_end = common_sql.get_daily_water_channel(list_of_abonents_2[x], electric_data_end)
            data_table_temp = [[data_table_temp_start[0][0],data_table_temp_start[0][1],data_table_temp_start[0][2],data_table_temp_start[0][3],data_table_temp_end[0][2],data_table_temp_end[0][2]-data_table_temp_start[0][2]]]

            data_table.extend(data_table_temp)

    elif(bool(is_object_level_1.search(obj_key))):
        
        list_of_objects_2 = common_sql.list_of_objects(common_sql.return_parent_guid_by_abonent_name(meters_name)) #Список квартир для объекта с пульсарами
        data_table = []
        for x in range(len(list_of_objects_2)):
            data_table_temp = [(list_of_objects_2[x][0],)]
            list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(meters_name), list_of_objects_2[x][0])
            for y in range(len(list_of_abonents_2)):
                data_table_temp2_end = common_sql.get_daily_water_channel(list_of_abonents_2[y], electric_data_end)
                data_table_temp2_start = common_sql.get_daily_water_channel(list_of_abonents_2[y], electric_data_start)
                #print data_table_temp2_end
                if bool(data_table_temp2_end) and bool(data_table_temp2_start):
                
                    data_table_temp2 = [[data_table_temp2_start[0][0],data_table_temp2_start[0][1],data_table_temp2_start[0][2],data_table_temp2_start[0][3],data_table_temp2_end[0][2],data_table_temp2_end[0][2]-data_table_temp2_start[0][2]]]
                else:
                    data_table_temp2 = [[list_of_abonents_2[y][0], 'Н/Д', '-', '-', '-']]                

                data_table_temp.extend(data_table_temp2)
            data_table.extend(data_table_temp)

    else:
        data_table = []
                                                     
    
    args['data_table'] = data_table
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name


    return render(request, "data_table/water/11.html", args)

def num_from_name(name):
    start = name.find('№')
    num = name[start+1:]
    return num
    
    
def add_numbers(request):
    g =  Abonents.objects.values_list("guid")
    for x in range(len(g)):
        t = Abonents.objects.get(guid = g[x][0])
        if num_from_name(t.name):
            t.account_2 = num_from_name(t.name)  # change field
            t.save() # this will update only
        else:
            next
    html = 'Готово'
    return HttpResponse(html)


def electric_potreblenie_3_zones_v2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = '1'
    electric_data_start = ''
    electric_data_end = ''
    
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            res='Электричество'

            if (is_electric_monthly=="1"):
                dm='monthly'
            else:
                dm='daily'
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                    isAbon=True                    
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
                
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                    isAbon=False
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
            #*********************************************************************************************************************************************************************
            elif (is_electric_delta == '1') &(bool(is_group_level.search(obj_key))):
                    data_table=common_sql.get_data_table_electric_period_for_group(obj_title,obj_parent_title, electric_data_start, electric_data_end, res)
                    request.session["data_table_export"] = data_table
            else:
                pass
        else:
            obj_title = 'Не выбран'
            obj_parent_title = 'Не выбран'
            obj_key = 'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
            is_electric_delta = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    return render(request, "data_table/electric/17.html", args)
    

def electric_potreblenie_2_zones_v2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = '1'
    electric_data_start = ''
    electric_data_end = ''
    
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            res='Электричество'
            if (is_electric_monthly=="1"):
                dm='monthly'
            else:
                dm='daily'
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                    isAbon=True                    
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
                
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                    isAbon=False
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
            #*********************************************************************************************************************************************************************
            elif (is_electric_delta == '1') &(bool(is_group_level.search(obj_key))):
                    data_table=common_sql.get_data_table_electric_period_for_group(obj_title,obj_parent_title, electric_data_start, electric_data_end, res)
                    request.session["data_table_export"] = data_table
            else:
                pass
        else:
            obj_title = 'Не выбран'
            obj_parent_title = 'Не выбран'
            obj_key = 'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
            is_electric_delta = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    return render(request, "data_table/electric/31.html", args)



def electric_between_3_zones(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    dates = None
    is_electric_period = None
    data_table=[]
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_daily == '1') & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                params=['T0 A+','T1 A+','T2 A+','T3 A+', 'Электричество']
                data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
            else:
                pass
            
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    
    return render(request, "data_table/electric/29.html", args)

def electric_between_2_zones(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    dates = None
    is_electric_period = None
    data_table=[]
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_daily == '1') & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                params=['T0 A+','T1 A+','T2 A+','T3 A+', 'Электричество']
                data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
            else:
                pass
            
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    
    return render(request, "data_table/electric/27.html", args)

def electric_between(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    dates = None
    is_electric_period = None
    data_table=[]
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_daily == '1') & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                params=['T0 A+','T1 A+','T2 A+','T3 A+', 'Электричество']
                data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
            else:
                pass
            
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render(request, "data_table/electric/25.html", args)

def electric_simple_2_zones_v2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'monthly')

                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'daily')


            elif (is_electric_current == "1") & (bool(is_abonent_level.search(obj_key))):
                pass
                            
            elif (is_electric_period == "1") & (is_electric_daily =="1") & (bool(is_abonent_level.search(obj_key))): # pokazaniya za period
                pass
                #------------

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # показания на начало месяца для объекта
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д']]        

#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for abonents group
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д']]

            elif (is_electric_daily == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д']]
              
            elif (is_electric_monthly == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д']]

            elif (is_electric_current == '1') & (bool(is_object_level.search(obj_key))): # текущие для объекта учёта
                    pass
            else:
                pass
        else:
            obj_title = 'Не выбран'
            obj_parent_title = 'Не выбран'
            obj_key = 'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    
    return render(request, "data_table/electric/14.html", args)


def electric_simple_3_zones_v2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'monthly')

                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'daily')
                

            elif (is_electric_current == "1") & (bool(is_abonent_level.search(obj_key))):
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'current')
                            
            elif (is_electric_period == "1") & (is_electric_daily =="1") & (bool(is_abonent_level.search(obj_key))): # pokazaniya za period
                pass
                #------------

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # показания на начало месяца для объекта
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    #print len(data_table)
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д']]        

#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for abonents group
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:                       
                        data_table = [[electric_data_end, obj_title, 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д']]

            elif (is_electric_daily == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д']]
              
            elif (is_electric_monthly == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д']]

            elif (is_electric_current == '1') & (bool(is_object_level.search(obj_key))): # текущие для объекта учёта
                     data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'current')

#*********************************************************************************************************************************************************************
            else:
                pass
        else:
            obj_title = 'Не выбран'
            obj_parent_title = 'Не выбран'
            obj_key = 'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render(request, "data_table/electric/16.html", args)
    
def electric_current_3_zones_v2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'current')         
                            
            elif (bool(is_object_level.search(obj_key))): # показания на начало месяца для объекта
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'current')
                   
     
        else:
            obj_title = 'Не выбран'
            obj_parent_title = 'Не выбран'
            obj_key = 'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render(request, "data_table/electric/16.html", args)
    
def electric_simple_3_zones_v3(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'monthly')

                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'daily')


            elif (is_electric_current == "1") & (bool(is_abonent_level.search(obj_key))):
                pass
                            
            elif (is_electric_period == "1") & (is_electric_daily =="1") & (bool(is_abonent_level.search(obj_key))): # pokazaniya za period
                pass
                #------------

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # показания на начало месяца для объекта
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v3(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д']]        

#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for abonents group
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v3(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д', 'Н/Д']]

            elif (is_electric_daily == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    pass
                    # data_table= common_sql.get_data_table_by_date_for_group_3_zones_v3(obj_title, electric_data_end, 'daily')
                    # if not data_table:
                    #     data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]
              
            elif (is_electric_monthly == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    pass
                    # data_table= common_sql.get_data_table_by_date_for_group_3_zones_v3(obj_title, electric_data_end, 'monthly')
                    # if not data_table:
                    #     data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]

            elif (is_electric_current == '1') & (bool(is_object_level.search(obj_key))): # текущие для объекта учёта
                    pass

#*********************************************************************************************************************************************************************
            else:
                pass
        else:
            obj_title = 'Не выбран'
            obj_parent_title = 'Не выбран'
            obj_key = 'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0

                    
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render(request, "data_table/electric/72.html", args)
    
    
#________________-


def pokazaniya_heat_v2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    data_table = []
    list_except = []
    if (bool(is_abonent_level.search(obj_key))):     
        data_table = common_sql.get_data_table_by_date_heat_v2(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_heat_v2(meters_name, parent_name, electric_data_end, False)
        for row in data_table:
            for x in list_except:
                if x==row[2]:
                    data_table.remove(x)


    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 
    return render(request, "data_table/heat/18.html", args)


#--------------------------------------------
def potreblenie_heat_v2(request): 
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']
    electric_data_start = request.GET['electric_data_start']
    obj_key             = request.GET['obj_key']
    list_except = []
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
#                     
#    if (bool(is_abonent_level.search(obj_key))):        
#        data_table = common_sql.get_data_table_for_period_for_abon_heat_v2(meters_name, parent_name, electric_data_start, electric_data_end)
#
#    elif (bool(is_object_level_2.search(obj_key))):
#        data_table = common_sql.get_data_table_for_period_heat_v2(meters_name, parent_name, electric_data_start, electric_data_end)

                 
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_for_period_v3(meters_name, parent_name, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_for_period_v3(meters_name, parent_name, electric_data_start, electric_data_end, False)
    else:
        data_table = []
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    args['obj_title'] = meters_name 


    return render(request, "data_table/heat/19.html", args)
    
def pokazaniya_heat_current_v2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key'] 
    list_except = []
    data_table=[]
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']

    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_current_heat_v2(meters_name, parent_name, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_current_heat_v2(meters_name, parent_name, False)
        for row in data_table:
            for x in list_except:
                if x==row[2]:
                    data_table.remove(x)

    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 
    return render(request, "data_table/heat/20.html", args)

# Test SPG
def pokazaniya_spg(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = get_data_table_by_date_spg(meters_name, parent_name, electric_data_end)
    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents_heat(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []
        for x in range(len(list_of_abonents_2)):
            data_table_temp = get_data_table_by_date_spg(list_of_abonents_2[x], parent_name, electric_data_end)
            if data_table_temp:            
                data_table.extend(data_table_temp)
            else:
                data_table.extend([[electric_data_end,list_of_abonents_2[x][0],'Н/Д','Н/Д','Н/Д']])
                
              
    else:
        data_table = [] 
        
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 

    return render(request, "data_table/gas/22.html", args)
    
    
def pokazaniya_sayany_v2(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, False)
    
    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 
      
    return render(request, "data_table/heat/30.html", args)
    
def pokazaniya_sayany_last(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, False)


    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        if (data_table[i][3] is None):
            #print data_table[i][1], meters_name
            data_table[i][0]=electric_data_end
            dt=common_sql.get_data_table_by_date_heat_sayany_v2(data_table[i][1], meters_name, None, True)
            if (len(dt)>0):
                data_table[i]=dt[0]
        data_table[i]=tuple(data_table[i])
    
    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 
      
    return render(request, "data_table/heat/32.html", args)
    
def heat_potreblenie_sayany(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']

    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_period_heat_sayany(meters_name, parent_name,electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_period_heat_sayany(meters_name, parent_name,electric_data_start, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    args['obj_title'] = meters_name 
      
    return render(request, "data_table/heat/33.html", args)

def electric_check_factory_numbers(request):
    args= {}
    
    data_table = []
    data_table = common_sql.get_data_table_diferent_numbers()
    
    args['data_table'] = data_table
    
    return render(request, "data_table/electric/40.html", args)

def water_by_date(request):
    args= {}
    is_abonent_level = re.compile(r'level2')
    is_object_level_2 = re.compile(r'level1')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    # is_electric_daily   = request.GET['is_electric_daily']
    # is_electric_current = request.GET['is_electric_current']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            # request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            # request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
    # dc - daily or current
    dc='daily'
    # if is_electric_current == "1":
    #     dc=u'current'
    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_water_by_date(meters_name, parent_name, electric_data_end, True,dc)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_water_by_date(meters_name, parent_name, electric_data_end, False,dc)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 
      
    return render(request, "data_table/water/38.html", args)
    
def water_potreblenie_pulsar(request):
    args= {}
    is_abonent_level = re.compile(r'level2')
    is_object_level_2 = re.compile(r'level1')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    electric_data_start   = request.GET['electric_data_start']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_end"]   = electric_data_start   = request.GET['electric_data_start']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']

    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_water_period_pulsar(meters_name, parent_name,electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_water_period_pulsar(meters_name, parent_name,electric_data_start, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
    # for i in range(len(data_table)):
    #     data_table[i]=list(data_table[i])
    #     num=data_table[i][3]
    #     if ('ХВС, №' in num) or ('ГВС, №' in num):
    #         num=num.replace(u'ХВС, №', ' ')
    #         num=num.replace(u'ГВС, №', ' ')
    #         data_table[i][3]=num
    #         #print num
    #     data_table[i]=tuple(data_table[i])
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    args['obj_title'] = meters_name 
      
    return render(request, "data_table/water/39.html", args)


def pokazaniya_water_hvs_tekon(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']            
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, 'Канал 1',  'Tekon_hvs',True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, 'Канал 1',  'Tekon_hvs', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 
      
    return render(request, "data_table/water/34.html", args)

def water_potreblenie_hvs_tekon(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name, electric_data_start, electric_data_end, 'Канал 1',  'Tekon_hvs', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name,electric_data_start, electric_data_end, 'Канал 1',  'Tekon_hvs', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    args['obj_title'] = meters_name 
      
    return render(request, "data_table/water/35.html", args)

def pokazaniya_water_gvs_tekon(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']            
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, 'Канал 2',  'Tekon_gvs', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, 'Канал 2',  'Tekon_gvs', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 
      
    return render(request, "data_table/water/36.html", args)

def water_potreblenie_gvs_tekon(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name, electric_data_start, electric_data_end, 'Канал 2',  'Tekon_gvs', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name,electric_data_start, electric_data_end, 'Канал 2',  'Tekon_gvs', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    args['obj_title'] = meters_name 
      
    return render(request, "data_table/water/37.html", args)

def tekon_heat_by_date(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']    
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']            
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_heat_daily(meters_name, parent_name, electric_data_end, 'Канал 3',  'Tekon_heat', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_heat_daily(meters_name, parent_name, electric_data_end, 'Канал 3',  'Tekon_heat', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 
      
    return render(request, "data_table/heat/50.html", args)

def tekon_heat_potreblenie(request):
    args= {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name, electric_data_start, electric_data_end, 'Канал 3',  'Tekon_heat', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name,electric_data_start, electric_data_end, 'Канал 3',  'Tekon_heat', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    args['obj_title'] = meters_name 
      
    return render(request, "data_table/heat/51.html", args)
    
def resources_all(request):
    args= {}
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']            
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']
            data_table = common_sql.get_data_table_report_all_res_period3(electric_data_start, electric_data_end)
            #data_table = common_sql.get_data_table_report_all_res_period(u'10.02.2017', u'20.02.2017')

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
#        
#    #удаляем из номеров счётчиков лишнее
    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        num=data_table[i][3]
        if ('ХВС, №' in num) or ('ГВС, №' in num):
            num=num.replace('ХВС, №', ' ')
            num=num.replace('ГВС, №', ' ')
            data_table[i][3]=num
            #print num
        data_table[i]=tuple(data_table[i])

    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
      
    return render(request, "data_table/9.html", args)
    
def resources_all_by_date(request):
    args= {}
    electric_data_end   = request.GET['electric_data_end']            
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            data_table = common_sql.get_data_table_report_all_res_by_date(electric_data_end)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
      
    return render(request, "data_table/42.html", args)
    
def resources_electric_by_date(request):
    args= {}
    electric_data_end   = request.GET['electric_data_end']            
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            data_table = common_sql.get_data_table_report_electric_res_by_date(electric_data_end)

    #zamenyem None na N/D vezde
#    if len(data_table)>0: 
#        data_table=common_sql.ChangeNull(data_table, None)

    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
      
    return render(request, "data_table/44.html", args)
    
def resources_water_by_date(request):
    args= {}
    electric_data_end   = request.GET['electric_data_end']            
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            data_table = common_sql.get_data_table_report_water_res_by_date(electric_data_end)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
      
    return render(request, "data_table/46.html", args)
    
def resources_heat_by_date_2(request):
    args= {}
    electric_data_end   = request.GET['electric_data_end']            
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            data_table = common_sql.get_data_table_report_heat_res_by_date(electric_data_end)
         
    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        if (data_table[i][5] is None):            
            dt=common_sql.get_data_table_by_date_heat_sayany_for_buhgaltery(data_table[i][0], data_table[i][8],electric_data_end)         
            if (len(dt)>0):                
                data_table[i]=dt[0]
        data_table[i]=tuple(data_table[i])
            
   #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    
      
    return render(request, "data_table/48.html", args)

def test_test(request):
    args={}
    args['test_test'] = 10
    countAll=300
    return render(request, "data_table/test/23.html", args)


def forma_80020_v2(request):
    args= {}
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']            
    
    data_table = []
    data_table_check_data_header = []
    data_table_check_data = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]    = electric_data_end    = request.GET['electric_data_end']
            request.session["electric_data_start"]  = electric_data_start  = request.GET['electric_data_start']
            request.session["obj_title"]            = group_name           = request.GET['obj_title']

            CLEAN_DOUBLE_30 = getattr(settings, 'CLEAN_DOUBLE_30', 'False')
            #Удаляем дубли получасовок, если такие имеются, рекомендовано для СЭТов!
            # Изменять в settings
            if CLEAN_DOUBLE_30:
                common_sql.del_double_30_by_dates(electric_data_start,electric_data_end)

            # Запрашиваем данные для первой таблицы с процентами опроса получасовок за указанный период
            data_table=common_sql.get_80020_statistic(group_name,electric_data_start,electric_data_end)
            if len(data_table)>0: 
                data_table=common_sql.ChangeNull(data_table,None)
            #___________________________________________________________________________________________
       
            #Заполняем list со значениями нужных параметров
            
            list_of_taken_params=common_sql.get_header_taken_params(group_name)
             #___________________________________________________________________________________________                   
            k=0
            data_table_check_data=[]
            for row in list_of_taken_params:
                factory_number_manual = row[0]                
                name_param = row[1]
                k+=1
                if k==1:                    
                    data_table_check_data=common_sql.get_A_R_energy_by_factory_number_period(factory_number_manual,electric_data_start,electric_data_end, name_param)                
                else:
                    temp_dt=[]
                    temp_dt =  common_sql.get_A_R_energy_by_factory_number_period(factory_number_manual,electric_data_start,electric_data_end, name_param)
                    if len(temp_dt)>0:
                        temp_dt=common_sql.ChangeNull(temp_dt,None)
                        data_table_check_data=add_1columns_to_dt(data_table_check_data,temp_dt,1)
                    # print factory_number_manual                    
                    # print temp_dt
                    # print k
                    # print temp_dt
               
            #Заголовок
            #В заголовок добавляем дату чуть позже, чтобы спокойно пройти по номерам и параметрам
            list_of_taken_params.insert(0,['Дата'])
            if len(list_of_taken_params)>1:
                data_table_check_data_header = list(list_of_taken_params)
            

    args['data_table'] = data_table
    args['data_table_check_data_header'] = data_table_check_data_header
    args['data_table_check_data'] = data_table_check_data
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    
    return render(request, "data_table/electric/41_2.html", args)
    
def pulsar_heat_period(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')

    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            
#*********************************************************************************************************************************************************************
            if (bool(is_abonent_level.search(obj_key))):
                data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title,electric_data_end, electric_data_start, True)
            elif (bool(is_object_level_2.search(obj_key))):
                data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title, electric_data_end,electric_data_start, False)
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render(request, "data_table/heat/59.html", args)

def pulsar_heat_period_2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')

    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            
#*********************************************************************************************************************************************************************
            if (bool(is_abonent_level.search(obj_key))):
                data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title,electric_data_end, electric_data_start, True)
            elif (bool(is_object_level_2.search(obj_key))):
                data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title, electric_data_end,electric_data_start, False)
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render(request, "data_table/heat/61.html", args)

def pulsar_heat_daily(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']


#*********************************************************************************************************************************************************************
            if (bool(is_abonent_level.search(obj_key))):
                data_table = common_sql.get_data_table_by_date_daily_pulsar_teplo(obj_parent_title, obj_title, electric_data_end, True)
            elif (bool(is_object_level_2.search(obj_key))):
                data_table = common_sql.get_data_table_by_date_daily_pulsar_teplo(obj_parent_title, obj_title, electric_data_end, False)
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render(request, "data_table/heat/56.html", args)
    
def water_pulsar_potreblenie_skladochnaya(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            #request.session["obj_parent_title"]    = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']              
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_for_period_Skladochnaya(obj_parent_title, obj_title, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_for_period_Skladochnaya(obj_parent_title, obj_title, electric_data_start,electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    

    return render(request, "data_table/water/67.html", args)
    
def pulsar_heat_daily_2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']


#*********************************************************************************************************************************************************************
            if (bool(is_abonent_level.search(obj_key))):
                data_table = common_sql.get_data_table_by_date_daily_pulsar_teplo(obj_parent_title, obj_title, electric_data_end, True)
            elif (bool(is_object_level_2.search(obj_key))):
                data_table = common_sql.get_data_table_by_date_daily_pulsar_teplo(obj_parent_title, obj_title, electric_data_end, False)
                
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    

    return render(request, "data_table/heat/62.html", args)

def pulsar_water_daily(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            #request.session["obj_parent_title"]    = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']            
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_daily(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_daily(obj_parent_title, obj_title, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull_and_LeaveEmptyCol(data_table, None, 7)
        
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
   

    return render(request, "data_table/water/58.html", args)
    
def pulsar_water_period(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            #request.session["obj_parent_title"]    = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']              
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_for_period(obj_parent_title, obj_title, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_for_period(obj_parent_title, obj_title, electric_data_start,electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
   

    return render(request, "data_table/water/57.html", args)
    
def pulsar_water_daily_row(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            #request.session["obj_parent_title"]    = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']            
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_daily_row(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_daily_row(obj_parent_title, obj_title, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull_for_pulsar(data_table)
        
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
   

    return render(request, "data_table/water/60.html", args)
    
def heat_elf_daily(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            #request.session["obj_parent_title"]    = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']            
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_elf_heat_daily(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_elf_heat_daily(obj_parent_title, obj_title, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)
        
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
   

    return render(request, "data_table/heat/64.html", args)
    
def heat_elf_period(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            #request.session["obj_parent_title"]    = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']   
            request.session["electric_data_end"]   = electric_data_start   = request.GET['electric_data_start'] 
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_elf_period(obj_parent_title, obj_title, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_elf_period(obj_parent_title, obj_title, electric_data_start,electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)
        
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
   

    return render(request, "data_table/heat/63.html", args)
    
def heat_water_elf_daily(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            #request.session["obj_parent_title"]    = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']            
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_elf_heat_water_daily(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_elf_heat_water_daily(obj_parent_title, obj_title, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)
        
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
   

    return render(request, "data_table/66.html", args)
def translate(name):
 
    #Заменяем пробелы и преобразуем строку к нижнему регистру
    name = str(name)
    name = name.replace(' ','-').lower()
 
    #
    transtable = (
        ## Большие буквы
        ("Щ", "Sch"),
        ("Щ", "SCH"),
        # two-symbol
        ("Ё", "Yo"),
        ("Ё", "YO"),
        ("Ж", "Zh"),
        ("Ж", "ZH"),
        ("Ц", "Ts"),
        ("Ц", "TS"),
        ("Ч", "Ch"),
        ("Ч", "CH"),
        ("Ш", "Sh"),
        ("Ш", "SH"),
        ("Ы", "Yi"),
        ("Ы", "YI"),
        ("Ю", "Yu"),
        ("Ю", "YU"),
        ("Я", "Ya"),
        ("Я", "YA"),
        # one-symbol
        ("А", "A"),
        ("Б", "B"),
        ("В", "V"),
        ("Г", "G"),
        ("Д", "D"),
        ("Е", "E"),
        ("З", "Z"),
        ("И", "I"),
        ("Й", "J"),
        ("К", "K"),
        ("Л", "L"),
        ("М", "M"),
        ("Н", "N"),
        ("О", "O"),
        ("П", "P"),
        ("Р", "R"),
        ("С", "S"),
        ("Т", "T"),
        ("У", "U"),
        ("Ф", "F"),
        ("Х", "H"),
        ("Э", "E"),
        ("Ъ", "`"),
        ("Ь", "'"),
        ## Маленькие буквы
        # three-symbols
        ("щ", "sch"),
        # two-symbols
        ("ё", "yo"),
        ("ж", "zh"),
        ("ц", "ts"),
        ("ч", "ch"),
        ("ш", "sh"),
        ("ы", "yi"),
        ("ю", "yu"),
        ("я", "ya"),
        # one-symbol
        ("а", "a"),
        ("б", "b"),
        ("в", "v"),
        ("г", "g"),
        ("д", "d"),
        ("е", "e"),
        ("з", "z"),
        ("и", "i"),
        ("й", "j"),
        ("к", "k"),
        ("л", "l"),
        ("м", "m"),
        ("н", "n"),
        ("о", "o"),
        ("п", "p"),
        ("р", "r"),
        ("с", "s"),
        ("т", "t"),
        ("у", "u"),
        ("ф", "f"),
        ("х", "h"),
        ("ъ", "`"),
        ("ь", "'"),
        ("э", "e"),
    )
    #перебираем символы в таблице и заменяем
    for symb_in, symb_out in transtable:
        name = name.replace(symb_in, symb_out)
    #возвращаем переменную
    return name
    
def makeOneCoords(graphic_data,numField1):
    labels=[]
    for i in range(len(graphic_data)):
        graphic_data[i]=list(graphic_data[i]) 
        date=graphic_data[i][numField1]   
        #print numField1, date         
        #print date 
        if (date=='Н/Д' or date is None or date==None): 
            labels.append(str(0))
        else:
            #print type(date)
            if type(date)==datetime.date:
                labels.append(date.strftime("%d-%m-%Y"))
            elif type(date)==float or type(date)==decimal.Decimal:
                labels.append(str(date))            
            elif type(date)==datetime.datetime:
                labels.append(date.strftime("%d-%m-%Y %H:%M"))
            elif type(date)==str:
                labels.append(str(translate(date)))
            else:
                labels.append(str(translate(date)))
            
    #print('labels',labels)
    return labels
  
def get_rgba_color(num_color):
    transp=str("0.3")
    if num_color == 1: return  str("rgba(255, 0, 0,"+ transp+")") #red
    if num_color == 2: return  str("rgba(50, 205, 50,"+ transp+")") #limeGreen
    if num_color == 3: return  str("rgba(255, 255, 0,"+ transp+")") #yellow
    if num_color == 4: return  str("rgba(0, 255, 255,"+ transp+")") #aqua
    if num_color == 5: return  str("rgba(70, 130, 180,"+ transp+")") #steelBlue
    if num_color == 6: return  str("rgba(0, 0, 128,"+ transp+")") #darkBlue
    if num_color == 7: return  str("128, 0, 0,"+ transp+")") #darkRed
    if num_color == 8: return  str("rgba(255, 0, 255,"+ transp+")") #fuksiya
    if num_color == 9: return  str("rgba(128, 0, 128,"+ transp+")") #purple
    if num_color == 10: return  str("rgba(255, 165, 0,"+ transp+")") #orange
    if num_color == 11: return  str("rgba(0, 255, 0,"+ transp+")") #lime
    if num_color == 12: return  str("rgba(255, 215, 0,"+ transp+")") #gold
    if num_color == 13: return  str("rgba(255, 20, 147,"+ transp+")") #deepPink
    if num_color == 14: return  str("rgba(127, 255, 212,"+ transp+")") #aquamarine
    if num_color == 15: return  str("rgba(210, 105, 30,"+ transp+")") #chocolate
    else: return str("rgba(0, 0, 0,"+ transp+")")
            
        
        
def electric_daily_graphic(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''

    if request.is_ajax():        
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
              
            if (is_electric_daily == '1') & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                #print obj_title, obj_parent_title,electric_data_start, electric_data_end
                params=['T0 A+','T1 A+','T2 A+','T3 A+', 'Электричество']
                data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
            else:
                pass
            
    AllData=[]
    Xcoord=[]
    
    if (len(data_table) >0):
        Xcoord=makeOneCoords(data_table,0) #label 
    
        AllData=[{str("data"):makeOneCoords(data_table,12), str("label"):str("delta T0"), str("backgroundColor"): get_rgba_color(5)},
             {str("data"):makeOneCoords(data_table,13), str("label"):str("delta T1"),  str("backgroundColor"): get_rgba_color(1)},
             {str("data"):makeOneCoords(data_table,14), str("label"):str("delta T2"),  str("backgroundColor"): get_rgba_color(10)},
             {str("data"):makeOneCoords(data_table,15), str("label"):str("delta T3"),  str("backgroundColor"): get_rgba_color(8)}]       
    
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['label'] = Xcoord
    args['AllData']=AllData
   

    return render(request, "data_table/electric/69.html", args)
    
def electric_potreblenie_3_zones_v3(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u'1'
    electric_data_start = u''
    electric_data_end = u''
    data_table_graphic = []
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            res='Электричество'
            
            if (is_electric_monthly=="1"):
                dm='monthly'
            else:
                dm='daily'
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                    isAbon=True                    
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table      
                    params=[u'T0 A+',u'T1 A+',u'T2 A+',u'T3 A+', u'Электричество']
                    data_table_graphic = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
                       
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                    isAbon=False
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table

                    data_table_graphic = common_sql.get_data_table_electric_between_for_obj(obj_title, obj_parent_title,electric_data_start, electric_data_end)

            #*********************************************************************************************************************************************************************
            elif (is_electric_delta == '1') &(bool(is_group_level.search(obj_key))):
                    data_table=common_sql.get_data_table_electric_period_for_group(obj_title,obj_parent_title, electric_data_start, electric_data_end, res)
                    #print data_table
                    request.session["data_table_export"] = data_table
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
            is_electric_delta = 0
    
    AllData=[]
    Xcoord=[]    
    if len( data_table_graphic) >0:
        Xcoord=makeOneCoords(data_table_graphic,0) #label 
    
        AllData=[{str("data"):makeOneCoords(data_table_graphic,12), str("label"):str("potreblenie T0"), str("backgroundColor"): get_rgba_color(5)},
             {str("data"):makeOneCoords(data_table_graphic,13), str("label"):str("potreblenie T1"),  str("backgroundColor"): get_rgba_color(1)},
             {str("data"):makeOneCoords(data_table_graphic,14), str("label"):str("potreblenie T2"),  str("backgroundColor"): get_rgba_color(10)},
             {str("data"):makeOneCoords(data_table_graphic,15), str("label"):str("potreblenie T3"),  str("backgroundColor"): get_rgba_color(8)}]
    
    #print AllData
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    args['label'] = Xcoord
    args['AllData']=AllData
    return render(request, "data_table/electric/91.html", args)
    
def pulsar_water_period_2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    data_table_graphic =[]
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']            
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']              
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_for_period(obj_parent_title, obj_title, electric_data_start, electric_data_end, True)
        data_table_graphic = common_sql.get_data_table_water_pulsar1_between_dates(obj_title, obj_parent_title,electric_data_start, electric_data_end, True)
                      
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_for_period(obj_parent_title, obj_title, electric_data_start,electric_data_end, False)
        data_table_graphic = common_sql.get_data_table_water_pulsar1_between_dates(obj_title, obj_parent_title,electric_data_start, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
    
    AllData=[]
    Xcoord=[]
    #print data_table_graphic
    if (len( data_table_graphic) >0):
        Xcoord=makeOneCoords(data_table_graphic,0) #label 
        
        AllData=[{str("data"):makeOneCoords(data_table_graphic,5), str("label"):str("GVS"), str("backgroundColor"): get_rgba_color(1)},
             {str("data"):makeOneCoords(data_table_graphic,6), str("label"):str("HVS"),  str("backgroundColor"): get_rgba_color(5)}]
                 
    
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['label'] = Xcoord
    args['AllData']=AllData

    return render(request, "data_table/water/73.html", args)
    

    
def heat_karat_daily(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    data_table=[]
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']       
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']            
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_karat_heat_water_daily(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_karat_heat_water_daily(obj_parent_title, obj_title, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)
        
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    return render(request, "data_table/heat/74.html", args)
    
def heat_karat_potreblenie(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']   
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start'] 
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_karat_potreblenie(obj_parent_title, obj_title, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_karat_potreblenie(obj_parent_title, obj_title,electric_data_start, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)
        
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    return render(request, "data_table/heat/75.html", args)
    
    
def balance_daily_electric(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']       
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']  
    #print obj_title
    type_resource='Электричество'
    if not(bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_balance_electric_daily(obj_parent_title, obj_title, electric_data_end,type_resource)
                  
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)
        
    AllData=[]
    Xcoord=[]
    data_table_graphic=data_table
    if len( data_table_graphic) > 0:
        Xcoord=makeOneCoords(data_table_graphic,2) #label 
    
        AllData=[{str("data"):makeOneCoords(data_table_graphic,3), str("label"):str("summa T0"), str("backgroundColor"): get_rgba_color(5)}]#,
#             {str("data"):makeOneCoords(data_table_graphic,13), str("label"):str("potreblenie T1"),  str("backgroundColor"): get_rgba_color(1)},
#             {str("data"):makeOneCoords(data_table_graphic,14), str("label"):str("potreblenie T2"),  str("backgroundColor"): get_rgba_color(10)},
#             {str("data"):makeOneCoords(data_table_graphic,15), str("label"):str("potreblenie T3"),  str("backgroundColor"): get_rgba_color(8)}]
             
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['label'] = Xcoord
    args['AllData']=AllData
    return render(request, "data_table/electric/76.html", args)

def IsEmptyTable(data_table, empty_field):
    isEmpty = True
    for row in data_table:
        if not(row[empty_field] == None  or row[empty_field] == 'None'):

            isEmpty = False
    return isEmpty

def balance_period_electric(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    decimal.getcontext().prec = 3
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']       
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']  
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']  
   
    AllData=[]
    Xcoord=[]    
    dtAll=[]
    dt_type_abon=common_sql.GetSimpleTable('types_abonents',"","")
    
    for i in range(0,len(dt_type_abon)):
         guid_type_abon=dt_type_abon[i][0]         
             
         if not(bool(is_abonent_level.search(obj_key))):
             #print(guid_type_abon, dt_type_abon[i][1])
             data_table = common_sql.get_data_table_balance_electric_perid(obj_parent_title, obj_title,electric_data_start, electric_data_end,guid_type_abon)
             if i==1:
                     Xcoord=makeOneCoords(data_table,5)
                     
             type_abon=translate(dt_type_abon[i][1])
             if IsEmptyTable(data_table, 0):
                 continue
             #print type_abon             
             if len(data_table)>0: 
                 data_table[0]=list(data_table[0])
                 data_table[0][6]="-"
                 data_table[0]=tuple(data_table[0])
                 data_table=common_sql.ChangeNull(data_table, None)                 
                 dtAll.append(data_table)
                 AllData.append({str("data"):makeOneCoords(data_table,6), str("label"):str(type_abon), str("backgroundColor"): get_rgba_color(i+2)})
                 
    dt_delta=[]   
    
    #print len(dtAll)
    if len(dtAll)>0:
        for j in range(1,len(dtAll[0])):
            sumD=0
            vv=0
            for i in range(0,len(dtAll)):
                #print i, j
                #print dtAll[i][j][1], dtAll[i][j][2],  dtAll[i][j][5], dtAll[i][j][6], dtAll[i][j][8]
                if (dtAll[i][j][6] == 'Н/Д' or dtAll[i][j][6] == None  or dtAll[i][j][6] == 'None'): 
                    break
                              
                if dtAll[i][j][1] == True:                   
                    sumD+=decimal.Decimal(dtAll[i][j][6])
                    vv=decimal.Decimal(dtAll[i][j][6])                    
                else:
                    sumD-=decimal.Decimal(dtAll[i][j][6])
            #считаем проценты
            percent=0           
            if (vv > decimal.Decimal(0)):
                percent=sumD*100/vv
               
            dt_delta.append([Xcoord[j],sumD, decimal.Decimal(percent)])

    #добавляем диаграмму с суммой по всем потребителям
    if len(dtAll)>2:
        type_abon='potrebiteli'
        data_table = common_sql.get_data_table_balance_electric_perid_potrebiteli(obj_parent_title, obj_title,electric_data_start, electric_data_end, type_abon) 
                  
        if len(data_table)>0: 
            data_table[0]=list(data_table[0])
            data_table[0][6]="-"
            data_table[0]=tuple(data_table[0])
            data_table=common_sql.ChangeNull(data_table, None)
            dtAll.append(data_table)
            AllData.append({str("data"):makeOneCoords(data_table,6), str("label"):str(type_abon), str("backgroundColor"): get_rgba_color(1)}) 

    args['dt_delta'] = dt_delta #небаланс по группе        
    args['data_table'] = dtAll
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['label'] = Xcoord
    args['AllData']=AllData
    return render(request, "data_table/electric/77.html", args)


def normalaize_data(data_table):
    new_dt =[]
    i = 0
    while i < len(data_table):
        if data_table[i][0] == "null" or data_table[i][0] == None:
            new_dt.append(data_table[i])
            if i+1 == len(data_table):
                break            
            else: 
                i=i+1
                new_dt.append([None, None, None,0,None, data_table[i][5],None,None,None])
        else:
            if i-1 > 0:
                if data_table[i-1][0] == "null" or data_table[i-1][0] == None:
                    new_dt.append([None, None, None,0,None, data_table[i][5],None,None,None])
                else:
                    new_dt.append(data_table[i])                    
        i+=1
    j=0
    while j<len(new_dt):
        #print(j, len(new_dt))
        if new_dt[j][6] != None:
            if new_dt[j][6] < 0 and j-1>=0:
                new_dt[j] = list(new_dt[j])
                new_dt[j][6] = None
                new_dt[j] = tuple(new_dt[j])
                new_dt[j-1] = list(new_dt[j-1])
                new_dt[j-1][6] = None
                new_dt[j-1] = tuple(new_dt[j-1])
        j+=1
 
    return new_dt

def balance_period_electric_2(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    decimal.getcontext().prec = 3
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']       
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']  
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']  
   
    AllData=[]
    Xcoord=[]    
    dtAll=[]
    dt_type_abon=common_sql.GetSimpleTable('types_abonents',"","")
    
    for i in range(0,len(dt_type_abon)):
         guid_type_abon=dt_type_abon[i][0]         
             
         if not(bool(is_abonent_level.search(obj_key))):
             #print(guid_type_abon, dt_type_'abon[i][1])
             data_table = common_sql.get_data_table_balance_electric_perid(obj_parent_title, obj_title,electric_data_start, electric_data_end,guid_type_abon)
             data_table = normalaize_data(data_table)
             if i==1:
                     Xcoord=makeOneCoords(data_table,5)
                     
             type_abon=translate(dt_type_abon[i][1])
             if IsEmptyTable(data_table, 0):
                 continue
             #print type_abon             
             if len(data_table)>0: 
                 data_table[0]=list(data_table[0])
                 data_table[0][6]="-"
                 data_table[0]=tuple(data_table[0])
                 data_table=common_sql.ChangeNull(data_table, None)                 
                 dtAll.append(data_table)
                 AllData.append({str("data"):makeOneCoords(data_table,6), str("label"):str(type_abon), str("backgroundColor"): get_rgba_color(i+2)})
                 
    dt_delta=[]   
    to_del =[]
    #print len(dtAll)
    if len(dtAll)>0:
        for j in range(1,len(dtAll[0])):
            sumD=0
            vv=0
            for i in range(0,len(dtAll)):
                #print(i, j)
                print (Xcoord[j], dtAll[i][j][1], dtAll[i][j][2],  dtAll[i][j][5], dtAll[i][j][6], dtAll[i][j][8])
                if (dtAll[i][j][6] == 'Н/Д' or dtAll[i][j][6] == None  or dtAll[i][j][6] == 'None' or dtAll[i][j][6] == 'null'):
                    print('break')
                    to_del.append(dtAll[i][j][5])
                    break
                #print('ok')
                if dtAll[i][j][1] == True:                   
                    sumD+=decimal.Decimal(dtAll[i][j][6])
                    vv=decimal.Decimal(dtAll[i][j][6])                    
                else:
                    sumD-=decimal.Decimal(dtAll[i][j][6])

            #считаем проценты
            percent=0           
            if (vv > decimal.Decimal(0)):
                percent=sumD*100/vv
            #print([Xcoord[j],sumD, decimal.Decimal(percent)])
            dt_delta.append([Xcoord[j],sumD, decimal.Decimal(percent)])

    # for row in dt_delta:
    #     for d in to_del:
    #         ddd = datetime.datetime.strptime(row[0], '%d-%m-%Y')
    #         #print(type(row[0]),type(ddd), type(d))
    #         #print(ddd.date(),d,ddd.date() == d)
    #         if ddd.date() == d:
    #             print(row)
    #             dt_delta.remove(row)

    #добавляем диаграмму с суммой по всем потребителям
    if len(dtAll)>2:
        type_abon='potrebiteli'
        data_table = common_sql.get_data_table_balance_electric_perid_potrebiteli(obj_parent_title, obj_title,electric_data_start, electric_data_end, type_abon) 
                  
        if len(data_table)>0: 
            data_table[0]=list(data_table[0])
            data_table[0][6]="-"
            data_table[0]=tuple(data_table[0])
            data_table=common_sql.ChangeNull(data_table, None)
            dtAll.append(data_table)
            AllData.append({str("data"):makeOneCoords(data_table,6), str("label"):str(type_abon), str("backgroundColor"): get_rgba_color(1)}) 

    args['dt_delta'] = dt_delta #небаланс по группе        
    args['data_table'] = dtAll
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['label'] = Xcoord
    args['AllData']=AllData
    return render(request, "data_table/electric/77.html", args)

def all_res_by_date(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    decimal.getcontext().prec = 3
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']       
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']  
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']  
   
    data_table=[]
    obj_parent_title_water=""
    obj_parent_title_electric=""
    if (bool(is_abonent_level.search(obj_key))):      
            obj_parent_title_water=str(obj_parent_title)+" Вода"
            obj_parent_title_electric=obj_parent_title
            #print  obj_parent_title_water, obj_parent_title_electric, len( obj_parent_title_electric)
            data_table = common_sql.get_data_table_all_res_for_abon(obj_parent_title_water, obj_parent_title_electric, obj_title, electric_data_end)
    if (bool(is_object_level.search(obj_key))):
        #здесь условно на уровне абонента
        n=str(obj_parent_title).find('Вода')
        #print obj_parent_title, n
        if (n>0):
            obj_parent_title_water=obj_parent_title
            obj_parent_title_electric=obj_parent_title[0:n-1]
            #print  obj_parent_title_water, obj_parent_title_electric, len( obj_parent_title_electric)
        data_table = common_sql.get_data_table_all_res_for_abon(obj_parent_title_water, obj_parent_title_electric, obj_title, electric_data_end)
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)         
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['data_table'] = data_table
    
    return render(request, "data_table/76.html", args)
    
def water_potreblenie_pulsar_with_graphic(request):
    args= {}
    is_abonent_level = re.compile(r'level2')
    is_object_level_2 = re.compile(r'level1')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    electric_data_start   = request.GET['electric_data_start']            
    obj_key             = request.GET['obj_key']
    
    data_table = []
    data_table_graphic=[]
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_end"]   = electric_data_start   = request.GET['electric_data_start']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']

    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_water_period_pulsar(meters_name, parent_name,electric_data_start, electric_data_end, True)
        data_table_graphic = common_sql.get_data_table_water_between(meters_name, parent_name,electric_data_start, electric_data_end,True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_water_period_pulsar(meters_name, parent_name,electric_data_start, electric_data_end, False)
        data_table_graphic = common_sql.get_data_table_water_between(meters_name, parent_name,electric_data_start, electric_data_end,False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        num=data_table[i][3]
        if ('ХВС, №' in num) or ('ГВС, №' in num):
            num=num.replace('ХВС, №', ' ')
            num=num.replace('ГВС, №', ' ')
            data_table[i][3]=num
            #print num
        data_table[i]=tuple(data_table[i])      
           
    AllData=[]
    Xcoord=[]
    
    if (len( data_table_graphic) >0):
        Xcoord=makeOneCoords(data_table_graphic,0) #label 
    
        AllData=[{str("data"):makeOneCoords(data_table_graphic,6), str("label"):str("potreblenie GVS"), str("backgroundColor"): get_rgba_color(1)},
             {str("data"):makeOneCoords(data_table_graphic,7), str("label"):str("potreblenie HVS"),  str("backgroundColor"): get_rgba_color(5)}]
             
    args['data_table'] = data_table
    args['obj_title'] = meters_name
    args['obj_key'] = obj_key
    args['obj_parent_title'] = parent_name
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    #args['dates'] = dates
    args['label'] = Xcoord
    args['AllData']=AllData
    return render(request, "data_table/water/79.html", args)
    
def water_consumption_impuls(request):
    args= {}
    is_abonent_level = re.compile(r'level2')
    is_object_level_2 = re.compile(r'level1')    
    obj_parent_title         = request.GET['obj_parent_title']
    obj_title         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    electric_data_start   = request.GET['electric_data_start']            
    obj_key             = request.GET['obj_key']    
    data_table = []

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_end"]   = electric_data_start   = request.GET['electric_data_start']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']

    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_dt_water_impulse_consumption(obj_title, obj_parent_title,electric_data_start, electric_data_end, True)        
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_dt_water_impulse_consumption(obj_title, obj_parent_title,electric_data_start, electric_data_end, False)
        
    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
             
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    return render(request, "data_table/water/101.html", args)


def pulsar_heat_period_with_graphic(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    
    data_table_graphic=[]
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']


#*********************************************************************************************************************************************************************
           
            if (bool(is_abonent_level.search(obj_key))):
                data_table = common_sql.(obj_parent_title, obj_title,electric_data_end, electric_data_start, True)
                data_table_graphic = common_sql.get_data_table_heat_between(obj_parent_title, obj_title,electric_data_start, electric_data_end,True)
            elif (bool(is_object_level_2.search(obj_key))):
                data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title, electric_data_end,electric_data_start, False)
                data_table_graphic = common_sql.get_data_table_heat_between(obj_parent_title, obj_title,electric_data_start, electric_data_end,False)
    AllData=[]
    Xcoord=[]
    
    if (len( data_table_graphic) >0):
        Xcoord=makeOneCoords(data_table_graphic,0) #label 
    
        AllData=[{str("data"):makeOneCoords(data_table_graphic,6), str("label"):str("potreblenie Energii"), str("backgroundColor"): get_rgba_color(12)},
             {str("data"):makeOneCoords(data_table_graphic,7), str("label"):str("potreblenie Ob'ema"),  str("backgroundColor"): get_rgba_color(14)}]
             
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['label'] = Xcoord
    args['AllData']=AllData
    

    return render(request, "data_table/heat/81.html", args)

 
def instruction_user(request):
    from django.contrib.staticfiles import finders
    result_url = finders.find('User_manual_Prizmer.pdf')

    with open(result_url, 'rb') as pdf:
        response = HttpResponse(pdf.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename=instruction.pdf'
        return response
    pdf.closed    
#    


def instruction_admin(request):
    from django.contrib.staticfiles import finders
    result_url = finders.find('Admin_manual_Prizmer.pdf')

    with open(result_url, 'rb') as pdf:
        response = HttpResponse(pdf.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename=instruction.pdf'
        return response
    pdf.closed       
    
    
#    with open('static/Admin_manual_Prizmer.pdf', 'rb') as pdf:
#    from django.contrib.staticfiles.storage import staticfiles_storage
#    url = staticfiles_storage.url('User_manual_Prizmer.pdf')
#    #print url
#    with open(url, 'rb') as pdf:
#        response = HttpResponse(pdf.read(), content_type='application/pdf')
#        response['Content-Disposition'] = 'inline; filename=instruction.pdf'
#        return response
#    pdf.closed

def add_3columns_to_dt(data_table,data_range,n1,n2,n3):
#    print 'len(data_table) in function ', len(data_table)
#    print 'len(data_range) in function ', len(data_range)
    for i in range(0,len(data_table)):
        data_table[i]=list(data_table[i]) 
#        print i
#        print data_range[i][n1]
        data_table[i].append(data_range[i][n1])
        data_table[i].append(data_range[i][n2])
        data_table[i].append(data_range[i][n3])
        #print data_table[i]
        data_table[i]=tuple(data_table[i])
    return data_table    

def add_2columns_to_dt(data_table,data_range,n1,n2):
#    print 'len(data_table) in function ', len(data_table)
#    print 'len(data_range) in function ', len(data_range)
    for i in range(0,len(data_table)):
        data_table[i]=list(data_table[i]) 
#        print i
#        print data_range[i][n1]
        data_table[i].append(data_range[i][n1])
        data_table[i].append(data_range[i][n2])        
        #print data_table[i]
        data_table[i]=tuple(data_table[i])
    return data_table                  

def add_1columns_to_dt(data_table,data_range,n1):
#    print 'len(data_table) in function ', len(data_table)
#    print 'len(data_range) in function ', len(data_range)
    for i in range(0,len(data_table)):
        data_table[i]=list(data_table[i]) 
#        print i
#        print data_range[i][n1]
        data_table[i].append(data_range[i][n1])                
        #print data_table[i]
        data_table[i]=tuple(data_table[i])
    return data_table 

def water_elf_potreblenie_monthly_with_delta(request):
    args = {}
#    is_abonent_level = re.compile(r'abonent')
#    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title'] 
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']  
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']  
    
    dt_date=[]
    dt_range=[]
    dt_date=common_sql.generate_monthly_range(electric_data_start,electric_data_end)
    double_dates=[]
   
    for row in range(0,len(dt_date)):
       data_start = dt_date[row][0].strftime("%d.%m.%Y")
       #print 'len(dt_date)', len(dt_date)
       #print 'row', row
       if (row+1)<len(dt_date):
           data_end = dt_date[row+1][0].strftime("%d.%m.%Y")
       else:
           data_end = dt_date[row][0].strftime("%d.%m.%Y")
       #print data_start, data_end
       double_dates.append('Значение на ' + data_start)
       double_dates.append('Значение на ' + data_end)
       double_dates.append('Разница')
       dt_range = common_sql.get_data_table_elf_period_monthly(data_start, data_end)       
       if row == 0:         
           data_table=dt_range           
       else:
                          
           data_table=add_3columns_to_dt(data_table,dt_range,4,5,6)
    val_num=0 
    count_month=[]            
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)    
        val_num= len(data_table[0]) - 6   
        
    count_month=list(range(1,len(dt_date)-1))
    if len(data_table)>0o3:
        double_dates.pop()
        double_dates.pop()
        double_dates.pop()
    
    
    #print 'val_num', val_num
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta  
    
    args['electric_data_end'] = dt_date[-1][0].strftime("%d.%m.%Y")
    args['electric_data_start'] = dt_date[0][0].strftime("%d.%m.%Y")
    args['dt_dates'] = double_dates
    args['count_month'] =  count_month
    args['val_num'] =  val_num

    return render(request, "data_table/water/83.html", args)
    
def water_elf_potreblenie(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title'] 
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']  
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']  
    
    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_water_period_elf(obj_title, obj_parent_title, electric_data_start, electric_data_end, True)
      
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_water_period_elf(obj_title, obj_parent_title, electric_data_start, electric_data_end, False)
       

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
    
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta  
    
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start

    return render(request, "data_table/water/85.html", args)
    
def water_elf_daily(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title'] 
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']  
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']  
    
    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_water_daily_elf(obj_title, obj_parent_title,  electric_data_end, True)
      
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_water_daily_elf(obj_title, obj_parent_title,  electric_data_end, False)
       

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
    
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta      
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start

    return render(request, "data_table/water/84.html", args)
    
def electric_res_status(request):
    args = {}

    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title'] 
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']  
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']  
    
    dt_objects = common_sql.get_res_objects('electric')
    #print 'print len(dt_objects) ', len(dt_objects)
    dtAll_statistic=[]
    dtAll_no_data_meters=[]
    for obj in dt_objects:
        #print obj[0]
        dt_statistic= common_sql.get_electric_count(obj[0],  electric_data_end)
        dt_no_data_meters=common_sql.get_electric_no_data(obj[0],  electric_data_end)
        dtAll_statistic.append(dt_statistic)
        if len(dt_no_data_meters)>0: 
            dt_no_data_meters=common_sql.ChangeNull(dt_no_data_meters, None)
            dtAll_no_data_meters.append(dt_no_data_meters)
      
    
    args['dtAll_statistic'] = dtAll_statistic
    args['dtAll_no_data_meters'] = dtAll_no_data_meters
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta      
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start

    return render(request, "data_table/86.html", args)
    
def heat_digital_res_status(request):
    args = {}

    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']             
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']  
    
    dt_objects = common_sql.get_res_objects('heat')
    #print 'print len(dt_objects) ', len(dt_objects)
    dtAll_statistic=[]
    dtAll_no_data_meters=[]
    for obj in dt_objects:
        #print obj[0]
        dt_statistic= common_sql.get_heat_count(obj[0],  electric_data_end)
        dt_no_data_meters=common_sql.get_heat_no_data(obj[0],  electric_data_end)
        dtAll_statistic.append(dt_statistic)
        if len(dt_no_data_meters)>0: 
            dt_no_data_meters=common_sql.ChangeNull(dt_no_data_meters, None)
            dtAll_no_data_meters.append(dt_no_data_meters)
      
   
    
    
    args['dtAll_statistic'] = dtAll_statistic
    args['dtAll_no_data_meters'] = dtAll_no_data_meters
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta      
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start

    return render(request, "data_table/88.html", args)
    
def water_impulse_res_status(request):
    args = {}

    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']             
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']  
    
    dt_objects = common_sql.get_water_impulse_objects()
    
    dtAll_statistic=[]
    dtAll_no_data_meters=[]
    for obj in dt_objects:
        #print electric_data_end
        dt_statistic= common_sql.get_water_impulse_count(obj[0],  electric_data_end)
        dt_no_data_meters=common_sql.get_water_impulse_no_data(obj[0],  electric_data_end)
        dtAll_statistic.append(dt_statistic)
        if len(dt_no_data_meters)>0: 
            dt_no_data_meters=common_sql.ChangeNull(dt_no_data_meters, None)
            dtAll_no_data_meters.append(dt_no_data_meters)
      
   
    
    
    args['dtAll_statistic'] = dtAll_statistic
    args['dtAll_no_data_meters'] = dtAll_no_data_meters
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta      
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start

    return render(request, "data_table/90.html", args)
    
def water_digital_pulsar_res_status(request):
    args = {}

    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']             
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']  
    
    dt_objects = common_sql.get_water_digital_pulsar_objects()
    
    dtAll_statistic=[]
    dtAll_no_data_meters=[]
    for obj in dt_objects:
        #print electric_data_end
        dt_statistic= common_sql.get_water_digital_pulsar_count(obj[0],  electric_data_end)
        dt_no_data_meters=common_sql.get_water_digital_pulsar_no_data(obj[0],  electric_data_end)
        dtAll_statistic.append(dt_statistic)
        if len(dt_no_data_meters)>0: 
            dt_no_data_meters=common_sql.ChangeNull(dt_no_data_meters, None)
            dtAll_no_data_meters.append(dt_no_data_meters)
      
   
    
    
    args['dtAll_statistic'] = dtAll_statistic
    args['dtAll_no_data_meters'] = dtAll_no_data_meters
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta      
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start

    return render(request, "data_table/94.html", args)


def balance_period_water_impulse(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')   
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    decimal.getcontext().prec = 3
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']       
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']  
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']  
   
    AllData=[]
    Xcoord=[]    
     
    if not(bool(is_abonent_level.search(obj_key))):
         data_table = common_sql.get_data_table_balance_water_impulse_perid(obj_parent_title, obj_title,electric_data_start, electric_data_end)
              
         
         
    if (len( data_table) >0):
        Xcoord=makeOneCoords(data_table,0) #label 
#        print data_table[0][1],data_table[0][2],data_table[0][3],data_table[0][4],data_table[0][5],data_table[0][6]
#        print data_table[1][1],data_table[1][2],data_table[1][3],data_table[1][4],data_table[1][5],data_table[1][6]
#        print data_table[2][1],data_table[2][2],data_table[2][3],data_table[2][4],data_table[2][5],data_table[2][6]
        AllData=[{str("data"):makeOneCoords(data_table,5), str("label"):str("potrebiteli(-)"), str("backgroundColor"): get_rgba_color(8)},
             {str("data"):makeOneCoords(data_table,6), str("label"):str("vvod(+)"),  str("backgroundColor"): get_rgba_color(14)}]
        
        data_table=common_sql.ChangeNull(data_table, None)
      
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['label'] = Xcoord
    args['AllData']=AllData
    return render(request, "data_table/water/87.html", args)

    
def all_res_status_monthly(request):
    args = {}    
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    decimal.getcontext().prec = 3
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title'] 
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']  
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']  
    #print electric_data_end
    dt_objects = common_sql.get_res_objects('electric')
    #print 'print len(dt_objects) ', len(dt_objects)
        
    dt_date=common_sql.get_date_month_range_by_date(electric_data_end)
    dtAll_el = []
    for dd in dt_date:
        dtAll_statistic=[]
        for obj in dt_objects:            
            dt_statistic= common_sql.get_electric_count(obj[0],  dd[0])           
            dtAll_statistic.append(dt_statistic)
        dtAll_el.append( dtAll_statistic)

    #heat
    dt_objects = common_sql.get_res_objects('heat')
    dtAll_h=[]
    for dd in dt_date:
        dtAll_statistic=[]
        for obj in dt_objects:        
            dt_statistic= common_sql.get_heat_count(obj[0],  dd[0])        
            dtAll_statistic.append(dt_statistic)
        dtAll_h.append( dtAll_statistic)
  

    #water
    dt_objects = common_sql.get_water_impulse_objects()
    dtAll_w=[]
    for dd in dt_date:
        dtAll_statistic=[]   
        for obj in dt_objects:        
            dt_statistic= common_sql.get_water_impulse_count(obj[0],  dd[0])       
            dtAll_statistic.append(dt_statistic)
        dtAll_w.append(dtAll_statistic)

        
    args['dtAll_el'] = dtAll_el
    args['dtAll_h'] = dtAll_h
    args['dtAll_w'] = dtAll_w
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta      
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    
    return render(request, "data_table/92.html", args)
    
def del_comment(request):
    if request.is_ajax():
        if request.method == 'GET':
            guid_comment       = request.GET['id']
            comm = Comments.objects.get(guid=guid_comment)
            comm.delete()

    #print data_table
    return redirect('..')


def comment(request):
    args = {}

    data_table = []

    if request.is_ajax():
        if request.method == 'GET':
            guid_abonent       = request.GET['id']
            resource           = request.GET['resource']
            guid_resource = ''
            if resource =='electric':
                guid_resource = 'ba710cff-e390-48ca-b442-70141c9864f7'
            if resource == 'heat':
                guid_resource = 'c0491ede-e00b-4e1d-a8ba-1ef61dba1cd3'
            if resource == 'water':
                guid_resource = '47f0b64c-2bf6-45b4-972b-601f473a3752'
            if resource == 'gvs':
                guid_resource = '57ec8f42-69c6-4f79-81bb-8ea139407aa9'
            if resource == 'impulse':
                guid_resource = '12574d66-2034-4c8e-8c8c-249757736858'
    if (not(guid_abonent is None) and not(guid_abonent=="")):
        data_table = common_sql.get_data_table_comments_for_abon(guid_abonent, guid_resource)    
    
    args['data_table'] = data_table
    if len(data_table)>0:
        args['object'] = data_table[0][5]
        args['abonent'] = data_table[0][4]
        args['guid_comment'] = data_table[0][0]
    #print data_table
    return render(request, "data_table/comment.html", args)

class AddCommentForm(forms.ModelForm):
    class Meta:
        model = Comments
        fields = ('comment','guid_abonents',)
 
@csrf_protect
@csrf_exempt       
def add_comment(request):
    args={}
    comment_status = 'Добавление нового комментария'
    #print('1111111')
    if request.method == "GET":        
        form=AddCommentForm()
        guid_abonent        =  request.GET['id']
        resource            =  request.GET['resource']
        #print('is get')
    args['form'] = form
    args['comment_status'] = comment_status
    args['guid_abonent'] = guid_abonent
    args['resource'] = resource

    return render(request, "data_table/add_comment.html", args)

@csrf_protect
@csrf_exempt       
def load_comment(request):
    comment_status = 'Добавление нового комментария'
    if request.method == "POST":      
        form = AddCommentForm(data=request.POST)  
        guid_abonent     =  request.POST['guid_abonents']
        resource         =  request.POST['resource']#request.GET.get('resource') #request.GET['resource'] #request.POST.get('resource', '')
        guid_resource = ''
        if resource =='electric':
            guid_resource = 'ba710cff-e390-48ca-b442-70141c9864f7'
        if resource == 'heat':
            guid_resource = 'c0491ede-e00b-4e1d-a8ba-1ef61dba1cd3'
        if resource == 'water':
            guid_resource = '47f0b64c-2bf6-45b4-972b-601f473a3752'
        if resource == 'gvs':
            guid_resource = '57ec8f42-69c6-4f79-81bb-8ea139407aa9'
        if resource == 'impulse':
            guid_resource = '12574d66-2034-4c8e-8c8c-249757736858'
        #print(resource)
        #print(guid_resource)
        if form.is_valid():
            Comments = form.save(commit=False)
            Comments.guid_abonents = Abonents.objects.get(guid=guid_abonent)
            if len(guid_resource) > 0:
                Comments.guid_resources = Resources.objects.get(guid=guid_resource)  
            Comments.date=datetime.datetime.now()          
            Comments.save()           
            comment_status = 'Комментарий добавлен'
    #return redirect('..')
    if resource =='electric':    return redirect('../electric')
    if resource =='water' or resource == 'impulse':    return redirect('../water')
    if resource =='heat':    return redirect('../heat')

#Разработка формы 80040 ___________________---------------------------------------_______________________________
    
def forma_80040(request):
    args= {}
    electric_data_start = request.GET['electric_data_start']
    electric_data_end   = request.GET['electric_data_end']            
    
    data_table = []
    data_table_check_data_header = []
    data_table_check_data = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]    = electric_data_end    = request.GET['electric_data_end']
            request.session["electric_data_start"]  = electric_data_start  = request.GET['electric_data_start']
            request.session["obj_title"]            = group_name           = request.GET['obj_title']
            
            # Формируем список дат на основе начальной и конечной даты полученной от web-календарей
            end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
            start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
            list_of_dates = [x for x in common_sql.daterange(start_date,
                          end_date,
                          step=datetime.timedelta(days=1),
                          inclusive=True)]

            # Делаем информационную табличку со счётчиками входящими в данную группу
            data_table = common_sql.get_info_group_80020_meters(group_name)
            #print group_name
            #print u'Это наш data_TABLE --> ', data_table
        
        # Делаем проверку на сумму профилей мощности и размности показаний счётчика
            # Узнаем начальные и конечные показания по счётчику
            # T0 A+ Начальная дата
                
            for y in range(len(data_table)):
                data_table[y] = list(data_table[y])
                my_parametr = 'T0 A+'
                result = common_sql.get_data_table_electric_parametr_daily_by_meters_number(data_table[y][3], electric_data_start, my_parametr)

                if result:
                    data_table[y].append(result[0][0])
                else:
                    data_table[y].append('Н/Д')
            # T0 A+ Конечная  дата
            for z in range(len(data_table)):
                data_table[z] = list(data_table[z])
                my_parametr = 'T0 A+'
                result = common_sql.get_data_table_electric_parametr_daily_by_meters_number(data_table[z][3], electric_data_end, my_parametr)
               
                if result:
                    data_table[z].append(result[0][0])
                else:
                    data_table[z].append('Н/Д')
                    
            for w in range(len(data_table)):
                try:
                    data_table[w].append(data_table[w][8]-data_table[w][7])
                except:
                    data_table[w].append('Н/Д')
                    
            #считаем сумму получасовок А+
            for m in range(len(data_table)):
                sum_of_30_min_a = 0
                for date in range(len(list_of_dates)-1):
                    try:
                        sum_of_30_min_a = sum_of_30_min_a + common_sql.get_sum_of_30_profil_by_meter_number(list_of_dates[date], data_table[m][3], 'A+ Профиль')
                    except:
                        pass
                try:
                    data_table[m].append(sum_of_30_min_a)
                except:
                    data_table[m].append('Н/Д')
                
                # Вычисляем процент сбора получасовок
                sum_of_count_30_min = 0
                for j in range (len(list_of_dates)-1):
                    sum_of_count_30_min = sum_of_count_30_min + common_sql.get_count_of_30_profil_by_meter_number(list_of_dates[j], data_table[m][3], 'A+ Профиль')
                try: # Если начальная и конечная даты совпадают, то получается деление на нуль
                    data_table[m].append((sum_of_count_30_min*100.0)/((len(list_of_dates)-1)*48))
                except: # В случае деления на нуль ничего не делаем... Ждем ввода двух разных дат
                    pass
                    
                                         
                    
            #Заполняем list со значениями нужных параметров
            list_of_taken_params = []
            for x in range(len(data_table)):
                #Получаем считываемые параметры по заводскому номеру прибора.
                 #A+
                name_of_type_meters = common_sql.get_name_of_type_meter_by_serial_number(data_table[x][3])
                guid_params = ''
                if name_of_type_meters[0][0] == 'Меркурий 230-УМ':
                    guid_params = '922ad57c-8f5e-4f00-a78d-e3ba89ef859f'
                elif name_of_type_meters[0][0] == 'Меркурий 230':
                    guid_params = '6af9ddce-437a-4e07-bd70-6cf9dcc10b31'
                else:
                    pass
                result = common_sql.get_taken_param_by_meters_number_and_guid_params(data_table[x][3], guid_params)
                list_of_taken_params.append(str(result[0][0]) + ' ' + str(result[0][1]))
                 #R+
                if name_of_type_meters[0][0] == 'Меркурий 230-УМ':
                     guid_params = '61101fa3-a96a-4934-9482-e32036c12829'
                elif name_of_type_meters[0][0] == 'Меркурий 230':
                     guid_params = '66e997c0-8128-40a7-ae65-7e8993fbea61'
                else:
                    pass
                result = common_sql.get_taken_param_by_meters_number_and_guid_params(data_table[x][3], guid_params)
                list_of_taken_params.append(str(result[0][0]) + ' ' + str(result[0][1]))
                                
            # Добавляем дату в лист с параметрами и делаем таблицу для шапки таблицы 
            list_of_taken_params.insert(0, 'Дата')
            data_table_check_data_header = list_of_taken_params
            
                     
        
        #Проверяем сколько получасовок имеем за каждые сутки промежутка по каждому считываемому параметру
            for x in range(len(list_of_dates)):
                list_of_one_date_check = []
                list_of_one_date_check.append(list_of_dates[x])
                for y in range(1, len(list_of_taken_params)):
                    my_split_params = list_of_taken_params[y].split()
                    my_names_param = my_split_params[1] + ' ' + my_split_params[2]
                    
                    list_of_one_date_check.append(common_sql.get_count_of_30_profil_by_meter_number(list_of_dates[x], my_split_params[0], my_names_param))
                data_table_check_data.append(list_of_one_date_check)
            
            data_table_check_data.pop()
                                                
            
    args['data_table'] = data_table
    args['data_table_check_data_header'] = data_table_check_data_header
    args['data_table_check_data'] = data_table_check_data
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    
    return render(request, "data_table/electric/71.html", args)

def electric_report_for_c300(request):
    args = {}    
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    decimal.getcontext().prec = 3
    data_table=[]

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title'] 
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']  
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']  
    
    if (bool(is_abonent_level.search(obj_key))): 
        pass
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_electric_period_c300(obj_parent_title, obj_title ,electric_data_start, electric_data_end)

    i=0
    while i<len(data_table):
        #print data_table[i][6]
        data_table[i]=list(data_table[i])
        data_table[i+1]=list(data_table[i+1])
        data_table[i+2]=list(data_table[i+2])

        if data_table[i][8] == '-' or data_table[i+1][8] =='-' or data_table[i+2][8] == '-':
            data_table[i][8]     = '-'
            data_table[i+1][8]   = '-'
            data_table[i+2][8]   = '-'
            data_table[i][10]     = '-'
            data_table[i+1][10]   = '-'
            data_table[i+2][10]   = '-'
        
        if data_table[i][9] == '-' or data_table[i+1][9] =='-' or data_table[i+2][9] == '-':
            data_table[i][9]     = '-'
            data_table[i+1][9]   = '-'
            data_table[i+2][9]   = '-'
            data_table[i][10]     = '-'
            data_table[i+1][10]   = '-'
            data_table[i+2][10]   = '-'

        data_table[i]   = tuple(data_table[i])
        data_table[i+1] = tuple(data_table[i+1])
        data_table[i+2] = tuple(data_table[i+2])
        i+=3   

    # zamenyem None na N/D vezde
    # if len(data_table)>0: 
    #     data_table=common_sql.ChangeNull(data_table, None) 
        
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta      
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    
    return render(request, "data_table/89.html", args)

def water_impulse_report_for_c300(request):
    args = {}    
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    is_object_level_1 = re.compile(r'level')
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    decimal.getcontext().prec = 3
    data_table=[]
  
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title'] 
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']  
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']  
    
    # print (bool(is_object_level_2.search(obj_key)))
    # print (bool(is_object_level_1.search(obj_key))) 
    # print bool(is_abonent_level.search(obj_key))
    if (bool(is_object_level_1.search(obj_key))) : 
         data_table = common_sql.get_data_table_water_period_c300(obj_parent_title, obj_title ,electric_data_start, electric_data_end)
    else:
        pass
       
          
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta      
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    
    return render(request, "data_table/93.html", args)


def electric_period_graphic_activ_reactiv(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''

    if request.is_ajax():        
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   = request.GET['is_electric_delta']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
              
            if (is_electric_daily == '1') & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                #print obj_title, obj_parent_title,electric_data_start, electric_data_end
                params=['T0 A+','T0 R+', 'Электричество']
                data_table= common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
            else:
                pass
            
    AllData=[]
    Xcoord=[]
    
    if (len(data_table) >0):
        Xcoord=makeOneCoords(data_table,0) #label 
    
        AllData=[{str("data"):makeOneCoords(data_table,10), str("label"):str("delta T0 A+"), str("backgroundColor"): get_rgba_color(5)},
             {str("data"):makeOneCoords(data_table,11), str("label"):str("delta T0 R+"),  str("backgroundColor"): get_rgba_color(1)}]       
    
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['label'] = Xcoord
    args['AllData']=AllData
   

    return render(request, "data_table/electric/95.html", args)


def water_current_impulse(request):
    args= {}
    is_abonent_level = re.compile(r'level2')
    is_object_level_2 = re.compile(r'level1')
    
    parent_name         = request.GET['obj_parent_title']
    meters_name         = request.GET['obj_title']
    electric_data_end   = request.GET['electric_data_end']            
    obj_key             = request.GET['obj_key']    
    
    data_table = []
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
           
    # dc - daily or current
    dc='current'

    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_water_current(meters_name, parent_name, electric_data_end, True,dc)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_water_current(meters_name, parent_name, electric_data_end, False,dc)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)
    
    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = meters_name 
      
    return render(request, "data_table/water/96.html", args)

def electric_restored_activ_reactiv_daily(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''    
    electric_data_end = ''

    if request.is_ajax():        
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']            
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
        
        d= datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
        electric_data_start=datetime.date(d.year, d.month, 1)
        if (bool(is_abonent_level.search(obj_key))):   #             
            params=['T0 A+','T0 R+', 'Электричество']
            dt = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
            #print dt
            i=len(dt) - 1

            if not(dt[i][5] == 'Н/Д') and not(dt[i][6] == 'Н/Д'):
                data_table=[dt[i]] # на дату есть срез, просто выводим daily_value        
            
            else:
                for row in reversed(dt):
                    #print row
                    #print row[5],row[6]
                    if (row[5] == 'Н/Д') or (row[6] == 'Н/Д'): #activ-5, reactiv-6                        
                        continue
                    else: #начинаем суммировать получасовки
                        date = row[0]
                        activ = row[5]
                        reactiv = row[6]
                        date2 = d - datetime.timedelta(days=1)
                        data_table = common_sql.get_restored_activ_reactiv(obj_title, obj_parent_title, date, activ,reactiv,date2,electric_data_end)
                        break
                #если на 1 число нет суточных, то берем месячные
                if ((dt[0][5] == 'Н/Д') or (dt[0][6] == 'Н/Д')) and len(data_table)<1:
                    #print 'monthly'
                    dt_monthly = common_sql.get_dt_monthly_activ_reactiv(obj_title, obj_parent_title, electric_data_end)
                    if len(dt_monthly)>0:
                        activ = dt_monthly[4]
                        reactiv = dt_monthly[5]
                        date2 = d - datetime.timedelta(days=1)
                        data_table = common_sql.get_restored_activ_reactiv(obj_title, obj_parent_title, date, activ,reactiv,date2,electric_data_end)

            if(datetime.datetime.now().date() < d.date()):
                data_table = []
        else:
            pass
    

    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = obj_title 
      
    return render(request, "data_table/electric/98.html", args)

def heat_danfoss_period(request):
       
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''    
    electric_data_end = ''

    if request.is_ajax():        
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']            
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']
        
        dc='current'

        if (bool(is_abonent_level.search(obj_key))): 
            data_table = common_sql.get_data_table_danfoss_period(obj_parent_title, obj_title, electric_data_start, electric_data_end, True,dc)
        elif (bool(is_object_level_2.search(obj_key))):
            data_table = common_sql.get_data_table_danfoss_period(obj_parent_title, obj_title, electric_data_start, electric_data_end, False,dc)
            
    

    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['electric_data_start'] = electric_data_start
    args['obj_title'] = obj_title 
      
    return render(request, "data_table/heat/97.html", args)

def heat_danfoss_daily(request):
    # 
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''    
    electric_data_end = ''

    if request.is_ajax():        
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title   = request.GET['obj_parent_title']            
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_monthly"]   = is_electric_monthly   = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"]   = is_electric_current   = request.GET['is_electric_current']
                
        #print 'Danfoss'
        dc = 'daily'
        
        if (bool(is_abonent_level.search(obj_key))): 
            data_table = common_sql.get_data_table_heat_danfos_daily(obj_parent_title, obj_title, electric_data_end, True,dc)
        elif (bool(is_object_level_2.search(obj_key))):
            data_table = common_sql.get_data_table_heat_danfos_daily(obj_parent_title, obj_title, electric_data_end, False,dc)

    args['data_table'] = data_table
    args['electric_data_end'] = electric_data_end
    args['obj_title'] = obj_title 
      
    return render(request, "data_table/heat/100.html", args)

def electric_period_30(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table = []
    obj_title = 'Не выбран'
    obj_key = 'Не выбран'
    obj_parent_title = 'Не выбран'
    is_electric_monthly = ''
    is_electric_daily = ''
    is_electric_current = ''
    is_electric_delta = ''
    electric_data_start = ''
    electric_data_end = ''
    data_table = []
    if request.is_ajax():        
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']            
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
              
            if (bool(is_abonent_level.search(obj_key))):   #  Получасовки по абоненту
                params=['A+ Профиль','R+ Профиль']
                data_table= common_sql.get_electric_30_by_abonent_for_period(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
            else:
                pass

     #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
    #for row in data_table:
        #print row[0], row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10]
    AllData=[]
    Xcoord=[]   
    #print data_table 
    if (len(data_table) > 0):
        Xcoord=makeOneCoords(data_table,5) #дата и время     
        #AllData=[{str("data"):makeOneCoords(data_table,6), str("label"):str("A+"), str("backgroundColor"): get_rgba_color(5)}] 
        AllData=[{str("data"):makeOneCoords(data_table,6), str("label"):str("A+"), str("backgroundColor"): get_rgba_color(12)},
              {str("data"):makeOneCoords(data_table,7), str("label"):str("R+"),  str("backgroundColor"): get_rgba_color(9)}]       

    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['label'] = Xcoord
    args['AllData']=AllData
    return render(request, "data_table/electric/99.html", args)

def electric_3_zones(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
                
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # месячные для абонента
                data_table = common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'monthly', True)
                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # суточные для абонента
                data_table = common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'daily', True)

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # месячные для объекта
                    data_table = common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'monthly', False)
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]        

#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for object
                    data_table= common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'daily', False)
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]

            elif (is_electric_daily == '1') & (bool(is_group_level.search(obj_key))): # показания по баланскной группе                    
                    data_table = common_sql.get_electric_by_date_balance(obj_parent_title, obj_title, electric_data_end, 'daily')
                   
              
            elif (is_electric_monthly == '1') & (bool(is_group_level.search(obj_key))): # показания по баланскной группе месячные
                    data_table = common_sql.get_electric_by_date_balance(obj_parent_title, obj_title, electric_data_end, 'monthly')

#*********************************************************************************************************************************************************************
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 

    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates    

    return render(request,"data_table/electric/102.html", args)

def electric_2_zones(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']

            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # месячные для абонента
                data_table = common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'monthly', True)
                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # суточные для абонента
                data_table = common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'daily', True)

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # месячные для объекта
                    data_table= common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'monthly', False)
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]        

#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for object
                    data_table= common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'daily', False)
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]

            elif (is_electric_daily == '1') & (bool(is_group_level.search(obj_key))): # показания по баланскной группе                    
                    data_table = common_sql.get_electric_by_date_balance(obj_parent_title, obj_title, electric_data_end, 'daily')
                   
              
            elif (is_electric_monthly == '1') & (bool(is_group_level.search(obj_key))): # показания по баланскной группе месячные
                    data_table = common_sql.get_electric_by_date_balance(obj_parent_title, obj_title, electric_data_end, 'monthly')

#*********************************************************************************************************************************************************************
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 

    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates    

    return render(request,"data_table/electric/104.html", args)

def electric_1_zones(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']

            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # месячные для абонента
                data_table = common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'monthly', True)
                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # суточные для абонента
                data_table = common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'daily', True)

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # месячные для объекта
                    data_table= common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'monthly', False)
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]        

#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for object
                    data_table= common_sql.get_electric_by_date(obj_parent_title, obj_title, electric_data_end, 'daily', False)
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]

            elif (is_electric_daily == '1') & (bool(is_group_level.search(obj_key))): # показания по баланскной группе                    
                    data_table = common_sql.get_electric_by_date_balance(obj_parent_title, obj_title, electric_data_end, 'daily')
                   
              
            elif (is_electric_monthly == '1') & (bool(is_group_level.search(obj_key))): # показания по баланскной группе месячные
                    data_table = common_sql.get_electric_by_date_balance(obj_parent_title, obj_title, electric_data_end, 'monthly')

#*********************************************************************************************************************************************************************
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 

    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates    

    return render(request,"data_table/electric/106.html", args)

def electric_consumption_2_zones(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u'1'
    electric_data_start = u''
    electric_data_end = u''
    data_table_graphic = []
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            res='Электричество'
            
            if (is_electric_monthly=="1"):
                dm='monthly'
            else:
                dm='daily'
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                    isAbon=True                    
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table      
                    params=[u'T0 A+',u'T1 A+',u'T2 A+',u'T3 A+', u'Электричество']
                    data_table_graphic = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
                       
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                    isAbon=False
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table

                    data_table_graphic = common_sql.get_data_table_electric_between_for_obj(obj_title, obj_parent_title,electric_data_start, electric_data_end)

            #*********************************************************************************************************************************************************************
            elif (is_electric_delta == '1') &(bool(is_group_level.search(obj_key))):
                    data_table=common_sql.get_data_table_electric_period_for_group(obj_title,obj_parent_title, electric_data_start, electric_data_end, res)
                    #print data_table
                    request.session["data_table_export"] = data_table
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
            is_electric_delta = 0
    
    AllData=[]
    Xcoord=[]    
    if len( data_table_graphic) >0:
        Xcoord=makeOneCoords(data_table_graphic,0) #label 
    
        AllData=[{str("data"):makeOneCoords(data_table_graphic,12), str("label"):str("potreblenie T0"), str("backgroundColor"): get_rgba_color(5)},
             {str("data"):makeOneCoords(data_table_graphic,13), str("label"):str("potreblenie T1"),  str("backgroundColor"): get_rgba_color(1)},
             {str("data"):makeOneCoords(data_table_graphic,14), str("label"):str("potreblenie T2"),  str("backgroundColor"): get_rgba_color(10)},
             {str("data"):makeOneCoords(data_table_graphic,15), str("label"):str("potreblenie T3"),  str("backgroundColor"): get_rgba_color(8)}]
    
    #print AllData
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    args['label'] = Xcoord
    args['AllData']=AllData
    return render(request,"data_table/electric/103.html", args)

def electric_consumption_1_zone(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u'1'
    electric_data_start = u''
    electric_data_end = u''
    data_table_graphic = []
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            res='Электричество'
            
            if (is_electric_monthly=="1"):
                dm='monthly'
            else:
                dm='daily'
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                    isAbon=True                    
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table      
                    params=[u'T0 A+',u'T1 A+',u'T2 A+',u'T3 A+', u'Электричество']
                    data_table_graphic = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
                       
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                    isAbon=False
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table

                    data_table_graphic = common_sql.get_data_table_electric_between_for_obj(obj_title, obj_parent_title,electric_data_start, electric_data_end)

            #*********************************************************************************************************************************************************************
            elif (is_electric_delta == '1') &(bool(is_group_level.search(obj_key))):
                    data_table=common_sql.get_data_table_electric_period_for_group(obj_title,obj_parent_title, electric_data_start, electric_data_end, res)
                    #print data_table
                    request.session["data_table_export"] = data_table
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
            is_electric_delta = 0
    
    AllData=[]
    Xcoord=[]    
    if len( data_table_graphic) >0:
        Xcoord=makeOneCoords(data_table_graphic,0) #label 
    
        AllData=[{str("data"):makeOneCoords(data_table_graphic,12), str("label"):str("potreblenie T0"), str("backgroundColor"): get_rgba_color(5)},
             {str("data"):makeOneCoords(data_table_graphic,13), str("label"):str("potreblenie T1"),  str("backgroundColor"): get_rgba_color(1)},
             {str("data"):makeOneCoords(data_table_graphic,14), str("label"):str("potreblenie T2"),  str("backgroundColor"): get_rgba_color(10)},
             {str("data"):makeOneCoords(data_table_graphic,15), str("label"):str("potreblenie T3"),  str("backgroundColor"): get_rgba_color(8)}]
    
    #print AllData
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    args['label'] = Xcoord
    args['AllData']=AllData
    return render(request,"data_table/electric/105.html", args)


def electric_by_date_podolsk(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']

            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # месячные для абонента
                data_table = common_sql.get_electric_by_date_podolsk(obj_parent_title, obj_title, electric_data_end, 'monthly', True)
                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # суточные для абонента
                data_table = common_sql.get_electric_by_date_podolsk(obj_parent_title, obj_title, electric_data_end, 'daily', True)

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # месячные для объекта
                    data_table= common_sql.get_electric_by_date_podolsk(obj_parent_title, obj_title, electric_data_end, 'monthly', False)
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]        

#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # суточные для объекта
                    data_table= common_sql.get_electric_by_date_podolsk(obj_parent_title, obj_title, electric_data_end, 'daily', False)
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 

    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates    

    return render(request,"data_table/electric/108.html", args)

def electric_consumption_podolsk(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u'1'
    electric_data_start = u''
    electric_data_end = u''
    data_table_graphic = []
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            res='Электричество'
            
            if (is_electric_monthly=="1"):
                dm='monthly'
            else:
                dm='daily'
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                    isAbon=True                    
                    data_table=common_sql.get_data_table_electric_period_podolsk(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table      
                       
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                    isAbon=False
                    data_table=common_sql.get_data_table_electric_period_podolsk(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table

            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
            is_electric_delta = 0
        
    #print AllData
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    return render(request,"data_table/electric/107.html", args)

def meter_info(request):
    args = {}
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    tcp_status = u''
    com_status = u''

    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            
            if (obj_key.find('meter') >= 0):
                #print('is meter!!!!!!!!!!!!!!!!')
                data_table = common_sql.get_meter_info_by_number(obj_parent_title, obj_title)
                dt_tcp = common_sql.get_tcp_ip_info_by_meter(obj_title)
                dt_com = common_sql.get_com_info_by_meter(obj_title)
                if len(dt_tcp) > 1 : tcp_status = 'Больше одного подключения!'
                elif len(dt_tcp) == 0:
                    tcp_status = 'Нет привязок'
                    dt_tcp = [[u'-', u'-', u'-', u'-', u'-', u'-', u'-', u'-', u'-', u'-', u'-', u'-', u'-', u'-']]
                if len(dt_com) > 1 : com_status = 'Больше одного подключения!'
                elif len(dt_com) == 0: 
                    com_status = 'Нет привязок'
                    dt_com = [[u'-', u'-', u'-', u'-', u'-', u'-', u'-', u'-', u'-', u'-', u'-', u'-', u'-', u'-']]

    args['factory_number'] = obj_title
    args['data_table'] = data_table
    args['dt_tcp'] = dt_tcp
    args['dt_com'] = dt_com
    args['com_status'] = com_status
    args['tcp_status'] = tcp_status
    return render(request,"data_table/meter_info.html", args)

def water_tem104_daily(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    #is_electric_monthly = u''
    is_electric_daily = u''
    electric_data_start = u''
    electric_data_end = u''
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            #request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            #request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            isWater = True
            if (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # суточные для абонента
                data_table = common_sql.get_tem104_by_date(obj_parent_title, obj_title, electric_data_end, isWater, True)
                args['obj_title'] = obj_parent_title
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # суточные для объекта
                data_table= common_sql.get_tem104_by_date(obj_parent_title, obj_title, electric_data_end, isWater, False)
                args['obj_title'] = obj_title

    args['data_table'] = data_table
    #args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    #args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    return render(request,"data_table/water/110.html", args)

def water_tem104_consumption(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    #is_electric_monthly = u''
    is_electric_daily = u''
    electric_data_start = u''
    electric_data_end = u''
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            #request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            isWater = True
            if (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # суточные для абонента
                data_table = common_sql.get_tem104_consumption(obj_parent_title, obj_title,electric_data_start, electric_data_end, isWater, True)
                args['obj_title'] = obj_parent_title
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # суточные для объекта
                data_table= common_sql.get_tem104_consumption(obj_parent_title, obj_title, electric_data_start, electric_data_end, isWater, False)
                args['obj_title'] = obj_title

    args['data_table'] = data_table
    #args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    #args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    return render(request,"data_table/water/109.html", args)

def heat_tem104_daily(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    #is_electric_monthly = u''
    is_electric_daily = u''
    electric_data_start = u''
    electric_data_end = u''
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            #request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            #request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            isWater = False
            if (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # суточные для абонента
                data_table = common_sql.get_tem104_by_date(obj_parent_title, obj_title, electric_data_end, isWater, True)
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # суточные для объекта
                data_table= common_sql.get_tem104_by_date(obj_parent_title, obj_title, electric_data_end, isWater, False)

    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    #args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    return render(request,"data_table/heat/112.html", args)

def heat_tem104_consumption(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    #is_electric_monthly = u''
    is_electric_daily = u''
    electric_data_start = u''
    electric_data_end = u''
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            #request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            isWater = True
            if (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # суточные для абонента
                data_table = common_sql.get_tem104_consumption(obj_parent_title, obj_title,electric_data_start, electric_data_end, isWater, True)
                args['obj_title'] = obj_parent_title
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # суточные для объекта
                data_table= common_sql.get_tem104_consumption(obj_parent_title, obj_title, electric_data_start, electric_data_end, isWater, False)
                args['obj_title'] = obj_title

    args['data_table'] = data_table
    #args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    #args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    return render(request,"data_table/heat/111.html", args)

def normalaize_data_consumption(data_table, col=12):
    new_dt =[]
    
    i = 0
    while i < len(data_table):
        if data_table[i][col] == 'Н/Д':
            new_dt.append(data_table[i])
            i+=1
            continue
        if float(data_table[i][col]) >= 0:
            new_dt.append(data_table[i])
        else:
            new_dt.append([data_table[i][0], None, None, None, None, None, None, None, None, None, None, None, None, None, None, None])
            if i-1 > 0:
                new_dt[i-1] = list(new_dt[i-1])
                new_dt[i-1] = [data_table[i-1][0], None, None, None, None, None, None, None, None, None, None, None, None, None, None, None]
                new_dt[i-1] = tuple(new_dt[i-1])
        i+=1

    return new_dt

def get_aver_for_7_days(data_table_graphic):
    i = len(data_table_graphic)-1
    sum=0
    j = 0
    while i > 0:
        if len(data_table_graphic) - i > 7:
            #print('break', len(data_table_graphic), i)
            break
        if data_table_graphic[i][12] == None or data_table_graphic[i][12] == 'Н/Д' or data_table_graphic[i][12] == 0:
            #print('none', data_table_graphic[i][12])
            i-=1
        else:
            if float(data_table_graphic[i][12]) > 0:
                sum +=data_table_graphic[i][12]
                #print(i, data_table_graphic[i][12])
                j+=1
            i-=1
            
    #print(sum, 'sum')
    if j > 0: div = j
    else: div = 7
    #print(div,'div')
    aver = sum/div
    #print(aver, 'aver')
    return aver

def analize_consumption(data_table, electric_data_start, electric_data_end, obj_title, obj_parent_title ):
    new_dt = []
    params=[u'T0 A+',u'T1 A+',u'T2 A+',u'T3 A+', u'Электричество']
    for row in data_table:
        new_row = []
        new_row = list(row)
        #print(row[0], row[1], obj_title, obj_parent_title)
        if len(data_table)==1:
            obj_title = obj_parent_title
        data_table_graphic = common_sql.get_data_table_electric_between(row[0],obj_title,electric_data_start, electric_data_end, params)
        data_table_graphic = normalaize_data_consumption(data_table_graphic)
        meter = row[1]
        #print(meter)
        #print(data_table_graphic)
        aver = get_aver_for_7_days(data_table_graphic)
        max = aver*3
        min = aver/3
        #print('min',min, '   max', max)
        status = 0 # 1-больше, 2 -меньше
        date_status = ''
        for x in data_table_graphic:
            if x[12] == None or x[12] == 'Н/Д' or x[12] == 0:
                continue
            if x[12] > max:
                status = 1
                date_status = x[0]
                break
            if x[12]<min:
                status = 2
                date_status = x[0]
                break

        new_row.append(status)
        new_row.append(date_status)
        new_row.append(min)
        new_row.append(max)
        new_row = tuple(new_row)
        new_dt.append(new_row)
        #print(new_row)
    return new_dt

def electric_consumption_3_zones_with_limit(request):
    args = {}
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    data_table = []
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    is_electric_monthly = u''
    is_electric_daily = u''
    is_electric_current = u''
    is_electric_delta = u'1'
    electric_data_start = u''
    electric_data_end = u''
    data_table_graphic = []
    dates = None
    is_electric_period = None
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
            request.session["is_electric_current"] = is_electric_current = request.GET['is_electric_current']
            request.session["is_electric_delta"]   = is_electric_delta   
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["is_electric_period"]  = is_electric_period  = request.GET['is_electric_period']
            
            res='Электричество'
            
            if (is_electric_monthly=="1"):
                dm='monthly'
            else:
                dm='daily'
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                    isAbon=True                    
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table      
                    params=[u'T0 A+',u'T1 A+',u'T2 A+',u'T3 A+', u'Электричество']
                    data_table_graphic = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
                       
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                    isAbon=False
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table

                    data_table_graphic = common_sql.get_data_table_electric_between_for_obj(obj_title, obj_parent_title,electric_data_start, electric_data_end)

            #*********************************************************************************************************************************************************************
            elif (is_electric_delta == '1') &(bool(is_group_level.search(obj_key))):
                    data_table=common_sql.get_data_table_electric_period_for_group(obj_title,obj_parent_title, electric_data_start, electric_data_end, res)
                    #print data_table
                    request.session["data_table_export"] = data_table
            else:
                pass
        else:
            obj_title = u'Не выбран'
            obj_parent_title = u'Не выбран'
            obj_key = u'Не выбран'
            is_electric_monthly = 0
            is_electric_daily = 0 
            is_electric_current = 0
            is_electric_delta = 0
    
    AllData=[]
    Xcoord=[]    
    if len( data_table_graphic) >0:
        data_table_graphic = normalaize_data_consumption(data_table_graphic)
        Xcoord=makeOneCoords(data_table_graphic,0) #label 
    
        AllData=[{str("data"):makeOneCoords(data_table_graphic,12), str("label"):str("potreblenie T0"), str("backgroundColor"): get_rgba_color(5)},
             {str("data"):makeOneCoords(data_table_graphic,13), str("label"):str("potreblenie T1"),  str("backgroundColor"): get_rgba_color(1)},
             {str("data"):makeOneCoords(data_table_graphic,14), str("label"):str("potreblenie T2"),  str("backgroundColor"): get_rgba_color(10)},
             {str("data"):makeOneCoords(data_table_graphic,15), str("label"):str("potreblenie T3"),  str("backgroundColor"): get_rgba_color(8)}]
    
    data_table = analize_consumption(data_table, electric_data_start, electric_data_end,obj_title,obj_parent_title)
    
    #print(obj_parent_title, obj_title)
    #print('len', len(data_table))
    if (len(data_table) == 1): args['obj_name'] = obj_parent_title
    else: args['obj_name'] = obj_title
    #print AllData
    args['data_table'] = data_table
    args['obj_title'] = obj_title
    args['obj_key'] = obj_key
    args['obj_parent_title'] = obj_parent_title
    args['is_electric_monthly'] = is_electric_monthly
    args['is_electric_daily'] = is_electric_daily
    args['is_electric_current'] = is_electric_current
    args['is_electric_delta'] = is_electric_delta
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['is_electric_period'] = is_electric_period
    args['dates'] = dates
    args['label'] = Xcoord
    args['AllData']=AllData
    return render(request, "data_table/electric/113.html", args)

def makeLimit(iter, coord):
    labels = []
    for i in range(iter):
        labels.append(coord)
    return labels

def extended_info(request):
    args = {}
    obj_title = u'Не выбран'
    obj_parent_title = u'Не выбран'
    electric_data_start = u''
    electric_data_end = u''
    data_table = []
    #rint('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
    if request.is_ajax():        
        if request.method == 'GET':
            request.session["obj_title"]           = obj_title           = request.GET['obj_title']
            request.session["obj_parent_title"]    = obj_parent_title    = request.GET['obj_parent_title']
            request.session["electric_data_start"] = electric_data_start = request.GET['electric_data_start']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["mini"]                = mini                = request.GET['mini']
            request.session["maxi"]                = maxi                = request.GET['maxi']
            print(obj_title, obj_parent_title,electric_data_start, electric_data_end)
            #print(mini, maxi)
            params=['T0 A+','T1 A+','T2 A+','T3 A+', 'Электричество']
            data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
            
            # if  (bool(is_abonent_level.search(obj_key))): # delta for abonents
            #     data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end, params)
                       
            # elif (bool(is_object_level.search(obj_key))): # daily delta for abonents group
            #     data_table = common_sql.get_data_table_electric_between_for_obj(obj_title, obj_parent_title,electric_data_start, electric_data_end)
            
    AllData=[]
    Xcoord=[]
    
    if (len(data_table) >0):
        Xcoord=makeOneCoords(data_table,0) #label
        #print(mini, maxi)
        mini = mini.replace(',','.')
        maxi = maxi.replace(',','.')
        #print(mini, maxi)
        AllData=[{str("data"):makeOneCoords(data_table,12), str("label"):str("delta T0"), str("backgroundColor"): get_rgba_color(5)},
             {str("data"):makeOneCoords(data_table,13), str("label"):str("delta T1"),  str("backgroundColor"): get_rgba_color(1)},
             {str("data"):makeOneCoords(data_table,14), str("label"):str("delta T2"),  str("backgroundColor"): get_rgba_color(10)},
             {str("data"):makeOneCoords(data_table,15), str("label"):str("delta T3"),  str("backgroundColor"): get_rgba_color(8)},
             {str("data"):makeLimit(len(data_table), mini), str("label"):str("min"),  str("backgroundColor"): 'rgba(255, 159, 64, 0.0)', str("borderColor"):'rgba(255, 99, 132, 1)'},
             {str("data"):makeLimit(len(data_table), maxi), str("label"):str("max"),  str("backgroundColor"): 'rgba(255, 159, 64, 0.0)', str("borderColor"):'rgba(255, 99, 132, 1)'}]      
    
    args['data_table'] = data_table
    args['abonent'] = obj_title
    args['object'] = obj_parent_title
    args['electric_data_start'] = electric_data_start
    args['electric_data_end'] = electric_data_end
    args['label'] = Xcoord
    args['AllData']=AllData  

    return render(request, "data_table/extended_info.html", args)

    